from flask import Flask, jsonify, request, render_template_string, session, redirect, url_for
import json, os, threading
from datetime import datetime, timedelta

# Railway Volume: si existe /data usar ese directorio persistente
_BASE_DIR = '/data' if os.path.isdir('/data') else os.environ.get('DATA_DIR', os.path.dirname(os.path.abspath(__file__)))
os.makedirs(_BASE_DIR, exist_ok=True)
CONFIG_FILE   = os.path.join(_BASE_DIR, 'config.json')
DATA_FILE     = os.path.join(_BASE_DIR, 'data.json')
HISTORIAL_FILE = os.path.join(_BASE_DIR, 'historial.json')
VIDEOS_DIR    = os.path.join(_BASE_DIR, 'static_videos')
STATE_FILE    = os.path.join(_BASE_DIR, 'state.json')
os.makedirs(VIDEOS_DIR, exist_ok=True)

CAJA_NOMBRES = {1: 'Abajo', 2: 'Extendido', 3: 'VIP'}

DEFAULT_CONFIG = {
    'password': '1212',
    'pin_manager': '4321',
    'pin_cajaabajo': '1234',
    'pin_cajaextendido': '1234',
    'pin_cajavip': '1234',
    'pin_tarjetas': '1234',
}

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                cfg = json.load(f)
                for k, v in DEFAULT_CONFIG.items():
                    if k not in cfg:
                        cfg[k] = v
                return cfg
        except Exception:
            pass
    return dict(DEFAULT_CONFIG)

def save_config(cfg):
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'jagger_vip_secret_k9x_2024')
app.permanent_session_lifetime = timedelta(hours=12)
lock = threading.Lock()

_db = {'transactions': [], 'tx_id_counter': 0, 'tarjetas': {}, 'tarjetas_conf': [], 'menu': [], 'menu_id_counter': 0}

# Cargar datos guardados al iniciar
if os.path.exists(DATA_FILE):
    try:
        with open(DATA_FILE, 'r', encoding='utf-8') as _f:
            _loaded = json.load(_f)
            _db.update(_loaded)
    except Exception as _e:
        print(f'[data] Error cargando data.json: {_e}')

# Estado compartido entre todos los dispositivos
_state = {
    'hora_fin': '05:30',
    'premio': '',
    'winner_show': False,
    'winner_ts': 0,
    'cartel_show': False,
    'cartel_ts': 0,
    'cartel_data': {},
    'cartel_precios': {'virtual': 0, 'fisico': 0, 'combo': 0},
    'publicidad_activa': False,
    'publicidad_url': '',
    'publicidad_frecuencia': 15,
    'publicidad_mostrar_ts': 0,
    'design': {
        'tema': 'default',
        'colores': {
            '--gold': '#c9a227', '--gold-light': '#e8c84a', '--gold-dim': '#7a6010',
            '--black': '#080808', '--surface': '#111111', '--surface-gold': '#0d0b00',
            '--border': '#2a2a2a', '--text': '#f0ece0', '--text-dim': '#555555',
            '--white': '#ffffff', '--green': '#2ecc71', '--danger': '#a83030',
        },
        'logo': 'RANKING',
        'vip': 'VIP',
        'tagline': 'JAGGER CLUB',
        'tagline_color': '#555555',
        'tagline_glow': '0',
        'tagline_font': "'Rajdhani',sans-serif",
        'winner_msg': '¡EL GANADOR DE LA NOCHE!',
        'winner_sub': '',
        'premio': '',
        'premio_size': '22',
        'hora_fin': '05:30',
        'efecto': 'ninguno',
        'tipo_particula': 'confetti',
        'deco_activa': True,
        'petals_activos': True,
        'falling_gloves': True,
    },
}

# Claves del state que se persisten en disco (excluye timestamps efímeros)
_STATE_PERSIST_KEYS = {
    'hora_fin', 'premio', 'cartel_precios',
    'publicidad_activa', 'publicidad_url', 'publicidad_frecuencia',
    'design',
}

def save_state():
    """Persiste las claves importantes del _state en disco."""
    try:
        data = {k: _state[k] for k in _STATE_PERSIST_KEYS if k in _state}
        with open(STATE_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def load_state():
    """Carga el state persistido del disco al arrancar."""
    if not os.path.exists(STATE_FILE):
        return
    try:
        with open(STATE_FILE, 'r', encoding='utf-8') as f:
            saved = json.load(f)
        for k, v in saved.items():
            if k in _state:
                if k == 'design' and isinstance(_state['design'], dict) and isinstance(v, dict):
                    _state['design'].update(v)
                else:
                    _state[k] = v
    except Exception:
        pass

load_state()  # Cargar al iniciar

def load_data():
    return _db

def save_data(data):
    try:
        with open(DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False)
    except Exception as e:
        print(f'[data] Error guardando: {e}')

def load_historial():
    if os.path.exists(HISTORIAL_FILE):
        try:
            with open(HISTORIAL_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return []

def save_historial(h):
    try:
        with open(HISTORIAL_FILE, 'w', encoding='utf-8') as f:
            json.dump(h, f, ensure_ascii=False)
    except Exception as e:
        print(f'[historial] Error guardando: {e}')

HTML = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Ranking VIP</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=Oswald:wght@400;500;600;700&family=Rajdhani:wght@400;500;600&display=swap');
:root{
  --gold:#c9a227;--gold-light:#e8c84a;--gold-dim:#7a6010;
  --black:#080808;--surface:#111;--border:#2a2a2a;
  --text:#f0ece0;--text-dim:#555;--danger:#a83030;
  --white:#fff;--green:#2ecc71;
  --surface-gold:#0d0b00;
}
*{box-sizing:border-box;margin:0;padding:0;}
body{background:var(--black);color:var(--text);font-family:'Rajdhani',sans-serif;min-height:100vh;}

/* TABS */
.tabs-bar{display:flex;background:#0a0a0a;border-bottom:1px solid #222;position:sticky;top:0;z-index:100;}
.tab-btn{flex:1;padding:13px 6px;text-align:center;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:12px;font-weight:600;letter-spacing:2px;text-transform:uppercase;color:#888;background:none;border:none;border-bottom:2px solid transparent;transition:all .2s;}
.tab-btn:hover{color:#999;}
.tab-btn.active{color:var(--gold);border-bottom:2px solid var(--gold);}
.tab-btn .dot{display:inline-block;width:5px;height:5px;border-radius:50%;background:var(--gold);margin-left:5px;vertical-align:middle;animation:blink 2s infinite;}
@keyframes blink{0%,100%{opacity:1}50%{opacity:.2}}
body.modo-presentacion .tab-btn-caja{display:none;}
body.modo-presentacion .status-bar{display:none;}
body.modo-presentacion .config-panel{display:none;}
body.modo-presentacion .tabs-bar{display:none;}

.screen{display:none;padding:22px 26px 70px;}
.screen.active{display:block;}

/* CONFIG PANEL */
.config-panel{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:14px 18px;margin-bottom:20px;}
.config-title{color:var(--gold-dim);font-size:10px;letter-spacing:2px;text-transform:uppercase;margin-bottom:10px;}
.config-row{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin-bottom:8px;}
.config-row:last-child{margin-bottom:0;}
.config-label{color:#bbb;font-size:13px;white-space:nowrap;}
.config-input{background:#0d0d0d;border:1px solid var(--border);border-radius:6px;color:var(--text);padding:8px 12px;font-family:'Rajdhani',sans-serif;font-size:14px;}
.config-input:focus{outline:none;border-color:var(--gold);}
.config-input.wide{flex:1;min-width:180px;}
.config-input.narrow{width:110px;}
.btn-pres{background:var(--gold);color:#000;border:none;border-radius:6px;padding:9px 20px;font-family:'Rajdhani',sans-serif;font-size:13px;font-weight:700;letter-spacing:1px;cursor:pointer;white-space:nowrap;transition:background .15s;}
.btn-pres:hover{background:var(--gold-light);}
.btn-reset{background:transparent;color:#555;border:1px solid #222;border-radius:6px;padding:9px 16px;font-family:'Rajdhani',sans-serif;font-size:13px;cursor:pointer;white-space:nowrap;transition:all .15s;}
.btn-reset:hover{border-color:var(--danger);color:#cc4444;}

/* PANTALLA HEADER */
.pres-header-wrap{position:relative;margin-bottom:28px;}
.pres-clock{position:absolute;top:0;left:0;text-align:left;}
.pres-clock-hora{font-family:'Oswald',sans-serif;font-size:13px;color:var(--text-dim);letter-spacing:1px;text-transform:uppercase;}
.pres-clock-time{font-family:'Oswald',sans-serif;font-size:32px;color:#ffffff;font-weight:700;line-height:1.1;}
.pres-clock-fin{font-size:11px;color:var(--gold-dim);letter-spacing:1px;margin-top:6px;text-transform:uppercase;}
.pres-clock-fin-val{font-family:'Oswald',sans-serif;font-size:26px;color:var(--gold);font-weight:700;line-height:1.1;display:block;}
.pres-header{text-align:center;padding-top:8px;}
.pres-logo{font-family:'Oswald',sans-serif;font-size:52px;font-weight:700;color:var(--white);letter-spacing:10px;text-transform:uppercase;display:block;width:100%;}
.pres-logo .vip{color:var(--gold);}
.pres-logo .club{font-size:26px;font-weight:600;color:#888;letter-spacing:10px;display:block;margin-top:2px;text-transform:uppercase;width:100%;}
.pres-line{height:1px;background:linear-gradient(to right,transparent,var(--gold),transparent);margin:12px auto;max-width:100%;}
.live-badge{display:inline-flex;align-items:center;gap:6px;border:1px solid #2a2a2a;border-radius:20px;padding:4px 14px;font-size:11px;color:#777;letter-spacing:1px;margin-top:6px;}
.live-dot{width:6px;height:6px;border-radius:50%;background:#3a9a5a;animation:blink 1.5s infinite;}

/* RANKING */
.ranking-wrap{max-width:100%;margin:0;padding:0 10px;}
.rank-header{display:grid;grid-template-columns:100px 1fr 130px 180px;background:#0d0d0d;border:1px solid #2a2a2a;border-radius:8px 8px 0 0;padding:12px 30px;margin-bottom:3px;}
.rank-header span{font-family:'Oswald',sans-serif;font-size:13px;font-weight:500;letter-spacing:2px;text-transform:uppercase;color:var(--gold-dim);}
.rank-header .col-r{text-align:right;}
.rank-rows{display:flex;flex-direction:column;gap:4px;}
.rank-row{display:grid;grid-template-columns:100px 1fr 130px 180px;align-items:center;background:var(--surface);border:1px solid #1e1e1e;border-radius:6px;padding:20px 30px;transition:border-color .3s, background .4s, box-shadow .4s;}
.rank-row.rank-1{background:var(--surface-gold);border-color:var(--gold-dim);}
/* Nueva tarjeta que entra */
.rank-row.nueva{animation:entradaFila .55s cubic-bezier(.22,1,.36,1) both;}
@keyframes entradaFila{from{opacity:0;transform:translateX(-32px)}to{opacity:1;transform:none}}
/* Highlight flash al subir al #1 */
@keyframes crownGlow{
  0%  {box-shadow:0 0 0px rgba(201,162,39,0);}
  40% {box-shadow:0 0 32px rgba(201,162,39,.55);}
  100%{box-shadow:none;}
}
.rank-row.rank-1.ascendio{animation:crownGlow .9s ease both;}
/* Guante que golpea hacia la derecha */
@keyframes gloveKnockout{
  0%  {left:-60px;opacity:1;transform:scale(1.2) rotate(-15deg);}
  35% {left:60%;opacity:1;transform:scale(1.3) rotate(-10deg);}
  50% {left:70%;opacity:0.8;transform:scale(1.1) rotate(-5deg);}
  70% {left:110%;opacity:0;transform:scale(0.8) rotate(0deg);}
  100%{left:110%;opacity:0;}
}
/* Punch hit */
@keyframes punchHit{
  0%  {transform:translateX(-60px) translateY(22px) rotate(-5deg);filter:brightness(2);}
  30% {transform:translateX(8px) translateY(0) rotate(1deg);filter:brightness(1.5);}
  60% {transform:translateX(-3px) rotate(-0.5deg);}
  100%{transform:none;filter:none;}
}
.col-puesto{font-family:'Oswald',sans-serif;font-size:28px;font-weight:700;color:#444;}
.rank-row.rank-1 .col-puesto{color:var(--gold);font-size:34px;}
.rank-row.rank-2 .col-puesto{color:#aaa;}
.rank-row.rank-3 .col-puesto{color:#8a6a40;}
.col-nombre{font-family:'Oswald',sans-serif;font-size:30px;font-weight:600;color:var(--white);}
.rank-row.rank-1 .col-nombre{font-size:36px;}
.col-mesa{font-family:'Oswald',sans-serif;font-size:24px;font-weight:700;color:#e8e8e8;letter-spacing:1px;}
.rank-row.rank-1 .col-mesa{font-size:28px;color:#ffffff;}
.col-total{font-family:'Oswald',sans-serif;font-size:30px;font-weight:700;color:var(--gold);text-align:right;}
.rank-row.rank-1 .col-total{font-size:38px;}
.miles-lbl{font-size:0.45em;opacity:0.55;letter-spacing:2px;margin-left:5px;vertical-align:middle;font-weight:600;}

.btn-exit-float{display:none;position:fixed;top:12px;left:16px;z-index:9999;background:transparent;color:#1c1c1c;border:none;font-size:22px;font-family:'Oswald',sans-serif;font-weight:300;cursor:pointer;padding:4px 8px;transition:color .3s;line-height:1;}
.btn-exit-float:hover{color:#555;}
body.modo-presentacion .btn-exit-float{display:block;}
.empty-msg{text-align:center;color:#222;font-size:15px;padding:70px 20px;letter-spacing:2px;font-family:'Oswald',sans-serif;}

/* PREMIO */
.premio-wrap{text-align:center;margin-top:36px;padding-bottom:20px;}
.premio-box{display:inline-block;background:var(--gold);color:#000;font-family:'Rajdhani',sans-serif;font-size:22px;font-weight:700;letter-spacing:.5px;padding:14px 44px;border-radius:6px;border:none;}
.premio-box:empty{display:none;}
@keyframes bottleBounce{
  0%,100%{transform:translateY(0) rotate(-5deg) scale(1);}
  20%{transform:translateY(-18px) rotate(5deg) scale(1.05);}
  40%{transform:translateY(-6px) rotate(-3deg) scale(0.98);}
  60%{transform:translateY(-14px) rotate(4deg) scale(1.03);}
  80%{transform:translateY(-4px) rotate(-2deg) scale(0.99);}
}

/* CAJAS */
.caja-header-row{display:flex;align-items:center;gap:12px;margin-bottom:20px;}
.caja-badge{background:var(--gold);color:#000;border-radius:6px;padding:5px 14px;font-family:'Oswald',sans-serif;font-size:13px;letter-spacing:2px;font-weight:600;}
.caja-title{font-family:'Oswald',sans-serif;font-size:20px;color:var(--white);letter-spacing:1px;}
.modo-tabs{display:flex;gap:8px;margin-bottom:16px;}
.modo-tab{flex:1;padding:10px;text-align:center;background:var(--surface);border:1px solid var(--border);border-radius:8px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;font-weight:600;letter-spacing:1px;color:var(--text-dim);transition:all .2s;}
.modo-tab.active{background:#0f0e05;border-color:var(--gold-dim);color:var(--gold);}
.modo-content{display:none;}
.modo-content.active{display:block;}
.scan-hint{display:flex;align-items:center;gap:10px;background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:14px 18px;margin-bottom:16px;color:var(--text-dim);font-size:13px;letter-spacing:1px;}
.scan-hint.esperando{border-color:var(--gold-dim);color:var(--gold);}
.scan-icon{font-size:20px;opacity:.4;}
.scan-hint.esperando .scan-icon{opacity:1;animation:blink .6s infinite;}
.tarjeta-card{background:#0a1200;border:1px solid #1a3a00;border-radius:10px;padding:16px 20px;margin-bottom:16px;display:none;}
.tarjeta-card.visible{display:block;}
.tarjeta-card.sin-saldo{background:#120000;border-color:#3a0000;}
.tarjeta-top{display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;}
.tarjeta-mesa-label{color:#3a6a00;font-size:11px;letter-spacing:2px;text-transform:uppercase;}
.tarjeta-card.sin-saldo .tarjeta-mesa-label{color:#6a2000;}
.tarjeta-mesa-num{font-family:'Oswald',sans-serif;font-size:32px;color:#ffffff;font-weight:700;}
.tarjeta-card.sin-saldo .tarjeta-mesa-num{color:#ff8888;}
.tarjeta-saldo-wrap{text-align:right;}
.tarjeta-saldo-label{color:#3a6a00;font-size:10px;letter-spacing:1px;text-transform:uppercase;}
.tarjeta-saldo{font-family:'Oswald',sans-serif;font-size:26px;color:var(--green);font-weight:700;}
.tarjeta-card.sin-saldo .tarjeta-saldo{color:var(--danger);}
.tarjeta-bar-wrap{height:4px;background:#1a1a1a;border-radius:2px;}
.tarjeta-bar{height:4px;background:var(--green);border-radius:2px;transition:width .5s ease;}
.tarjeta-card.sin-saldo .tarjeta-bar{background:var(--danger);}
.tarjeta-nombre{margin-top:10px;font-size:13px;color:#3a5a00;}
.tarjeta-nombre span{color:#aaddaa;font-weight:600;}
.form-card{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:20px;margin-bottom:18px;}
.field-label{color:#888;font-size:11px;letter-spacing:2px;text-transform:uppercase;margin-bottom:6px;display:block;font-weight:600;}
.field-input{width:100%;background:#0d0d0d;border:1px solid var(--border);border-radius:7px;color:var(--text);padding:11px 13px;font-family:'Rajdhani',sans-serif;font-size:15px;font-weight:500;transition:border-color .2s;margin-bottom:14px;}
.field-input:focus{outline:none;border-color:var(--gold);}
.field-input::placeholder{color:#333;}
.field-input.amount-input{font-family:'Oswald',sans-serif;font-size:28px;font-weight:700;color:var(--gold);padding:13px 16px;letter-spacing:1px;}
.field-input.amount-input::placeholder{color:#222;font-size:20px;}
.hint-miles{color:#666;font-size:13px;margin-top:-10px;margin-bottom:10px;letter-spacing:1px;}
.hint-miles.ok{color:#3a6a00;font-family:'Oswald',sans-serif;font-size:15px;font-weight:600;}
.btn-row{display:flex;gap:10px;}
.btn-add{background:var(--gold);color:#000;border:none;border-radius:7px;padding:12px 22px;font-family:'Rajdhani',sans-serif;font-size:15px;font-weight:700;letter-spacing:1px;text-transform:uppercase;cursor:pointer;flex:1;transition:background .15s,transform .1s;}
.btn-add:hover{background:var(--gold-light);}
.btn-add:active{transform:scale(.97);}
.btn-add:disabled{background:#2a2a2a;color:#555;cursor:not-allowed;}
.section-label{color:#888;font-size:11px;letter-spacing:2px;text-transform:uppercase;margin-bottom:10px;}
.tx-list{display:flex;flex-direction:column;gap:6px;max-height:240px;overflow-y:auto;}
.tx-list::-webkit-scrollbar{width:3px;}
.tx-list::-webkit-scrollbar-thumb{background:#2a2a2a;border-radius:3px;}
.tx-item{display:flex;justify-content:space-between;align-items:center;background:var(--surface);border:1px solid #1a1a1a;border-radius:7px;padding:10px 14px;}
.tx-info{flex:1;}
.tx-name{font-size:15px;font-weight:700;color:#f0ece0;}
.tx-meta{font-size:12px;color:#888;margin-top:1px;}
.tx-right{display:flex;align-items:center;gap:10px;}
.tx-amount{font-family:'Oswald',sans-serif;font-size:18px;color:var(--gold);font-weight:700;}
.btn-del{background:none;border:1px solid #2a1a1a;color:#4a2a2a;border-radius:5px;padding:4px 8px;font-size:12px;cursor:pointer;transition:all .15s;}
.btn-del:hover{border-color:var(--danger);color:#cc4444;}
.btn-edit{background:none;border:1px solid #1a2a1a;color:#2a4a2a;border-radius:5px;padding:4px 8px;font-size:12px;cursor:pointer;transition:all .15s;margin-right:4px;}
.btn-edit:hover{border-color:#3a6a10;color:#6a9a30;}
.caja-total-bar{background:var(--surface);border:1px solid #1e1a00;border-radius:8px;padding:13px 18px;display:flex;justify-content:space-between;align-items:center;margin-top:14px;}
.caja-total-label{color:#aaa;font-size:11px;letter-spacing:2px;text-transform:uppercase;}
.caja-total-val{font-family:'Oswald',sans-serif;font-size:24px;color:var(--gold);}
.no-tx{color:#555;font-size:13px;padding:14px 0;letter-spacing:1px;}

/* CONFIGURACION TARJETAS */
.conf-header{margin-bottom:24px;}
.conf-title{font-family:'Oswald',sans-serif;font-size:22px;color:var(--white);letter-spacing:1px;margin-bottom:6px;}
.conf-sub{color:#aaa;font-size:13px;line-height:1.7;}
.tarjetas-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(240px,1fr));gap:10px;margin-bottom:24px;}
.tarjeta-conf{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:14px 16px;}
.tarjeta-conf.configurada{border-color:#2a3a00;background:#0a0f00;}
.tc-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;}
.tc-num{font-family:'Oswald',sans-serif;font-size:16px;color:var(--gold-dim);font-weight:600;}
.tarjeta-conf.configurada .tc-num{color:var(--gold);}
.tc-status{font-size:11px;letter-spacing:1px;color:#777;text-transform:uppercase;}
.tarjeta-conf.configurada .tc-status{color:#3a6a00;}
.tc-btns{display:flex;gap:6px;align-items:center;}
.tc-scan-btn{background:#1a1a1a;color:#555;border:1px solid #2a2a2a;border-radius:5px;padding:4px 10px;font-size:11px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-weight:600;letter-spacing:1px;transition:all .15s;}
.tc-scan-btn:hover{border-color:var(--gold-dim);color:var(--gold);}
.tc-scan-btn.activo{border-color:var(--gold);color:var(--gold);animation:blink .6s infinite;}
.tc-confirm-btn{background:var(--gold);color:#000;border:none;border-radius:5px;padding:4px 10px;font-size:11px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-weight:700;letter-spacing:1px;transition:background .15s;}
.tc-confirm-btn:hover{background:var(--gold-light);}
.tc-clear-btn{background:none;color:#333;border:1px solid #1e1e1e;border-radius:5px;padding:4px 8px;font-size:11px;cursor:pointer;transition:all .15s;}
.tc-clear-btn:hover{border-color:var(--danger);color:#cc4444;}
.tc-field{display:flex;flex-direction:column;gap:4px;margin-bottom:8px;}
.tc-label{font-size:11px;letter-spacing:1px;color:#aaa;text-transform:uppercase;font-weight:600;}
.tc-input{background:#0d0d0d;border:1px solid var(--border);border-radius:5px;color:var(--text);padding:7px 10px;font-family:'Rajdhani',sans-serif;font-size:14px;font-weight:500;width:100%;}
.tc-input:focus{outline:none;border-color:var(--gold);}
.tc-input::placeholder{color:#333;}
.tc-code{font-size:11px;color:#666;margin-top:4px;font-family:monospace;}
.tarjeta-conf.configurada .tc-code{color:#2a4a00;}
.tc-saldo-bar{margin-top:8px;}
.tc-saldo-info{display:flex;justify-content:space-between;font-size:11px;margin-bottom:4px;}
.tc-saldo-used{color:var(--danger);}
.tc-saldo-left{color:var(--green);}
.saldo-bajo-warn{background:#1a0800;border:1px solid #6a2a00;border-radius:7px;padding:8px 12px;font-size:12px;color:#ff7733;letter-spacing:1px;margin-top:6px;display:flex;align-items:center;gap:7px;animation:warnPulse 1.5s ease-in-out infinite alternate;}
@keyframes warnPulse{0%{border-color:#6a2a00;color:#ff7733;}100%{border-color:#ff7733;color:#ffaa66;}}
.tc-bar-wrap{height:3px;background:#1a1a1a;border-radius:2px;}
.tc-bar-fill{height:3px;background:var(--green);border-radius:2px;transition:width .5s;}
.conf-actions{display:flex;gap:10px;flex-wrap:wrap;}
.btn-guardar-conf{background:var(--gold);color:#000;border:none;border-radius:7px;padding:12px 30px;font-family:'Rajdhani',sans-serif;font-size:15px;font-weight:700;letter-spacing:1px;cursor:pointer;transition:background .15s;}
.btn-guardar-conf:hover{background:var(--gold-light);}
.btn-limpiar-conf{background:transparent;color:#555;border:1px solid #222;border-radius:7px;padding:12px 20px;font-family:'Rajdhani',sans-serif;font-size:14px;cursor:pointer;transition:all .15s;}
.btn-limpiar-conf:hover{border-color:var(--danger);color:#cc4444;}

/* STATUS BAR */

.status-ok{color:#3a6a3a;}
.status-err{color:var(--danger);}
.toast{position:fixed;bottom:40px;left:50%;transform:translateX(-50%);background:#1a2a00;border:1px solid #3a6a00;color:var(--green);padding:10px 24px;border-radius:8px;font-size:14px;font-weight:600;z-index:9999;opacity:0;transition:opacity .3s;pointer-events:none;white-space:nowrap;}
.toast.show{opacity:1;}
.toast.error{background:#2a0000;border-color:#6a0000;color:#ff6666;}

/* ══════════════════════════════════════════
   STATS
══════════════════════════════════════════ */
.stats-header{margin-bottom:24px;}
.stats-title{font-family:'Oswald',sans-serif;font-size:22px;color:var(--white);letter-spacing:1px;margin-bottom:4px;}
.stats-sub{font-size:13px;color:var(--text-dim);letter-spacing:1px;}
.kpi-row{display:grid;grid-template-columns:repeat(auto-fill,minmax(180px,1fr));gap:12px;margin-bottom:24px;}
.kpi-card{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:18px 20px;}
.kpi-label{font-size:10px;color:var(--text-dim);letter-spacing:2px;text-transform:uppercase;margin-bottom:8px;}
.kpi-val{font-family:'Oswald',sans-serif;font-size:28px;font-weight:700;color:var(--gold);}
.charts-row{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:24px;}
@media(max-width:700px){.charts-row{grid-template-columns:1fr;}}
.chart-box{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:18px 20px;}
.chart-title{font-family:'Oswald',sans-serif;font-size:14px;color:var(--gold-dim);letter-spacing:2px;text-transform:uppercase;margin-bottom:14px;}
.chart-legend{display:flex;flex-wrap:wrap;gap:10px;margin-top:12px;}
.legend-item{display:flex;align-items:center;gap:6px;font-size:12px;color:var(--text-dim);}
.legend-dot{width:10px;height:10px;border-radius:50%;flex-shrink:0;}
.stats-section-title{font-family:'Oswald',sans-serif;font-size:14px;color:var(--gold-dim);letter-spacing:2px;text-transform:uppercase;margin:0 0 12px;padding-bottom:6px;border-bottom:1px solid #1e1e1e;}
/* Top clientes */
.top-list{display:flex;flex-direction:column;gap:8px;margin-bottom:24px;}
.top-item{display:flex;align-items:center;gap:12px;background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:12px 16px;}
.top-pos{font-family:'Oswald',sans-serif;font-size:20px;font-weight:700;color:#333;width:30px;text-align:center;flex-shrink:0;}
.top-item.pos-1 .top-pos{color:var(--gold);}
.top-item.pos-2 .top-pos{color:#aaa;}
.top-item.pos-3 .top-pos{color:#8a6a40;}
.top-name{font-family:'Oswald',sans-serif;font-size:20px;color:var(--white);flex:1;font-weight:600;}
.top-mesa{font-size:12px;color:#888;letter-spacing:1px;}
.top-bar-wrap{flex:2;height:6px;background:#1a1a1a;border-radius:3px;}
.top-bar-fill{height:6px;background:var(--gold);border-radius:3px;transition:width .6s ease;}
.top-amount{font-family:'Oswald',sans-serif;font-size:18px;color:var(--gold);text-align:right;min-width:90px;}
/* Detalle cajas */
.cajas-detail-row{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:24px;}
@media(max-width:600px){.cajas-detail-row{grid-template-columns:1fr;}}
.caja-stat-card{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:16px 18px;}
.caja-stat-badge{background:var(--gold);color:#000;font-family:'Oswald',sans-serif;font-size:11px;letter-spacing:2px;font-weight:700;padding:3px 10px;border-radius:4px;display:inline-block;margin-bottom:10px;}
.caja-stat-total{font-family:'Oswald',sans-serif;font-size:26px;color:var(--gold);font-weight:700;margin-bottom:4px;}
.caja-stat-ops{font-size:12px;color:var(--text-dim);}
.caja-stat-list{margin-top:10px;display:flex;flex-direction:column;gap:4px;max-height:140px;overflow-y:auto;}
.caja-stat-item{display:flex;justify-content:space-between;font-size:13px;color:var(--text-dim);padding:4px 0;border-bottom:1px solid #1a1a1a;}
.caja-stat-item:last-child{border-bottom:none;}
.caja-stat-item span:last-child{color:var(--gold);}

/* ══════════════════════════════════════════
   PANEL DE PERSONALIZACION
══════════════════════════════════════════ */
.custom-section{margin-bottom:28px;}
.custom-section-title{font-family:'Oswald',sans-serif;font-size:16px;color:var(--gold);letter-spacing:2px;text-transform:uppercase;margin-bottom:14px;padding-bottom:6px;border-bottom:1px solid #1e1e1e;}
.color-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:12px;margin-bottom:10px;}
.color-item{display:flex;align-items:center;gap:10px;background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:10px 14px;}
.color-swatch{width:36px;height:36px;border-radius:6px;border:2px solid #333;cursor:pointer;flex-shrink:0;position:relative;overflow:hidden;}
.color-swatch input[type=color]{position:absolute;inset:-4px;width:calc(100%+8px);height:calc(100%+8px);border:none;cursor:pointer;opacity:0;}
.color-label{font-size:12px;color:var(--text-dim);letter-spacing:1px;flex:1;}
.color-hex{font-size:11px;color:#444;font-family:monospace;}
.custom-text-row{display:flex;gap:12px;align-items:center;background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:12px 16px;margin-bottom:10px;}
.custom-text-label{font-size:12px;color:var(--text-dim);letter-spacing:1px;white-space:nowrap;min-width:130px;}
.custom-text-input{flex:1;background:#0d0d0d;border:1px solid var(--border);border-radius:6px;color:var(--text);padding:9px 12px;font-family:'Rajdhani',sans-serif;font-size:14px;}
.custom-text-input:focus{outline:none;border-color:var(--gold);}
.btn-custom-save{background:var(--gold);color:#000;border:none;border-radius:7px;padding:12px 32px;font-family:'Rajdhani',sans-serif;font-size:15px;font-weight:700;letter-spacing:1px;cursor:pointer;transition:background .15s;margin-right:10px;}
.btn-custom-save:hover{background:var(--gold-light);}
.btn-custom-reset{background:transparent;color:#555;border:1px solid #222;border-radius:7px;padding:12px 20px;font-family:'Rajdhani',sans-serif;font-size:14px;cursor:pointer;transition:all .15s;}
.btn-custom-reset:hover{border-color:var(--danger);color:#cc4444;}
.preview-bar{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:16px 20px;margin-bottom:20px;text-align:center;}
.preview-label{font-size:10px;color:var(--text-dim);letter-spacing:2px;text-transform:uppercase;margin-bottom:8px;}
.preview-logo{font-family:'Oswald',sans-serif;font-size:32px;font-weight:700;letter-spacing:6px;color:var(--white);}
.preview-logo .vip-prev{color:var(--gold);}
.preview-logo .club-prev{font-size:13px;font-weight:400;color:#5a5a5a;letter-spacing:5px;display:block;margin-top:2px;}

/* ══════════════════════════════════════════
   ANIMACION GANADOR
══════════════════════════════════════════ */
#winner-overlay{
  display:none;
  position:fixed;inset:0;z-index:10000;
  background:rgba(0,0,0,0.97);
  flex-direction:column;align-items:center;justify-content:center;
  overflow:hidden;
}
#winner-overlay.show{display:flex;}

/* Confetti particles */
.confetti-wrap{position:absolute;inset:0;pointer-events:none;overflow:hidden;}
.confetti-piece{
  position:absolute;top:-20px;
  width:10px;height:16px;
  border-radius:2px;
  animation:confettiFall linear infinite;
}
@keyframes confettiFall{
  0%  {transform:translateY(-20px) rotate(0deg);opacity:1;}
  100%{transform:translateY(110vh) rotate(720deg);opacity:.2;}
}

/* Destellos radiales */
.winner-rays{
  position:absolute;inset:0;
  background:conic-gradient(from 0deg, transparent 0deg, rgba(201,162,39,0.04) 10deg, transparent 20deg,
    transparent 40deg, rgba(201,162,39,0.04) 50deg, transparent 60deg,
    transparent 80deg, rgba(201,162,39,0.04) 90deg, transparent 100deg,
    transparent 120deg, rgba(201,162,39,0.04) 130deg, transparent 140deg,
    transparent 160deg, rgba(201,162,39,0.04) 170deg, transparent 180deg,
    transparent 200deg, rgba(201,162,39,0.04) 210deg, transparent 220deg,
    transparent 240deg, rgba(201,162,39,0.04) 250deg, transparent 260deg,
    transparent 280deg, rgba(201,162,39,0.04) 290deg, transparent 300deg,
    transparent 320deg, rgba(201,162,39,0.04) 330deg, transparent 340deg,
    transparent 360deg);
  animation:raysRotate 12s linear infinite;
  pointer-events:none;
}
@keyframes raysRotate{from{transform:rotate(0deg)}to{transform:rotate(360deg)}}

.winner-content{
  position:relative;z-index:2;
  text-align:center;
  animation:winnerEntrada 1s cubic-bezier(.22,1,.36,1) both;
}
@keyframes winnerEntrada{
  0%  {opacity:0;transform:scale(.5) translateY(60px);}
  60% {transform:scale(1.05) translateY(-10px);}
  100%{opacity:1;transform:scale(1) translateY(0);}
}

.winner-corona{font-size:80px;animation:coronaPulse 1.5s ease-in-out infinite alternate;display:block;margin-bottom:10px;}
@keyframes coronaPulse{from{transform:scale(1) rotate(-5deg)}to{transform:scale(1.15) rotate(5deg)}}


.winner-titulo{
  font-family:'Oswald',sans-serif;
  font-size:22px;font-weight:500;
  letter-spacing:10px;text-transform:uppercase;
  color:var(--gold-dim);margin-bottom:6px;
}
.winner-nombre{
  font-family:'Oswald',sans-serif;
  font-size:clamp(52px,8vw,110px);
  font-weight:700;
  color:#fff;
  letter-spacing:4px;
  text-transform:uppercase;
  line-height:1;
  margin-bottom:10px;
  text-shadow:0 0 60px rgba(201,162,39,.5);
  animation:nombreGlow 2s ease-in-out infinite alternate;
}
@keyframes nombreGlow{
  from{text-shadow:0 0 40px rgba(201,162,39,.3);}
  to  {text-shadow:0 0 80px rgba(201,162,39,.8), 0 0 120px rgba(201,162,39,.3);}
}
.winner-line{
  height:2px;
  background:linear-gradient(to right,transparent,var(--gold),transparent);
  margin:16px auto;width:80%;
  animation:lineExpand 1s ease both;animation-delay:.5s;
  transform-origin:center;
}
@keyframes lineExpand{from{transform:scaleX(0)}to{transform:scaleX(1)}}

.winner-info-row{
  display:flex;gap:60px;justify-content:center;align-items:center;
  margin:20px 0;flex-wrap:wrap;
}
.winner-info-block{text-align:center;}
.winner-info-label{
  font-size:12px;letter-spacing:3px;text-transform:uppercase;
  color:var(--gold-dim);margin-bottom:4px;
}
.winner-info-val{
  font-family:'Oswald',sans-serif;
  font-size:clamp(28px,4vw,52px);
  font-weight:700;color:var(--gold);
}

.winner-mensaje{
  font-family:'Rajdhani',sans-serif;
  font-size:clamp(20px,3vw,36px);
  font-weight:600;
  color:#fff;
  margin-top:24px;
  padding:16px 40px;
  border:2px solid var(--gold);
  border-radius:8px;
  background:rgba(201,162,39,0.08);
  max-width:90vw;
  line-height:1.3;
  animation:mensajeAppear 1s ease both;animation-delay:.8s;
  opacity:0;
}
@keyframes mensajeAppear{from{opacity:0;transform:translateY(20px)}to{opacity:1;transform:none}}

.winner-close{
  position:fixed;top:20px;right:28px;z-index:10001;
  background:transparent;color:#333;border:none;
  font-size:28px;cursor:pointer;font-family:'Oswald',sans-serif;
  transition:color .2s;line-height:1;
}
.winner-close:hover{color:#888;}

/* Boton manual ganador */
.btn-show-winner{
  background:linear-gradient(135deg,#c9a227,#e8c84a);
  color:#000;border:none;border-radius:7px;
  padding:10px 24px;
  font-family:'Rajdhani',sans-serif;font-size:14px;font-weight:700;
  letter-spacing:1px;cursor:pointer;
  transition:all .2s;
  box-shadow:0 4px 20px rgba(201,162,39,.3);
}
.btn-show-winner:hover{transform:translateY(-1px);box-shadow:0 6px 28px rgba(201,162,39,.4);}

@keyframes shimmerGold{0%{background-position:200% center}100%{background-position:-200% center}}
@keyframes shimmerRed{0%{background-position:200% center}100%{background-position:-200% center}}
@keyframes emojiGlow{from{filter:drop-shadow(0 0 20px rgba(201,162,39,0.4))}to{filter:drop-shadow(0 0 50px rgba(201,162,39,0.9))}}

/* ══ TEMA JAGGER 12 AÑOS (B&W + PODIO DORADO/PLATEADO/BRONCE) ══ */
body.tema-jagger12 .rank-row{border-color:#1e1e1e;}
body.tema-jagger12 .rank-row.rank-1{background:#1a1200;border-color:#c9a227;}
body.tema-jagger12 .rank-row.rank-2{background:#0e0e0e;border-color:#777;}
body.tema-jagger12 .rank-row.rank-3{background:#0e0800;border-color:#7a4a20;}
body.tema-jagger12 .rank-row.rank-1 .col-puesto{color:#c9a227;font-size:34px;}
body.tema-jagger12 .rank-row.rank-2 .col-puesto{color:#aaaaaa;}
body.tema-jagger12 .rank-row.rank-3 .col-puesto{color:#cd7f32;}
body.tema-jagger12 .rank-row.rank-1 .col-total{color:#e8c84a;}
body.tema-jagger12 .rank-row.rank-2 .col-total{color:#cccccc;}
body.tema-jagger12 .rank-row.rank-3 .col-total{color:#cd7f32;}
body.tema-jagger12 .col-total{color:#ccc;}
body.tema-jagger12 .pres-line{background:linear-gradient(to right,transparent,#c9a227,transparent);}
body.tema-jagger12 .live-dot{background:#c9a227;}
body.tema-jagger12 .col-nombre{color:#fff !important;}
body.tema-jagger12 .col-mesa{color:#ddd !important;}



/* ══ TEMA A TOUCH OF PINK ══ */
body.tema-touchofpink .rank-header{background:#2d0022;border-color:#6a2050;}
body.tema-touchofpink .rank-header span{color:#f472b6;text-shadow:0 0 8px rgba(244,114,182,0.4);}
body.tema-touchofpink .rank-row{border-color:#8a3070;background:#3d002c;}
body.tema-touchofpink .rank-row.rank-1{background:#5a0042;border-color:#f472b6;box-shadow:0 0 20px rgba(244,114,182,.2);}
body.tema-touchofpink .rank-row.rank-1 .col-puesto{color:#f472b6;font-size:34px;}
body.tema-touchofpink .rank-row.rank-2 .col-puesto{color:#ffffff;}
body.tema-touchofpink .rank-row.rank-3 .col-puesto{color:#fbb6ce;}
body.tema-touchofpink .rank-row.rank-1 .col-total{color:#fce7f3;}
body.tema-touchofpink .col-total{color:#fbb6ce;}
body.tema-touchofpink .col-nombre{color:#ffffff !important;}
body.tema-touchofpink .col-mesa{color:#fce7f3 !important;}
body.tema-touchofpink .col-puesto{color:#eeaad8;}
body.tema-touchofpink .pres-line{background:linear-gradient(to right,transparent,#f472b6,transparent);}
body.tema-touchofpink .live-dot{background:#f472b6;}
body.tema-touchofpink .live-badge{border-color:#8a3070;color:#f472b6;}
body.tema-touchofpink .rank-row.rank-1.ascendio{animation:crownGlow .9s ease both;}
/* A Touch of Pink — modo rose medio */
body.tema-touchofpink.pink-claro{--black:#3a0028;--surface:#580040;--border:#a04080;--gold:#f472b6;--gold-light:#fbb6ce;--gold-dim:#e896cc;--text:#ffe8f5;--text-dim:#ddaacc;--white:#ffffff;}
body.tema-touchofpink.pink-claro .rank-header{background:#4a0035;border-color:#a04080;}
body.tema-touchofpink.pink-claro .rank-header span{color:#f472b6;text-shadow:0 0 8px rgba(244,114,182,0.4);}
body.tema-touchofpink.pink-claro .rank-row{background:#580040;border-color:#a04080;}
body.tema-touchofpink.pink-claro .rank-row.rank-1{background:#7a0058;border-color:#f472b6;box-shadow:0 0 24px rgba(244,114,182,.3);}
body.tema-touchofpink.pink-claro .rank-row.rank-1 .col-puesto{color:#f472b6;}
body.tema-touchofpink.pink-claro .rank-row.rank-2 .col-puesto{color:#ffffff;}
body.tema-touchofpink.pink-claro .rank-row.rank-3 .col-puesto{color:#fbb6ce;}
body.tema-touchofpink.pink-claro .rank-row.rank-1 .col-total{color:#fce7f3;}
body.tema-touchofpink.pink-claro .col-total{color:#fbb6ce;}
body.tema-touchofpink.pink-claro .col-nombre{color:#ffffff !important;}
body.tema-touchofpink.pink-claro .col-mesa{color:#fce7f3 !important;}
body.tema-touchofpink.pink-claro .col-puesto{color:#eeaad8;}

/* ══ CARTEL OVERLAY ══ */
#cartel-overlay.show{display:flex !important;}

</style>
</head>
<body>
<div id="tema-overlay" style="position:fixed;inset:0;pointer-events:none;z-index:2;overflow:hidden;transition:opacity .5s;opacity:0;"></div>
<div id="efectos-overlay" style="position:fixed;inset:0;pointer-events:none;z-index:3;overflow:hidden;"></div>
<div id="copos-wrap" style="position:fixed;inset:0;pointer-events:none;z-index:1;overflow:hidden;display:none;"></div>
<div id="petalos-wrap" style="position:fixed;inset:0;pointer-events:none;z-index:1;overflow:hidden;display:none;"></div>

<div class="tabs-bar">
  <button class="tab-btn" onclick="showTab('pantalla')" id="tbtn-pantalla">📺 Pantalla</button>
  <button class="tab-btn active" onclick="showTab('menu-mgr')" id="tbtn-menu-mgr">🍽 Menú</button>
  <button class="tab-btn" onclick="showTab('publicidad')" id="tbtn-publicidad">📺 Publicidad</button>
  <button class="tab-btn" onclick="showTab('stats')" id="tbtn-stats">📊 Stats</button>
  <button class="tab-btn" onclick="showTab('pines')" id="tbtn-pines">🔑 PINs</button>
  <button class="tab-btn" onclick="showTab('diseno')" id="tbtn-diseno">🎨 Diseño</button>
  <a href="/" style="display:flex;align-items:center;padding:0 14px;color:#333;font-family:'Rajdhani',sans-serif;font-size:11px;letter-spacing:1px;text-decoration:none;transition:color .2s;" onmouseover="this.style.color='#888'" onmouseout="this.style.color='#333'">← Hub</a>
</div>

<!-- PANTALLA: solo controles, sin preview de ranking -->
<div id="tab-pantalla" class="screen">
  <!-- Tutorial desplegable -->
  <div style="margin-bottom:14px;">
    <button onclick="(function(b,p){b._open=!b._open;p.style.display=b._open?'block':'none';b.querySelector('.tut-arrow').textContent=b._open?'▲':'▼';})(this,document.getElementById('tut-pantalla'))" style="width:100%;display:flex;justify-content:space-between;align-items:center;background:#0a0b00;border:1px solid #2a2a00;border-radius:8px;padding:10px 16px;font-family:'Rajdhani',sans-serif;font-size:13px;font-weight:700;color:#c9a227;letter-spacing:1px;cursor:pointer;text-align:left;">
      <span>📖 Cómo usar — Control de Pantalla</span><span class="tut-arrow" style="font-size:12px;color:#556;">▼</span>
    </button>
    <div id="tut-pantalla" style="display:none;background:#080800;border:1px solid #1a1a00;border-radius:0 0 8px 8px;padding:14px 18px;font-family:'Rajdhani',sans-serif;font-size:13px;color:#aaa;line-height:1.8;">
      <div style="display:flex;flex-direction:column;gap:8px;">
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">①</span><span><strong style="color:#fff;">Hora de finalización:</strong> Configurá a qué hora termina la noche. Se muestra en el contador de la pantalla principal.</span></div>
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">②</span><span><strong style="color:#fff;">Mostrar Ganador:</strong> Muestra en pantalla el cliente con mayor consumo, con animación de confetti y corona.</span></div>
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">③</span><span><strong style="color:#fff;">Resetear noche:</strong> Borra todos los consumos del día y devuelve los saldos de tarjetas a su valor inicial. Esta acción no se puede deshacer.</span></div>
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">④</span><span><strong style="color:#fff;">Cerrar noche:</strong> Guarda los datos en el historial permanente antes de resetear.</span></div>
      </div>
    </div>
  </div>
  <div class="config-panel">
    <div class="config-title">Manejo de pantalla</div>
    <div class="config-row">
      <span class="config-label">Hora de finalizacion:</span>
      <input class="config-input narrow" id="hora-fin-input" type="time" value="05:30" oninput="updateHoraFin()" />
      <button class="btn-show-winner" onclick="mostrarGanadorManual()">🏆 Mostrar Ganador Ahora</button>
    </div>
    <div class="config-row" style="margin-top:8px;">
      <button class="btn-reset" onclick="resetNoche()">🔄 Resetear noche</button>
      <button onclick="cerrarNoche()" style="background:transparent;color:#3a9a5a;border:1px solid #1a3a2a;border-radius:6px;padding:9px 16px;font-family:'Rajdhani',sans-serif;font-size:13px;font-weight:700;letter-spacing:1px;cursor:pointer;white-space:nowrap;transition:all .15s;" onmouseover="this.style.borderColor='#3a9a5a'" onmouseout="this.style.borderColor='#1a3a2a'">💾 Cerrar noche</button>
    </div>
  </div>

</div>

<!-- CAJAS -->
<div id="tab-caja1" class="screen"><div id="caja-inner-1"></div></div>
<div id="tab-caja2" class="screen"><div id="caja-inner-2"></div></div>
<div id="tab-caja3" class="screen"><div id="caja-inner-3"></div></div>

<!-- PINES -->
<div id="tab-pines" class="screen">
  <!-- Tutorial desplegable -->
  <div style="margin-bottom:14px;">
    <button onclick="(function(b,p){b._open=!b._open;p.style.display=b._open?'block':'none';b.querySelector('.tut-arrow').textContent=b._open?'▲':'▼';})(this,document.getElementById('tut-pines'))" style="width:100%;display:flex;justify-content:space-between;align-items:center;background:#0a0b00;border:1px solid #2a2a00;border-radius:8px;padding:10px 16px;font-family:'Rajdhani',sans-serif;font-size:13px;font-weight:700;color:#c9a227;letter-spacing:1px;cursor:pointer;text-align:left;">
      <span>📖 Cómo usar — PINs de acceso</span><span class="tut-arrow" style="font-size:12px;color:#556;">▼</span>
    </button>
    <div id="tut-pines" style="display:none;background:#080800;border:1px solid #1a1a00;border-radius:0 0 8px 8px;padding:14px 18px;font-family:'Rajdhani',sans-serif;font-size:13px;color:#aaa;line-height:1.8;">
      <div style="display:flex;flex-direction:column;gap:8px;">
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">①</span><span><strong style="color:#fff;">Cada sección tiene su propio PIN</strong> de 4 dígitos: Manager, Caja Abajo, Caja Extendido, Caja VIP y Tarjetas.</span></div>
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">②</span><span><strong style="color:#fff;">Cambiar un PIN:</strong> Ingresá el nuevo PIN de 4 dígitos en el campo correspondiente y tocá <strong>💾 Guardar PINs</strong>.</span></div>
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">③</span><span><strong style="color:#fff;">Efectos del cambio:</strong> Al guardar, todas las sesiones activas en otros dispositivos se cierran automáticamente por seguridad.</span></div>
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">④</span><span>Podés dejar vacíos los campos que no querés cambiar.</span></div>
      </div>
    </div>
  </div>
  <div class="conf-header">
    <div class="conf-title">🔑 PINs de acceso</div>
    <div class="conf-sub">Cambiá los PINs de acceso a cada sección. Todos deben ser exactamente 4 dígitos.</div>
  </div>
  <div style="max-width:480px;">
    <div id="pines-list2" style="display:flex;flex-direction:column;gap:12px;margin-bottom:20px;"></div>
    <button onclick="guardarPines2()" style="background:#c9a227;color:#000;border:none;border-radius:7px;padding:12px 28px;font-family:'Rajdhani',sans-serif;font-size:15px;font-weight:700;letter-spacing:1px;cursor:pointer;transition:background .15s;" onmouseover="this.style.background='#e8c84a'" onmouseout="this.style.background='#c9a227'">💾 Guardar PINs</button>
    <div id="pines2-msg" style="font-family:'Rajdhani',sans-serif;font-size:13px;letter-spacing:1px;min-height:18px;margin-top:10px;"></div>
  </div>
</div>

<!-- ══ PANEL PUBLICIDAD (placeholder removed custom tab) ══ -->
<div id="tab-diseno" class="screen">
  <div class="conf-header">
    <div class="conf-title">🎨 Personalización</div>
    <div class="conf-sub">Cambiá colores, temas y textos. Los cambios se sincronizan en tiempo real con la pantalla.</div>
  </div>
  <!-- Tutorial desplegable -->
  <div style="margin-bottom:14px;">
    <button onclick="(function(b,p){b._open=!b._open;p.style.display=b._open?'block':'none';b.querySelector('.tut-arrow').textContent=b._open?'▲':'▼';})(this,document.getElementById('tut-diseno'))" style="width:100%;display:flex;justify-content:space-between;align-items:center;background:#0a0b00;border:1px solid #2a2a00;border-radius:8px;padding:10px 16px;font-family:'Rajdhani',sans-serif;font-size:13px;font-weight:700;color:#c9a227;letter-spacing:1px;cursor:pointer;text-align:left;">
      <span>📖 Cómo usar — Diseño y Personalización</span><span class="tut-arrow" style="font-size:12px;color:#556;">▼</span>
    </button>
    <div id="tut-diseno" style="display:none;background:#080800;border:1px solid #1a1a00;border-radius:0 0 8px 8px;padding:14px 18px;font-family:'Rajdhani',sans-serif;font-size:13px;color:#aaa;line-height:1.8;">
      <div style="display:flex;flex-direction:column;gap:8px;">
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">①</span><span><strong style="color:#fff;">Temas de noche:</strong> Elegí un tema especial (Jagger 12 Años, Touch of Pink, etc.) para cambiar el look completo de la pantalla de ranking.</span></div>
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">②</span><span><strong style="color:#fff;">Efectos de fondo:</strong> Activá burbujas, estrellas u otros efectos animados para el modo presentación.</span></div>
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">③</span><span><strong style="color:#fff;">Colores personalizados:</strong> Modificá cada color del sistema individualmente usando los selectores de color.</span></div>
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">④</span><span><strong style="color:#fff;">Textos y mensajes:</strong> Cambiá el texto del ganador, el tagline del club, el logo VIP y el mensaje del premio.</span></div>
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">⑤</span><span><strong style="color:#fff;">Aplicar cambios:</strong> Tocá <strong>✓ Aplicar cambios</strong> para sincronizar todo en tiempo real con la pantalla de ranking.</span></div>
      </div>
    </div>
  </div>

  <!-- Temas especiales -->
  <div class="custom-section">
    <div class="custom-section-title">🎉 Temas de noche especial</div>
    <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(180px,1fr));gap:10px;margin-bottom:14px;">
      <button onclick="aplicarTema('default')" style="background:#111;border:1px solid #2a2a2a;border-radius:10px;padding:16px 10px;cursor:pointer;text-align:center;color:#888;font-family:'Rajdhani',sans-serif;font-weight:600;letter-spacing:1px;font-size:13px;transition:all .2s;" onmouseover="this.style.borderColor='#555'" onmouseout="this.style.borderColor='#2a2a2a'">
        <div style="font-size:28px;margin-bottom:6px;">⬛</div>DEFAULT
      </button>
      <button onclick="aplicarTema('jagger12')" style="background:#000;border:2px solid #333;border-radius:10px;padding:16px 10px;cursor:pointer;text-align:center;color:#fff;font-family:'Rajdhani',sans-serif;font-weight:700;letter-spacing:1px;font-size:13px;transition:all .2s;" onmouseover="this.style.borderColor='#888'" onmouseout="this.style.borderColor='#333'">
        <div style="font-size:28px;margin-bottom:6px;">🥂</div>JAGGER 12 AÑOS
      </button>

      <button onclick="aplicarTema('touchofpink')" style="background:linear-gradient(135deg,#140010,#2a0020);border:2px solid #9d174d;border-radius:10px;padding:16px 10px;cursor:pointer;text-align:center;color:#f472b6;font-family:'Rajdhani',sans-serif;font-weight:700;letter-spacing:1px;font-size:13px;transition:all .2s;" onmouseover="this.style.borderColor='#f472b6'" onmouseout="this.style.borderColor='#9d174d'">
        <div style="font-size:28px;margin-bottom:6px;">🌸</div>TURNS PINK
      </button>
    </div>
    <!-- Toggle decoraciones para temas que las tienen -->
    <div id="tema-deco-toggle" style="display:none;background:#0a0a0a;border:1px solid #222;border-radius:8px;padding:12px 16px;margin-top:8px;align-items:center;gap:12px;flex-wrap:wrap;">
      <label id="deco-main-label" style="display:flex;align-items:center;gap:8px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;color:#888;letter-spacing:1px;">
        <input type="checkbox" id="toggle-deco" onchange="toggleDecoActual(this.checked)" style="width:16px;height:16px;accent-color:#c9a227;" checked />
        <span id="toggle-deco-label">Activar decoraciones animadas</span>
      </label>
      <div id="show-12-toggle" style="display:none;">
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;color:#888;letter-spacing:1px;">
          <input type="checkbox" id="toggle-12" onchange="mostrar12Fondo=this.checked;reiniciarDeco12();fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({mostrar12:this.checked})}).catch(()=>{});showToast(this.checked?'12 de fondo activado':'12 de fondo desactivado');" style="width:16px;height:16px;accent-color:#c9a227;" checked />
          <span>Mostrar "12" de fondo</span>
        </label>
      </div>

      <div id="falling-gloves-toggle" style="display:none;">
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;color:#888;letter-spacing:1px;">
          <input type="checkbox" id="toggle-falling-gloves" onchange="fallingGlovesActivos=this.checked;fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({falling_gloves:this.checked})}).catch(()=>{});showToast(this.checked?'✨ Lluvia activada':'Lluvia desactivada');" style="width:16px;height:16px;accent-color:#ff2222;" checked />
          <span>🥊 Guantes cayendo (lluvia de guantes)</span>
        </label>
      </div>

      <div id="pink-petalos-toggle" style="display:none;">
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;color:#cc88bb;letter-spacing:1px;">
          <input type="checkbox" id="toggle-pink-petalos" onchange="pinkPetalosActivos=this.checked;if(this.checked){iniciarPetalos();}else{const w=document.getElementById('petalos-wrap');if(w)w.innerHTML='';}fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({petals_activos:this.checked})}).catch(()=>{});" style="width:16px;height:16px;accent-color:#f472b6;" checked />
          <span>🌸 Pétalos cayendo</span>
        </label>
      </div>
      <div id="pink-modo-toggle" style="display:none;flex-direction:column;gap:6px;">
        <div style="font-family:'Rajdhani',sans-serif;font-size:11px;color:#cc88bb;letter-spacing:1px;text-transform:uppercase;margin-bottom:2px;">Fondo</div>
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;color:#cc88bb;letter-spacing:1px;">
          <input type="radio" name="pink-modo" value="oscuro" checked style="accent-color:#f472b6;" onchange="document.body.classList.remove('pink-claro');pinkModoClaro=false;" />
          <span>🌙 Rosa oscuro</span>
        </label>
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;color:#cc88bb;letter-spacing:1px;">
          <input type="radio" name="pink-modo" value="claro" style="accent-color:#f472b6;" onchange="document.body.classList.add('pink-claro');pinkModoClaro=true;" />
          <span>🌸 Rosa medio</span>
        </label>
      </div>
      <div id="ko-anim-toggle" style="display:none;">
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;color:#888;letter-spacing:1px;">
          <input type="checkbox" id="toggle-ko-anim" onchange="koAnimActiva=this.checked;showToast(this.checked?'🥊 Animación KO activada':'Animación KO desactivada');" style="width:16px;height:16px;accent-color:#ff2222;" checked />
          <span>🥊 Animación KO al cambiar de posición</span>
        </label>
      </div>
    </div>
  </div>

  <!-- Efectos de fondo -->
  <div class="custom-section">
    <div class="custom-section-title">✨ Efectos de fondo (modo presentación)</div>
    <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(140px,1fr));gap:8px;margin-bottom:10px;" id="efectos-grid">
      <button onclick="aplicarEfecto('ninguno')" id="efecto-btn-ninguno" style="background:#0a0a0a;border:1px solid #2a2a2a;border-radius:8px;padding:12px 8px;cursor:pointer;text-align:center;color:#555;font-family:'Rajdhani',sans-serif;font-weight:600;font-size:12px;letter-spacing:1px;transition:all .2s;">
        <div style="font-size:22px;margin-bottom:4px;">⬛</div>NINGUNO
      </button>
      <button onclick="aplicarEfecto('burbujas')" id="efecto-btn-burbujas" style="background:#0a0a0a;border:1px solid #2a2a2a;border-radius:8px;padding:12px 8px;cursor:pointer;text-align:center;color:#555;font-family:'Rajdhani',sans-serif;font-weight:600;font-size:12px;letter-spacing:1px;transition:all .2s;">
        <div style="font-size:22px;margin-bottom:4px;">🫧</div>BURBUJAS
      </button>
      <button onclick="aplicarEfecto('estrellas')" id="efecto-btn-estrellas" style="background:#0a0a0a;border:1px solid #2a2a2a;border-radius:8px;padding:12px 8px;cursor:pointer;text-align:center;color:#555;font-family:'Rajdhani',sans-serif;font-weight:600;font-size:12px;letter-spacing:1px;transition:all .2s;">
        <div style="font-size:22px;margin-bottom:4px;">⭐</div>ESTRELLAS
      </button>

    </div>
    <div class="preview-bar">
      <div class="preview-label">Vista previa del logo</div>
      <div class="preview-logo" id="prev-logo">RANKING <span class="vip-prev" id="prev-vip">VIP</span><span class="club-prev" id="prev-club">JAGGER CLUB</span></div>
    </div>
  </div>

  <!-- Tamaño reloj -->
  <div class="custom-section">
    <div class="custom-section-title">🕐 Tamaño del reloj (pantalla)</div>
    <div style="display:flex;gap:8px;margin-top:6px;">
      <button id="cs-chico" onclick="setClockSize('chico')" style="flex:1;background:#0a0a0a;border:1px solid #2a2a2a;border-radius:8px;padding:10px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-weight:600;font-size:13px;color:#555;letter-spacing:1px;transition:all .2s;">Chico</button>
      <button id="cs-mediano" onclick="setClockSize('mediano')" style="flex:1;background:#0a0a0a;border:1px solid var(--gold);border-radius:8px;padding:10px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-weight:600;font-size:13px;color:var(--gold);letter-spacing:1px;transition:all .2s;">Mediano</button>
      <button id="cs-grande" onclick="setClockSize('grande')" style="flex:1;background:#0a0a0a;border:1px solid #2a2a2a;border-radius:8px;padding:10px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-weight:600;font-size:13px;color:#555;letter-spacing:1px;transition:all .2s;">Grande</button>
    </div>
  </div>

  <!-- Textos -->
  <div class="custom-section">
    <div class="custom-section-title">Textos</div>
    <div class="custom-text-row">
      <span class="custom-text-label">Texto junto al nombre (ej: VIP, GOLD, PLUS...)</span>
      <input class="custom-text-input" id="ct-vip" type="text" placeholder="VIP" value="VIP" oninput="previewTextos()" />
    </div>
    <div style="margin-top:12px;background:#0a0a0a;border:1px solid #c9a227;border-radius:8px;padding:12px 14px;">
      <div style="font-family:'Rajdhani',sans-serif;font-size:11px;color:#c9a227;letter-spacing:2px;text-transform:uppercase;margin-bottom:8px;">🏆 Mensaje de Premio (se ve en todas las pantallas)</div>
      <input class="custom-text-input" id="ct-premio" type="text" placeholder="Ej: Botella de Jäger gratis" style="width:100%;margin-bottom:8px;" oninput="previewPremio()" />
      <div style="display:flex;align-items:center;gap:10px;">
        <span style="font-family:'Rajdhani',sans-serif;font-size:12px;color:#666;letter-spacing:1px;white-space:nowrap;">Tamaño</span>
        <input type="range" id="ct-premio-size" min="14" max="48" step="2" value="22" oninput="previewPremio()" style="flex:1;" />
        <span id="ct-premio-size-val" style="font-size:11px;color:#555;width:36px;flex-shrink:0;">22px</span>
      </div>
    </div>
    <div class="custom-text-row">
      <span class="custom-text-label">Mensaje ganador</span>
      <input class="custom-text-input" id="ct-winner-msg" type="text" placeholder="¡EL GANADOR DE LA NOCHE!" value="¡EL GANADOR DE LA NOCHE!" />
    </div>
    <div class="custom-text-row">
      <span class="custom-text-label">Subtítulo ganador</span>
      <input class="custom-text-input" id="ct-winner-sub" type="text" placeholder="Ej: ¡Se lleva la botella!" value="" />
    </div>
    <div style="border-top:1px solid #1a1a1a;margin-top:14px;padding-top:14px;">
      <div style="font-size:10px;color:#555;letter-spacing:2px;text-transform:uppercase;margin-bottom:10px;">Texto debajo de "RANKING VIP"</div>
      <div class="custom-text-row">
        <span class="custom-text-label">Texto</span>
        <input class="custom-text-input" id="ct-tagline" type="text" placeholder="JAGGER CLUB" value="JAGGER CLUB" oninput="previewTagline()" />
      </div>
      <div class="custom-text-row" style="align-items:center;">
        <span class="custom-text-label">Color</span>
        <input type="color" id="ct-tagline-color" value="#555555" oninput="previewTagline()" style="width:38px;height:28px;border:none;background:none;cursor:pointer;padding:0;flex-shrink:0;" />
      </div>
      <div class="custom-text-row" style="align-items:center;gap:8px;">
        <span class="custom-text-label">Brillo glow</span>
        <input type="range" id="ct-tagline-brightness" min="0" max="1" step="0.05" value="0" oninput="previewTagline();document.getElementById('ct-tagline-brightness-val').textContent=parseFloat(this.value).toFixed(2)" style="flex:1;" />
        <span id="ct-tagline-brightness-val" style="font-size:11px;color:#555;width:32px;text-align:right;flex-shrink:0;">0.00</span>
      </div>
      <div class="custom-text-row" style="align-items:center;">
        <span class="custom-text-label">Fuente</span>
        <select id="ct-tagline-font" onchange="previewTagline()" style="flex:1;background:#0d0d0d;border:1px solid #2a2a2a;border-radius:6px;color:#f0ece0;padding:7px 8px;font-size:13px;outline:none;">
          <option value="'Rajdhani',sans-serif">Rajdhani</option>
          <option value="'Oswald',sans-serif">Oswald</option>
          <option value="Impact,sans-serif">Impact</option>
          <option value="Arial,sans-serif">Arial</option>
        </select>
      </div>
    </div>
  </div>

  <!-- Opciones del ganador -->
  <div class="custom-section">
    <div class="custom-section-title">🏆 Opciones del ganador</div>
    <div style="display:flex;flex-direction:column;gap:8px;">
      <div style="background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:14px 16px;">
        <div style="font-family:'Rajdhani',sans-serif;font-size:12px;color:#666;letter-spacing:2px;text-transform:uppercase;margin-bottom:10px;">Efecto de partículas</div>
        <div style="display:flex;flex-direction:column;gap:8px;">
          <label style="display:flex;align-items:center;gap:10px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:14px;color:#aaa;letter-spacing:1px;">
            <input type="radio" name="tipo-particula" value="confetti" checked style="accent-color:#c9a227;" onchange="tipoParticula='confetti';confettiGanadorActivo=true;" />
            <span>🎊 Confetti de colores</span>
          </label>
          <label style="display:flex;align-items:center;gap:10px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:14px;color:#aaa;letter-spacing:1px;">
            <input type="radio" name="tipo-particula" value="confetti_dorado" style="accent-color:#c9a227;" onchange="tipoParticula='confetti_dorado';confettiGanadorActivo=true;" />
            <span>✨ Confetti dorado</span>
          </label>
          <label style="display:flex;align-items:center;gap:10px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:14px;color:#aaa;letter-spacing:1px;">
            <input type="radio" name="tipo-particula" value="billetes" style="accent-color:#c9a227;" onchange="tipoParticula='billetes';confettiGanadorActivo=true;" />
            <span>💵 Lluvia de billetes</span>
          </label>
          <label style="display:flex;align-items:center;gap:10px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:14px;color:#aaa;letter-spacing:1px;">
            <input type="radio" name="tipo-particula" value="champagne" style="accent-color:#c9a227;" onchange="tipoParticula='champagne';confettiGanadorActivo=true;" />
            <span>🍾 Lluvia de botellas</span>
          </label>
          <label style="display:flex;align-items:center;gap:10px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:14px;color:#aaa;letter-spacing:1px;">
            <input type="radio" name="tipo-particula" value="ninguno" style="accent-color:#c9a227;" onchange="tipoParticula='ninguno';confettiGanadorActivo=false;" />
            <span>Sin partículas</span>
          </label>
        </div>
      </div>
    </div>
  </div>

  <div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:8px;">
    <button class="btn-custom-save" onclick="aplicarPersonalizacion()">✓ Aplicar cambios</button>
  </div>


</div>

<!-- ══ PANEL PUBLICIDAD ══ -->
<div id="tab-publicidad" class="screen">
  <div class="conf-header">
    <div class="conf-title">📺 Sistema de publicidad</div>
    <div class="conf-sub">Subí un archivo de video .mp4 desde este dispositivo para mostrarlo en la pantalla de presentación.</div>
  </div>
  <!-- Tutorial desplegable -->
  <div style="margin-bottom:14px;">
    <button onclick="(function(b,p){b._open=!b._open;p.style.display=b._open?'block':'none';b.querySelector('.tut-arrow').textContent=b._open?'▲':'▼';})(this,document.getElementById('tut-pub'))" style="width:100%;display:flex;justify-content:space-between;align-items:center;background:#0a0b00;border:1px solid #2a2a00;border-radius:8px;padding:10px 16px;font-family:'Rajdhani',sans-serif;font-size:13px;font-weight:700;color:#c9a227;letter-spacing:1px;cursor:pointer;text-align:left;">
      <span>📖 Cómo usar — Publicidad</span><span class="tut-arrow" style="font-size:12px;color:#556;">▼</span>
    </button>
    <div id="tut-pub" style="display:none;background:#080800;border:1px solid #1a1a00;border-radius:0 0 8px 8px;padding:14px 18px;font-family:'Rajdhani',sans-serif;font-size:13px;color:#aaa;line-height:1.8;">
      <div style="display:flex;flex-direction:column;gap:8px;">
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">①</span><span><strong style="color:#fff;">Subir archivo:</strong> Elegí un archivo .mp4 guardado en este dispositivo y tocá "Subir y usar este video". El video queda guardado de forma local en el sistema.</span></div>
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">②</span><span><strong style="color:#fff;">Activar programa:</strong> Configurá cada cuántos minutos aparece el video en la pantalla de presentación y tocá "Activar programa".</span></div>
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">③</span><span><strong style="color:#fff;">Mostrar ahora:</strong> Envía el video inmediatamente a la pantalla de presentación sin esperar el intervalo configurado.</span></div>
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">④</span><span>La publicidad solo aparece en el modo presentación de la pantalla principal, no en las cajas.</span></div>
      </div>
    </div>
  </div>
  <div style="background:#1a0a00;border:1px solid #3a2000;border-radius:8px;padding:10px 16px;margin-bottom:16px;font-family:'Rajdhani',sans-serif;font-size:13px;color:#c9a227;letter-spacing:0.5px;">
    💡 Subí siempre un <strong>archivo .mp4 local</strong> desde este dispositivo. El video queda guardado en el sistema y no depende de internet ni de links externos.
  </div>
  <div style="margin-bottom:20px;">
    <!-- Subir archivo local -->
    <div style="background:var(--surface);border:1px solid #2a5a00;border-radius:10px;padding:18px 20px;">
      <div style="font-family:'Oswald',sans-serif;font-size:13px;color:var(--gold);letter-spacing:2px;text-transform:uppercase;margin-bottom:6px;">✅ Subir archivo de video .mp4</div>
      <div style="font-size:12px;color:#555;margin-bottom:12px;font-family:'Rajdhani',sans-serif;">Elegí un archivo .mp4 guardado en este dispositivo. Queda guardado de forma local en el sistema.</div>
      <!-- URL del video (se completa sola al subir el archivo) -->
      <input id="pub-url" type="hidden" />
      <label class="field-label">Seleccioná un archivo .mp4</label>
      <input id="pub-file" type="file" accept="video/mp4,.mp4" style="width:100%;background:#0d0d0d;border:1px solid var(--border);border-radius:7px;color:var(--text);padding:10px 12px;font-family:'Rajdhani',sans-serif;font-size:13px;margin-bottom:8px;" />
      <button onclick="subirVideoPublicidad()" class="btn-add">⬆ Subir y usar este video</button>
      <div id="pub-upload-status" style="margin:10px 0 4px;font-size:12px;color:#555;letter-spacing:1px;"></div>
      <label class="field-label" style="margin-top:10px;">Cada cuántos minutos aparece en pantalla</label>
      <input id="pub-frec" class="field-input" type="number" min="1" max="120" value="15" />
      <div class="btn-row" style="margin-top:8px;gap:10px;">
        <button onclick="activarPublicidad()" class="btn-add">▶ Activar programa</button>
        <button onclick="mostrarAhora()" style="background:#1a3a1a;color:#3a9a5a;border:1px solid #2a5a2a;border-radius:7px;padding:10px 16px;font-family:'Rajdhani',sans-serif;font-size:14px;font-weight:700;cursor:pointer;letter-spacing:1px;" onmouseover="this.style.background='#223a22'" onmouseout="this.style.background='#1a3a1a'">📺 Mostrar ahora en pantalla</button>
        <button onclick="desactivarPublicidad()" style="background:transparent;color:#555;border:1px solid #222;border-radius:7px;padding:10px 16px;font-family:'Rajdhani',sans-serif;font-size:14px;cursor:pointer;transition:all .15s;" onmouseover="this.style.borderColor='#c9a227';this.style.color='#c9a227'" onmouseout="this.style.borderColor='#222';this.style.color='#555'">■ Desactivar</button>
      </div>
    </div>
  </div>
  <div id="pub-estado-box" style="background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:14px 18px;display:flex;align-items:center;gap:14px;">
    <span style="font-size:22px;" id="pub-estado-icon">⏸</span>
    <div>
      <div style="font-family:'Oswald',sans-serif;font-size:14px;color:var(--text);letter-spacing:1px;">Estado: <span id="pub-estado-txt" style="color:var(--gold);">Inactiva</span></div>
      <div id="pub-estado-url" style="font-size:11px;color:#555;margin-top:3px;"></div>
    </div>
  </div>
</div>

<!-- ══ PANEL MENÚ MANAGER ══ -->
<div id="tab-menu-mgr" class="screen active">
  <div class="conf-header" style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:10px;">
    <div style="display:flex;align-items:center;gap:12px;">
      <button onclick="resetearMenuDefault()" style="background:linear-gradient(135deg,#1a1200,#2a1c00);border:1px solid #c9a22766;border-radius:8px;padding:10px 14px;font-family:'Rajdhani',sans-serif;font-size:18px;font-weight:700;color:#c9a227;cursor:pointer;line-height:1;box-shadow:0 0 12px rgba(201,162,39,0.12);" onmouseover="this.style.borderColor='#c9a227';" onmouseout="this.style.borderColor='#c9a22766';">↺</button>
      <div>
        <div class="conf-title">🍽 Menú de productos</div>
        <div class="conf-sub">Tocá nombre o precio para editarlo. Los cambios se guardan al instante.</div>
      </div>
    </div>
  </div>
  <!-- Tutorial desplegable -->
  <div style="margin-bottom:14px;">
    <button onclick="(function(b,p){b._open=!b._open;p.style.display=b._open?'block':'none';b.querySelector('.tut-arrow').textContent=b._open?'▲':'▼';})(this,document.getElementById('tut-menu'))" style="width:100%;display:flex;justify-content:space-between;align-items:center;background:#0a0b00;border:1px solid #2a2a00;border-radius:8px;padding:10px 16px;font-family:'Rajdhani',sans-serif;font-size:13px;font-weight:700;color:#c9a227;letter-spacing:1px;cursor:pointer;text-align:left;">
      <span>📖 Cómo usar — Menú de productos</span><span class="tut-arrow" style="font-size:12px;color:#556;">▼</span>
    </button>
    <div id="tut-menu" style="display:none;background:#080800;border:1px solid #1a1a00;border-radius:0 0 8px 8px;padding:14px 18px;font-family:'Rajdhani',sans-serif;font-size:13px;color:#aaa;line-height:1.8;">
      <div style="display:flex;flex-direction:column;gap:8px;">
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">①</span><span><strong style="color:#fff;">Editar nombre o precio:</strong> Tocá directamente sobre el nombre o precio de cualquier producto para editarlo en línea. Presioná Enter o hacé click fuera para guardar.</span></div>
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">②</span><span><strong style="color:#fff;">Agregar producto:</strong> Completá categoría, nombre y precio en el formulario de arriba y tocá <strong>+ Agregar</strong>.</span></div>
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">③</span><span><strong style="color:#fff;">Eliminar:</strong> Usá el botón ✕ al lado de cada producto para eliminarlo.</span></div>
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">④</span><span><strong style="color:#fff;">Restablecer menú oficial:</strong> El botón ↺ carga el menú predefinido del sistema, reemplazando todos los productos actuales.</span></div>
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">⑤</span><span><strong style="color:#fff;">Precios de cartel:</strong> Definí cuánto descuenta automáticamente cada tipo de cartel (virtual/físico/combo) del saldo de una tarjeta.</span></div>
      </div>
    </div>
  </div>
  <!-- Agregar producto -->
  <div style="background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:16px 20px;margin-bottom:20px;">
    <div style="font-family:'Oswald',sans-serif;font-size:14px;color:var(--gold);letter-spacing:2px;text-transform:uppercase;margin-bottom:14px;">+ Nuevo producto</div>
    <div style="display:flex;gap:10px;flex-wrap:wrap;align-items:flex-end;">
      <div style="display:flex;flex-direction:column;gap:4px;flex:1;min-width:140px;">
        <label class="field-label">Categoría</label>
        <select id="new-prod-cat" class="field-input" style="padding:10px 12px;">
          <option value="Champagne">Champagne</option>
          <option value="Bottle Service">Bottle Service</option>
          <option value="Importados">Importados</option>
          <option value="Bebidas">Bebidas</option>
          <option value="Tragos">Tragos</option>
          <option value="Shots">Shots</option>
        </select>
      </div>
      <div style="display:flex;flex-direction:column;gap:4px;flex:2;min-width:180px;">
        <label class="field-label">Nombre</label>
        <input id="new-prod-nombre" class="field-input" type="text" placeholder="Nombre del producto" />
      </div>
      <div style="display:flex;flex-direction:column;gap:4px;flex:1;min-width:120px;">
        <label class="field-label">Precio ($)</label>
        <input id="new-prod-precio" class="field-input" type="number" min="0" step="100" placeholder="0" />
      </div>
      <button onclick="agregarProducto()" class="btn-add" style="flex:none;padding:10px 20px;white-space:nowrap;">+ Agregar</button>
    </div>
  </div>
  <!-- Precios de Cartel -->
  <div style="background:var(--surface);border:1px solid #3a3000;border-radius:10px;padding:16px 20px;margin-bottom:20px;">
    <div style="font-family:'Oswald',sans-serif;font-size:14px;color:var(--gold);letter-spacing:2px;text-transform:uppercase;margin-bottom:4px;">🍾 Precios de Cartel</div>
    <div style="font-size:11px;color:#555;letter-spacing:1px;margin-bottom:14px;">Monto que se descuenta de la tarjeta al emitir cada tipo de cartel. Poner 0 para no descontar.</div>
    <div style="display:flex;gap:10px;flex-wrap:wrap;align-items:flex-end;">
      <div style="display:flex;flex-direction:column;gap:4px;flex:1;min-width:120px;">
        <label class="field-label">📺 Virtual ($)</label>
        <input id="cartel-precio-virtual" class="field-input" type="number" min="0" step="500" placeholder="0" style="padding:10px 12px;" />
      </div>
      <div style="display:flex;flex-direction:column;gap:4px;flex:1;min-width:120px;">
        <label class="field-label">🖨 Físico ($)</label>
        <input id="cartel-precio-fisico" class="field-input" type="number" min="0" step="500" placeholder="0" style="padding:10px 12px;" />
      </div>
      <div style="display:flex;flex-direction:column;gap:4px;flex:1;min-width:120px;">
        <label class="field-label">⭐ Combo ($)</label>
        <input id="cartel-precio-combo" class="field-input" type="number" min="0" step="500" placeholder="0" style="padding:10px 12px;" />
      </div>
      <button onclick="guardarCartelPrecios()" class="btn-add" style="flex:none;padding:10px 20px;white-space:nowrap;">Guardar</button>
    </div>
    <div id="cartel-precios-status" style="font-size:11px;color:#555;margin-top:8px;letter-spacing:1px;"></div>
  </div>
  <!-- Búsqueda + Listado -->
  <div style="margin-bottom:14px;">
    <input id="menu-mgr-search" class="field-input" type="text" placeholder="🔍 Buscar producto..." oninput="renderMenuMgr()" style="margin-bottom:0;" />
  </div>
  <div id="menu-mgr-content"></div>
</div>

<!-- ══ PANEL STATS ══ -->
<div id="tab-stats" class="screen">
  <!-- Tutorial desplegable -->
  <div style="margin-bottom:14px;">
    <button onclick="(function(b,p){b._open=!b._open;p.style.display=b._open?'block':'none';b.querySelector('.tut-arrow').textContent=b._open?'▲':'▼';})(this,document.getElementById('tut-stats'))" style="width:100%;display:flex;justify-content:space-between;align-items:center;background:#0a0b00;border:1px solid #2a2a00;border-radius:8px;padding:10px 16px;font-family:'Rajdhani',sans-serif;font-size:13px;font-weight:700;color:#c9a227;letter-spacing:1px;cursor:pointer;text-align:left;">
      <span>📖 Cómo leer las Estadísticas</span><span class="tut-arrow" style="font-size:12px;color:#556;">▼</span>
    </button>
    <div id="tut-stats" style="display:none;background:#080800;border:1px solid #1a1a00;border-radius:0 0 8px 8px;padding:14px 18px;font-family:'Rajdhani',sans-serif;font-size:13px;color:#aaa;line-height:1.8;">
      <div style="display:flex;flex-direction:column;gap:8px;">
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">①</span><span><strong style="color:#fff;">KPIs:</strong> Total facturado en la noche, número de operaciones y el promedio por operación.</span></div>
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">②</span><span><strong style="color:#fff;">Gráficos:</strong> Distribución de ventas por caja y evolución del consumo por hora durante la noche.</span></div>
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">③</span><span><strong style="color:#fff;">Top clientes:</strong> Ranking de clientes con mayor consumo de la noche actual.</span></div>
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">④</span><span><strong style="color:#fff;">Historial:</strong> Tocá <strong>📊 VER HISTORIAL</strong> para acceder a los registros de noches anteriores.</span></div>
        <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">⑤</span><span>Los datos se actualizan automáticamente cada 2 segundos.</span></div>
      </div>
    </div>
  </div>
  <div class="stats-header" style="display:flex;align-items:flex-start;justify-content:space-between;flex-wrap:wrap;gap:10px;">
    <div>
      <div class="stats-title">Estadísticas de la noche</div>
      <div class="stats-sub" id="stats-sub">—</div>
    </div>
    <a href="/historial" target="_blank" style="background:#c9a227;color:#000;border:none;border-radius:7px;padding:10px 22px;font-family:'Rajdhani',sans-serif;font-size:13px;font-weight:700;letter-spacing:2px;cursor:pointer;text-decoration:none;white-space:nowrap;flex-shrink:0;margin-top:4px;">📊 VER HISTORIAL</a>
  </div>

  <!-- KPIs -->
  <div class="kpi-row">
    <div class="kpi-card">
      <div class="kpi-label">Total de la noche</div>
      <div class="kpi-val" id="kpi-total">$0</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">Operaciones</div>
      <div class="kpi-val" id="kpi-ops">0</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">Promedio por operación</div>
      <div class="kpi-val" id="kpi-avg">$0</div>
    </div>
  </div>

  <!-- Fila de gráficos -->
  <div class="charts-row">
    <div class="chart-box">
      <div class="chart-title">Facturación por caja</div>
      <canvas id="chart-cajas" height="180"></canvas>
      <div class="chart-legend" id="legend-cajas"></div>
    </div>
    <div class="chart-box">
      <div class="chart-title">Consumo por hora</div>
      <canvas id="chart-horas" height="180"></canvas>
    </div>
  </div>

  <!-- Top clientes -->
  <div class="stats-section-title">Top clientes de la noche</div>
  <div id="stats-top-clientes"></div>

  <!-- Detalle por caja -->
  <div class="stats-section-title">Detalle por caja</div>
  <div class="cajas-detail-row" id="cajas-detail"></div>
</div>

<!-- ══ OVERLAY PUBLICIDAD ══ -->
<div id="pub-overlay" style="display:none;position:fixed;inset:0;z-index:10500;background:rgba(0,0,0,0.9);align-items:center;justify-content:center;">
  <div style="background:#000;border-radius:14px;overflow:hidden;width:min(90vw,960px);aspect-ratio:16/9;position:relative;box-shadow:0 0 80px rgba(0,0,0,0.9);">
    <button onclick="cerrarPublicidadOverlay(true)" style="position:absolute;top:10px;right:12px;z-index:10502;background:rgba(0,0,0,0.7);color:#ccc;border:1px solid #444;font-size:18px;cursor:pointer;border-radius:6px;width:32px;height:32px;line-height:1;transition:color .2s;" onmouseover="this.style.color='#fff'" onmouseout="this.style.color='#ccc'">✕</button>
    <video id="pub-video" autoplay playsinline style="width:100%;height:100%;object-fit:contain;display:none;background:#000;"></video>
    <iframe id="pub-iframe" style="width:100%;height:100%;border:none;display:none;" allowfullscreen allow="autoplay;fullscreen"></iframe>
  </div>
</div>

<!-- ══ OVERLAY GANADOR ══ -->
<div id="winner-overlay">
  <div class="winner-rays"></div>
  <div class="confetti-wrap" id="confetti-wrap"></div>
  <button class="winner-close" onclick="cerrarGanador()">✕</button>
  <div class="winner-content">
    <span class="winner-corona" id="winner-corona">👑</span>
    <div class="winner-titulo" id="winner-titulo">GANADOR DE LA NOCHE</div>
    <div class="winner-nombre" id="winner-nombre">—</div>
    <div class="winner-line"></div>
    <div class="winner-info-row">
      <div class="winner-info-block">
        <div class="winner-info-label">Mesa</div>
        <div class="winner-info-val" id="winner-mesa">—</div>
      </div>
      <div class="winner-info-block">
        <div class="winner-info-label">Total consumido</div>
        <div class="winner-info-val" id="winner-total">—</div>
      </div>
    </div>
    <div class="winner-mensaje" id="winner-mensaje"></div>
  </div>
</div>



<!-- ══ OVERLAY CARTEL ══ -->
<div id="cartel-overlay" style="display:none;position:fixed;inset:0;z-index:11000;background:rgba(0,0,0,0.98);flex-direction:column;align-items:center;justify-content:center;overflow:hidden;">
  <div id="cartel-tema-bg" style="position:absolute;inset:0;pointer-events:none;z-index:0;"></div>
  <!-- Efecto de rayos de fondo -->
  <div id="cartel-rays" style="position:absolute;inset:0;pointer-events:none;z-index:1;overflow:hidden;"></div>
  <button onclick="cerrarCartel()" style="position:fixed;top:20px;right:28px;z-index:11001;background:transparent;color:#333;border:none;font-size:28px;cursor:pointer;font-family:'Oswald',sans-serif;transition:color .2s;line-height:1;" onmouseover="this.style.color='#888'" onmouseout="this.style.color='#333'">✕</button>
  <div id="cartel-content" style="position:relative;z-index:2;text-align:center;max-width:90vw;width:100%;padding:0 24px;">
    <!-- Emoji animado grande -->
    <div id="cartel-emoji-big" style="font-size:110px;margin-bottom:6px;display:block;animation:bottleBounce 1.2s cubic-bezier(.36,.07,.19,.97) infinite,emojiGlow 2s ease-in-out infinite alternate;filter:drop-shadow(0 0 30px rgba(201,162,39,0.6));"></div>
    <div id="cartel-nombre-display" style="font-family:'Oswald',sans-serif;font-size:clamp(48px,8vw,100px);font-weight:700;color:#fff;letter-spacing:4px;text-transform:uppercase;line-height:1;text-shadow:0 0 40px rgba(201,162,39,.5);"></div>
    <div id="cartel-mesa-display" style="font-family:'Oswald',sans-serif;font-size:clamp(20px,3vw,36px);color:#888;letter-spacing:3px;margin-top:8px;"></div>
    <div style="height:2px;background:linear-gradient(to right,transparent,var(--gold),transparent);margin:18px auto;max-width:600px;width:80%;"></div>
    <div id="cartel-frase-display" style="font-family:'Rajdhani',sans-serif;font-size:clamp(26px,4.5vw,56px);font-weight:700;color:#fff;letter-spacing:2px;text-transform:uppercase;line-height:1.3;padding:18px 36px;border:2px solid var(--gold);border-radius:8px;background:rgba(201,162,39,0.1);max-width:85vw;display:inline-block;text-shadow:0 0 20px rgba(201,162,39,0.3);box-shadow:0 0 40px rgba(201,162,39,0.15),inset 0 0 40px rgba(201,162,39,0.05);"></div>
  </div>
</div>

<!-- ══ MODAL CARTEL EDITOR ══ -->
<div id="cartel-modal" style="display:none;position:fixed;inset:0;z-index:10500;background:rgba(0,0,0,0.92);align-items:center;justify-content:center;">
  <div style="background:#111;border:1px solid #2a2a2a;border-radius:14px;padding:28px 32px;width:min(480px,92vw);max-height:90vh;overflow-y:auto;">
    <div style="font-family:'Oswald',sans-serif;font-size:22px;color:#fff;letter-spacing:1px;margin-bottom:20px;">📣 Configurar Cartel</div>
    <div style="margin-bottom:14px;">
      <label style="font-size:10px;color:#555;letter-spacing:2px;text-transform:uppercase;display:block;margin-bottom:6px;">Nombre (opcional)</label>
      <input id="cartel-nombre" type="text" placeholder="Ej: MATI" style="width:100%;background:#0d0d0d;border:1px solid #2a2a2a;border-radius:7px;color:#f0ece0;padding:11px 13px;font-family:'Rajdhani',sans-serif;font-size:16px;font-weight:600;outline:none;" onfocus="this.style.borderColor='#c9a227'" onblur="this.style.borderColor='#2a2a2a'" />
    </div>
    <div style="margin-bottom:14px;">
      <label style="font-size:10px;color:#555;letter-spacing:2px;text-transform:uppercase;display:block;margin-bottom:6px;">Mesa (opcional)</label>
      <input id="cartel-mesa" type="text" placeholder="Ej: 5" style="width:100%;background:#0d0d0d;border:1px solid #2a2a2a;border-radius:7px;color:#f0ece0;padding:11px 13px;font-family:'Rajdhani',sans-serif;font-size:16px;outline:none;" onfocus="this.style.borderColor='#c9a227'" onblur="this.style.borderColor='#2a2a2a'" />
    </div>
    <div style="margin-bottom:14px;">
      <label style="font-size:10px;color:#555;letter-spacing:2px;text-transform:uppercase;display:block;margin-bottom:6px;">Frase del cartel</label>
      <input id="cartel-frase" type="text" placeholder="Ej: SACÓ UN NUVO CON BENGALAS" style="width:100%;background:#0d0d0d;border:1px solid #2a2a2a;border-radius:7px;color:#f0ece0;padding:11px 13px;font-family:'Rajdhani',sans-serif;font-size:15px;outline:none;" onfocus="this.style.borderColor='#c9a227'" onblur="this.style.borderColor='#2a2a2a'" />
    </div>
    <div style="margin-bottom:18px;">
      <label style="font-size:10px;color:#555;letter-spacing:2px;text-transform:uppercase;display:block;margin-bottom:5px;">Emojis <span style="color:#333;font-size:9px;letter-spacing:1px;">(hasta 3 · vacío = auto)</span></label>
      <div style="margin-bottom:8px;display:flex;gap:8px;align-items:center;">
        <input id="cartel-emoji-input" type="text" placeholder="🍾" maxlength="24" style="width:120px;background:#0d0d0d;border:1px solid #2a2a2a;border-radius:7px;color:#f0ece0;padding:10px 10px;font-size:22px;text-align:center;outline:none;letter-spacing:4px;" onfocus="this.style.borderColor='#c9a227'" onblur="this.style.borderColor='#2a2a2a'" oninput="limitarEmojis(this)" />
        <button onclick="document.getElementById('cartel-emoji-input').value=''" style="background:none;border:1px solid #222;border-radius:5px;color:#555;padding:8px 10px;cursor:pointer;font-size:13px;font-family:'Rajdhani',sans-serif;" title="Limpiar">✕</button>
      </div>
      <div style="display:flex;gap:6px;flex-wrap:wrap;">
        <span onclick="agregarEmojiCartel(this.textContent)" style="font-size:24px;cursor:pointer;padding:4px 6px;background:#0a0a0a;border-radius:6px;border:1px solid #1a1a1a;" title="Tap para agregar">🍾</span>
        <span onclick="agregarEmojiCartel(this.textContent)" style="font-size:24px;cursor:pointer;padding:4px 6px;background:#0a0a0a;border-radius:6px;border:1px solid #1a1a1a;" title="Tap para agregar">🎆</span>
        <span onclick="agregarEmojiCartel(this.textContent)" style="font-size:24px;cursor:pointer;padding:4px 6px;background:#0a0a0a;border-radius:6px;border:1px solid #1a1a1a;" title="Tap para agregar">🥊</span>
        <span onclick="agregarEmojiCartel(this.textContent)" style="font-size:24px;cursor:pointer;padding:4px 6px;background:#0a0a0a;border-radius:6px;border:1px solid #1a1a1a;" title="Tap para agregar">🔥</span>
        <span onclick="agregarEmojiCartel(this.textContent)" style="font-size:24px;cursor:pointer;padding:4px 6px;background:#0a0a0a;border-radius:6px;border:1px solid #1a1a1a;" title="Tap para agregar">💎</span>
        <span onclick="agregarEmojiCartel(this.textContent)" style="font-size:24px;cursor:pointer;padding:4px 6px;background:#0a0a0a;border-radius:6px;border:1px solid #1a1a1a;" title="Tap para agregar">🎉</span>
        <span onclick="agregarEmojiCartel(this.textContent)" style="font-size:24px;cursor:pointer;padding:4px 6px;background:#0a0a0a;border-radius:6px;border:1px solid #1a1a1a;" title="Tap para agregar">⭐</span>
        <span onclick="agregarEmojiCartel(this.textContent)" style="font-size:24px;cursor:pointer;padding:4px 6px;background:#0a0a0a;border-radius:6px;border:1px solid #1a1a1a;" title="Tap para agregar">🏆</span>
      </div>
    </div>
    <div id="cartel-precio-section" style="margin-bottom:18px;border-top:1px solid #1a1a1a;padding-top:14px;">
      <label style="font-size:10px;color:#555;letter-spacing:2px;text-transform:uppercase;display:block;margin-bottom:10px;">Tipo de cartel</label>
      <div style="display:flex;gap:8px;margin-bottom:10px;">
        <button id="cbtn-virtual" onclick="selCartelTipo('virtual')" style="flex:1;background:#0a0a0a;border:1px solid #333;border-radius:7px;padding:10px 6px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:12px;font-weight:700;color:#555;letter-spacing:1px;text-align:center;line-height:1.6;">💻 VIRTUAL<br><span id="cprecio-v" style="font-size:11px;color:#444;">—</span></button>
        <button id="cbtn-fisico" onclick="selCartelTipo('fisico')" style="flex:1;background:#0a0a0a;border:1px solid #333;border-radius:7px;padding:10px 6px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:12px;font-weight:700;color:#555;letter-spacing:1px;text-align:center;line-height:1.6;">🖨 FÍSICO<br><span id="cprecio-f" style="font-size:11px;color:#444;">—</span></button>
        <button id="cbtn-combo" onclick="selCartelTipo('combo')" style="flex:1;background:#0a0a0a;border:1px solid #333;border-radius:7px;padding:10px 6px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:12px;font-weight:700;color:#555;letter-spacing:1px;text-align:center;line-height:1.6;">📦 COMBO<br><span id="cprecio-c" style="font-size:11px;color:#444;">—</span></button>
      </div>
      <div id="cartel-descuento-mgr" style="display:none;background:#0a0a0a;border:1px solid #2a2a2a;border-radius:7px;padding:10px 14px;">
        <label style="display:flex;align-items:center;gap:10px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;color:#aaa;letter-spacing:1px;">
          <input type="checkbox" id="cartel-descontar-mgr" style="width:16px;height:16px;accent-color:#c9a227;" checked />
          <span>Descontar <strong id="cartel-monto-mgr" style="color:#c9a227;">$0</strong> de la tarjeta</span>
        </label>
      </div>
    </div>
    <div style="display:flex;gap:10px;flex-wrap:wrap;">
      <button onclick="mostrarCartel()" style="flex:1;background:#c9a227;color:#000;border:none;border-radius:7px;padding:13px 20px;font-family:'Rajdhani',sans-serif;font-size:15px;font-weight:700;letter-spacing:1px;cursor:pointer;transition:background .15s;" onmouseover="this.style.background='#e8c84a'" onmouseout="this.style.background='#c9a227'">📣 MOSTRAR</button>
      <button onclick="cerrarCartelModal()" style="background:transparent;color:#555;border:1px solid #222;border-radius:7px;padding:13px 16px;font-family:'Rajdhani',sans-serif;font-size:14px;cursor:pointer;transition:all .15s;" onmouseover="this.style.borderColor='#c9a227';this.style.color='#c9a227'" onmouseout="this.style.borderColor='#222';this.style.color='#555'">Cancelar</button>
    </div>
  </div>
</div>


<div class="toast" id="toast"></div>

<!-- ══ MODAL EDITAR TX ══ -->
<div id="edit-tx-modal" style="display:none;position:fixed;inset:0;z-index:10800;background:rgba(0,0,0,0.88);align-items:center;justify-content:center;">
  <div style="background:#111;border:1px solid #2a2a2a;border-radius:12px;padding:28px 32px;width:min(400px,92vw);">
    <div style="font-family:'Oswald',sans-serif;font-size:18px;color:#fff;letter-spacing:1px;margin-bottom:18px;">✎ Editar operación</div>
    <input type="hidden" id="edit-tx-id" />
    <label class="field-label">Nombre</label>
    <input class="field-input" id="edit-tx-name" type="text" autocomplete="off" />
    <label class="field-label">Monto ($)</label>
    <input class="field-input amount-input" id="edit-tx-amount" type="number" min="1" step="100" />
    <div style="display:flex;gap:10px;margin-top:6px;">
      <button onclick="confirmarEditTx()" style="flex:1;background:#c9a227;color:#000;border:none;border-radius:7px;padding:12px;font-family:'Rajdhani',sans-serif;font-size:15px;font-weight:700;letter-spacing:1px;cursor:pointer;">Guardar</button>
      <button onclick="document.getElementById('edit-tx-modal').style.display='none'" style="background:transparent;color:#555;border:1px solid #222;border-radius:7px;padding:12px 16px;font-family:'Rajdhani',sans-serif;font-size:14px;cursor:pointer;">Cancelar</button>
    </div>
  </div>
</div>

<!-- Login modal removido: auth via sesión de servidor -->

<script>
let txData = [];
let tarjetasData = {};
let confTarjetas = Array.from({length:30}, (_,i) => ({slot:i+1, codigo:'', saldo_inicial:''}));
let scanSlotActivo = null;
let cajaFocus = 0;
let globalBuffer = '';
let globalTimeout = null;
let horaFin = '05:30';
let ganadorMostrado = false;

// ══════════════════════════════════════════
//  PERSONALIZACION
// ══════════════════════════════════════════
const COLOR_DEFS = [
  {key:'--gold',       label:'Dorado principal — logo VIP, números del ranking, totales, botones principales y barra del ganador', default:'#c9a227'},
  {key:'--gold-light', label:'Dorado claro — color al pasar el mouse por encima de botones dorados', default:'#e8c84a'},
  {key:'--gold-dim',   label:'Dorado oscuro — títulos de sección, línea decorativa del header, texto secundario dorado', default:'#7a6010'},
  {key:'--black',      label:'Fondo principal — color de fondo de TODA la pantalla', default:'#080808'},
  {key:'--surface',    label:'Superficies — fondo de paneles, tarjetas, filas del ranking y formularios', default:'#111111'},
  {key:'--border',     label:'Bordes — líneas que rodean los paneles, tarjetas y separadores', default:'#2a2a2a'},
  {key:'--text',       label:'Texto principal — todo el texto de contenido, etiquetas e instrucciones', default:'#f0ece0'},
  {key:'--text-dim',   label:'Texto secundario — hints, subtítulos, labels de campos y texto apagado', default:'#555555'},
  {key:'--white',      label:'Blanco — nombres de clientes en el ranking y encabezados principales', default:'#ffffff'},
  {key:'--green',      label:'Verde — saldo disponible de tarjetas e indicador de conexión activa', default:'#2ecc71'},
  {key:'--danger',     label:'Rojo — errores, saldo insuficiente, botones de eliminar y resetear', default:'#a83030'},
];

let customColors = {};
COLOR_DEFS.forEach(c => customColors[c.key] = c.default);

// ══════════════════════════════════════════
//  TEMAS FESTIVOS
// ══════════════════════════════════════════
const TEMAS = {
  default: {
    colors: {}, // sin cambios, usa defaults
    overlay: '',
    bodyClass: ''
  },
  fullblack: {
    colors: {'--black':'#000000','--surface':'#080808','--border':'#1a1a1a','--gold':'#c9a227','--text':'#cccccc'},
    overlay: '',
    bodyClass: 'tema-fullblack'
  },
  navidad: {
    colors: {'--black':'#050f05','--surface':'#091209','--border':'#1a3a1a','--gold':'#e8c84a','--text':'#f0f0e0','--white':'#ffffff'},
    bodyClass: 'tema-navidad',
    overlay: `
      <!-- Estrellas de nieve -->
      <svg width="100%" height="100%" xmlns="http://www.w3.org/2000/svg" style="position:absolute;inset:0;">
        <defs>
          <radialGradient id="vign" cx="50%" cy="50%" r="70%">
            <stop offset="0%" stop-color="transparent"/>
            <stop offset="100%" stop-color="#020a02" stop-opacity="0.7"/>
          </radialGradient>
        </defs>
        <rect width="100%" height="100%" fill="url(#vign)"/>
      </svg>
      <!-- Copos de nieve animados -->
      <div id="copos-wrap" style="position:absolute;inset:0;overflow:hidden;"></div>
      <!-- Árbol esquina izquierda -->
      <svg width="180" height="280" viewBox="0 0 180 280" style="position:absolute;bottom:0;left:0;opacity:.18;" xmlns="http://www.w3.org/2000/svg">
        <polygon points="90,10 160,120 120,120 150,200 100,200 110,260 80,260 70,200 30,200 60,120 20,120" fill="#2d7a2d"/>
        <rect x="75" y="255" width="30" height="25" fill="#5c3a1a" rx="3"/>
        <circle cx="90" cy="8" r="8" fill="#ffd700"/>
        <circle cx="55" cy="100" r="5" fill="#ff4444"/><circle cx="120" cy="130" r="5" fill="#4488ff"/>
        <circle cx="75" cy="160" r="5" fill="#ffaa00"/><circle cx="105" cy="180" r="5" fill="#ff4444"/>
        <circle cx="60" cy="195" r="4" fill="#4488ff"/><circle cx="130" cy="155" r="4" fill="#ffaa00"/>
        <circle cx="45" cy="130" r="4" fill="#ff4444"/><circle cx="90" cy="90" r="4" fill="#4488ff"/>
      </svg>
      <!-- Árbol esquina derecha -->
      <svg width="140" height="220" viewBox="0 0 180 280" style="position:absolute;bottom:0;right:0;opacity:.14;transform:scaleX(-1);" xmlns="http://www.w3.org/2000/svg">
        <polygon points="90,10 160,120 120,120 150,200 100,200 110,260 80,260 70,200 30,200 60,120 20,120" fill="#2d7a2d"/>
        <rect x="75" y="255" width="30" height="25" fill="#5c3a1a" rx="3"/>
        <circle cx="90" cy="8" r="8" fill="#ffd700"/>
        <circle cx="55" cy="100" r="5" fill="#ff4444"/><circle cx="120" cy="130" r="5" fill="#4488ff"/>
        <circle cx="75" cy="160" r="5" fill="#ffaa00"/><circle cx="105" cy="180" r="5" fill="#ff4444"/>
      </svg>
      <!-- Guirnalda superior -->
      <svg width="100%" height="60" style="position:absolute;top:0;left:0;" xmlns="http://www.w3.org/2000/svg" preserveAspectRatio="none">
        <path d="M0,15 Q12,28 24,15 Q36,2 48,15 Q60,28 72,15 Q84,2 96,15 Q108,28 120,15 Q132,2 144,15 Q156,28 168,15 Q180,2 192,15 Q204,28 216,15 Q228,2 240,15 Q252,28 264,15 Q276,2 288,15 Q300,28 312,15 Q324,2 336,15 Q348,28 360,15 Q372,2 384,15 Q396,28 408,15 Q420,2 432,15 Q444,28 456,15 Q468,2 480,15 Q492,28 504,15 Q516,2 528,15 Q540,28 552,15 Q564,2 576,15 Q588,28 600,15" stroke="#2d7a2d" stroke-width="4" fill="none" opacity="0.5"/>
        <circle cx="24" cy="15" r="4" fill="#ff4444" opacity="0.7"/><circle cx="72" cy="15" r="4" fill="#ffd700" opacity="0.7"/>
        <circle cx="120" cy="15" r="4" fill="#4488ff" opacity="0.7"/><circle cx="168" cy="15" r="4" fill="#ff4444" opacity="0.7"/>
        <circle cx="216" cy="15" r="4" fill="#ffd700" opacity="0.7"/><circle cx="264" cy="15" r="4" fill="#4488ff" opacity="0.7"/>
        <circle cx="312" cy="15" r="4" fill="#ff4444" opacity="0.7"/><circle cx="360" cy="15" r="4" fill="#ffd700" opacity="0.7"/>
        <circle cx="408" cy="15" r="4" fill="#4488ff" opacity="0.7"/><circle cx="456" cy="15" r="4" fill="#ff4444" opacity="0.7"/>
        <circle cx="504" cy="15" r="4" fill="#ffd700" opacity="0.7"/><circle cx="552" cy="15" r="4" fill="#4488ff" opacity="0.7"/>
      </svg>`
  },
  anonuevo: {
    colors: {'--black':'#000510','--surface':'#050a18','--border':'#1a1a4a','--gold':'#ffe066','--text':'#e8e8ff','--white':'#ffffff'},
    bodyClass: 'tema-anonuevo',
    overlay: `
      <svg width="100%" height="100%" xmlns="http://www.w3.org/2000/svg" style="position:absolute;inset:0;">
        <defs>
          <radialGradient id="sky" cx="50%" cy="30%" r="80%">
            <stop offset="0%" stop-color="#050a28" stop-opacity="0.6"/>
            <stop offset="100%" stop-color="#000510" stop-opacity="0.9"/>
          </radialGradient>
        </defs>
        <rect width="100%" height="100%" fill="url(#sky)"/>
        <!-- Estrellas fijas -->
        <circle cx="5%" cy="8%" r="1.5" fill="white" opacity="0.8"/>
        <circle cx="12%" cy="20%" r="1" fill="white" opacity="0.6"/>
        <circle cx="20%" cy="5%" r="2" fill="white" opacity="0.9"/>
        <circle cx="32%" cy="15%" r="1" fill="white" opacity="0.7"/>
        <circle cx="45%" cy="3%" r="1.5" fill="white" opacity="0.8"/>
        <circle cx="55%" cy="18%" r="1" fill="white" opacity="0.6"/>
        <circle cx="67%" cy="7%" r="2" fill="white" opacity="0.9"/>
        <circle cx="78%" cy="12%" r="1" fill="white" opacity="0.7"/>
        <circle cx="88%" cy="4%" r="1.5" fill="white" opacity="0.8"/>
        <circle cx="95%" cy="22%" r="1" fill="white" opacity="0.6"/>
        <circle cx="15%" cy="35%" r="1" fill="#ffe066" opacity="0.5"/>
        <circle cx="40%" cy="28%" r="1.5" fill="#ffe066" opacity="0.4"/>
        <circle cx="72%" cy="30%" r="1" fill="#ffe066" opacity="0.5"/>
        <circle cx="90%" cy="38%" r="1.5" fill="#ffe066" opacity="0.4"/>
      </svg>
      <!-- Fuegos artificiales animados -->
      <div id="fuegos-wrap" style="position:absolute;inset:0;overflow:hidden;pointer-events:none;"></div>
      <!-- Luna / reloj medianoche -->
      <svg width="90" height="90" viewBox="0 0 90 90" style="position:absolute;top:20px;right:30px;opacity:.25;" xmlns="http://www.w3.org/2000/svg">
        <circle cx="45" cy="45" r="40" fill="none" stroke="#ffe066" stroke-width="2"/>
        <circle cx="45" cy="45" r="35" fill="none" stroke="#ffe066" stroke-width="0.5" opacity="0.5"/>
        <line x1="45" y1="10" x2="45" y2="45" stroke="#ffe066" stroke-width="2" stroke-linecap="round"/>
        <line x1="45" y1="45" x2="70" y2="45" stroke="#ffe066" stroke-width="1.5" stroke-linecap="round"/>
        <circle cx="45" cy="45" r="3" fill="#ffe066"/>
        <text x="45" y="78" text-anchor="middle" font-size="8" fill="#ffe066" font-family="Oswald">FELIZ AÑO</text>
      </svg>`
  },
  touchofpink: {
    colors: {
      '--black':      '#080005',
      '--surface':    '#140010',
      '--border':     '#3d1035',
      '--gold':       '#f472b6',
      '--gold-light': '#fbb6ce',
      '--gold-dim':   '#9d174d',
      '--text':       '#ffe0f0',
      '--white':      '#ffffff'
    },
    bodyClass: 'tema-touchofpink',
    overlay: `
      <svg width="100%" height="100%" xmlns="http://www.w3.org/2000/svg" style="position:absolute;inset:0;pointer-events:none;">
        <defs>
          <radialGradient id="pinkglow" cx="50%" cy="50%" r="70%">
            <stop offset="0%" stop-color="#3d0028" stop-opacity="0.45"/>
            <stop offset="100%" stop-color="#080005" stop-opacity="0.95"/>
          </radialGradient>
        </defs>
        <rect width="100%" height="100%" fill="url(#pinkglow)"/>
      </svg>
      <!-- Rosa izquierda -->
      <svg width="220" height="230" viewBox="0 0 220 230" style="position:absolute;bottom:0;left:0;opacity:.22;pointer-events:none;" xmlns="http://www.w3.org/2000/svg">
        <g transform="translate(75,155)">
          <ellipse cx="0" cy="-42" rx="20" ry="38" transform="rotate(0)"   fill="#f472b6"/>
          <ellipse cx="0" cy="-42" rx="20" ry="38" transform="rotate(45)"  fill="#ec4899"/>
          <ellipse cx="0" cy="-42" rx="20" ry="38" transform="rotate(90)"  fill="#f472b6"/>
          <ellipse cx="0" cy="-42" rx="20" ry="38" transform="rotate(135)" fill="#ec4899"/>
          <ellipse cx="0" cy="-42" rx="20" ry="38" transform="rotate(180)" fill="#f472b6"/>
          <ellipse cx="0" cy="-42" rx="20" ry="38" transform="rotate(225)" fill="#ec4899"/>
          <ellipse cx="0" cy="-42" rx="20" ry="38" transform="rotate(270)" fill="#f472b6"/>
          <ellipse cx="0" cy="-42" rx="20" ry="38" transform="rotate(315)" fill="#ec4899"/>
          <circle cx="0" cy="0" r="16" fill="#fbb6ce"/>
        </g>
        <line x1="75" y1="155" x2="50" y2="230" stroke="#4a7a40" stroke-width="3" opacity="0.5"/>
        <ellipse cx="64" cy="196" rx="14" ry="7" fill="#4a7a40" transform="rotate(-30,64,196)" opacity="0.4"/>
      </svg>
      <!-- Rosa derecha (más pequeña) -->
      <svg width="170" height="190" viewBox="0 0 220 230" style="position:absolute;bottom:0;right:0;opacity:.16;pointer-events:none;transform:scaleX(-1);" xmlns="http://www.w3.org/2000/svg">
        <g transform="translate(75,170)">
          <ellipse cx="0" cy="-36" rx="17" ry="32" transform="rotate(0)"   fill="#fbb6ce"/>
          <ellipse cx="0" cy="-36" rx="17" ry="32" transform="rotate(60)"  fill="#f9a8d4"/>
          <ellipse cx="0" cy="-36" rx="17" ry="32" transform="rotate(120)" fill="#fbb6ce"/>
          <ellipse cx="0" cy="-36" rx="17" ry="32" transform="rotate(180)" fill="#f9a8d4"/>
          <ellipse cx="0" cy="-36" rx="17" ry="32" transform="rotate(240)" fill="#fbb6ce"/>
          <ellipse cx="0" cy="-36" rx="17" ry="32" transform="rotate(300)" fill="#f9a8d4"/>
          <circle cx="0" cy="0" r="13" fill="#ffe0f0"/>
        </g>
        <line x1="75" y1="170" x2="95" y2="230" stroke="#4a7a40" stroke-width="2.5" opacity="0.4"/>
      </svg>
      <!-- Dress code watermark -->
      <div style="position:absolute;bottom:16px;left:50%;transform:translateX(-50%);font-family:'Oswald',sans-serif;font-size:11px;letter-spacing:7px;color:rgba(244,114,182,0.3);text-transform:uppercase;white-space:nowrap;pointer-events:none;">DRESS CODE · TURNS PINK</div>
      <!-- Pétalos animados -->
      <div id="petalos-wrap" style="position:absolute;inset:0;overflow:hidden;pointer-events:none;"></div>`
  },
  halloween: {
    colors: {'--black':'#050200','--surface':'#0a0500','--border':'#3a1a00','--gold':'#ff8c00','--gold-light':'#ffaa00','--gold-dim':'#7a4000','--text':'#f0d0a0','--white':'#fff0e0'},
    bodyClass: 'tema-halloween',
    overlay: `
      <!-- Niebla de fondo -->
      <svg width="100%" height="100%" xmlns="http://www.w3.org/2000/svg" style="position:absolute;inset:0;">
        <defs>
          <radialGradient id="fog" cx="50%" cy="100%" r="80%">
            <stop offset="0%" stop-color="#1a0a00" stop-opacity="0.5"/>
            <stop offset="100%" stop-color="#050200" stop-opacity="0.9"/>
          </radialGradient>
        </defs>
        <rect width="100%" height="100%" fill="url(#fog)"/>
        <!-- Luna llena -->
        <circle cx="85%" cy="12%" r="55" fill="#ff8c00" opacity="0.08"/>
        <circle cx="85%" cy="12%" r="48" fill="#ffaa00" opacity="0.06"/>
        <circle cx="85%" cy="12%" r="40" fill="#fff0c0" opacity="0.1"/>
      </svg>
      <!-- Telarañas SVG esquinas -->
      <svg width="260" height="200" viewBox="0 0 260 200" style="position:absolute;top:0;left:0;opacity:.35;" xmlns="http://www.w3.org/2000/svg">
        <g stroke="#888" stroke-width="0.8" fill="none" opacity="0.9">
          <!-- Radio lines from top-left corner -->
          <line x1="0" y1="0" x2="180" y2="0"/>
          <line x1="0" y1="0" x2="150" y2="50"/>
          <line x1="0" y1="0" x2="110" y2="90"/>
          <line x1="0" y1="0" x2="60" y2="130"/>
          <line x1="0" y1="0" x2="0" y2="180"/>
          <!-- Arcos concéntricos -->
          <path d="M40,0 Q20,20 0,40"/>
          <path d="M90,0 Q55,35 20,55 Q0,65 0,90"/>
          <path d="M140,0 Q95,45 55,75 Q25,100 0,140"/>
          <path d="M190,0 Q140,50 95,95 Q50,135 0,175"/>
        </g>
        <!-- Araña -->
        <g transform="translate(62,62)">
          <circle cx="0" cy="0" r="8" fill="#222" stroke="#555" stroke-width="0.5"/>
          <circle cx="0" cy="0" r="4" fill="#111"/>
          <circle cx="3" cy="-2" r="1.5" fill="#ff0000" opacity="0.6"/>
          <circle cx="-1" cy="-2" r="1.5" fill="#ff0000" opacity="0.6"/>
          <!-- Patas -->
          <line x1="-8" y1="-3" x2="-18" y2="-10" stroke="#444" stroke-width="1"/>
          <line x1="-8" y1="0" x2="-18" y2="0" stroke="#444" stroke-width="1"/>
          <line x1="-8" y1="3" x2="-18" y2="10" stroke="#444" stroke-width="1"/>
          <line x1="8" y1="-3" x2="18" y2="-10" stroke="#444" stroke-width="1"/>
          <line x1="8" y1="0" x2="18" y2="0" stroke="#444" stroke-width="1"/>
          <line x1="8" y1="3" x2="18" y2="10" stroke="#444" stroke-width="1"/>
          <!-- Hilo -->
          <line x1="0" y1="-8" x2="0" y2="-40" stroke="#555" stroke-width="0.6"/>
        </g>
      </svg>
      <!-- Telaraña esquina derecha (espejada) -->
      <svg width="220" height="180" viewBox="0 0 260 200" style="position:absolute;top:0;right:0;opacity:.3;transform:scaleX(-1);" xmlns="http://www.w3.org/2000/svg">
        <g stroke="#777" stroke-width="0.8" fill="none">
          <line x1="0" y1="0" x2="180" y2="0"/>
          <line x1="0" y1="0" x2="150" y2="50"/>
          <line x1="0" y1="0" x2="110" y2="90"/>
          <line x1="0" y1="0" x2="60" y2="130"/>
          <line x1="0" y1="0" x2="0" y2="180"/>
          <path d="M40,0 Q20,20 0,40"/>
          <path d="M90,0 Q55,35 20,55 Q0,65 0,90"/>
          <path d="M140,0 Q95,45 55,75 Q25,100 0,140"/>
        </g>
      </svg>
      <!-- Murciélagos animados -->
      <div id="murcielagos-wrap" style="position:absolute;inset:0;overflow:hidden;pointer-events:none;"></div>
      <!-- Calabazas decorativas abajo -->
      <svg width="100%" height="80" style="position:absolute;bottom:0;left:0;" xmlns="http://www.w3.org/2000/svg" viewBox="0 0 800 80" preserveAspectRatio="xMidYMax meet">
        <!-- Calabaza izq -->
        <g transform="translate(30,10)" opacity="0.25">
          <ellipse cx="25" cy="40" rx="22" ry="28" fill="#c84b00"/>
          <ellipse cx="25" cy="40" rx="16" ry="28" fill="#e05500"/>
          <ellipse cx="25" cy="40" rx="10" ry="28" fill="#c84b00"/>
          <rect x="22" y="8" width="6" height="12" rx="3" fill="#3a6a00"/>
          <!-- Cara -->
          <polygon points="17,32 14,38 20,38" fill="#050200"/>
          <polygon points="33,32 30,38 36,38" fill="#050200"/>
          <path d="M16,46 Q25,55 34,46" stroke="#050200" stroke-width="2" fill="none"/>
          <!-- Ojos brillan naranja -->
          <polygon points="17,32 14,38 20,38" fill="#ff8c00" opacity="0.5"/>
          <polygon points="33,32 30,38 36,38" fill="#ff8c00" opacity="0.5"/>
        </g>
        <!-- Calabaza derecha -->
        <g transform="translate(730,5)" opacity="0.2">
          <ellipse cx="25" cy="45" rx="26" ry="32" fill="#c84b00"/>
          <ellipse cx="25" cy="45" rx="18" ry="32" fill="#e05500"/>
          <ellipse cx="25" cy="45" rx="10" ry="32" fill="#c84b00"/>
          <rect x="22" y="10" width="6" height="14" rx="3" fill="#3a6a00"/>
          <polygon points="17,36 13,44 21,44" fill="#050200"/>
          <polygon points="33,36 29,44 37,44" fill="#050200"/>
          <path d="M15,52 Q25,63 35,52" stroke="#050200" stroke-width="2" fill="none"/>
        </g>
      </svg>`
  }
};

function aplicarTema(nombre) {
  const tema = TEMAS[nombre];
  if (!tema) return;

  // Limpiar clases anteriores
  document.body.classList.remove('tema-fullblack','tema-navidad','tema-anonuevo','tema-halloween','tema-touchofpink','pink-claro');
  pinkModoClaro = false;

  // Restaurar colores default primero
  if (nombre === 'default') {
    COLOR_DEFS.forEach(c => document.documentElement.style.setProperty(c.key, c.default));
    document.documentElement.style.setProperty('--surface-gold', '#0d0b00');
    customColors = {};
    COLOR_DEFS.forEach(c => customColors[c.key] = c.default);
  } else {
    // Aplicar colores del tema
    COLOR_DEFS.forEach(c => {
      const v = tema.colors[c.key] || c.default;
      document.documentElement.style.setProperty(c.key, v);
      customColors[c.key] = v;
    });
    if (tema.colors['--surface']) {
      document.documentElement.style.setProperty('--surface-gold', blendSurfaceGold(tema.colors['--surface']));
    }
    if (tema.bodyClass) document.body.classList.add(tema.bodyClass);
  }

  // Overlay decorativo
  const overlay = document.getElementById('tema-overlay');
  if (overlay) { overlay.innerHTML = tema.overlay || ''; overlay.style.opacity = tema.overlay ? '1' : '0'; }

  // Iniciar animaciones específicas
  if (nombre === 'navidad') iniciarCopos();
  if (nombre === 'anonuevo') iniciarFuegos();
  if (nombre === 'halloween') iniciarMurcielagos();
  if (nombre === 'touchofpink') iniciarPetalos();

  buildColorGrid();
  showToast('Tema ' + nombre.toUpperCase() + ' aplicado');
  try { localStorage.setItem('rankingVIP_tema', nombre); } catch(e) {}
  fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({tema:nombre,colores:customColors})}).catch(()=>{});
}

function iniciarCopos() {
  const wrap = document.getElementById('copos-wrap');
  if (!wrap) return;
  wrap.innerHTML = '';
  const copoSVGs = ['❄','❅','❆','✻','✼'];
  for (let i = 0; i < 40; i++) {
    const el = document.createElement('div');
    const sym = copoSVGs[Math.floor(Math.random() * copoSVGs.length)];
    el.textContent = sym;
    el.style.cssText = `position:absolute;top:-30px;left:${Math.random()*100}vw;font-size:${10+Math.random()*18}px;color:rgba(255,255,255,${0.3+Math.random()*0.4});animation:copoFall ${5+Math.random()*8}s linear ${Math.random()*8}s infinite;pointer-events:none;`;
    wrap.appendChild(el);
  }
  // Agregar keyframes si no existen
  if (!document.getElementById('kf-copo')) {
    const s = document.createElement('style');
    s.id = 'kf-copo';
    s.textContent = '@keyframes copoFall{0%{transform:translateY(-30px) rotate(0deg);opacity:1}100%{transform:translateY(105vh) rotate(360deg);opacity:0}}';
    document.head.appendChild(s);
  }
}

function iniciarFuegos() {
  const wrap = document.getElementById('fuegos-wrap');
  if (!wrap) return;
  wrap.innerHTML = '';
  const colores = ['#ffe066','#ff4488','#44aaff','#ff8844','#aaffaa','#ff44ff'];
  function lanzarFuego() {
    if (!document.getElementById('fuegos-wrap')) return;
    const x = 10 + Math.random() * 80;
    const y = 5 + Math.random() * 50;
    const color = colores[Math.floor(Math.random() * colores.length)];
    const burst = document.createElement('div');
    burst.style.cssText = `position:absolute;left:${x}%;top:${y}%;pointer-events:none;`;
    for (let i = 0; i < 14; i++) {
      const spark = document.createElement('div');
      const angle = (i / 14) * 360;
      const dist = 30 + Math.random() * 40;
      spark.style.cssText = `position:absolute;width:3px;height:3px;border-radius:50%;background:${color};box-shadow:0 0 4px ${color};animation:spark ${0.8+Math.random()*0.4}s ease-out forwards;--dx:${Math.cos(angle*Math.PI/180)*dist}px;--dy:${Math.sin(angle*Math.PI/180)*dist}px;`;
      burst.appendChild(spark);
    }
    wrap.appendChild(burst);
    setTimeout(() => burst.remove(), 1400);
    setTimeout(lanzarFuego, 800 + Math.random() * 2000);
  }
  if (!document.getElementById('kf-spark')) {
    const s = document.createElement('style');
    s.id = 'kf-spark';
    s.textContent = '@keyframes spark{0%{opacity:1;transform:translate(0,0)}100%{opacity:0;transform:translate(var(--dx),var(--dy))}}';
    document.head.appendChild(s);
  }
  lanzarFuego(); lanzarFuego(); lanzarFuego();
}

function iniciarMurcielagos() {
  const wrap = document.getElementById('murcielagos-wrap');
  if (!wrap) return;
  wrap.innerHTML = '';
  if (!document.getElementById('kf-bat')) {
    const s = document.createElement('style');
    s.id = 'kf-bat';
    s.textContent = `@keyframes batFly{0%{transform:translateX(-80px) translateY(0)}100%{transform:translateX(110vw) translateY(var(--dy))}}
    @keyframes batWing{0%,100%{transform:scaleY(1)}50%{transform:scaleY(-0.3)}}`;
    document.head.appendChild(s);
  }
  function lanzarMurcie() {
    if (!document.getElementById('murcielagos-wrap')) return;
    const el = document.createElement('div');
    const y = 5 + Math.random() * 60;
    const dur = 6 + Math.random() * 8;
    const dy = (Math.random() - 0.5) * 200;
    const sz = 18 + Math.random() * 20;
    el.style.cssText = `position:absolute;top:${y}%;left:-80px;font-size:${sz}px;animation:batFly ${dur}s linear forwards;--dy:${dy}px;opacity:0.5;`;
    el.textContent = '🦇';
    wrap.appendChild(el);
    setTimeout(() => el.remove(), dur * 1000);
    setTimeout(lanzarMurcie, 2000 + Math.random() * 5000);
  }
  lanzarMurcie(); setTimeout(lanzarMurcie, 2000); setTimeout(lanzarMurcie, 4000);
}

function iniciarPetalos() {
  const wrap = document.getElementById('petalos-wrap');
  if (!wrap) return;
  wrap.innerHTML = '';
  const shapes = ['🌸','🌸','🌺','🌷','💮'];
  for (let i = 0; i < 35; i++) {
    const el = document.createElement('div');
    const sym = shapes[Math.floor(Math.random() * shapes.length)];
    el.textContent = sym;
    el.style.cssText = `position:absolute;top:-40px;left:${Math.random()*100}vw;font-size:${12+Math.random()*16}px;opacity:${0.25+Math.random()*0.45};animation:copoFall ${6+Math.random()*9}s linear ${Math.random()*8}s infinite;pointer-events:none;`;
    wrap.appendChild(el);
  }
  if (!document.getElementById('kf-copo')) {
    const s = document.createElement('style');
    s.id = 'kf-copo';
    s.textContent = '@keyframes copoFall{0%{transform:translateY(-40px) rotate(0deg);opacity:1}100%{transform:translateY(105vh) rotate(360deg);opacity:0}}';
    document.head.appendChild(s);
  }
}

// cargarTemaGuardado redefined below

function buildColorGrid() {
  const grid = document.getElementById('color-grid');
  if (!grid) return;
  grid.innerHTML = COLOR_DEFS.map(c => {
    const val = customColors[c.key];
    const hexId = 'hex-' + c.key.replace(/--/g,'').replace(/-/g,'_');
    const inputId = 'inp-' + c.key.replace(/--/g,'').replace(/-/g,'_');
    const swatchId = 'sw-' + c.key.replace(/--/g,'').replace(/-/g,'_');
    return `<div class="color-item">
      <div class="color-swatch" id="${swatchId}" style="background:${val}">
        <input type="color" id="${inputId}" value="${val}" data-key="${c.key}" oninput="onColorChange(this)" />
      </div>
      <div style="flex:1">
        <div class="color-label">${c.label}</div>
        <div class="color-hex" id="${hexId}">${val}</div>
      </div>
      <button onclick="resetColorSingle('${c.key}')" title="Restablecer este color" style="background:none;border:1px solid #2a2a2a;color:#444;border-radius:5px;padding:4px 8px;font-size:11px;cursor:pointer;font-family:'Rajdhani',sans-serif;letter-spacing:1px;transition:all .15s;white-space:nowrap;" onmouseover="this.style.borderColor='#c9a227';this.style.color='#c9a227'" onmouseout="this.style.borderColor='#2a2a2a';this.style.color='#444'">↺ Default</button>
    </div>`;
  }).join('');
}

function resetColorSingle(key) {
  const def = COLOR_DEFS.find(c => c.key === key);
  if (!def) return;
  customColors[key] = def.default;
  document.documentElement.style.setProperty(key, def.default);
  // Update swatch bg
  const swatchId = 'sw-' + key.replace(/--/g,'').replace(/-/g,'_');
  const sw = document.getElementById(swatchId);
  if (sw) sw.style.background = def.default;
  // Update input value
  const inputId = 'inp-' + key.replace(/--/g,'').replace(/-/g,'_');
  const inp = document.getElementById(inputId);
  if (inp) inp.value = def.default;
  // Update hex label
  const hexId = 'hex-' + key.replace(/--/g,'').replace(/-/g,'_');
  const hexEl = document.getElementById(hexId);
  if (hexEl) hexEl.textContent = def.default;
  // Special: update preview gold
  if (key === '--gold') document.getElementById('prev-vip').style.color = def.default;
  if (key === '--surface') {
    const goldTint = blendSurfaceGold(def.default);
    document.documentElement.style.setProperty('--surface-gold', goldTint);
  }
  showToast('Color restablecido al default');
}

function blendSurfaceGold(surfaceHex) {
  // Mix the surface color with a subtle golden tint for rank-1
  try {
    const r = parseInt(surfaceHex.slice(1,3),16);
    const g = parseInt(surfaceHex.slice(3,5),16);
    const b = parseInt(surfaceHex.slice(5,7),16);
    // Add slight golden warmth: shift toward gold (#c9a227)
    const nr = Math.min(255, Math.round(r * 0.85 + 0xc9 * 0.15));
    const ng = Math.min(255, Math.round(g * 0.88 + 0xa2 * 0.12));
    const nb = Math.min(255, Math.round(b * 0.95 + 0x27 * 0.05));
    return '#' + nr.toString(16).padStart(2,'0') + ng.toString(16).padStart(2,'0') + nb.toString(16).padStart(2,'0');
  } catch(e) { return surfaceHex; }
}

function onColorChange(el) {
  const key = el.dataset.key;
  const val = el.value;
  customColors[key] = val;
  el.parentElement.style.background = val;
  const hexId = 'hex-' + key.replace(/--/g,'').replace(/-/g,'_');
  const hexEl = document.getElementById(hexId);
  if (hexEl) hexEl.textContent = val;
  // Aplicar en tiempo real
  document.documentElement.style.setProperty(key, val);
  if (key === '--surface') {
    const goldTint = blendSurfaceGold(val);
    document.documentElement.style.setProperty('--surface-gold', goldTint);
  }
  if (key === '--gold') {
    document.getElementById('prev-vip').style.color = val;
  }
  // Sincronizar con /pantalla inmediatamente
  clearTimeout(onColorChange._t);
  onColorChange._t = setTimeout(()=>{
    fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({colores:customColors})}).catch(()=>{});
  }, 300);
}

function previewTextos() {
  const vip  = document.getElementById('ct-vip').value  || 'VIP';
  document.getElementById('prev-vip').textContent  = vip;
  previewTagline();
}
function previewPremio() {
  const txt = document.getElementById('ct-premio') ? document.getElementById('ct-premio').value : '';
  const sz  = document.getElementById('ct-premio-size') ? document.getElementById('ct-premio-size').value : '22';
  const pb  = document.getElementById('premio-box');
  if(pb){ pb.textContent=txt; pb.style.fontSize=sz+'px'; }
  const sv  = document.getElementById('ct-premio-size-val');
  if(sv) sv.textContent=sz+'px';
}
function hexToRgb(hex){
  const h=hex.replace('#','');
  return {r:parseInt(h.slice(0,2),16),g:parseInt(h.slice(2,4),16),b:parseInt(h.slice(4,6),16)};
}
function buildTaglineShadow(color, glow) {
  const g=parseFloat(glow||0);
  if(g<=0) return 'none';
  const {r,g:gr,b}=hexToRgb(color||'#555555');
  return `0 0 14px rgba(${r},${gr},${b},${g}), 0 0 30px rgba(${r},${gr},${b},${(g*0.6).toFixed(2)}), 0 0 60px rgba(${r},${gr},${b},${(g*0.3).toFixed(2)})`;
}
function applyTaglineStyle(el, text, color, glow, font) {
  if (!el) return;
  if (temaActual === 'touchofpink') {
    const kids = el.querySelectorAll('div');
    kids.forEach(d => d.style.fontFamily = font);
    return;
  }
  el.textContent = text;
  el.style.color = color;
  el.style.filter = 'none';
  el.style.textShadow = buildTaglineShadow(color, glow);
  el.style.fontFamily = font;
  el.style.fontSize = '28px';
  el.style.fontWeight = '600';
  el.style.letterSpacing = '5px';
  el.style.textTransform = 'uppercase';
  el.style.textAlign = 'center';
}
function previewTagline() {
  const text = (document.getElementById('ct-tagline')||{}).value || 'JAGGER CLUB';
  const color = (document.getElementById('ct-tagline-color')||{}).value || '#555555';
  const glow = (document.getElementById('ct-tagline-brightness')||{}).value || '0';
  const font = (document.getElementById('ct-tagline-font')||{}).value || "'Rajdhani',sans-serif";
  const pv = document.getElementById('prev-club');
  if (pv) { pv.textContent=text; pv.style.color=color; pv.style.textShadow=buildTaglineShadow(color,glow); pv.style.fontFamily=font; }
  applyTaglineStyle(document.getElementById('tema-tagline'), text, color, glow, font);
}

function aplicarPersonalizacion() {
  // Colores
  COLOR_DEFS.forEach(c => {
    document.documentElement.style.setProperty(c.key, customColors[c.key]);
  });
  // Sync surface-gold tint
  if (customColors['--surface']) {
    document.documentElement.style.setProperty('--surface-gold', blendSurfaceGold(customColors['--surface']));
  }
  // Textos logo
  const vip  = document.getElementById('ct-vip').value  || 'VIP';
  const _lv = document.getElementById('logo-vip'); if(_lv) _lv.textContent = vip;
  // Tagline
  const taglineText = document.getElementById('ct-tagline').value || 'JAGGER CLUB';
  const taglineColor = document.getElementById('ct-tagline-color').value;
  const taglineGlow = document.getElementById('ct-tagline-brightness').value;
  const taglineFont = document.getElementById('ct-tagline-font').value;
  applyTaglineStyle(document.getElementById('tema-tagline'), taglineText, taglineColor, taglineGlow, taglineFont);
  // Guardar en localStorage para persistir
  const wm = document.getElementById('ct-winner-msg').value;
  const ws = document.getElementById('ct-winner-sub').value;
  try {
    localStorage.setItem('rankingVIP_colors', JSON.stringify(customColors));
    localStorage.setItem('rankingVIP_vip', vip);
    localStorage.setItem('rankingVIP_wmsg', wm);
    localStorage.setItem('rankingVIP_wsub', ws);
    localStorage.setItem('rankingVIP_tagline_text', taglineText);
    localStorage.setItem('rankingVIP_tagline_color', taglineColor);
    localStorage.setItem('rankingVIP_tagline_glow', taglineGlow);
    localStorage.setItem('rankingVIP_tagline_font', taglineFont);
  } catch(e){}
  const premioVal = document.getElementById('ct-premio') ? document.getElementById('ct-premio').value : '';
  const premioSize = document.getElementById('ct-premio-size') ? document.getElementById('ct-premio-size').value : '22';
  const tipoParticula = document.querySelector('input[name="tipo-particula"]:checked') ? document.querySelector('input[name="tipo-particula"]:checked').value : 'confetti';
  fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({
    colores: customColors,
    tema: temaActual || 'default',
    efecto: efectoActual || 'ninguno',
    vip: vip,
    tagline: taglineText,
    tagline_color: taglineColor,
    tagline_glow: taglineGlow,
    tagline_font: taglineFont,
    winner_msg: wm,
    winner_sub: ws,
    premio: premioVal,
    premio_size: premioSize,
    tipo_particula: tipoParticula,
  })}).then(()=>{ showToastConfirm('✅ Cambios confirmados en todas las pantallas'); }).catch(()=>{ showToast('Error al sincronizar', true); });
}

function resetPersonalizacion() {
  if (!confirm('Restaurar todos los colores y textos al default?')) return;
  COLOR_DEFS.forEach(c => {
    customColors[c.key] = c.default;
    document.documentElement.style.setProperty(c.key, c.default);
  });
  document.documentElement.style.setProperty('--surface-gold', '#0d0b00');
  document.getElementById('ct-vip').value  = 'VIP';
  const _rlv = document.getElementById('logo-vip'); if(_rlv) _rlv.textContent = 'VIP';
  document.getElementById('ct-tagline').value = 'JAGGER CLUB';
  document.getElementById('ct-tagline-color').value = '#555555';
  document.getElementById('ct-tagline-brightness').value = '0';
  document.getElementById('ct-tagline-brightness-val').textContent = '0.00';
  applyTaglineStyle(document.getElementById('tema-tagline'),'JAGGER CLUB','#555','0',"'Rajdhani',sans-serif");
  try { localStorage.removeItem('rankingVIP_colors'); localStorage.removeItem('rankingVIP_club'); localStorage.removeItem('rankingVIP_vip'); localStorage.removeItem('rankingVIP_tagline_text'); localStorage.removeItem('rankingVIP_tagline_color'); localStorage.removeItem('rankingVIP_tagline_glow'); localStorage.removeItem('rankingVIP_tagline_font'); } catch(e){}
  buildColorGrid();
  previewTextos();
  fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({
    tema:'default', colores:{}, vip:'VIP', tagline:'JAGGER CLUB',
    tagline_color:'#555555', tagline_glow:'0', winner_msg:'¡EL GANADOR DE LA NOCHE!', winner_sub:'',
  })}).catch(()=>{});
  showToast('Colores restaurados');
}

function cargarPersonalizacionGuardada() {
  // Cargar colores del server (fuente de verdad) y también localStorage como fallback
  fetch('/api/design').then(r=>r.json()).then(d=>{
    if(d.colores && Object.keys(d.colores).length>0){
      customColors = Object.assign({}, customColors, d.colores);
      Object.entries(d.colores).forEach(([k,v])=>document.documentElement.style.setProperty(k,v));
      buildColorGrid();
    }
    if(d.tema && d.tema!==temaActual) aplicarTema(d.tema);
  }).catch(()=>{});
  try {
    const savedColors = localStorage.getItem('rankingVIP_colors');
    if (savedColors) {
      const parsed = JSON.parse(savedColors);
      // Solo usar localStorage si el server no tiene colores (se sobreescribe después con fetch)
      COLOR_DEFS.forEach(c => {
        if (parsed[c.key] && !customColors[c.key]) { customColors[c.key]=parsed[c.key]; document.documentElement.style.setProperty(c.key, parsed[c.key]); }
      });
    }
    const club = localStorage.getItem('rankingVIP_club');
    const vip  = localStorage.getItem('rankingVIP_vip');
    if (vip)  { document.getElementById('logo-vip').textContent  = vip;  document.getElementById('ct-vip').value  = vip;  document.getElementById('prev-vip').textContent  = vip; }
    const wm = localStorage.getItem('rankingVIP_wmsg');
    const ws = localStorage.getItem('rankingVIP_wsub');
    if (wm) document.getElementById('ct-winner-msg').value = wm;
    if (ws) document.getElementById('ct-winner-sub').value = ws;
    const tlText  = localStorage.getItem('rankingVIP_tagline_text');
    const tlColor = localStorage.getItem('rankingVIP_tagline_color');
    const tlGlow  = localStorage.getItem('rankingVIP_tagline_glow');
    const tlFont  = localStorage.getItem('rankingVIP_tagline_font');
    if (tlText||tlColor||tlGlow||tlFont) {
      const inp    = document.getElementById('ct-tagline');
      const colInp = document.getElementById('ct-tagline-color');
      const brInp  = document.getElementById('ct-tagline-brightness');
      const brVal  = document.getElementById('ct-tagline-brightness-val');
      const fntInp = document.getElementById('ct-tagline-font');
      if (tlText  && inp)    inp.value    = tlText;
      if (tlColor && colInp) colInp.value = tlColor;
      if (tlGlow  && brInp)  { brInp.value=tlGlow; if(brVal) brVal.textContent=parseFloat(tlGlow).toFixed(2); }
      if (tlFont  && fntInp) fntInp.value = tlFont;
      applyTaglineStyle(
        document.getElementById('tema-tagline'),
        tlText  || 'JAGGER CLUB',
        tlColor || '#555555',
        tlGlow  || '0',
        tlFont  || "'Rajdhani',sans-serif"
      );
    }
  } catch(e){}
}

// ══════════════════════════════════════════
//  ANIMACION GANADOR
// ══════════════════════════════════════════
function generarParticulas(wrapId) {
  const wrap = document.getElementById(wrapId || 'confetti-wrap');
  wrap.innerHTML = '';
  if (tipoParticula === 'ninguno') return;
  if (tipoParticula === 'billetes') {
    for (let i = 0; i < 55; i++) {
      const el = document.createElement('span');
      el.style.position = 'absolute';
      el.style.top = '-60px';
      el.style.left = (Math.random() * 100) + 'vw';
      el.style.fontSize = (20 + Math.random() * 18) + 'px';
      el.style.animationName = 'confettiFall';
      el.style.animationDuration = (4 + Math.random() * 4) + 's';
      el.style.animationDelay = (Math.random() * 5) + 's';
      el.style.animationTimingFunction = 'linear';
      el.style.animationIterationCount = 'infinite';
      el.textContent = '💵';
      wrap.appendChild(el);
    }
  } else {
    const colors = temaActual === 'touchofpink'
      ? ['#f472b6','#fbb6ce','#ffffff','#f9a8d4','#ec4899','#ffffff','#fce7f3','#ff69b4','#fff0f5']
      : ['#c9a227','#e8c84a','#fff','#f0ece0','#2ecc71','#e74c3c','#3498db','#9b59b6','#ff9f43'];
    for (let i = 0; i < 120; i++) {
      const el = document.createElement('div');
      el.className = 'confetti-piece';
      el.style.left = Math.random() * 100 + 'vw';
      el.style.background = colors[Math.floor(Math.random() * colors.length)];
      el.style.width = (6 + Math.random() * 10) + 'px';
      el.style.height = (10 + Math.random() * 16) + 'px';
      el.style.animationDuration = (3 + Math.random() * 5) + 's';
      el.style.animationDelay = (Math.random() * 4) + 's';
      el.style.borderRadius = Math.random() > 0.5 ? '50%' : '2px';
      wrap.appendChild(el);
    }
  }
}


function mostrarGanador() {
  const totals = {}, mesas = {};
  txData.forEach(t => {
    totals[t.name] = (totals[t.name]||0) + t.amount;
    if (t.mesa && !mesas[t.name]) mesas[t.name] = t.mesa;
  });
  const nombres = Object.keys(totals).sort((a,b) => totals[b]-totals[a]);
  if (!nombres.length) { showToast('No hay consumos registrados aun', true); return; }

  const ganador = nombres[0];
  const mesa = mesas[ganador] || '—';
  const total = totals[ganador];

  
  const wmsg = (document.getElementById('ct-winner-msg').value || '¡EL GANADOR DE LA NOCHE!').toUpperCase();
  const wsub = document.getElementById('ct-winner-sub').value || '';
  const premio = document.getElementById('msg-input').value.trim();

  // Icono: boxeo = guantes + trofeo, default = corona
  const coronaEl = document.getElementById('winner-corona');
  if (coronaEl) coronaEl.textContent = '👑';

  const nombreEl = document.getElementById('winner-nombre');
  const totalEl  = document.getElementById('winner-total');

  const _wt=document.getElementById('winner-titulo'); if(_wt) _wt.textContent = wmsg;
  document.getElementById('winner-mesa').textContent = mesa;
  nombreEl.style.borderRight = 'none';
  nombreEl.textContent = ganador.toUpperCase();
  totalEl.textContent  = fmt(total);

  const msgEl = document.getElementById('winner-mensaje');
  const textoFinal = wsub || premio || '';
  msgEl.textContent = textoFinal;
  msgEl.style.display = textoFinal ? 'block' : 'none';

  if (confettiGanadorActivo) generarParticulas();
  document.getElementById('winner-overlay').classList.add('show');

}

async function mostrarGanadorManual() {
  ganadorMostrado = true;
  try {
    await fetch('/api/winner/show',{method:'POST'});
    showToast('🏆 Ganador enviado a pantalla');
  } catch(e) { showToast('Error de conexión', true); }
}

function cerrarGanador() {
  document.getElementById('winner-overlay').classList.remove('show');
  fetch('/api/winner/hide',{method:'POST'});
}

// ══════════════════════════════════════════
//  RELOJ Y DETECCION DE HORA LIMITE
// ══════════════════════════════════════════
function tickClock() {
  const now = new Date();
  const h = String(now.getHours()).padStart(2,'0');
  const m = String(now.getMinutes()).padStart(2,'0');
  const el = document.getElementById('clock-display');
  if (el) el.textContent = h + ':' + m;

  // Detectar hora fin
  if (horaFin && !ganadorMostrado && txData.length > 0) {
    const [fh, fm] = horaFin.split(':').map(Number);
    if (now.getHours() === fh && now.getMinutes() === fm && now.getSeconds() < 30) {
      ganadorMostrado = true;
      mostrarGanador();
      fetch('/api/winner/show',{method:'POST'});
    }
  }
}
setInterval(tickClock, 1000);
tickClock();



// ══════════════════════════════════════════
//  LECTOR DE TARJETA
// ══════════════════════════════════════════
let pendingCodigos = {};
let lastKeyTime = 0;
const READER_SPEED_MS = 80;

document.addEventListener('keydown', (e) => {
  ['caja1','caja2','caja3'].forEach((s,i) => {
    const tab = document.getElementById('tab-'+s);
    if (tab && tab.classList.contains('active')) cajaFocus = i+1;
  });
  const enConfig = document.getElementById('tab-config') && document.getElementById('tab-config').classList.contains('active');
  const active = document.activeElement;

  const now = Date.now();
  const timeSinceLast = now - lastKeyTime;
  const isReaderSpeed = globalBuffer.length > 0 && timeSinceLast < READER_SPEED_MS;

  if (active && active.tagName === 'INPUT' && !active.classList.contains('tc-input')) {
    if (!isReaderSpeed && e.key !== 'Enter') return;
  }

  if (e.key === 'Enter') {
    if (globalBuffer.length > 1) {
      const codigo = globalBuffer.trim();
      if (enConfig && scanSlotActivo !== null) {
        pendingCodigos[scanSlotActivo] = codigo;
        const btn = document.getElementById('scan-btn-'+scanSlotActivo);
        if (btn) { btn.textContent='LEER'; btn.classList.remove('activo'); }
        const codeEl = document.getElementById('tc-code-'+scanSlotActivo);
        if (codeEl) codeEl.textContent = 'Leida: ' + codigo;
        const confirmBtn = document.getElementById('confirm-btn-'+scanSlotActivo);
        if (confirmBtn) confirmBtn.style.display = 'inline-block';
        const slotNum = confTarjetas[scanSlotActivo] ? confTarjetas[scanSlotActivo].slot : '?';
        showToast('Tarjeta leida con exito para Mesa ' + slotNum + ' — presiona CONFIRMAR');
        scanSlotActivo = null;
      } else if (cajaFocus > 0) {
        procesarTarjetaEnCaja(codigo, cajaFocus);
      }
    }
    globalBuffer = '';
    lastKeyTime = 0;
    clearTimeout(globalTimeout);
  } else if (e.key.length === 1) {
    globalBuffer += e.key;
    lastKeyTime = now;
    clearTimeout(globalTimeout);
    globalTimeout = setTimeout(() => { globalBuffer = ''; lastKeyTime = 0; }, 500);
  }
});

function confirmarVinculo(idx) {
  const codigo = pendingCodigos[idx];
  if (!codigo) { showToast('Primero pasa la tarjeta por el lector', true); return; }
  // Verificar duplicado: misma tarjeta ya registrada en otro slot
  const dupIdx = confTarjetas.findIndex((t, i) => i !== idx && t.codigo === codigo);
  if (dupIdx !== -1) {
    showToast('⚠ Esta tarjeta ya está registrada en Mesa ' + confTarjetas[dupIdx].slot + '. No se puede vincular dos veces.', true);
    delete pendingCodigos[idx];
    const codeEl2 = document.getElementById('tc-code-'+idx);
    if (codeEl2) codeEl2.textContent = 'Presiona LEER y pasa la tarjeta';
    const confirmBtn2 = document.getElementById('confirm-btn-'+idx);
    if (confirmBtn2) confirmBtn2.style.display = 'none';
    const scanBtn2 = document.getElementById('scan-btn-'+idx);
    if (scanBtn2) { scanBtn2.textContent = 'LEER'; scanBtn2.classList.remove('activo'); }
    return;
  }
  confTarjetas[idx].codigo = codigo;
  delete pendingCodigos[idx];
  const codeEl = document.getElementById('tc-code-'+idx);
  if (codeEl) codeEl.textContent = 'Cod: ' + codigo;
  const tcEl = document.getElementById('tc-'+idx);
  if (tcEl) tcEl.classList.add('configurada');
  const confirmBtn = document.getElementById('confirm-btn-'+idx);
  if (confirmBtn) confirmBtn.style.display = 'none';
  const statusEl = document.getElementById('tc-status-'+idx);
  if (statusEl) { statusEl.textContent = 'Vinculada'; statusEl.style.color = '#3a9a5a'; }
  showToast('Tarjeta vinculada a Mesa ' + confTarjetas[idx].slot);
}

function procesarTarjetaEnCaja(codigo, caja) {
  const idx = confTarjetas.findIndex(t => t.codigo === codigo);
  if (idx === -1) { showToast('Tarjeta no configurada — vinculala primero en la pestana Tarjetas', true); return; }
  const conf = confTarjetas[idx];
  const mesa = String(conf.slot);
  const saldoInfo = tarjetasData[codigo];
  const saldoActual = (saldoInfo && saldoInfo.saldo_actual !== undefined) ? saldoInfo.saldo_actual : parseFloat(conf.saldo_inicial||0);
  const saldoInicial = parseFloat(conf.saldo_inicial||0);
  const nombre = saldoInfo ? (saldoInfo.nombre||conf.nombre_cliente||'') : (conf.nombre_cliente||'');
  window['cajaState'+caja] = { codigo, mesa, slot: conf.slot, saldo_inicial: saldoInicial, saldo_actual: saldoActual, nombre };
  setModo(caja, 'tarjeta');
  renderCajaInner(caja);
  showToast('Tarjeta leida con exito para Mesa ' + mesa + ' — Saldo: '+fmt(saldoActual));
  setTimeout(() => {
    const el = document.getElementById(nombre ? 'amount'+caja : 'name'+caja);
    if (el) el.focus();
  }, 100);
}

function setModo(caja, modo) {
  ['tarjeta','manual','recargar'].forEach(m => {
    const tab = document.getElementById('modo-tab-'+m+'-'+caja);
    const content = document.getElementById('modo-content-'+m+'-'+caja);
    if (tab) tab.classList.toggle('active', m===modo);
    if (content) content.classList.toggle('active', m===modo);
  });
  setTimeout(() => updateCajaGastoLabel(caja), 0);
}

function showTab(id) {
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
  document.querySelectorAll('.screen').forEach(s => s.classList.remove('active'));
  document.getElementById('tbtn-'+id).classList.add('active');
  document.getElementById('tab-'+id).classList.add('active');
  if (id === 'stats') { setTimeout(renderStats, 50); }
  if (id === 'pantalla') { setTimeout(renderPantalla, 50); }
  if (id === 'menu-mgr') cargarMenuMgr();
  if (id === 'pines') cargarPines2();
  if (id === 'diseno') { if(typeof buildColorGrid==='function') buildColorGrid(); if(typeof cargarTemaGuardado==='function') cargarTemaGuardado(); }
}
function activarPresentacion() { document.body.classList.add('modo-presentacion'); showTab('pantalla'); }
function salirPresentacion() { document.body.classList.remove('modo-presentacion'); }
function updateMsg() {
  const el = document.getElementById('msg-input');
  if (!el) return;
  const val = el.value.trim();
  const pb = document.getElementById('premio-box'); if (pb) pb.textContent = val;
  fetch('/api/state',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({premio:val})});
  fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({premio:val})}).catch(()=>{});
}
function updatePremioSize() {
  const size = document.getElementById('premio-size')?.value;
  const pb = document.getElementById('premio-box');
  if (size && pb) pb.style.fontSize = size;
}

function updateHoraFin() {
  horaFin = document.getElementById('hora-fin-input').value || '05:30';
  ganadorMostrado = false;
  const el = document.getElementById('clock-fin');
  if (el) el.textContent = horaFin;
  fetch('/api/state',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({hora_fin:horaFin})});
  fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({hora_fin:horaFin})}).catch(()=>{});
}function fmt(n) { return '$'+Number(n).toLocaleString('es-AR',{minimumFractionDigits:0,maximumFractionDigits:0}); }
function fmtLabel(n){
  if(n>=1000000){const m=n/1000000;return(m%1===0?m:m.toFixed(1))+(m<2?' MILLÓN':' MILLONES');}
  if(n>=1000){const k=n/1000;return(k%1===0?k:k.toFixed(1))+(k<2?' MIL':' MILES');}
  return '';
}
function fmtDisplay(n){
  const base='$'+Number(n).toLocaleString('es-AR',{minimumFractionDigits:0,maximumFractionDigits:0});
  const lbl=fmtLabel(n);
  return lbl?base+'<span class="miles-lbl">'+lbl+'</span>':base;
}
function showToastConfirm(msg) {
  let el = document.getElementById('_tc_confirm');
  if (!el) {
    el = document.createElement('div');
    el.id = '_tc_confirm';
    el.style.cssText = 'position:fixed;bottom:80px;left:20px;z-index:9999;background:#071507;border:1px solid #2ecc71;border-radius:10px;padding:14px 22px;font-family:"Rajdhani",sans-serif;font-size:15px;font-weight:700;color:#2ecc71;letter-spacing:1px;pointer-events:none;opacity:0;transition:opacity .35s;max-width:340px;display:flex;align-items:center;gap:10px;';
    document.body.appendChild(el);
  }
  el.innerHTML = '✅ ' + msg;
  el.style.opacity = '1';
  clearTimeout(showToastConfirm._t);
  showToastConfirm._t = setTimeout(() => { el.style.opacity = '0'; }, 4000);
}
function showToast(msg, error=false) {
  const t = document.getElementById('toast');
  t.textContent = msg; t.className='toast show'+(error?' error':'');
  setTimeout(()=>t.className='toast', 2800);
}

async function cerrarNoche() {
  if (!confirm('¿Cerrar la noche y guardar en el historial?\nLos datos quedan registrados para siempre.')) return;
  try {
    const r = await fetch('/api/cerrar_noche', {method:'POST'});
    const d = await r.json();
    if (d.ok) {
      showToast('✓ Noche guardada en el historial');
      if (confirm('¿Exportar la noche a Excel ahora?')) window.open('/api/export/excel','_blank');
    } else { showToast(d.error||'Error al cerrar la noche', true); }
  } catch(e) { showToast('Error de conexión', true); }
}

async function resetNoche() {
  if (!confirm('¿Resetear la noche? Todos los consumos se borrarán y los saldos vuelven al valor inicial.')) return;
  if (!confirm('⚠️ SEGUNDA CONFIRMACIÓN: Esta acción no se puede deshacer. ¿Confirmar reseteo?')) return;
  await fetch('/api/reset',{method:'POST'});
  ganadorMostrado = false;
  knownNames.clear();
  for(let c=1;c<=3;c++){window['cajaState'+c]=null;renderCajaInner(c);}
  await loadData();
  showToast('Noche reseteada');
}

function renderCajaInner(caja) {
  const container = document.getElementById('caja-inner-'+caja);
  const ta = window['cajaState'+caja];
  let tarjetaHTML = '';
  if (ta) {
    const techoReal = Math.max(ta.saldo_inicial, ta.saldo_actual);
    const pct = techoReal > 0 ? Math.max(0, Math.min(100, Math.round((ta.saldo_actual/techoReal)*100))) : 0;
    const sinSaldo = ta.saldo_actual <= 0;
    const pctSaldo = techoReal > 0 ? ta.saldo_actual / techoReal : 1;
    const warnSaldo = !sinSaldo && pctSaldo <= 0.2;
    tarjetaHTML = `<div class="tarjeta-card visible ${sinSaldo?'sin-saldo':''}">
      <div class="tarjeta-top">
        <div><div class="tarjeta-mesa-label">Mesa</div><div class="tarjeta-mesa-num">${ta.mesa}</div></div>
        <div class="tarjeta-saldo-wrap"><div class="tarjeta-saldo-label">Saldo disponible</div><div class="tarjeta-saldo">${fmt(ta.saldo_actual)}</div></div>
      </div>
      <div class="tarjeta-bar-wrap"><div class="tarjeta-bar" style="width:${pct}%"></div></div>
      ${ta.nombre?`<div class="tarjeta-nombre">Cliente: <span>${ta.nombre}</span></div>`:''}
      ${warnSaldo?`<div class="saldo-bajo-warn">⚠ Saldo bajo — queda ${Math.round(pctSaldo*100)}%</div>`:''}
      ${sinSaldo?`<div class="saldo-bajo-warn" style="border-color:#a83030;color:#ff4444;animation:none;">✕ Sin saldo disponible</div>`:''}
    </div>`;
  }

  container.innerHTML = `
    <div class="caja-header-row">
      <span class="caja-badge">CAJA ${caja}</span>
      <span class="caja-title">Registrar consumo</span>
      <button onclick="abrirCartelModal(${caja})" style="margin-left:auto;background:linear-gradient(135deg,#c9a227,#e8c84a);color:#000;border:none;border-radius:6px;padding:7px 16px;font-family:'Rajdhani',sans-serif;font-size:12px;font-weight:700;letter-spacing:2px;cursor:pointer;transition:all .15s;white-space:nowrap;" onmouseover="this.style.opacity='0.85'" onmouseout="this.style.opacity='1'">📣 CARTEL</button>
    </div>
    <div style="margin-bottom:12px;">
      <button onclick="this.nextElementSibling.style.display=this.nextElementSibling.style.display==='none'?'block':'none';this.textContent=this.nextElementSibling.style.display==='none'?'▸ Instrucciones rápidas':'▾ Instrucciones rápidas';" style="background:none;border:1px solid #2a2a2a;border-radius:6px;color:#888;font-size:11px;letter-spacing:1px;padding:4px 12px;cursor:pointer;">▸ Instrucciones rápidas</button>
      <div style="display:none;margin-top:8px;background:#0d0d0d;border:1px solid #2a2a2a;border-radius:8px;padding:12px 16px;font-size:12px;line-height:1.8;color:#aaa;">
        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;">
          <div>
            <div style="color:#e8c84a;font-weight:700;letter-spacing:1px;margin-bottom:4px;">▤ COBRAR CON TARJETA</div>
            <div>1. Pasá la tarjeta por el lector</div>
            <div>2. Verificá el saldo disponible</div>
            <div>3. Ingresá el nombre del cliente</div>
            <div>4. Ingresá el monto a cobrar</div>
            <div>5. Presioná <strong style="color:#fff;">+ Agregar</strong></div>
            <div style="color:#888;margin-top:4px;font-size:11px;">El sistema descuenta el monto del saldo de la tarjeta automáticamente.</div>
          </div>
          <div>
            <div style="color:#e8c84a;font-weight:700;letter-spacing:1px;margin-bottom:4px;">✎ COBRAR SIN TARJETA</div>
            <div>1. Seleccioná la pestaña <strong style="color:#fff;">Sin tarjeta</strong></div>
            <div>2. Ingresá el nombre del cliente</div>
            <div>3. Ingresá el número de mesa</div>
            <div>4. Ingresá el monto consumido</div>
            <div>5. Presioná <strong style="color:#fff;">+ Agregar</strong></div>
            <div style="color:#888;margin-top:4px;font-size:11px;">Usá este modo si el cliente no tiene tarjeta VIP o el lector no funciona.</div>
          </div>
          <div>
            <div style="color:#3a9a5a;font-weight:700;letter-spacing:1px;margin-bottom:4px;">⊕ CARGAR SALDO</div>
            <div>1. Seleccioná la pestaña <strong style="color:#fff;">Recargar</strong></div>
            <div>2. Pasá la tarjeta por el lector</div>
            <div>3. Verificá que sea la tarjeta correcta</div>
            <div>4. Ingresá el monto a agregar</div>
            <div>5. Presioná <strong style="color:#fff;">⊕ Recargar saldo</strong></div>
            <div style="color:#888;margin-top:4px;font-size:11px;">El saldo se suma al disponible. Guardá la config. de tarjetas luego.</div>
          </div>
        </div>
      </div>
    </div>
    <div class="modo-tabs">
      <div class="modo-tab ${!ta?'active':''}" id="modo-tab-tarjeta-${caja}" onclick="setModo(${caja},'tarjeta')">▤ Con tarjeta</div>
      <div class="modo-tab" id="modo-tab-manual-${caja}" onclick="setModo(${caja},'manual')">✎ Sin tarjeta</div>
      <div class="modo-tab" id="modo-tab-recargar-${caja}" onclick="setModo(${caja},'recargar')" style="color:#3a9a5a;">⊕ Recargar</div>
    </div>
    <div class="modo-content active" id="modo-content-tarjeta-${caja}">
      <div class="scan-hint ${!ta?'esperando':''}">
        <span class="scan-icon">▤</span>
        ${ta ? 'Tarjeta activa — pasa otra para cambiar' : 'Pasa la tarjeta por el lector para continuar'}
      </div>
      ${tarjetaHTML}
      <div class="form-card" ${!ta?'style="opacity:.35;pointer-events:none"':''}>
        <label class="field-label">Nombre del cliente</label>
        <input class="field-input" id="name${caja}" type="text" placeholder="Nombre..." autocomplete="off" list="nl${caja}" value="${ta&&ta.nombre?ta.nombre:''}" />
        <datalist id="nl${caja}"></datalist>
        <label class="field-label">Monto ($)</label>
        <input class="field-input amount-input" id="amount${caja}" type="number" min="0" step="100" placeholder="0" />
        <div class="hint-miles" id="hint${caja}">Ingresá el monto completo</div>
        <div class="btn-row" style="margin-top:4px">
          <button class="btn-add" id="btnadd${caja}" onclick="addTx(${caja},true)" ${!ta?'disabled':''}>+ Agregar</button>
        </div>
      </div>
    </div>
    <div class="modo-content" id="modo-content-manual-${caja}">
      <div class="form-card">
        <label class="field-label">Nombre del cliente</label>
        <input class="field-input" id="mname${caja}" type="text" placeholder="Nombre..." autocomplete="off" list="mnl${caja}" />
        <datalist id="mnl${caja}"></datalist>
        <label class="field-label">Mesa (opcional)</label>
        <div style="display:flex;gap:8px;align-items:center;position:relative;">
          <input class="field-input" id="mmesa${caja}" type="text" placeholder="Ej: 5" style="flex:1;" />
          <button onclick="toggleMesasPopup(${caja})" id="btn-mesas-popup-${caja}" title="Ver mesas de esta noche" style="background:#0f0e05;border:1px solid #c9a22766;border-radius:8px;padding:10px 13px;font-size:16px;cursor:pointer;color:#c9a227;flex-shrink:0;transition:border-color .15s;" onmouseover="this.style.borderColor='#c9a227'" onmouseout="this.style.borderColor='#c9a22766'">🗂</button>
          <div id="mesas-popup-${caja}" style="display:none;position:absolute;top:100%;right:0;z-index:200;margin-top:6px;background:#111;border:1px solid #c9a22744;border-radius:10px;min-width:200px;max-width:260px;box-shadow:0 8px 32px rgba(0,0,0,0.7);overflow:hidden;">
            <div style="padding:10px 14px 6px;font-family:'Oswald',sans-serif;font-size:11px;color:#c9a227;letter-spacing:2px;text-transform:uppercase;border-bottom:1px solid #222;">Mesas esta noche</div>
            <div id="mesas-popup-list-${caja}" style="max-height:220px;overflow-y:auto;padding:6px 0;"></div>
          </div>
        </div>
        <label class="field-label">Monto ($)</label>
        <input class="field-input amount-input" id="mamount${caja}" type="number" min="0" step="100" placeholder="0" />
        <div class="hint-miles" id="mhint${caja}">Ingresá el monto completo</div>
        <div class="btn-row" style="margin-top:4px">
          <button class="btn-add" onclick="addTx(${caja},false)">+ Agregar</button>
        </div>
      </div>
    </div>
    <div class="modo-content" id="modo-content-recargar-${caja}">
      <div class="scan-hint ${!ta?'esperando':''}">
        <span class="scan-icon">▤</span>
        ${ta ? 'Tarjeta activa — pasa otra para cambiar' : 'Pasa la tarjeta para recargar'}
      </div>
      ${tarjetaHTML}
      <div class="form-card" ${!ta?'style="opacity:.35;pointer-events:none"':''}>
        <label class="field-label" style="color:#3a9a5a;">Monto a recargar ($)</label>
        <input class="field-input amount-input" id="recarga${caja}" type="number" min="0" step="100" placeholder="0" style="color:#3a9a5a;" />
        <div class="hint-miles" id="rhint${caja}">Ingresá el monto a agregar</div>
        <div class="btn-row" style="margin-top:4px">
          <button class="btn-add" style="background:#2a6a3a;border-color:#3a9a5a;" onclick="recargarTarjeta(${caja})" ${!ta?'disabled':''}>⊕ Recargar saldo</button>
        </div>
      </div>
    </div>
    <div class="section-label" style="margin-top:16px">Ultimas operaciones</div>
    <div id="txlist${caja}" class="tx-list"></div>
    <div class="caja-total-bar" style="flex-direction:column;align-items:stretch;gap:2px;">
      <span class="caja-total-label" style="font-size:10px;">GASTADO ESTA NOCHE</span>
      <div style="display:flex;justify-content:space-between;align-items:center;width:100%;">
        <span id="caja-gasto-label-${caja}" style="font-family:'Rajdhani',sans-serif;font-size:13px;color:#888;letter-spacing:1px;max-width:60%;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;">—</span>
        <span class="caja-total-val" id="ctotal${caja}">$0</span>
      </div>
    </div>
    <div style="padding:6px 0 4px;">
      <button onclick="cerrarGanadorDesdeCaja()" style="width:100%;background:#0d0800;border:1px solid #3a2a0066;border-radius:8px;padding:8px 16px;font-family:'Rajdhani',sans-serif;font-size:12px;font-weight:700;color:#555;letter-spacing:2px;cursor:pointer;text-transform:uppercase;transition:all .15s;" onmouseover="this.style.borderColor='#c9a227';this.style.color='#c9a227'" onmouseout="this.style.borderColor='#3a2a0066';this.style.color='#555'">🏆 Cerrar mensaje de ganador</button>
    </div>`;

  const amEl = document.getElementById('amount'+caja);
  const hintEl = document.getElementById('hint'+caja);
  if (amEl && hintEl) {
    amEl.addEventListener('input', () => {
      const v = parseFloat(amEl.value);
      if (v>0){
        const label = v>=1000000 ? ' · '+( v/1000000).toFixed(v%1000000===0?0:1)+' millón' :
                      v>=1000 ? ' · '+(v/1000).toFixed(v%1000===0?0:1)+' mil' :
                      v>=100 ? ' · '+Math.round(v/100)+' cientos' : '';
        hintEl.textContent='= '+fmt(v)+label;
        hintEl.className='hint-miles ok';
      } else {
        hintEl.textContent='Ingresá el monto completo';
        hintEl.className='hint-miles';
      }
    });
    // Enter en monto con tarjeta → agregar
    amEl.addEventListener('keydown', (e) => {
      if (e.key === 'Enter') { e.preventDefault(); addTx(caja, true); }
    });
  }
  // Enter en nombre con tarjeta → pasar al foco al monto
  const nameEl = document.getElementById('name'+caja);
  if (nameEl) {
    nameEl.addEventListener('keydown', (e) => {
      if (e.key === 'Enter') { e.preventDefault(); if (amEl) amEl.focus(); }
    });
  }

  const mamEl = document.getElementById('mamount'+caja);
  const mhintEl = document.getElementById('mhint'+caja);
  if (mamEl && mhintEl) {
    mamEl.addEventListener('input', () => {
      const v = parseFloat(mamEl.value);
      if (v>0){
        const label = v>=1000000 ? ' · '+(v/1000000).toFixed(v%1000000===0?0:1)+' millón' :
                      v>=1000 ? ' · '+(v/1000).toFixed(v%1000===0?0:1)+' mil' :
                      v>=100 ? ' · '+Math.round(v/100)+' cientos' : '';
        mhintEl.textContent='= '+fmt(v)+label;
        mhintEl.className='hint-miles ok';
      } else {
        mhintEl.textContent='Ingresá el monto completo';
        mhintEl.className='hint-miles';
      }
    });
    // Enter en monto manual → agregar
    mamEl.addEventListener('keydown', (e) => {
      if (e.key === 'Enter') { e.preventDefault(); addTx(caja, false); }
    });
  }
  // Enter en nombre manual → foco a mesa; Enter en mesa → foco a monto
  const mnameEl = document.getElementById('mname'+caja);
  const mmesaEl = document.getElementById('mmesa'+caja);
  if (mnameEl) {
    mnameEl.addEventListener('keydown', (e) => {
      if (e.key === 'Enter') { e.preventDefault(); if (mmesaEl) mmesaEl.focus(); }
    });
  }
  if (mmesaEl) {
    mmesaEl.addEventListener('keydown', (e) => {
      if (e.key === 'Enter') { e.preventDefault(); if (mamEl) mamEl.focus(); }
    });
  }
  // Hint monto recarga
  const reEl = document.getElementById('recarga'+caja);
  const rhintEl = document.getElementById('rhint'+caja);
  if (reEl && rhintEl) {
    reEl.addEventListener('input', () => {
      const v = parseFloat(reEl.value);
      if (v>0){ const lbl=fmtLabel(v); rhintEl.textContent='= '+fmt(v)+(lbl?' · '+lbl:''); rhintEl.className='hint-miles ok'; }
      else { rhintEl.textContent='Ingresá el monto a agregar'; rhintEl.className='hint-miles'; }
    });
    reEl.addEventListener('keydown', e=>{ if(e.key==='Enter'){e.preventDefault();recargarTarjeta(caja);} });
  }
  updateNamelists();
  renderCajaList(caja);
}

async function recargarTarjeta(caja) {
  const ta = window['cajaState'+caja];
  if (!ta) { showToast('Pasa una tarjeta primero', true); return; }
  const monto = parseFloat(document.getElementById('recarga'+caja).value);
  if (!monto || monto <= 0) { showToast('Ingresá un monto válido', true); return; }
  if (!confirm(`¿Agregar ${fmt(monto)} al saldo de Mesa ${ta.mesa}?`)) return;
  try {
    const r = await fetch('/api/tarjetas/recargar', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({codigo:ta.codigo, monto})});
    const d = await r.json();
    if (d.ok) {
      // Actualizar estado local ANTES de cualquier render
      ta.saldo_actual = d.nuevo_saldo;
      window['cajaState'+caja] = ta;
      // Sincronizar tarjetasData localmente para que el próximo render sea correcto
      if (!tarjetasData[ta.codigo]) tarjetasData[ta.codigo] = {};
      tarjetasData[ta.codigo].saldo_actual = d.nuevo_saldo;
      document.getElementById('recarga'+caja).value = '';
      showToast(`✓ Recargado ${fmt(monto)} — Nuevo saldo: ${fmt(d.nuevo_saldo)}`);
      await loadData();
      // Forzar el saldo correcto del server (autoritativo) post-loadData
      if (window['cajaState'+caja]) window['cajaState'+caja].saldo_actual = d.nuevo_saldo;
      if (!tarjetasData[ta.codigo]) tarjetasData[ta.codigo] = {};
      tarjetasData[ta.codigo].saldo_actual = d.nuevo_saldo;
      renderCajaInner(caja);
      setModo(caja, 'recargar');
    } else { showToast(d.error||'Error', true); }
  } catch(e) { showToast('Error de conexión', true); }
}

async function addTx(caja, conTarjeta) {
  let name, amount, mesa, tarjeta_codigo;
  if (conTarjeta) {
    const ta = window['cajaState'+caja];
    if (!ta) { showToast('Pasa una tarjeta primero', true); return; }
    name = document.getElementById('name'+caja).value.trim();
    amount = parseFloat(document.getElementById('amount'+caja).value);
    mesa = ta.mesa;
    tarjeta_codigo = ta.codigo;
    if (!name) { document.getElementById('name'+caja).focus(); showToast('Falta el nombre', true); return; }
    if (!amount || amount<=0) { document.getElementById('amount'+caja).focus(); showToast('Falta el monto', true); return; }
    if (amount > ta.saldo_actual) { showToast('Saldo insuficiente! Disponible: '+fmt(ta.saldo_actual), true); return; }
    try {
      const res = await fetch('/api/tx',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({name,amount,caja,mesa,tarjeta_codigo,client_time:new Date().toLocaleTimeString('es-AR',{hour:'2-digit',minute:'2-digit',hour12:false})})});
      if (res.ok) {
        ta.saldo_actual = ta.saldo_actual - amount;
        ta.nombre = name;
        window['cajaState'+caja] = ta;
        document.getElementById('amount'+caja).value='';
        showToast(fmt(amount)+' descontado de Mesa '+ta.mesa+' — Saldo restante: '+fmt(ta.saldo_actual));
        await loadData();
        if (window['cajaState'+caja]) window['cajaState'+caja].saldo_actual = ta.saldo_actual;
        renderCajaInner(caja);
        setModo(caja,'tarjeta');
      } else {
        const err = await res.json().catch(()=>({error:'Error del servidor'}));
        showToast(err.error||'Error del servidor', true);
      }
    } catch(e){ showToast('Error de conexion',true); }
  } else {
    name = document.getElementById('mname'+caja).value.trim();
    amount = parseFloat(document.getElementById('mamount'+caja).value);
    mesa = document.getElementById('mmesa'+caja).value.trim();
    tarjeta_codigo = '';
    if (!name) { document.getElementById('mname'+caja).focus(); showToast('Falta el nombre',true); return; }
    if (!amount||amount<=0) { document.getElementById('mamount'+caja).focus(); showToast('Falta el monto',true); return; }
    try {
      const res = await fetch('/api/tx',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({name,amount,caja,mesa,tarjeta_codigo,client_time:new Date().toLocaleTimeString('es-AR',{hour:'2-digit',minute:'2-digit',hour12:false})})});
      if (res.ok) {
        document.getElementById('mamount'+caja).value='';
        showToast('Registrado — '+fmt(amount)+' para '+name);
        await loadData(); renderCajaInner(caja);
        setModo(caja,'manual');
      }
    } catch(e){ showToast('Error de conexion',true); }
  }
}

async function deleteTx(id) {
  await fetch('/api/tx/'+id,{method:'DELETE'});
  await loadData();
  // Refrescar saldo de tarjeta activa sin forzar re-scan
  if (cajaState && cajaState.codigo) {
    const td = tarjetasData[cajaState.codigo];
    if (td && td.saldo_actual !== undefined) {
      cajaState.saldo_actual = td.saldo_actual;
      renderTarjeta();
    }
  }
}
async function cerrarGanadorDesdeCaja() {
  try {
    const st = await fetch('/api/state').then(r=>r.json()).catch(()=>({}));
    if (!st.winner_show) { showToast('El ganador todavía no apareció en pantalla', false); return; }
    await fetch('/api/winner/hide', {method:'POST'});
    showToast('Mensaje de ganador cerrado ✓');
  } catch(e) { showToast('Error de conexión', true); }
}
function editTx(id, amount, name) {
  document.getElementById('edit-tx-id').value = id;
  document.getElementById('edit-tx-amount').value = amount;
  document.getElementById('edit-tx-name').value = name;
  document.getElementById('edit-tx-modal').style.display = 'flex';
  setTimeout(()=>document.getElementById('edit-tx-amount').select(), 50);
}
async function confirmarEditTx() {
  const id     = parseInt(document.getElementById('edit-tx-id').value);
  const amount = parseFloat(document.getElementById('edit-tx-amount').value);
  const name   = document.getElementById('edit-tx-name').value.trim();
  if (!amount || amount <= 0) { showToast('Monto inválido', true); return; }
  try {
    const r = await fetch('/api/tx/'+id, {method:'PUT', headers:{'Content-Type':'application/json'}, body:JSON.stringify({amount, name})});
    const d = await r.json();
    if (d.ok) {
      document.getElementById('edit-tx-modal').style.display = 'none';
      await loadData();
      showToast('Operación actualizada');
    } else { showToast(d.error||'Error', true); }
  } catch(e) { showToast('Error de conexión', true); }
}

let lastWinnerTs = 0;
let lastCartelTs = 0;

async function loadData() {
  try {
    const [r1,r2,r3] = await Promise.all([
      fetch('/api/tx').then(r=>{ if(!r.ok) throw new Error('tx'); return r; }),
      fetch('/api/tarjetas').then(r=>{ if(!r.ok) throw new Error('tarjetas'); return r; }),
      fetch('/api/state').then(r=>{ if(!r.ok) throw new Error('state'); return r; })
    ]);
    txData = await r1.json(); tarjetasData = await r2.json();
    const st = await r3.json();
    setStatus(true); render();
    // Sincronizar estado compartido en todas las pantallas
    if (st.hora_fin && st.hora_fin !== horaFin) {
      horaFin = st.hora_fin;
      const hfi = document.getElementById('hora-fin-input');
      if (hfi) hfi.value = horaFin;
      const hfe = document.getElementById('clock-fin');
      if (hfe) hfe.textContent = horaFin;
    }
    if (st.premio !== undefined) {
      const msgEl = document.getElementById('msg-input');
      if (msgEl && msgEl.value !== st.premio) msgEl.value = st.premio;
      const pb = document.getElementById('premio-box');
      if (pb) pb.textContent = st.premio;
    }
    // El ganador y el cartel solo se muestran en /pantalla, no en el manager
    for (let c = 1; c <= 3; c++) {
      const ta = window['cajaState'+c];
      if (ta && ta.codigo && tarjetasData[ta.codigo] !== undefined) {
        ta.saldo_actual = tarjetasData[ta.codigo].saldo_actual;
        window['cajaState'+c] = ta;
      }
    }
  } catch(e){ setStatus(false); }
}

async function sincronizarConfTarjetas() {
  const tabAbierta = document.getElementById('tab-config') && document.getElementById('tab-config').classList.contains('active');
  if (tabAbierta) return;
  try {
    const res = await fetch('/api/tarjetas/config');
    if (res.ok) { const d = await res.json(); if (d && d.length) confTarjetas = d; }
  } catch(e){}
}

function setStatus(ok) {
  const el = document.getElementById('status-txt');
  if (!el) return;
  if (ok){el.textContent='Conexión';el.style.color='var(--green)';el.className='status-ok';const lu=document.getElementById('last-update');if(lu)lu.textContent=new Date().toLocaleTimeString('es-AR',{hour:'2-digit',minute:'2-digit',second:'2-digit',hour12:false});}
  else{el.textContent='Sin conexión';el.style.color='var(--danger)';el.className='status-err';}
}

function render() {
  renderPantalla();
  for(let c=1;c<=3;c++) renderCajaList(c);
  updateNamelists();
  renderStats();
}

let knownNames = new Set();
let prevRankOrder = [];
let koAnimEnCurso = false;

function renderPantalla() {
  if (koAnimEnCurso) return;

  const header = document.getElementById('rank-header');
  const rows   = document.getElementById('rank-rows');
  const empty  = document.getElementById('empty-msg');

  const totals = {}, mesas = {};
  txData.forEach(t => {
    totals[t.name] = (totals[t.name] || 0) + t.amount;
    if (t.mesa && !mesas[t.name]) mesas[t.name] = t.mesa;
  });
  const names = Object.keys(totals).sort((a,b) => totals[b] - totals[a]).slice(0, 5);

  if (!names.length) {
    header.style.display = 'none';
    rows.innerHTML = '';
    empty.style.display = 'block';
    knownNames.clear();
    prevRankOrder = [];
    return;
  }
  header.style.display = 'grid';
  empty.style.display  = 'none';

  const esc = s => String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');

  // ── Detectar KO animación (solo tema boxeo) ──
  if (false && koAnimActiva && prevRankOrder.length > 0) {
    const suben = names.filter((n,i) => { const p = prevRankOrder.indexOf(n); return p !== -1 && p > i; });
    const bajan = names.filter((n,i) => { const p = prevRankOrder.indexOf(n); return p !== -1 && p < i; });
    if (suben.length >= 1 && bajan.length >= 1) {
      const nameUp = suben[0], nameKO = bajan[0];
      const cardUp = document.querySelector(`.rank-row[data-name="${CSS.escape(nameUp)}"]`);
      const cardKO = document.querySelector(`.rank-row[data-name="${CSS.escape(nameKO)}"]`);
      if (cardUp && cardKO) {
        koAnimEnCurso = true;
        const onComplete = () => {
          rows.innerHTML = names.map((name,i) => {
            const pos = i+1, rc = pos<=3?'rank-'+pos:'';
            return `<div class="rank-row ${rc}" data-name="${esc(name)}" data-pos="${pos}"><div class="col-puesto">#${pos}</div><div class="col-nombre">${esc(name)}</div><div class="col-mesa">${esc(mesas[name]||'—')}</div><div class="col-total">${fmt(totals[name])}</div></div>`;
          }).join('');
          prevRankOrder = [...names];
          knownNames    = new Set(names);
          setTimeout(() => { koAnimEnCurso = false; }, 700);
        };
        animateRankSwap(cardUp, cardKO, onComplete);
        return;
      }
    }
  }

  // ══════════════════════════════════════════
  //  FLIP: animación orgánica para todos los modos
  // ══════════════════════════════════════════

  // 1. FIRST — capturar posiciones actuales de cada tarjeta existente
  const firstRects = {};
  rows.querySelectorAll('.rank-row[data-name]').forEach(el => {
    firstRects[el.dataset.name] = el.getBoundingClientRect();
  });

  const prevNames = new Set(prevRankOrder);
  const hasMoves  = names.some((n,i) => prevRankOrder[i] !== n && prevNames.has(n));

  // 2. LAST — actualizar el DOM al nuevo orden (números ya correctos)
  rows.innerHTML = names.map((name, i) => {
    const pos    = i + 1;
    const rc     = pos <= 3 ? 'rank-' + pos : '';
    const esNuevo = !knownNames.has(name);
    return `<div class="rank-row ${rc}${esNuevo?' nueva':''}" data-name="${esc(name)}" data-pos="${pos}" style="will-change:transform,opacity;">
      <div class="col-puesto">#${pos}</div>
      <div class="col-nombre">${esc(name)}</div>
      <div class="col-mesa">${esc(mesas[name]||'—')}</div>
      <div class="col-total">${fmt(totals[name])}</div>
    </div>`;
  }).join('');

  prevRankOrder = [...names];
  knownNames    = new Set(names);

  // 3. INVERT + PLAY — animar solo si hay movimientos reales
  if (hasMoves) {
    // Leer posiciones AFTER el re-render para calcular deltas reales
    rows.querySelectorAll('.rank-row[data-name]').forEach(el => {
      const name  = el.dataset.name;
      const first = firstRects[name];
      if (!first || el.classList.contains('nueva')) return;

      const last = el.getBoundingClientRect();
      const dy   = first.top - last.top;

      if (Math.abs(dy) < 2) return;

      const subeAlPrimero = el.classList.contains('rank-1') && dy > 0;

      // Invert: colocar la tarjeta en su posición VIEJA visualmente
      // El número ya muestra el valor nuevo (correcto) — solo el bloque se mueve
      el.style.transition = 'none';
      el.style.transform  = `translateY(${dy}px)`;
      // Opacidad inicial: la que sube arranca un poco más tenue
      el.style.opacity    = dy > 0 ? '0.6' : '0.9';

      void el.offsetHeight; // force reflow

      // Play: movimiento lento y orgánico hacia la posición final
      // Más lento cuanto mayor es el desplazamiento
      const dur   = Math.min(1.1, 0.65 + Math.abs(dy) / 900);
      const delay = dy > 0 ? 0 : 0.06; // las que bajan con mínimo delay

      el.style.transition = `transform ${dur}s cubic-bezier(.25,.46,.45,.94) ${delay}s,
                              opacity   ${dur * 0.6}s ease ${delay}s`;
      el.style.transform  = 'translateY(0)';
      el.style.opacity    = '1';

      // Glow dorado suave al llegar al #1
      if (subeAlPrimero) {
        setTimeout(() => {
          el.classList.add('ascendio');
          setTimeout(() => el.classList.remove('ascendio'), 950);
        }, (dur + delay) * 1000 - 80);
      }

      // Cleanup estilos inline
      setTimeout(() => {
        el.style.transition = '';
        el.style.transform  = '';
        el.style.opacity    = '';
      }, (dur + delay) * 1000 + 80);
    });
  }

  // Limpiar clase nueva después de la animación
  rows.querySelectorAll('.rank-row.nueva').forEach(el => {
    el.addEventListener('animationend', () => el.classList.remove('nueva'), { once: true });
  });
}

// ════════════════════════════════════════════
//  animateRankSwap — ANIMACIÓN KO COMPLETA
// ════════════════════════════════════════════
// Glove element (shared, created once)
(function() {
  if (!document.getElementById('ko-glove')) {
    const g = document.createElement('div');
    g.id = 'ko-glove';
    g.textContent = '🥊';
    g.style.cssText = `
      position:fixed;font-size:60px;z-index:9999;pointer-events:none;
      display:none;transform-origin:center center;transform:rotate(90deg);
      filter:drop-shadow(0 0 18px rgba(255,50,50,0.9));
      will-change:transform,opacity;
    `;
    document.body.appendChild(g);
  }
  // Keyframes for glove jab (in rotated space: translateY = horizontal movement)
  if (!document.getElementById('kf-ko-system')) {
    const s = document.createElement('style'); s.id = 'kf-ko-system';
    s.textContent = `
      @keyframes jabVibrate {
        0%   { transform: translateX(0); }
        20%  { transform: translateX(-9px) rotate(-1deg); }
        40%  { transform: translateX(7px) rotate(0.5deg); }
        60%  { transform: translateX(-5px); }
        80%  { transform: translateX(4px); }
        100% { transform: translateX(0); }
      }
      @keyframes cardFlyRight {
        0%   { transform: translateX(0); opacity:1; }
        100% { transform: translateX(115vw); opacity:0; }
      }
      @keyframes flashImpact {
        0%   { background:#3a0000; border-color:#ff2222; filter:brightness(2); }
        100% { background:inherit; border-color:inherit; filter:none; }
      }
    `;
    document.head.appendChild(s);
  }
})();

/**
 * animateRankSwap(cardOvertaking, cardBeingOvertaken, onComplete)
 * - cardOvertaking     : el elemento DOM de la tarjeta que SUBE
 * - cardBeingOvertaken : el elemento DOM de la tarjeta que BAJA (recibe el KO)
 * - onComplete         : callback que hace el swap de datos + re-render
 *
 * Coreografía:
 *  1. Guante → aparece horizontal a la izquierda de la víctima
 *  2. Jab 1 + Jab 2 (vibración)
 *  3. Wind-up (retroceso)
 *  4. Power Punch → tarjeta vuela por la derecha (position:fixed para libertad total)
 *  5. onComplete() → swap datos + renderRanking() → DOM actualizado
 *  6. Tarjeta que bajó: entra desde la izquierda a su NUEVO slot correcto
 *  7. Tarjeta que subió: se desliza verticalmente desde donde estaba
 */
function animateRankSwap(cardOvertaking, cardBeingOvertaken, onComplete) {
  const glove = document.getElementById('ko-glove');

  // ── Capturar posiciones ANTES de cualquier cambio ──
  const rectKO = cardBeingOvertaken.getBoundingClientRect();
  const rectUp = cardOvertaking.getBoundingClientRect();

  // Guardar data-names antes de que el DOM cambie
  const nameKO = cardBeingOvertaken.dataset.name;
  const nameUp = cardOvertaking.dataset.name;

  // Posicionar guante horizontal (→) a la izquierda de la víctima
  const gloveY = rectKO.top + rectKO.height / 2 - 34;
  const gloveX = rectKO.left - 90;

  glove.style.cssText = `
    position:fixed; display:block;
    top:${gloveY}px; left:${gloveX}px;
    font-size:60px; z-index:9999; pointer-events:none;
    transform:rotate(90deg) translateY(0px) scaleX(1);
    opacity:1; transition:none;
    filter:drop-shadow(0 0 18px rgba(255,50,50,0.9));
    transform-origin:center center;
    will-change:transform,opacity;
  `;

  // ── Timings ──
  const T_JAB1      = 80;
  const T_JAB1_RET  = 220;
  const T_JAB2      = 390;
  const T_JAB2_RET  = 530;
  const T_WINDUP    = 650;
  const T_PUNCH     = 840;
  const T_GLOVE_OUT = 1020;
  const T_RERENDER  = 1080;

  function jabCard() {
    cardBeingOvertaken.style.animation = 'none';
    void cardBeingOvertaken.offsetWidth;
    cardBeingOvertaken.style.animation = 'jabVibrate 0.16s ease-in-out';
  }

  // JAB 1
  setTimeout(() => {
    glove.style.transition = 'transform 0.10s ease-out';
    glove.style.transform  = 'rotate(90deg) translateY(-58px) scaleX(1.15)';
    jabCard();
  }, T_JAB1);
  setTimeout(() => {
    glove.style.transition = 'transform 0.09s ease-in';
    glove.style.transform  = 'rotate(90deg) translateY(-6px) scaleX(0.92)';
  }, T_JAB1_RET);

  // JAB 2
  setTimeout(() => {
    glove.style.transition = 'transform 0.10s ease-out';
    glove.style.transform  = 'rotate(90deg) translateY(-58px) scaleX(1.15)';
    jabCard();
  }, T_JAB2);
  setTimeout(() => {
    glove.style.transition = 'transform 0.09s ease-in';
    glove.style.transform  = 'rotate(90deg) translateY(-6px) scaleX(0.92)';
  }, T_JAB2_RET);

  // WIND-UP
  setTimeout(() => {
    glove.style.transition = 'transform 0.17s cubic-bezier(.4,0,.2,1)';
    glove.style.transform  = 'rotate(90deg) translateY(55px) scaleX(0.72)';
  }, T_WINDUP);

  // POWER PUNCH → tarjeta sale volando por la derecha
  setTimeout(() => {
    glove.style.transition = 'transform 0.14s cubic-bezier(.1,0,.5,1)';
    glove.style.transform  = 'rotate(90deg) translateY(-110px) scaleX(1.35)';

    // Flash impacto
    cardBeingOvertaken.style.animation = 'flashImpact 0.25s ease-out';

    // Sacar la tarjeta del flujo normal → position:fixed en su lugar exacto
    cardBeingOvertaken.style.position = 'fixed';
    cardBeingOvertaken.style.top      = rectKO.top + 'px';
    cardBeingOvertaken.style.left     = rectKO.left + 'px';
    cardBeingOvertaken.style.width    = rectKO.width + 'px';
    cardBeingOvertaken.style.zIndex   = '200';
    cardBeingOvertaken.style.margin   = '0';
    void cardBeingOvertaken.offsetWidth;

    cardBeingOvertaken.style.transition = 'transform 0.28s cubic-bezier(.4,0,1,1), opacity 0.2s 0.08s ease';
    cardBeingOvertaken.style.transform  = 'translateX(115vw)';
    cardBeingOvertaken.style.opacity    = '0';
  }, T_PUNCH);

  // Guante fade out
  setTimeout(() => {
    glove.style.transition = 'opacity 0.2s ease';
    glove.style.opacity    = '0';
  }, T_GLOVE_OUT);

  // ── RE-RENDER + animaciones de entrada ──
  setTimeout(() => {
    // 1. Ejecutar onComplete: actualiza datos y re-renderiza el DOM
    onComplete();

    // 2. Buscar las nuevas tarjetas en el DOM ya re-renderizado
    const newCardKO = document.querySelector(`.rank-row[data-name="${CSS.escape(nameKO)}"]`);
    const newCardUp = document.querySelector(`.rank-row[data-name="${CSS.escape(nameUp)}"]`);

    // 3. Preparar estado inicial ANTES de que el browser pinte
    //    — Tarjeta que bajó: invisible, fuera de pantalla a la izquierda
    //    — Tarjeta que subió: desplazada hacia abajo (donde estaba antes)
    if (newCardUp) {
      const destUp = newCardUp.getBoundingClientRect();
      const deltaY = rectUp.top - destUp.top;
      newCardUp.style.transition = 'none';
      newCardUp.style.transform  = `translateY(${deltaY}px)`;
      newCardUp.style.zIndex     = '80';
    }
    if (newCardKO) {
      newCardKO.style.transition = 'none';
      newCardKO.style.opacity    = '0';
      newCardKO.style.transform  = 'translateX(-115vw)';
    }

    // 4. Forzar un reflow para que el browser registre el estado inicial
    if (newCardUp) void newCardUp.offsetHeight;
    if (newCardKO) void newCardKO.offsetHeight;

    // 5. Activar transiciones en el siguiente frame de pintura
    requestAnimationFrame(() => {
      requestAnimationFrame(() => {
        // Tarjeta que SUBIÓ: desliza hacia arriba a su posición final
        if (newCardUp) {
          newCardUp.style.transition = 'transform 0.45s cubic-bezier(.22,1,.36,1)';
          newCardUp.style.transform  = 'translateY(0)';
        }
        // Tarjeta que BAJÓ: entra desde la izquierda a su slot correcto
        if (newCardKO) {
          newCardKO.style.transition = 'transform 0.38s cubic-bezier(.22,1,.36,1), opacity 0.15s ease';
          newCardKO.style.transform  = 'translateX(0)';
          newCardKO.style.opacity    = '1';
        }
      });
    });

    // 6. Cleanup: quitar todos los estilos inline cuando terminen las transiciones
    setTimeout(() => {
      [newCardUp, newCardKO].forEach(c => {
        if (!c) return;
        c.style.transition = '';
        c.style.transform  = '';
        c.style.opacity    = '';
        c.style.zIndex     = '';
      });
      glove.style.display    = 'none';
      glove.style.opacity    = '';
      glove.style.transform  = '';
      glove.style.transition = '';
      koAnimEnCurso = false;
    }, 550);

  }, T_RERENDER);
}

// ── Helper: vibraciónde jab ──
function jabCard(card) {
  card.style.animation = 'none';
  void card.offsetWidth;
  card.style.animation = 'jabVibrate 0.16s ease-in-out';
}

function renderCajaList(caja) {
  const txs=txData.filter(t=>t.caja===caja).slice().reverse();
  const total=txs.reduce((s,t)=>s+t.amount,0);
  const tel=document.getElementById('ctotal'+caja);if(tel)tel.innerHTML=fmtDisplay(total);
  updateCajaGastoLabel(caja);
  const list=document.getElementById('txlist'+caja);if(!list)return;
  if(!txs.length){list.innerHTML='<div class="no-tx">Sin operaciones aun</div>';return;}
  list.innerHTML=txs.map(t=>`
    <div class="tx-item">
      <div class="tx-info">
        <div class="tx-name">${t.name}${t.mesa?' <span style="color:#3a3a3a;font-size:11px">M'+t.mesa+'</span>':''}</div>
        <div class="tx-meta">${t.time}</div>
      </div>
      <div class="tx-right"><span class="tx-amount">${fmtDisplay(t.amount)}</span><button class="btn-edit" onclick="editTx(${t.id},${t.amount},'${String(t.name).replace(/'/g,"\\'")}')">✎</button><button class="btn-del" onclick="deleteTx(${t.id})">✕</button></div>
    </div>`).join('');
}

function updateNamelists() {
  const names=[...new Set(txData.map(t=>t.name))];
  for(let c=1;c<=3;c++){
    ['nl','mnl'].forEach(prefix=>{
      const dl=document.getElementById(prefix+c);
      if(dl)dl.innerHTML=names.map(n=>`<option value="${n}">`).join('');
    });
  }
}

// CONFIGURACION TARJETAS
function renderConfigTarjetas() {
  const grid=document.getElementById('tarjetas-grid');
  grid.innerHTML=confTarjetas.map((t,i)=>{
    const configurada=t.codigo;
    const saldoInfo=tarjetasData[t.codigo];
    const saldoIni=parseFloat(t.saldo_inicial||0);
    const saldoAct=saldoInfo!==undefined?saldoInfo.saldo_actual:saldoIni;
    const pct=saldoIni>0?Math.max(0,Math.round((saldoAct/saldoIni)*100)):0;
    const hasPending = pendingCodigos[i] !== undefined;
    return `<div class="tarjeta-conf ${configurada?'configurada':''}" id="tc-${i}">
      <div class="tc-header">
        <span class="tc-num">Tarjeta ${t.slot} — Mesa ${t.slot}</span>
        <div class="tc-btns">
          <span class="tc-status" id="tc-status-${i}" style="${configurada?'color:#3a9a5a':''}">${configurada?'Vinculada':'Sin vincular'}</span>
          <button class="tc-scan-btn" id="scan-btn-${i}" onclick="iniciarScan(${i})">LEER</button>
          <button class="tc-confirm-btn" id="confirm-btn-${i}" onclick="confirmarVinculo(${i})" style="display:${hasPending?'inline-block':'none'}">CONFIRMAR</button>
          <button class="tc-clear-btn" onclick="clearSlot(${i})" title="Borrar tarjeta">✕</button>
        </div>
      </div>
      <div class="tc-field">
        <span class="tc-label">Nombre del cliente</span>
        <input class="tc-input" id="tc-nombre-${i}" type="text" placeholder="Nombre..." value="${t.nombre_cliente||''}"
          oninput="confTarjetas[${i}].nombre_cliente=this.value" />
      </div>
      <div class="tc-field">
        <span class="tc-label">Saldo inicial</span>
        <input class="tc-input" id="tc-saldo-${i}" type="number" placeholder="" value="${t.saldo_inicial}"
          oninput="confTarjetas[${i}].saldo_inicial=this.value"
          onkeydown="if(event.key==='Enter'){event.preventDefault();const v=this.value;if(v>0){showToast('Monto $'+fmt(v).replace('$','')+' agregado a Tarjeta ${t.slot}');}}" />
      </div>
      <div class="tc-code" id="tc-code-${i}">${hasPending?'Leida: '+pendingCodigos[i]:(t.codigo?'Cod: '+t.codigo:'Presiona LEER y pasa la tarjeta')}</div>
      ${saldoIni>0?`<div style="margin-top:10px;">
        <div style="display:flex;justify-content:space-between;font-size:13px;font-family:'Rajdhani',sans-serif;letter-spacing:.5px;margin-bottom:6px;">
          <span style="color:#c84a4a;font-weight:700;">Gastado: ${fmt(Math.max(0,saldoIni-saldoAct))}</span>
          <span style="color:#3a9a5a;font-weight:700;">Disponible: ${fmt(Math.max(0,saldoAct))}</span>
        </div>
        <div style="height:8px;background:#111;border-radius:4px;overflow:hidden;display:flex;">
          <div style="height:100%;width:${Math.max(0,Math.min(100,100-pct))}%;background:linear-gradient(to right,#6b1010,#c84a4a);"></div>
          <div style="height:100%;flex:1;background:linear-gradient(to right,#1a5a2a,#3a9a5a);"></div>
        </div>
      </div>`:''}
    </div>`;
  }).join('');
}

function iniciarScan(idx) {
  if(scanSlotActivo!==null){const pb=document.getElementById('scan-btn-'+scanSlotActivo);if(pb){pb.textContent='LEER';pb.classList.remove('activo');}}
  scanSlotActivo=idx;
  const btn=document.getElementById('scan-btn-'+idx);
  btn.textContent='ESPERANDO...';btn.classList.add('activo');
  showToast('Pasa la tarjeta por el lector');
  setTimeout(()=>{if(scanSlotActivo===idx){scanSlotActivo=null;btn.textContent='LEER';btn.classList.remove('activo');}},10000);
}
function clearSlot(idx){confTarjetas[idx].codigo='';renderConfigTarjetas();}

async function guardarTarjetas() {
  try {
    const res=await fetch('/api/tarjetas/config',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(confTarjetas)});
    if(res.ok){showToast('Configuracion guardada');await loadData();renderConfigTarjetas();}
    else{
      const err=await res.json();
      showToast(err.error||'Error al guardar',true);
    }
  } catch(e){showToast('Error al guardar',true);}
}
async function limpiarTarjetas() {
  if(!confirm('Limpiar toda la configuracion?'))return;
  confTarjetas=Array.from({length:30},(_,i)=>({slot:i+1,codigo:'',saldo_inicial:''}));
  await fetch('/api/tarjetas/limpiar',{method:'POST'});
  renderConfigTarjetas();showToast('Tarjetas limpiadas');
}

async function cargarConfTarjetas() {
  try{const res=await fetch('/api/tarjetas/config');if(res.ok){const d=await res.json();if(d&&d.length)confTarjetas=d;}}catch(e){}
}

// ══════════════════════════════════════════
//  STATS
// ══════════════════════════════════════════
function renderStats() {
  if (!document.getElementById('tab-stats').classList.contains('active')) return;
  if (!txData.length) {
    document.getElementById('stats-sub').textContent = 'Sin datos todavía — registrá consumos para ver las estadísticas.';
    document.getElementById('kpi-total').textContent = '$0';
    document.getElementById('kpi-ops').textContent = '0';
    document.getElementById('kpi-avg').textContent = '$0';
    document.getElementById('stats-top-clientes').innerHTML = '<div style="color:#333;font-size:13px;padding:14px 0;">Sin datos aún</div>';
    document.getElementById('cajas-detail').innerHTML = '';
    return;
  }

  const total = txData.reduce((s,t)=>s+t.amount,0);
  const ops = txData.length;
  const avg = Math.round(total/ops);

  // Consumo por hora
  const porHora = {};
  txData.forEach(t => {
    const h = t.time ? t.time.split(':')[0] : '??';
    porHora[h] = (porHora[h]||0) + t.amount;
  });
  const horaMax = Object.entries(porHora).sort((a,b)=>b[1]-a[1])[0];

  // Primer y último registro
  const horas = txData.map(t=>t.time).filter(Boolean).sort();
  const subTxt = horas.length ? 'Primera operación: ' + horas[0] + '  •  Última: ' + horas[horas.length-1] : '';
  document.getElementById('stats-sub').textContent = subTxt;
  document.getElementById('kpi-total').innerHTML = fmtDisplay(total);
  document.getElementById('kpi-ops').textContent = ops;
  document.getElementById('kpi-avg').innerHTML = fmtDisplay(avg);

  // ─── Gráfico por caja ───
  const porCaja = {1:0, 2:0, 3:0};
  const opsCaja = {1:0, 2:0, 3:0};
  txData.forEach(t=>{ if(t.caja>=1&&t.caja<=3){porCaja[t.caja]+=t.amount;opsCaja[t.caja]++;} });
  const cajasColors = ['#c9a227','#e8c84a','#7a6010'];
  const cajasLabels = ['Abajo','Extendido','VIP'];
  drawBarChart('chart-cajas', [porCaja[1],porCaja[2],porCaja[3]], cajasLabels, cajasColors);
  document.getElementById('legend-cajas').innerHTML = cajasLabels.map((l,i)=>
    `<div class="legend-item"><div class="legend-dot" style="background:${cajasColors[i]}"></div>Caja ${l}: ${fmt(porCaja[i+1])} (${opsCaja[i+1]} ops)</div>`
  ).join('');

  // ─── Gráfico por hora ───
  // Ordenar horas cronológicamente (noche: 20-06)
  const allHoras = Array.from({length:12},(_,i)=>String((20+i)%24).padStart(2,'0'));
  const valHoras = allHoras.map(h=>porHora[h]||0);
  const horasLabels = allHoras.map(h=>h+'h');
  drawBarChart('chart-horas', valHoras, horasLabels, valHoras.map((_,i)=>allHoras[i]===horaMax?.[0]?'#c9a227':'#2a2a2a'));

  // ─── Top clientes ───
  const totals={}, mesas={};
  txData.forEach(t=>{totals[t.name]=(totals[t.name]||0)+t.amount; if(t.mesa&&!mesas[t.name])mesas[t.name]=t.mesa;});
  const sorted = Object.entries(totals).sort((a,b)=>b[1]-a[1]).slice(0,8);
  const maxVal = sorted[0]?.[1]||1;
  const esc = s=>String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
  document.getElementById('stats-top-clientes').innerHTML = '<div class="top-list">' +
    sorted.map(([name,val],i)=>`
      <div class="top-item pos-${i+1}">
        <div class="top-pos">#${i+1}</div>
        <div style="flex:1">
          <div class="top-name">${esc(name)}</div>
          ${mesas[name]?`<div class="top-mesa">Mesa ${esc(mesas[name])}</div>`:''}
        </div>
        <div class="top-bar-wrap"><div class="top-bar-fill" style="width:${Math.round(val/maxVal*100)}%"></div></div>
        <div class="top-amount">${fmtDisplay(val)}</div>
      </div>`).join('') + '</div>';

  // ─── Detalle por caja ───
  const cajaNombres = {1:'Abajo',2:'Extendido',3:'VIP'};
  document.getElementById('cajas-detail').innerHTML = [1,2,3].map(c=>{
    const txs = txData.filter(t=>t.caja===c);
    const tot = txs.reduce((s,t)=>s+t.amount,0);
    const allTxs = txs.slice().reverse();
    return `<div class="caja-stat-card">
      <div class="caja-stat-badge">${cajaNombres[c].toUpperCase()}</div>
      <div class="caja-stat-total">${fmtDisplay(tot)}</div>
      <div class="caja-stat-ops">${txs.length} operaciones</div>
      <div class="caja-stat-list" style="max-height:220px;">${
        allTxs.length
          ? allTxs.map(t=>{
              const itemsStr = t.items && t.items.length
                ? t.items.map(i=>`${i.qty}x ${esc(i.nombre)}`).join(', ')
                : '';
              return `<div class="caja-stat-item" style="flex-direction:column;align-items:flex-start;gap:2px;">
                <div style="display:flex;justify-content:space-between;width:100%;">
                  <span style="color:#d0cdc0;font-weight:600;">${esc(t.name)} <span style="color:#666;font-size:11px;">${t.time||''}</span></span>
                  <span style="color:var(--gold)">${fmtDisplay(t.amount)}</span>
                </div>
                ${itemsStr ? `<span style="font-size:11px;color:#888;">${itemsStr}</span>` : ''}
              </div>`;
            }).join('')
          : '<div style="color:#222;font-size:12px;padding:8px 0;">Sin operaciones</div>'
      }</div>
    </div>`;
  }).join('');
}

function drawBarChart(canvasId, values, labels, colors) {
  const canvas = document.getElementById(canvasId);
  if (!canvas) return;
  const ctx = canvas.getContext('2d');
  const W = canvas.offsetWidth || 300;
  canvas.width = W;
  const H = canvas.height;
  ctx.clearRect(0,0,W,H);
  const max = Math.max(...values, 1);
  const n = values.length;
  const padL=6, padR=6, padT=10, padB=28;
  const barW = Math.floor((W - padL - padR) / n);
  const gap = Math.max(2, Math.floor(barW*0.15));
  const bw = barW - gap;
  values.forEach((v,i)=>{
    const x = padL + i*barW + Math.floor(gap/2);
    const barH = Math.round((v/max)*(H-padT-padB));
    const y = H - padB - barH;
    const col = Array.isArray(colors)?colors[i%colors.length]:colors;
    // Barra
    ctx.fillStyle = v>0 ? col : '#1a1a1a';
    ctx.beginPath();
    ctx.roundRect ? ctx.roundRect(x, y, bw, barH, [3,3,0,0]) : ctx.rect(x, y, bw, barH);
    ctx.fill();
    // Label abajo
    ctx.fillStyle = '#444';
    ctx.font = `${Math.min(10, Math.floor(barW*0.55))}px Arial`;
    ctx.textAlign = 'center';
    ctx.fillText(labels[i], x + bw/2, H-padB+14);
    // Valor encima si la barra es visible
    if (v>0 && barH>18) {
      ctx.fillStyle = '#888';
      ctx.font = '9px Arial';
      const label = v>=1000000?'$'+(v/1000000).toFixed(1)+'M':v>=1000?'$'+(v/1000).toFixed(0)+'k':'$'+v;
      ctx.fillText(label, x+bw/2, y-3);
    }
  });
}


// ══════════════════════════════════════════
//  TEMAS NUEVOS: Jagger 12 años y Velada Boxeo
// ══════════════════════════════════════════
const TEMAS_EXTRA = {
  jagger12: {
    colors: {'--black':'#000000','--surface':'#0a0a0a','--border':'#333333','--gold':'#ffffff','--gold-light':'#dddddd','--gold-dim':'#888888','--text':'#e8e8e8','--text-dim':'#666666','--white':'#ffffff'},
    bodyClass: 'tema-jagger12',
    particleLabel: 'Activar burbujas de champagne'
  },

  touchofpink: {
    colors: {'--black':'#2d0020','--surface':'#480035','--border':'#8a3070','--gold':'#f472b6','--gold-light':'#fbb6ce','--gold-dim':'#e896cc','--text':'#ffe8f5','--text-dim':'#ddaacc','--white':'#ffffff'},
    bodyClass: 'tema-touchofpink',
    particleLabel: 'Activar pétalos animados'
  }
};

let temaActual = 'default';
let decoActiva = true;
let punchAnimActiva = false;
let koAnimActiva = true;
let fallingGlovesActivos = true;
let mostrar12Fondo = true;
let svg12Opacity = 0.13;
let svg12Color = 'white';
let svg12GlowBlur = 18;
let confettiGanadorActivo = true;
let tipoParticula = 'confetti'; // 'confetti' | 'billetes' | 'ninguno'
let pinkPetalosActivos = true;
let pinkModoClaro = false;

function aplicarTema(nombre) {
  const temaExtra = TEMAS_EXTRA[nombre];
  document.body.classList.remove('tema-fullblack','tema-navidad','tema-anonuevo','tema-halloween','tema-jagger12','tema-touchofpink','pink-claro');
  pinkModoClaro = false;
  const overlay = document.getElementById('tema-overlay');
  overlay.innerHTML = ''; overlay.style.opacity = '0';
  const tl = document.getElementById('tema-tagline');

  if (temaExtra) {
    COLOR_DEFS.forEach(c => {
      const v = temaExtra.colors[c.key] || c.default;
      document.documentElement.style.setProperty(c.key, v);
      customColors[c.key] = v;
    });
    // surface-gold tinted by new surface
    document.documentElement.style.setProperty('--surface-gold', blendSurfaceGold(temaExtra.colors['--surface'] || '#111'));
    if (temaExtra.bodyClass) document.body.classList.add(temaExtra.bodyClass);
    temaActual = nombre;
    // Show/hide punch animation toggle
    const fallingToggle = document.getElementById('falling-gloves-toggle');
    if (fallingToggle) fallingToggle.style.display = 'none';
    const koToggle = document.getElementById('ko-anim-toggle');
    if (koToggle) koToggle.style.display = 'none';
    const show12Toggle = document.getElementById('show-12-toggle');
    if (show12Toggle) show12Toggle.style.display = nombre === 'jagger12' ? 'block' : 'none';
    const pinkPetTog = document.getElementById('pink-petalos-toggle');
    if (pinkPetTog) pinkPetTog.style.display = nombre === 'touchofpink' ? 'block' : 'none';
    const pinkModoTog = document.getElementById('pink-modo-toggle');
    if (pinkModoTog) pinkModoTog.style.display = nombre === 'touchofpink' ? 'flex' : 'none';
    // Update logo
    const logoVip = document.getElementById('logo-vip');
    const savedVip = (()=>{ try { return localStorage.getItem('rankingVIP_vip') || 'VIP'; } catch(e){ return 'VIP'; } })();
    // Restore standard RANKING VIP logo (no club span)
    const mainLogo = document.getElementById('main-logo');
    if (mainLogo) mainLogo.innerHTML = `RANKING <span class="vip" id="logo-vip">${savedVip}</span>`;
    if (tl) {
      if (nombre === 'touchofpink') {
        tl.innerHTML = `<div style="text-align:center;line-height:1.3;">
          <div style="font-family:'Oswald',sans-serif;font-size:clamp(18px,2.8vw,30px);font-weight:700;letter-spacing:6px;color:#ffffff;text-shadow:0 0 12px rgba(255,255,255,0.9),0 0 28px rgba(255,255,255,0.5),0 0 50px rgba(244,114,182,0.4);">JAGGER CLUB</div>
          <div style="font-family:'Oswald',sans-serif;font-size:clamp(13px,2vw,22px);font-weight:600;letter-spacing:5px;color:#f472b6;margin-top:3px;text-shadow:0 0 10px rgba(244,114,182,1),0 0 24px rgba(244,114,182,0.7),0 0 50px rgba(244,114,182,0.4);">TURNS PINK</div>
        </div>`;
      } else {
        tl.innerHTML = `<div style="text-align:center;line-height:1.3;">
          <div style="font-family:'Oswald',sans-serif;font-size:clamp(14px,2.2vw,24px);font-weight:700;letter-spacing:4px;color:#e8c84a;text-shadow:0 0 14px rgba(232,200,74,0.9),0 0 30px rgba(201,162,39,0.6);">JAGGER CLUB · 12 AÑOS</div>
          <div style="font-family:'Oswald',sans-serif;font-size:clamp(11px,1.6vw,17px);font-weight:600;letter-spacing:3px;color:#e8c84a;opacity:0.75;text-shadow:0 0 10px rgba(201,162,39,0.5);margin-top:2px;">12 AÑOS DE HISTORIA NO SON PARA CUALQUIERA</div>
        </div>`;
      }
    }
    // Toggle deco
    const toggleWrap = document.getElementById('tema-deco-toggle');
    if (toggleWrap) {
      toggleWrap.style.display = 'flex';
      const decoLabel = nombre === 'jagger12' ? 'Activar burbujas de champagne' :
                        
                        'Activar decoraciones animadas';
      document.getElementById('toggle-deco-label').textContent = decoLabel;
      document.getElementById('toggle-deco').checked = decoActiva;
      const decoMainLabel = document.getElementById('deco-main-label');
      if (decoMainLabel) decoMainLabel.style.display = nombre === 'touchofpink' ? 'none' : '';
    }
    if (decoActiva) iniciarDecoTema(nombre);
    buildColorGrid();
    try { localStorage.setItem('rankingVIP_tema', nombre); } catch(e) {}
    showToast('Tema ' + nombre.toUpperCase() + ' aplicado');
    fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({tema:nombre,colores:customColors})}).catch(()=>{});
    return;
  }

  if (nombre === 'default') {
    COLOR_DEFS.forEach(c => document.documentElement.style.setProperty(c.key, c.default));
    document.documentElement.style.setProperty('--surface-gold', '#0d0b00');
    customColors = {};
    COLOR_DEFS.forEach(c => customColors[c.key] = c.default);
    const lv = document.getElementById('logo-vip'); if(lv) lv.textContent='VIP';
    // Restore main logo without club
    const mainLogo2 = document.getElementById('main-logo');
    if (mainLogo2) mainLogo2.innerHTML = `RANKING <span class="vip" id="logo-vip">VIP</span>`;
    if (tl) { tl.textContent = 'JAGGER CLUB'; tl.style.fontSize='28px'; tl.style.fontWeight='600'; tl.style.letterSpacing='5px'; tl.style.color='#555'; }
  }
  temaActual = nombre;
  const toggleWrap = document.getElementById('tema-deco-toggle');
  if (toggleWrap) toggleWrap.style.display = 'none';
  buildColorGrid();
  try { localStorage.setItem('rankingVIP_tema', nombre); } catch(e) {}
  showToast('Tema ' + nombre.toUpperCase() + ' aplicado');
  fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({tema:nombre,colores:customColors})}).catch(()=>{});
}

function toggleDecoActual(checked) {
  decoActiva = checked;
  fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({deco_activa:checked})}).catch(()=>{});
  if (temaActual === 'jagger12') {
    // En aniversario, solo controla las burbujas animadas (no el 12 de fondo)
    const wrap = document.getElementById('jagger12-particles');
    if (!checked) {
      if (wrap) wrap.innerHTML = ''; // detener burbujas
    } else {
      // Reiniciar solo burbujas
      if (wrap) {
        wrap.innerHTML = '';
        function lanzarBurbuja() {
          if (!decoActiva) return; // Stop when deactivated
          if (!document.getElementById('jagger12-particles')) return;
          const el = document.createElement('div');
          const sz = 4 + Math.random() * 12;
          const isGold = Math.random() > 0.7;
          el.style.cssText = `position:absolute;bottom:-20px;left:${Math.random()*100}%;width:${sz}px;height:${sz}px;border-radius:50%;border:1px solid rgba(${isGold?'201,162,39':'255,255,255'},${isGold?0.35:0.2});background:rgba(${isGold?'201,162,39':'255,255,255'},${isGold?0.06:0.03});animation:bubbleRise ${4+Math.random()*6}s ease-in forwards;pointer-events:none;`;
          wrap.appendChild(el);
          setTimeout(()=>el.remove(), 11000);
          setTimeout(lanzarBurbuja, 200+Math.random()*600);
        }
        for(let i=0;i<8;i++) setTimeout(lanzarBurbuja, i*120);
      }
    }
  } else {
    const overlay = document.getElementById('tema-overlay');
    if (!checked) { overlay.innerHTML = ''; overlay.style.opacity = '0'; }
    else iniciarDecoTema(temaActual);
  }
}

// Actualizar solo el SVG del 12 sin tocar las burbujas ni las decoraciones
function actualizar12Overlay() {
  if (temaActual === 'jagger12') {
    const overlay = document.getElementById('tema-overlay');
    const existing12 = overlay ? overlay.querySelector('svg:not([id])') : null;
    // Remove any existing 12 SVGs (identified by having the "12" text)
    if (overlay) {
      overlay.querySelectorAll('svg').forEach(svg => {
        if (svg.textContent && svg.textContent.trim() === '12') svg.remove();
      });
    }
    if (mostrar12Fondo && overlay) {
      const svgNS = 'http://www.w3.org/2000/svg';
      const svg = document.createElementNS(svgNS, 'svg');
      svg.setAttribute('width','100%'); svg.setAttribute('height','100%');
      svg.setAttribute('viewBox','0 0 1000 600');
      svg.style.cssText = `position:absolute;inset:0;pointer-events:none;opacity:${svg12Opacity};`;
      svg.setAttribute('preserveAspectRatio','xMidYMid meet');
      const defs = document.createElementNS(svgNS,'defs');
      defs.innerHTML = `<filter id="blur12j"><feGaussianBlur stdDeviation="8"/></filter><filter id="glow12j"><feGaussianBlur stdDeviation="${svg12GlowBlur}" result="blur"/><feMerge><feMergeNode in="blur"/><feMergeNode in="SourceGraphic"/></feMerge></filter>`;
      svg.appendChild(defs);
      const t1 = document.createElementNS(svgNS,'text');
      t1.setAttribute('x','500'); t1.setAttribute('y','380'); t1.setAttribute('text-anchor','middle');
      t1.setAttribute('dominant-baseline','middle'); t1.setAttribute('font-family','Oswald,Arial');
      t1.setAttribute('font-weight','700'); t1.setAttribute('font-size','560');
      t1.setAttribute('fill',`rgba(201,162,39,0.15)`); t1.setAttribute('letter-spacing','-10');
      t1.setAttribute('filter','url(#blur12j)'); t1.textContent = '12';
      const t2 = document.createElementNS(svgNS,'text');
      t2.setAttribute('x','500'); t2.setAttribute('y','380'); t2.setAttribute('text-anchor','middle');
      t2.setAttribute('dominant-baseline','middle'); t2.setAttribute('font-family','Oswald,Arial');
      t2.setAttribute('font-weight','700'); t2.setAttribute('font-size','560');
      t2.setAttribute('fill',svg12Color); t2.setAttribute('letter-spacing','-10');
      t2.setAttribute('filter','url(#glow12j)'); t2.textContent = '12';
      svg.appendChild(t1); svg.appendChild(t2);
      // Insert after particles wrap
      const particles = overlay.querySelector('#jagger12-particles');
      if (particles && particles.nextSibling) overlay.insertBefore(svg, particles.nextSibling);
      else overlay.appendChild(svg);
    }
  }
}

function iniciarDecoTema(nombre) {
  const overlay = document.getElementById('tema-overlay');
  overlay.style.opacity = '1';
  if (nombre === 'jagger12') {
    overlay.innerHTML = '';
    iniciarJagger12Deco();
  }
  if (nombre === 'touchofpink') {
    overlay.innerHTML = `
      <!-- Detalles blancos y rosados: destellos -->
      <div style="position:absolute;top:12%;left:8%;font-size:14px;color:rgba(255,255,255,0.45);pointer-events:none;animation:goldTwinkle 3s ease-in-out infinite;">✦</div>
      <div style="position:absolute;top:18%;right:10%;font-size:11px;color:rgba(255,255,255,0.38);pointer-events:none;animation:goldTwinkle 4s ease-in-out 1s infinite;">✦</div>
      <div style="position:absolute;top:50%;left:5%;font-size:12px;color:rgba(255,255,255,0.35);pointer-events:none;animation:goldTwinkle 5s ease-in-out 2s infinite;">✦</div>
      <div style="position:absolute;top:65%;right:6%;font-size:15px;color:rgba(255,255,255,0.4);pointer-events:none;animation:goldTwinkle 3.5s ease-in-out .5s infinite;">✦</div>
      <div style="position:absolute;top:33%;left:14%;font-size:9px;color:rgba(251,182,206,0.55);pointer-events:none;animation:goldTwinkle 4.5s ease-in-out 1.5s infinite;">✦</div>
      <div style="position:absolute;top:42%;right:16%;font-size:10px;color:rgba(251,182,206,0.5);pointer-events:none;animation:goldTwinkle 3.8s ease-in-out .8s infinite;">✦</div>
      <div style="position:absolute;top:75%;left:20%;font-size:8px;color:rgba(255,255,255,0.3);pointer-events:none;animation:goldTwinkle 6s ease-in-out 3s infinite;">✦</div>
      <div style="position:absolute;top:28%;right:22%;font-size:13px;color:rgba(244,114,182,0.4);pointer-events:none;animation:goldTwinkle 5.5s ease-in-out 2.5s infinite;">✦</div>
      <div id="petalos-wrap" style="position:absolute;inset:0;overflow:hidden;pointer-events:none;"></div>`;
    if (!document.getElementById('kf-bubble')) {
      const s = document.createElement('style'); s.id = 'kf-pink-twinkle';
      s.textContent = `@keyframes goldTwinkle{0%,100%{opacity:0.15;transform:scale(1)}50%{opacity:0.55;transform:scale(1.4)}}`;
      document.head.appendChild(s);
    }
    if (pinkPetalosActivos) iniciarPetalos();
  }
}

function iniciarJagger12Deco() {
  const overlay = document.getElementById('tema-overlay');
  const svg12 = mostrar12Fondo ? `
    <!-- 12 grande de fondo, centrado -->
    <svg width="100%" height="100%" viewBox="0 0 1000 600" style="position:absolute;inset:0;pointer-events:none;opacity:0.13;" xmlns="http://www.w3.org/2000/svg" preserveAspectRatio="xMidYMid meet">
      <defs>
        <filter id="blur12j"><feGaussianBlur stdDeviation="8"/></filter>
        <filter id="glow12j">
          <feGaussianBlur stdDeviation="18" result="blur"/>
          <feMerge><feMergeNode in="blur"/><feMergeNode in="SourceGraphic"/></feMerge>
        </filter>
      </defs>
      <text x="500" y="380" text-anchor="middle" dominant-baseline="middle" font-family="Oswald,Arial" font-weight="700" font-size="560" fill="rgba(201,162,39,0.15)" letter-spacing="-10" filter="url(#blur12j)">12</text>
      <text x="500" y="380" text-anchor="middle" dominant-baseline="middle" font-family="Oswald,Arial" font-weight="700" font-size="560" fill="white" letter-spacing="-10" filter="url(#glow12j)">12</text>
    </svg>` : '';
  overlay.innerHTML = `
    <div id="jagger12-particles" style="position:absolute;inset:0;overflow:hidden;pointer-events:none;"></div>
    ${svg12}
    <!-- Líneas doradas sutiles en esquinas -->
    <svg width="200" height="200" viewBox="0 0 200 200" style="position:absolute;top:0;left:0;opacity:0.18;pointer-events:none;" xmlns="http://www.w3.org/2000/svg">
      <line x1="0" y1="0" x2="120" y2="0" stroke="#c9a227" stroke-width="1"/>
      <line x1="0" y1="0" x2="0" y2="120" stroke="#c9a227" stroke-width="1"/>
      <line x1="0" y1="0" x2="60" y2="60" stroke="#c9a227" stroke-width="0.5"/>
      <circle cx="0" cy="0" r="3" fill="#c9a227"/>
    </svg>
    <svg width="200" height="200" viewBox="0 0 200 200" style="position:absolute;top:0;right:0;opacity:0.18;pointer-events:none;transform:scaleX(-1);" xmlns="http://www.w3.org/2000/svg">
      <line x1="0" y1="0" x2="120" y2="0" stroke="#c9a227" stroke-width="1"/>
      <line x1="0" y1="0" x2="0" y2="120" stroke="#c9a227" stroke-width="1"/>
      <line x1="0" y1="0" x2="60" y2="60" stroke="#c9a227" stroke-width="0.5"/>
      <circle cx="0" cy="0" r="3" fill="#c9a227"/>
    </svg>
    <!-- Detalles dorados: pequeñas estrellas/destellos -->
    <div style="position:absolute;top:15%;left:5%;font-size:10px;color:#c9a227;opacity:0.2;pointer-events:none;animation:goldTwinkle 3s ease-in-out infinite;">✦</div>
    <div style="position:absolute;top:25%;right:7%;font-size:8px;color:#c9a227;opacity:0.18;pointer-events:none;animation:goldTwinkle 4s ease-in-out 1s infinite;">✦</div>
    <div style="position:absolute;top:60%;left:3%;font-size:12px;color:#c9a227;opacity:0.15;pointer-events:none;animation:goldTwinkle 5s ease-in-out 2s infinite;">✦</div>
    <div style="position:absolute;top:70%;right:4%;font-size:9px;color:#c9a227;opacity:0.18;pointer-events:none;animation:goldTwinkle 3.5s ease-in-out 0.5s infinite;">✦</div>`;
  if (!document.getElementById('kf-bubble')) {
    const s = document.createElement('style'); s.id = 'kf-bubble';
    s.textContent = `@keyframes bubbleRise{0%{opacity:0.6;transform:translateY(0) scale(1)}50%{opacity:0.3}100%{opacity:0;transform:translateY(-100vh) scale(0.5)}}
    @keyframes goldTwinkle{0%,100%{opacity:0.1;transform:scale(1)}50%{opacity:0.25;transform:scale(1.4)}}`;
    document.head.appendChild(s);
  }
  if (decoActiva) {
    function lanzarBurbuja() {
      if (!decoActiva) return; // Stop when deactivated
      const wrap = document.getElementById('jagger12-particles'); if(!wrap) return;
      const el = document.createElement('div');
      const sz = 4 + Math.random() * 12;
      const isGold = Math.random() > 0.7;
      el.style.cssText = `position:absolute;bottom:-20px;left:${Math.random()*100}%;width:${sz}px;height:${sz}px;border-radius:50%;border:1px solid rgba(${isGold?'201,162,39':'255,255,255'},${isGold?0.35:0.2});background:rgba(${isGold?'201,162,39':'255,255,255'},${isGold?0.06:0.03});animation:bubbleRise ${4+Math.random()*6}s ease-in forwards;pointer-events:none;`;
      wrap.appendChild(el);
      setTimeout(()=>el.remove(), 11000);
      setTimeout(lanzarBurbuja, 200+Math.random()*600);
    }
    for(let i=0;i<8;i++) setTimeout(lanzarBurbuja, i*120);
  }
}

function reiniciarDeco12() {
  if (temaActual === 'jagger12') iniciarJagger12Deco();

}

// ══════════════════════════════════════════
//  CARTEL
// ══════════════════════════════════════════
const FRASES_RAPIDAS = [
  '🍾 SACÓ UN NUVO CON BENGALAS',
  '🥂 PIDIÓ CHAMPAGNE',
  '🎉 SACÓ BOTELLA VIP',
  '🔥 ARRANCÓ EL SHOW',
  '💎 MODO VIP ACTIVADO',
  '🚀 NIVEL ÉLITE'
];

function contarEmojis(str) {
  return [...new Intl.Segmenter('es',{granularity:'grapheme'}).segment(str)].length;
}
function limitarEmojis(input) {
  const segs = [...new Intl.Segmenter('es',{granularity:'grapheme'}).segment(input.value)];
  if (segs.length > 3) input.value = segs.slice(0,3).map(s=>s.segment).join('');
}
function agregarEmojiCartel(emoji) {
  const inp = document.getElementById('cartel-emoji-input');
  const segs = [...new Intl.Segmenter('es',{granularity:'grapheme'}).segment(inp.value)];
  if (segs.length < 3) inp.value = segs.map(s=>s.segment).join('') + emoji;
  else showToast('Máximo 3 emojis', true);
}
let cartelCajaActiva = 0;
let cartelPreciosMgr = {virtual:0, fisico:0, combo:0};
let cartelTipoMgr = 'virtual';

function abrirCartelModal(caja) {
  cartelCajaActiva = caja || 0;
  const ta = caja ? window['cajaState'+caja] : null;
  const nombreEl = document.getElementById('cartel-nombre');
  const mesaEl = document.getElementById('cartel-mesa');
  const fraseEl = document.getElementById('cartel-frase');
  const emojiEl = document.getElementById('cartel-emoji-input');
  if (nombreEl) nombreEl.value = ta?.nombre || '';
  if (mesaEl) mesaEl.value = ta?.slot ? String(ta.slot) : '';
  if (fraseEl) fraseEl.value = '';
  if (emojiEl) emojiEl.value = '';
  fetch('/api/cartel/precios').then(r=>r.json()).then(p=>{
    cartelPreciosMgr = p;
    const fv = n => n>0 ? '$'+Number(n).toLocaleString('es-AR') : 'Sin cargo';
    const pv=document.getElementById('cprecio-v'); if(pv) pv.textContent=fv(p.virtual||0);
    const pf=document.getElementById('cprecio-f'); if(pf) pf.textContent=fv(p.fisico||0);
    const pc=document.getElementById('cprecio-c'); if(pc) pc.textContent=fv(p.combo||0);
    selCartelTipo('virtual');
  }).catch(()=>selCartelTipo('virtual'));
  document.getElementById('cartel-modal').style.display = 'flex';
  setTimeout(()=>{ (ta?.nombre ? fraseEl : nombreEl)?.focus(); }, 120);
}

function selCartelTipo(tipo) {
  cartelTipoMgr = tipo;
  ['virtual','fisico','combo'].forEach(t=>{
    const b=document.getElementById('cbtn-'+t); if(!b) return;
    const sel=t===tipo;
    b.style.borderColor=sel?'#c9a227':'#333';
    b.style.color=sel?'#c9a227':'#555';
    b.style.background=sel?'#0d0b00':'#0a0a0a';
  });
  const precio=cartelPreciosMgr[tipo]||0;
  const descRow=document.getElementById('cartel-descuento-mgr');
  const montoLbl=document.getElementById('cartel-monto-mgr');
  const ta=cartelCajaActiva?window['cajaState'+cartelCajaActiva]:null;
  if(descRow){ descRow.style.display=(precio>0&&ta)?'block':'none'; if(montoLbl) montoLbl.textContent='$'+Number(precio).toLocaleString('es-AR'); }
}

function cerrarCartelModal() { document.getElementById('cartel-modal').style.display = 'none'; }

async function mostrarCartel() {
  const nombre = document.getElementById('cartel-nombre').value.trim().toUpperCase();
  const mesa = document.getElementById('cartel-mesa').value.trim();
  const frase = document.getElementById('cartel-frase').value.trim().toUpperCase();
  const ta = cartelCajaActiva ? window['cajaState'+cartelCajaActiva] : null;
  if (!nombre && !ta) { showToast('Ingresá un nombre o pasá una tarjeta primero', true); return; }
  if (!frase) { showToast('Escribí una frase para el cartel', true); return; }
  const emojiManual = document.getElementById('cartel-emoji-input').value.trim();
  const emojis = emojiManual || (
    frase.includes('NUVO') || frase.includes('BOTELLA') ? '🍾' :
    frase.includes('BENGALA') ? '🎆' :
    frase.includes('CHAMPAGNE') ? '🍾' :
    frase.includes('BOXEO') || frase.includes('FIGHT') ? '🥊' : '🍾'
  );
  const tipo = cartelTipoMgr;
  const precio = cartelPreciosMgr[tipo] || 0;
  const descontar = document.getElementById('cartel-descontar-mgr')?.checked;
  cerrarCartelModal();
  // Descontar de tarjeta si corresponde
  if (precio > 0 && descontar && ta && cartelCajaActiva) {
    if (ta.saldo_actual < precio) {
      showToast('Saldo insuficiente para el cartel ($'+Number(precio).toLocaleString('es-AR')+')', true); return;
    }
    try {
      const r = await fetch('/api/tx', {method:'POST', headers:{'Content-Type':'application/json'},
        body: JSON.stringify({
          name: nombre || ta.nombre || 'Cartel',
          amount: precio, caja: cartelCajaActiva,
          mesa: mesa || String(ta.slot||''),
          tarjeta_codigo: ta.codigo,
          items: [{nombre:'Cartel '+tipo, cantidad:1, precio}]
        })});
      const d = await r.json();
      if (!d.ok) { showToast(d.error||'Error descontando', true); return; }
      ta.saldo_actual -= precio;
      renderCajaInner(cartelCajaActiva);
      await loadData();
    } catch(e) { showToast('Error descontando de tarjeta', true); return; }
  }
  // Mostrar en pantalla (solo virtual o combo, no físico)
  if (tipo !== 'fisico') {
    try {
      await fetch('/api/cartel/show', {method:'POST', headers:{'Content-Type':'application/json'},
        body: JSON.stringify({nombre, mesa, frase, emoji: emojis})});
      showToast('📣 Cartel enviado a pantalla');
    } catch(e) { showToast('Error enviando cartel', true); }
  } else {
    showToast('🖨 Cartel físico — Mesa '+(mesa||String(ta?.slot||'—')));
  }
}

function cerrarCartel() {
  fetch('/api/cartel/hide',{method:'POST'}).then(()=>showToast('Cartel cerrado en pantalla')).catch(()=>{});
}

// ══════════════════════════════════════════
//  EFECTOS DE FONDO
// ══════════════════════════════════════════
let efectoActual = 'ninguno';
let efectoInterval = null;

let currentClockSize = 'mediano';
function setClockSize(size) {
  currentClockSize = size;
  ['chico','mediano','grande'].forEach(s => {
    const btn = document.getElementById('cs-'+s);
    if (!btn) return;
    btn.style.borderColor = s===size ? 'var(--gold)' : '#2a2a2a';
    btn.style.color = s===size ? 'var(--gold)' : '#555';
  });
  fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({clock_size:size})}).catch(()=>{});
  showToast('Reloj: '+size);
}

function aplicarEfecto(nombre) {
  efectoActual = nombre;
  const overlay = document.getElementById('efectos-overlay');
  if (overlay) { overlay.innerHTML = ''; }
  if (efectoInterval) { clearInterval(efectoInterval); efectoInterval = null; }
  // Highlight button activo
  document.querySelectorAll('[id^="efecto-btn-"]').forEach(b => {
    b.style.borderColor = '#2a2a2a'; b.style.color = '#555';
  });
  const activeBtn = document.getElementById('efecto-btn-' + nombre);
  if (activeBtn) { activeBtn.style.borderColor = '#c9a227'; activeBtn.style.color = '#c9a227'; }
  try { localStorage.setItem('rankingVIP_efecto', nombre); } catch(e){}
  fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({efecto:nombre})}).catch(()=>{});
  if (nombre === 'ninguno') return;
  if (nombre === 'burbujas') iniciarEfectoBurbujas();
  if (nombre === 'estrellas') iniciarEfectoEstrellas();
  if (nombre === 'lluvia_dorada') iniciarEfectoLluviaDorada();
  if (nombre === 'confetti') iniciarEfectoConfetti();
  showToast('Efecto ' + nombre.replace('_',' ').toUpperCase() + ' activado');
}

function iniciarEfectoBurbujas() {
  const overlay = document.getElementById('efectos-overlay'); if(!overlay) return;
  if (!document.getElementById('kf-efecto-bubble')) {
    const s = document.createElement('style'); s.id = 'kf-efecto-bubble';
    s.textContent = `@keyframes efectoBubble{0%{opacity:0.7;transform:translateY(0) scale(1)}50%{opacity:0.4}100%{opacity:0;transform:translateY(-100vh) scale(0.3)}}`;
    document.head.appendChild(s);
  }
  function crear() {
    if (!document.getElementById('efectos-overlay')) return;
    const el = document.createElement('div');
    const sz = 5 + Math.random() * 20;
    const isGold = Math.random() > 0.6;
    el.style.cssText = `position:absolute;bottom:-30px;left:${Math.random()*100}%;width:${sz}px;height:${sz}px;border-radius:50%;border:1px solid rgba(${isGold?'201,162,39':'255,255,255'},${isGold?0.4:0.2});background:rgba(${isGold?'201,162,39':'255,255,255'},${isGold?0.06:0.03});animation:efectoBubble ${5+Math.random()*8}s ease-in forwards;`;
    overlay.appendChild(el);
    setTimeout(()=>el.remove(), 14000);
  }
  for(let i=0;i<10;i++) setTimeout(crear, i*200);
  efectoInterval = setInterval(crear, 350);
}

function iniciarEfectoEstrellas() {
  const overlay = document.getElementById('efectos-overlay'); if(!overlay) return;
  if (!document.getElementById('kf-efecto-star')) {
    const s = document.createElement('style'); s.id = 'kf-efecto-star';
    s.textContent = `@keyframes starFloat{0%{opacity:0;transform:translateY(0) rotate(0deg) scale(0)}20%{opacity:0.8}80%{opacity:0.5}100%{opacity:0;transform:translateY(-80vh) rotate(360deg) scale(0.5)}}`;
    document.head.appendChild(s);
  }
  const syms = ['★','✦','✧','✶','✸','✺'];
  function crear() {
    if (!document.getElementById('efectos-overlay')) return;
    const el = document.createElement('div');
    const sym = syms[Math.floor(Math.random()*syms.length)];
    const sz = 10 + Math.random()*18;
    const colors = ['#c9a227','#e8c84a','#ffffff','#f0ece0'];
    const color = colors[Math.floor(Math.random()*colors.length)];
    el.textContent = sym;
    el.style.cssText = `position:absolute;bottom:-30px;left:${Math.random()*100}%;font-size:${sz}px;color:${color};opacity:0;animation:starFloat ${6+Math.random()*6}s ease-in-out forwards;pointer-events:none;`;
    overlay.appendChild(el);
    setTimeout(()=>el.remove(), 13000);
  }
  for(let i=0;i<8;i++) setTimeout(crear, i*300);
  efectoInterval = setInterval(crear, 500);
}

function iniciarEfectoLluviaDorada() {
  const overlay = document.getElementById('efectos-overlay'); if(!overlay) return;
  if (!document.getElementById('kf-efecto-gold')) {
    const s = document.createElement('style'); s.id = 'kf-efecto-gold';
    s.textContent = `@keyframes goldFall{0%{opacity:0.9;transform:translateY(-10px) rotate(0deg)}100%{opacity:0;transform:translateY(100vh) rotate(180deg)}}`;
    document.head.appendChild(s);
  }
  function crear() {
    if (!document.getElementById('efectos-overlay')) return;
    const el = document.createElement('div');
    const w = 2 + Math.random()*3, h = 12 + Math.random()*20;
    const gold = Math.random() > 0.3 ? '#c9a227' : '#e8c84a';
    el.style.cssText = `position:absolute;top:-30px;left:${Math.random()*100}%;width:${w}px;height:${h}px;background:${gold};border-radius:1px;opacity:0;animation:goldFall ${3+Math.random()*4}s linear forwards;pointer-events:none;`;
    overlay.appendChild(el);
    setTimeout(()=>el.remove(), 8000);
  }
  for(let i=0;i<15;i++) setTimeout(crear, i*100);
  efectoInterval = setInterval(crear, 120);
}

function iniciarEfectoConfetti() {
  const overlay = document.getElementById('efectos-overlay'); if(!overlay) return;
  if (!document.getElementById('kf-efecto-confetti')) {
    const s = document.createElement('style'); s.id = 'kf-efecto-confetti';
    s.textContent = `@keyframes confettiFall{0%{opacity:1;transform:translateY(-10px) rotate(0deg)}100%{opacity:0.3;transform:translateY(100vh) rotate(720deg)}}`;
    document.head.appendChild(s);
  }
  const colors = ['#c9a227','#e8c84a','#ffffff','#ff4444','#44aaff','#44ff88','#ff44ff'];
  function crear() {
    if (!document.getElementById('efectos-overlay')) return;
    const el = document.createElement('div');
    const color = colors[Math.floor(Math.random()*colors.length)];
    const w = 6 + Math.random()*10, h = 4 + Math.random()*8;
    el.style.cssText = `position:absolute;top:-20px;left:${Math.random()*100}%;width:${w}px;height:${h}px;background:${color};border-radius:${Math.random()>0.5?'50%':'2px'};animation:confettiFall ${4+Math.random()*5}s ease-in forwards;pointer-events:none;`;
    overlay.appendChild(el);
    setTimeout(()=>el.remove(), 10000);
  }
  for(let i=0;i<20;i++) setTimeout(crear, i*80);
  efectoInterval = setInterval(crear, 200);
}

function cargarEfectoGuardado() {
  try {
    const e = localStorage.getItem('rankingVIP_efecto');
    if (e && e !== 'ninguno') { setTimeout(()=>aplicarEfecto(e), 500); }
  } catch(e) {}
}

// ══════════════════════════════════════════
//  CARTEL – Override cargarTemaGuardado to support new themes
function cargarTemaGuardado() {
  try {
    const t = localStorage.getItem('rankingVIP_tema');
    if (t) aplicarTema(t);
  } catch(e) {}
}

// ══════════════════════════════════════════
//  AUTH — Cambio de PINs desde manager
// ══════════════════════════════════════════
async function cambiarPassword() {
  const actual = document.getElementById('pwd-actual').value;
  const nueva  = document.getElementById('pwd-nueva').value;
  const conf   = document.getElementById('pwd-confirm').value;
  const msg    = document.getElementById('pwd-msg');
  if (!actual||!nueva) { msg.style.color='#a83030'; msg.textContent='Completá todos los campos'; return; }
  if (!/^\d{4}$/.test(nueva)) { msg.style.color='#a83030'; msg.textContent='El nuevo PIN debe tener exactamente 4 dígitos'; return; }
  if (nueva !== conf)  { msg.style.color='#a83030'; msg.textContent='Los PINs no coinciden'; return; }
  try {
    const r = await fetch('/api/auth/change',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({current:actual,new:nueva})});
    const d = await r.json();
    if (d.ok) {
      msg.style.color='#2ecc71'; msg.textContent='Contraseña actualizada correctamente';
      document.getElementById('pwd-actual').value='';
      document.getElementById('pwd-nueva').value='';
      document.getElementById('pwd-confirm').value='';
    } else {
      msg.style.color='#a83030'; msg.textContent=d.error||'Error al cambiar contraseña';
    }
  } catch(e) { msg.style.color='#a83030'; msg.textContent='Error de conexión'; }
}

// ══════════════════════════════════════════
//  PRECIOS CARTEL
// ══════════════════════════════════════════
async function cargarCartelPrecios() {
  try {
    const r = await fetch('/api/cartel/precios');
    const d = await r.json();
    const vi = document.getElementById('cartel-precio-virtual');
    const fi = document.getElementById('cartel-precio-fisico');
    const co = document.getElementById('cartel-precio-combo');
    if (vi) vi.value = d.virtual || 0;
    if (fi) fi.value = d.fisico || 0;
    if (co) co.value = d.combo || 0;
  } catch(e) {}
}

async function guardarCartelPrecios() {
  const virtual = parseFloat(document.getElementById('cartel-precio-virtual')?.value || 0);
  const fisico  = parseFloat(document.getElementById('cartel-precio-fisico')?.value || 0);
  const combo   = parseFloat(document.getElementById('cartel-precio-combo')?.value || 0);
  try {
    await fetch('/api/cartel/precios', {method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({virtual, fisico, combo})});
    const st = document.getElementById('cartel-precios-status');
    if (st) { st.style.color='#2ecc71'; st.textContent='✓ Precios guardados'; setTimeout(()=>st.textContent='',3000); }
    showToast('Precios de cartel guardados');
  } catch(e) { showToast('Error guardando precios', true); }
}

// ══════════════════════════════════════════
//  MENÚ MANAGER
// ══════════════════════════════════════════
let menuMgrData = {{ menu_json | safe }};
const CATS_ORDER = ['Champagne','Bottle Service','Importados','Bebidas','Tragos','Shots'];

async function cargarMenuMgr() {
  try {
    const r = await fetch('/api/menu');
    menuMgrData = await r.json();
    renderMenuMgr();
  } catch(e) { showToast('Error cargando menú', true); }
}

async function resetearMenuDefault() {
  if (!confirm('¿Cargar el menú oficial? Se reemplazarán todos los productos actuales.')) return;
  try {
    const r = await fetch('/api/menu/reset', {method:'POST'});
    const d = await r.json();
    if (d.ok) { showToast('✓ Menú oficial cargado — '+d.count+' productos'); await cargarMenuMgr(); }
    else showToast('Error', true);
  } catch(e) { showToast('Error de conexión', true); }
}
function renderMenuMgr() {
  const q = (document.getElementById('menu-mgr-search')?.value||'').toLowerCase();
  const filtered = q ? menuMgrData.filter(p => p.nombre.toLowerCase().includes(q) || p.categoria.toLowerCase().includes(q)) : menuMgrData;
  const cats = CATS_ORDER.filter(c => filtered.some(p => p.categoria === c));
  const extra = [...new Set(filtered.map(p=>p.categoria))].filter(c => !CATS_ORDER.includes(c));
  const allCats = [...cats, ...extra];
  const esc = s => String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
  const el = document.getElementById('menu-mgr-content');
  if (!el) return;
  el.innerHTML = allCats.map(cat => {
    const prods = filtered.filter(p => p.categoria === cat);
    return `<div style="margin-bottom:20px;">
      <div style="font-family:'Oswald',sans-serif;font-size:15px;color:var(--gold);letter-spacing:3px;text-transform:uppercase;margin-bottom:10px;padding-bottom:8px;border-bottom:1px solid #2a1c00;">${esc(cat)}</div>
      <div style="display:flex;flex-direction:column;gap:6px;">
        ${prods.map(p => `<div style="display:flex;align-items:center;gap:10px;background:#0e0e0e;border:1px solid #222;border-radius:8px;padding:12px 16px;">
          <span class="nombre-editable" data-id="${p.id}" onclick="editarNombreInline(this)" title="Tocá para editar nombre" style="flex:1;font-size:16px;font-weight:600;color:#f0ece0;cursor:pointer;padding:3px 6px;border-radius:5px;border:1px solid transparent;" onmouseover="this.style.borderColor='#333'" onmouseout="this.style.borderColor='transparent'">${esc(p.nombre)}</span>
          <span class="precio-editable" data-id="${p.id}" onclick="editarPrecioInline(this)" title="Tocá para editar" style="font-family:'Oswald',sans-serif;font-size:18px;font-weight:700;color:var(--gold);cursor:pointer;padding:4px 10px;border-radius:5px;border:1px solid transparent;transition:border-color .2s;" onmouseover="this.style.borderColor='var(--gold-dim)'" onmouseout="if(!this.classList.contains('editing'))this.style.borderColor='transparent'">$${Number(p.precio).toLocaleString('es-AR')}</span>
          <button onclick="eliminarProducto(${p.id})" class="btn-del" title="Eliminar">✕</button>
        </div>`).join('')}
      </div>
    </div>`;
  }).join('') || '<div style="color:#333;padding:20px;text-align:center;letter-spacing:2px;">Sin productos en el menú</div>';
}

function editarNombreInline(el) {
  if (el.classList.contains('editing')) return;
  el.classList.add('editing');
  const id = parseInt(el.dataset.id);
  const prod = menuMgrData.find(p => p.id === id);
  if (!prod) return;
  el.innerHTML = `<input type="text" value="${prod.nombre.replace(/"/g,'&quot;')}" maxlength="60" style="width:100%;min-width:120px;background:#0d0d0d;border:none;border-radius:4px;color:var(--text);font-family:'Rajdhani',sans-serif;font-size:14px;padding:2px 6px;outline:none;" onblur="guardarNombreInline(this,${id})" onkeydown="if(event.key==='Enter')this.blur();" />`;
  el.querySelector('input').focus(); el.querySelector('input').select();
}
async function guardarNombreInline(inp, id) {
  const nuevo = inp.value.trim();
  if (!nuevo) { await cargarMenuMgr(); return; }
  try {
    await fetch('/api/menu/'+id,{method:'PUT',headers:{'Content-Type':'application/json'},body:JSON.stringify({nombre:nuevo})});
    await cargarMenuMgr(); showToast('Nombre actualizado ✓ — Recordá refrescar las cajas (F5)', false, 5000);
  } catch(e) { await cargarMenuMgr(); }
}
function editarPrecioInline(el) {
  if (el.classList.contains('editing')) return;
  el.classList.add('editing');
  const id = parseInt(el.dataset.id);
  const prod = menuMgrData.find(p => p.id === id);
  if (!prod) return;
  const orig = el.innerHTML;
  el.innerHTML = `<input type="number" value="${prod.precio}" step="100" min="0" style="width:100px;background:#0d0d0d;border:none;border-radius:4px;color:var(--gold);font-family:'Oswald',sans-serif;font-size:15px;padding:2px 6px;outline:none;" onblur="guardarPrecioInline(this,${id})" onkeydown="if(event.key==='Enter')this.blur();if(event.key==='Escape'){this.closest('.precio-editable').innerHTML='${orig}';this.closest('.precio-editable').classList.remove('editing');}" />`;
  el.querySelector('input').focus();
  el.querySelector('input').select();
}

async function guardarPrecioInline(inp, id) {
  const nuevo = parseFloat(inp.value);
  if (isNaN(nuevo) || nuevo < 0) { cargarMenuMgr(); return; }
  try {
    await fetch('/api/menu/'+id, {method:'PUT', headers:{'Content-Type':'application/json'}, body:JSON.stringify({precio: nuevo})});
    await cargarMenuMgr();
    showToast('Precio actualizado ✓ — Recordá refrescar las cajas (F5)', false, 5000);
  } catch(e) { cargarMenuMgr(); }
}

async function agregarProducto() {
  const cat = document.getElementById('new-prod-cat').value;
  const nombre = document.getElementById('new-prod-nombre').value.trim();
  const precio = parseFloat(document.getElementById('new-prod-precio').value);
  if (!nombre || isNaN(precio) || precio < 0) { showToast('Completá todos los campos', true); return; }
  try {
    await fetch('/api/menu', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({categoria:cat, nombre, precio})});
    document.getElementById('new-prod-nombre').value = '';
    document.getElementById('new-prod-precio').value = '';
    await cargarMenuMgr();
    showToast('Producto agregado');
  } catch(e) { showToast('Error al agregar', true); }
}

async function eliminarProducto(id) {
  if (!confirm('¿Eliminar este producto del menú?')) return;
  try {
    await fetch('/api/menu/'+id, {method:'DELETE'});
    await cargarMenuMgr();
    showToast('Producto eliminado');
  } catch(e) { showToast('Error al eliminar', true); }
}

// ══════════════════════════════════════════
//  PUBLICIDAD MANAGER
// ══════════════════════════════════════════
let pubPollingMgr = null;

async function cargarEstadoPub() {
  try {
    const r = await fetch('/api/publicidad/estado');
    const d = await r.json();
    const ico = document.getElementById('pub-estado-icon');
    const txt = document.getElementById('pub-estado-txt');
    const urlEl = document.getElementById('pub-estado-url');
    if (ico) ico.textContent = d.activa ? '▶' : '⏸';
    if (txt) { txt.textContent = d.activa ? 'Activa' : 'Inactiva'; txt.style.color = d.activa ? '#2ecc71' : 'var(--gold)'; }
    if (urlEl) urlEl.textContent = d.publicidad_url ? d.publicidad_url.slice(0,60)+(d.publicidad_url.length>60?'…':'') : '';
    if (document.getElementById('pub-url') && !document.getElementById('pub-url').value && d.publicidad_url) {
      document.getElementById('pub-url').value = d.publicidad_url;
    }
    if (document.getElementById('pub-frec') && d.frecuencia) {
      document.getElementById('pub-frec').value = d.frecuencia;
    }
  } catch(e) {}
}

async function activarPublicidad() {
  const url = (document.getElementById('pub-url')?.value||'').trim();
  const frec = parseInt(document.getElementById('pub-frec')?.value||15);
  if (!url) { showToast('Primero subí un video .mp4', true); return; }
  try {
    await fetch('/api/publicidad/activar', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({url, frecuencia:frec})});
    await cargarEstadoPub();
    showToast('Publicidad activada');
  } catch(e) { showToast('Error', true); }
}

async function mostrarAhora() {
  try {
    await fetch('/api/publicidad/mostrar-ahora', {method:'POST'});
    showToast('📺 Video enviado a pantalla');
  } catch(e) { showToast('Error', true); }
}

async function desactivarPublicidad() {
  try {
    await fetch('/api/publicidad/desactivar', {method:'POST'});
    await cargarEstadoPub();
    showToast('Publicidad desactivada');
  } catch(e) { showToast('Error', true); }
}

async function subirVideoPublicidad() {
  const fileInput = document.getElementById('pub-file');
  const statusEl = document.getElementById('pub-upload-status');
  if (!fileInput?.files?.length) { showToast('Seleccioná un archivo mp4', true); return; }
  const file = fileInput.files[0];
  if (!file.type.includes('mp4') && !file.name.endsWith('.mp4')) { showToast('Solo se permiten archivos mp4', true); return; }
  const fd = new FormData();
  fd.append('file', file);
  if (statusEl) statusEl.textContent = 'Subiendo...';
  try {
    const r = await fetch('/api/publicidad/upload', {method:'POST', body:fd});
    const d = await r.json();
    if (d.ok) {
      document.getElementById('pub-url').value = d.url;
      if (statusEl) statusEl.textContent = '✓ Subido: ' + d.url;
      showToast('Video subido correctamente');
    } else { if (statusEl) statusEl.textContent = 'Error: ' + (d.error||'?'); showToast(d.error||'Error', true); }
  } catch(e) { if (statusEl) statusEl.textContent = 'Error de conexión'; showToast('Error de conexión', true); }
}

// ══════════════════════════════════════════
//  PUBLICIDAD — Polling en pantalla de presentación
// ══════════════════════════════════════════
let pubLastActiva = false;
let pubCloseTimer = null;

async function pollPublicidad() {
  try {
    const r = await fetch('/api/publicidad/estado');
    const d = await r.json();
    const overlay = document.getElementById('pub-overlay');
    if (!overlay) return;
    if (d.publicidad_activa && !pubLastActiva) {
      // Activar
      pubLastActiva = true;
      const video = document.getElementById('pub-video');
      const iframe = document.getElementById('pub-iframe');
      const url = d.publicidad_url || '';
      const seg = parseInt(d.publicidad_segundos) || 30;
      overlay.style.display = 'flex';
      // Solo MP4
      if (iframe) { iframe.style.display = 'none'; iframe.src = ''; }
      if (video) {
        video.src = url; video.style.display = 'block';
        video.onended = () => cerrarPublicidadOverlay(false);
        video.play().catch(()=>{});
      }
      if (pubCloseTimer) clearTimeout(pubCloseTimer);
      pubCloseTimer = setTimeout(() => cerrarPublicidadOverlay(false), seg * 1000);
    } else if (!d.publicidad_activa && pubLastActiva) {
      cerrarPublicidadOverlay(false);
    }
  } catch(e) {}
}

function cerrarPublicidadOverlay(manual) {
  pubLastActiva = false;
  if (pubCloseTimer) { clearTimeout(pubCloseTimer); pubCloseTimer = null; }
  const overlay = document.getElementById('pub-overlay');
  if (overlay) overlay.style.display = 'none';
  const video = document.getElementById('pub-video');
  if (video) { video.pause(); video.src = ''; video.style.display = 'none'; }
  const iframe = document.getElementById('pub-iframe');
  if (iframe) { iframe.src = ''; iframe.style.display = 'none'; }
  if (manual) fetch('/api/publicidad/desactivar', {method:'POST'}).catch(()=>{});
}

// ══════════════════════════════════════════
//  GESTIÓN DE PINES
// ══════════════════════════════════════════
// PINs tab (dedicated tab)
const PIN_ROLES = [
  {key:'pin_manager', label:'Manager'},
  {key:'pin_tarjetas', label:'Tarjetas'},
  {key:'pin_cajaabajo', label:'Caja Abajo'},
  {key:'pin_cajaextendido', label:'Caja Extendido'},
  {key:'pin_cajavip', label:'Caja VIP'},
];
async function cargarPines2() {
  const el = document.getElementById('pines-list2');
  if (!el) return;
  try {
    const r = await fetch('/api/config/pines');
    const d = await r.json();
    el.innerHTML = PIN_ROLES.map(ro => `<div style="display:flex;align-items:center;gap:12px;">
      <span style="min-width:140px;font-size:13px;color:#aaa;letter-spacing:1px;">${ro.label}</span>
      <input type="password" id="pin2-${ro.key}" class="custom-text-input" maxlength="4" inputmode="numeric" pattern="[0-9]*" placeholder="Nuevo PIN (4 dígitos)" style="flex:1;max-width:180px;" />
      <span style="font-size:11px;color:#555;">Actual: ****</span>
    </div>`).join('');
  } catch(e) {}
}
async function guardarPines2() {
  const updates = {};
  for (const ro of PIN_ROLES) {
    const val = document.getElementById('pin2-'+ro.key)?.value?.trim();
    if (val) {
      if (!/^\d{4}$/.test(val)) { showToast('PIN debe tener 4 dígitos: '+ro.label, true); return; }
      updates[ro.key] = val;
    }
  }
  if (!Object.keys(updates).length) { showToast('Ingresá al menos un PIN para cambiar', true); return; }
  try {
    const r = await fetch('/api/config/pines', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify(updates)});
    const d = await r.json();
    if (d.ok) {
      const msg = document.getElementById('pines2-msg');
      if (msg) { msg.style.color='#2ecc71'; msg.textContent='✓ PINs actualizados — todos los dispositivos serán deslogueados'; setTimeout(()=>msg.textContent='',5000); }
      PIN_ROLES.forEach(ro => { const inp=document.getElementById('pin2-'+ro.key); if(inp) inp.value=''; });
      showToast('PINs guardados — sesiones cerradas en todos los dispositivos');
      setTimeout(()=>{ fetch('/api/logout',{method:'POST',headers:{'Content-Type':'application/json'},body:'{}'}); setTimeout(()=>location.href='/',1500); }, 2500);
    } else { showToast(d.error||'Error', true); }
  } catch(e) { showToast('Error de conexión', true); }
}

// ══════════════════════════════════════════
//  CAMBIO DE CONTRASEÑA (legacy compatibilidad)
// ══════════════════════════════════════════

// ══════════════════════════════════════════
//  POPUP MESAS (modo manual)
// ══════════════════════════════════════════
function toggleMesasPopup(caja) {
  const popup = document.getElementById('mesas-popup-'+caja);
  if (!popup) return;
  const visible = popup.style.display !== 'none';
  // Cerrar todos los otros
  [1,2,3].forEach(c => {
    const p = document.getElementById('mesas-popup-'+c);
    if (p) p.style.display = 'none';
  });
  if (!visible) {
    buildMesasPopup(caja);
    popup.style.display = 'block';
    // Cerrar al hacer click fuera
    setTimeout(() => {
      document.addEventListener('click', function closePop(e) {
        if (!popup.contains(e.target) && e.target.id !== 'btn-mesas-popup-'+caja) {
          popup.style.display = 'none';
          document.removeEventListener('click', closePop);
        }
      });
    }, 10);
  }
}

function buildMesasPopup(caja) {
  const list = document.getElementById('mesas-popup-list-'+caja);
  if (!list) return;
  // Obtener mesas únicas de esta noche con sus nombres y totales
  const mesaMap = {};
  txData.forEach(t => {
    if (t.mesa) {
      if (!mesaMap[t.mesa]) mesaMap[t.mesa] = { names: new Set(), total: 0 };
      mesaMap[t.mesa].names.add(t.name);
      mesaMap[t.mesa].total += t.amount;
    }
  });
  const mesas = Object.keys(mesaMap).sort((a,b) => Number(a)-Number(b) || a.localeCompare(b));
  if (!mesas.length) {
    list.innerHTML = '<div style="padding:12px 14px;color:#444;font-size:12px;letter-spacing:1px;">Sin mesas esta noche</div>';
    return;
  }
  list.innerHTML = mesas.map(m => {
    const info = mesaMap[m];
    const nombres = [...info.names].join(', ');
    return `<div onclick="seleccionarMesaPopup(${caja},'${m}')" style="padding:10px 14px;cursor:pointer;border-bottom:1px solid #1a1a1a;transition:background .15s;" onmouseover="this.style.background='#1a1600'" onmouseout="this.style.background='transparent'">
      <div style="font-family:'Oswald',sans-serif;font-size:18px;color:#fff;font-weight:700;">Mesa ${m}</div>
      <div style="font-size:11px;color:#666;letter-spacing:1px;margin-top:1px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;">${nombres}</div>
    </div>`;
  }).join('');
}

function seleccionarMesaPopup(caja, mesa) {
  const input = document.getElementById('mmesa'+caja);
  if (input) { input.value = mesa; }
  const popup = document.getElementById('mesas-popup-'+caja);
  if (popup) popup.style.display = 'none';
}

// ══════════════════════════════════════════
//  LABEL GASTO EN BARRA INFERIOR
// ══════════════════════════════════════════
function updateCajaGastoLabel(caja) {
  const el = document.getElementById('caja-gasto-label-'+caja);
  if (!el) return;
  const ta = window['cajaState'+caja];
  // Modo tarjeta: mostrar lo gastado en esta sesión de tarjeta
  const modoTarjetaActivo = document.getElementById('modo-content-tarjeta-'+caja)?.classList.contains('active');
  if (ta && modoTarjetaActivo) {
    const gastado = ta.saldo_inicial - ta.saldo_actual;
    el.textContent = gastado > 0 ? 'Tarjeta · ' + fmt(gastado) + ' gastado' : 'Tarjeta · sin cobros';
    el.style.color = '#c9a227';
  } else {
    // Modo manual: sumar lo registrado manualmente en esta caja hoy
    const totalManual = txData.filter(t => t.caja === caja && !t.tarjeta_codigo).reduce((s,t) => s+t.amount, 0);
    el.textContent = totalManual > 0 ? 'Manual · ' + fmt(totalManual) : 'Sin cobros manuales';
    el.style.color = '#888';
  }
}

// Init
buildColorGrid();
cargarPersonalizacionGuardada();
cargarTemaGuardado();
cargarEfectoGuardado();
cargarConfTarjetas().then(()=>{loadData();setInterval(loadData,2000);setInterval(sincronizarConfTarjetas,120000);});
setInterval(pollPublicidad, 5000);
renderMenuMgr();
cargarCartelPrecios();
cargarEstadoPub();
setInterval(cargarEstadoPub, 10000);
</script>
</body>
</html>
"""

HUB_HTML = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Jagger VIP — Hub</title>
<link href="https://fonts.googleapis.com/css2?family=Oswald:wght@400;600;700&family=Rajdhani:wght@400;600;700&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0;}
body{background:#060606;color:#e0e0e0;font-family:'Rajdhani',sans-serif;min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;padding:24px;}
.hub-logo{font-family:'Oswald',sans-serif;font-size:clamp(36px,7vw,72px);font-weight:700;letter-spacing:8px;color:#c9a227;text-transform:uppercase;text-align:center;text-shadow:0 0 50px rgba(201,162,39,0.3);line-height:1;}
.hub-tagline{font-size:12px;color:#666;letter-spacing:5px;text-transform:uppercase;text-align:center;margin-top:8px;margin-bottom:3px;}
.hub-club{font-size:18px;color:#bbb;letter-spacing:4px;text-transform:uppercase;text-align:center;margin-bottom:3px;font-family:'Rajdhani',sans-serif;font-weight:600;}
.hub-version{font-size:10px;color:#444;letter-spacing:3px;text-align:center;margin-bottom:20px;}
.hub-session{display:inline-flex;align-items:center;gap:7px;background:#0d0d0d;border:1px solid #1e1e1e;border-radius:24px;padding:9px 22px;font-family:'Rajdhani',sans-serif;font-size:13px;letter-spacing:2px;color:#666;transition:all .4s;}
.hub-session.active{border-color:#c9a22766;color:#e8c84a;background:#0d0b00;}
.hub-session .dot{width:8px;height:8px;border-radius:50%;background:#333;flex-shrink:0;transition:background .4s;}
.hub-session.active .dot{background:#2ecc71;box-shadow:0 0 10px #2ecc71;}
.hub-divider{width:100%;max-width:560px;height:1px;background:linear-gradient(to right,transparent,#1a1a1a,transparent);margin-bottom:20px;}
.hub-section-label{font-size:10px;color:#555;letter-spacing:4px;text-transform:uppercase;width:100%;max-width:560px;margin-bottom:10px;padding-left:2px;}
.hub-grid{display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;width:100%;max-width:560px;margin-bottom:16px;}
@media(max-width:480px){.hub-grid{grid-template-columns:1fr 1fr;}}
.hub-btn{display:flex;flex-direction:column;align-items:center;justify-content:center;gap:5px;padding:22px 10px;border:1px solid #161616;border-radius:14px;background:#0b0b0b;cursor:pointer;text-decoration:none;color:#e0e0e0;transition:all .25s;font-family:'Oswald',sans-serif;position:relative;overflow:hidden;}
.hub-btn::after{content:'';position:absolute;inset:0;background:linear-gradient(135deg,rgba(255,255,255,0.02),transparent);pointer-events:none;}
.hub-btn:hover{transform:translateY(-3px);box-shadow:0 8px 28px rgba(0,0,0,0.5);}
.hub-btn .icon{font-size:30px;filter:drop-shadow(0 2px 6px rgba(0,0,0,0.5));}
.hub-btn .label{font-size:14px;letter-spacing:2px;font-weight:700;text-transform:uppercase;}
.hub-btn .sublabel{font-size:9px;color:#333;letter-spacing:2px;text-transform:uppercase;}
.hub-btn.gold{border-color:#2a2200;background:linear-gradient(135deg,#0d0b00,#111000);}
.hub-btn.gold .label{color:#c9a227;}
.hub-btn.gold:hover{border-color:#c9a22766;box-shadow:0 8px 28px rgba(201,162,39,0.12);}
.hub-btn.pantalla{border-color:#0d200d;background:linear-gradient(135deg,#08100a,#0a110a);}
.hub-btn.pantalla .label{color:#2ecc71;}
.hub-btn.pantalla:hover{border-color:#2ecc7155;box-shadow:0 8px 28px rgba(46,204,113,0.1);}
.hub-btn.caja{border-color:#181818;background:linear-gradient(135deg,#0e0e0e,#0b0b0b);}
.hub-btn.caja .label{color:#aaa;}
.hub-btn.caja:hover{border-color:#444;box-shadow:0 8px 24px rgba(255,255,255,0.04);}
</style>
</head>
<body>
<div class="hub-logo">RANKING VIP</div>
<div class="hub-tagline">Sistema de gestión</div>
<div class="hub-club">JAGGER CLUB</div>
<div class="hub-version">v1.1</div>
<div style="display:flex;align-items:center;gap:10px;margin-bottom:28px;flex-wrap:wrap;justify-content:center;">
  <div class="hub-session" id="hub-session"><span class="dot"></span><span id="hub-session-txt">Verificando sesión...</span></div>
  <button id="hub-logout-btn" onclick="cerrarSesion()" style="display:none;align-items:center;gap:6px;background:transparent;border:1px solid #6a2a00;border-radius:24px;padding:9px 20px;font-family:'Rajdhani',sans-serif;font-size:12px;letter-spacing:2px;color:#cc5522;cursor:pointer;transition:all .2s;font-weight:700;" onmouseover="this.style.background='#2a1200';this.style.borderColor='#cc4400';this.style.color='#ff7744';" onmouseout="this.style.background='#1a0a00';this.style.borderColor='#6a2a00';this.style.color='#cc5522';">⏻ Cerrar sesión</button>
</div>
<div class="hub-section-label">Administración</div>
<div class="hub-grid" style="grid-template-columns:1fr 1fr 1fr;">
  <a href="/manager" class="hub-btn gold">
    <span class="icon">⚙️</span>
    <span class="label">Manager</span>
    <span class="sublabel">Control total</span>
  </a>
  <a href="/tarjetas" class="hub-btn gold">
    <span class="icon">💳</span>
    <span class="label">Tarjetas</span>
    <span class="sublabel">Config. mesas</span>
  </a>
  <a href="/pantalla" class="hub-btn pantalla">
    <span class="icon">📺</span>
    <span class="label">Pantalla</span>
    <span class="sublabel">Display vivo</span>
  </a>
</div>
<div class="hub-divider"></div>
<div class="hub-section-label">Cajas</div>
<div class="hub-grid" style="grid-template-columns:1fr 1fr 1fr;">
  <a href="/cajaabajo" class="hub-btn caja">
    <span class="icon">🎫</span>
    <span class="label">Abajo</span>
    <span class="sublabel">Caja 1</span>
  </a>
  <a href="/cajaextendido" class="hub-btn caja">
    <span class="icon">🎫</span>
    <span class="label">Extendido</span>
    <span class="sublabel">Caja 2</span>
  </a>
  <a href="/cajavip" class="hub-btn caja">
    <span class="icon">💎</span>
    <span class="label">VIP</span>
    <span class="sublabel">Caja 3</span>
  </a>
</div>
<div class="hub-divider"></div>
<div class="hub-grid" style="grid-template-columns:1fr;">
  <a id="btn-celular" href="/celular" class="hub-btn" style="opacity:.35;pointer-events:none;filter:grayscale(1);grid-column:1;background:linear-gradient(135deg,#0a0a14,#101020);border-color:#1a1a3a;">
    <span class="icon">📱</span>
    <span class="label">Celular</span>
    <span class="sublabel" id="celular-sublabel">Requiere sesión Manager</span>
  </a>
</div>
<script>
function cerrarSesion() {
  fetch('/api/logout',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({})})
    .then(()=>{ location.reload(); })
    .catch(()=>location.reload());
}
fetch('/api/session').then(r=>r.json()).then(s=>{
  const el=document.getElementById('hub-session');
  const txt=document.getElementById('hub-session-txt');
  const btn=document.getElementById('hub-logout-btn');
  const btnCelular=document.getElementById('btn-celular');
  const celularSub=document.getElementById('celular-sublabel');
  if(s.manager){
    el.classList.add('active');
    txt.textContent='Sesión activa — Manager';
    if(btn) btn.style.display='inline-flex';
    // Desbloquear celular
    if(btnCelular){
      btnCelular.style.opacity='1';
      btnCelular.style.pointerEvents='auto';
      btnCelular.style.filter='none';
      btnCelular.style.borderColor='#1a2a5a';
      btnCelular.style.background='linear-gradient(135deg,#08081a,#101028)';
    }
    if(celularSub) celularSub.textContent='Monitor en vivo';
  } else {
    const cajas=[];
    if(s.cajaabajo) cajas.push('Caja Abajo');
    if(s.cajaextendido) cajas.push('Caja Extendido');
    if(s.cajavip) cajas.push('Caja VIP');
    if(cajas.length>0){
      el.classList.add('active');
      txt.textContent='Sesión activa — '+cajas.join(' · ');
      if(btn) btn.style.display='inline-flex';
    } else {
      txt.textContent='Sin sesión activa';
      if(btn) btn.style.display='none';
    }
  }
}).catch(()=>{
  document.getElementById('hub-session-txt').textContent='Sin sesión activa';
});
</script>
<div style="position:fixed;bottom:12px;right:16px;font-family:'Rajdhani',sans-serif;font-size:11px;color:#333;letter-spacing:2px;pointer-events:none;">v1.1</div>
</body>
</html>"""

PANTALLA_MGR_HTML = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Control de Pantalla — Jagger VIP</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=Oswald:wght@400;500;600;700&family=Rajdhani:wght@400;500;600&display=swap');
:root{--gold:#c9a227;--gold-light:#e8c84a;--gold-dim:#7a6010;--black:#080808;--surface:#111;--border:#2a2a2a;--text:#f0ece0;--text-dim:#555;--danger:#a83030;--white:#fff;--green:#2ecc71;--surface-gold:#0d0b00;}
*{box-sizing:border-box;margin:0;padding:0;}
body{background:var(--black);color:var(--text);font-family:'Rajdhani',sans-serif;min-height:100vh;}
.tabs-bar{display:flex;background:#0a0a0a;border-bottom:1px solid #222;position:sticky;top:0;z-index:100;align-items:stretch;}
.tab-btn{flex:1;padding:13px 6px;text-align:center;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:12px;font-weight:600;letter-spacing:2px;text-transform:uppercase;color:var(--text-dim);background:none;border:none;border-bottom:2px solid transparent;transition:all .2s;}
.tab-btn:hover{color:#999;}.tab-btn.active{color:var(--gold);border-bottom:2px solid var(--gold);}
.tab-btn .dot{display:inline-block;width:5px;height:5px;border-radius:50%;background:var(--gold);margin-left:5px;vertical-align:middle;animation:blink 2s infinite;}
@keyframes blink{0%,100%{opacity:1}50%{opacity:.2}}
.nav-link{display:flex;align-items:center;padding:0 14px;color:#444;font-family:'Rajdhani',sans-serif;font-size:11px;letter-spacing:1px;text-decoration:none;white-space:nowrap;transition:color .2s;border-bottom:2px solid transparent;}
.nav-link:hover{color:#888;}
.screen{display:none;padding:22px 26px 70px;}.screen.active{display:block;}
.config-panel{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:14px 18px;margin-bottom:20px;}
.config-title{color:var(--gold-dim);font-size:10px;letter-spacing:2px;text-transform:uppercase;margin-bottom:10px;}
.config-row{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin-bottom:8px;}.config-row:last-child{margin-bottom:0;}
.config-label{color:var(--text-dim);font-size:12px;white-space:nowrap;}
.config-input{background:#0d0d0d;border:1px solid var(--border);border-radius:6px;color:var(--text);padding:8px 12px;font-family:'Rajdhani',sans-serif;font-size:14px;}
.config-input:focus{outline:none;border-color:var(--gold);}.config-input.wide{flex:1;min-width:180px;}.config-input.narrow{width:110px;}
.btn-pres{background:var(--gold);color:#000;border:none;border-radius:6px;padding:9px 20px;font-family:'Rajdhani',sans-serif;font-size:13px;font-weight:700;letter-spacing:1px;cursor:pointer;white-space:nowrap;transition:background .15s;}
.btn-pres:hover{background:var(--gold-light);}
.btn-reset{background:transparent;color:#555;border:1px solid #222;border-radius:6px;padding:9px 16px;font-family:'Rajdhani',sans-serif;font-size:13px;cursor:pointer;white-space:nowrap;transition:all .15s;}
.btn-reset:hover{border-color:var(--danger);color:#cc4444;}
.btn-show-winner{background:transparent;color:#c9a227;border:1px solid #4a3a00;border-radius:6px;padding:9px 16px;font-family:'Rajdhani',sans-serif;font-size:13px;cursor:pointer;white-space:nowrap;transition:all .15s;}
.btn-show-winner:hover{background:#1a1400;border-color:#c9a227;}
.pres-header-wrap{position:relative;margin-bottom:28px;}
.pres-clock{position:absolute;top:0;left:0;text-align:left;}
.pres-clock-hora{font-family:'Oswald',sans-serif;font-size:13px;color:var(--text-dim);letter-spacing:1px;text-transform:uppercase;}
.pres-clock-time{font-family:'Oswald',sans-serif;font-size:32px;color:#fff;font-weight:700;line-height:1.1;}
.pres-clock-fin{font-size:11px;color:var(--gold-dim);letter-spacing:1px;margin-top:6px;text-transform:uppercase;}
.pres-clock-fin-val{font-family:'Oswald',sans-serif;font-size:26px;color:var(--gold);font-weight:700;line-height:1.1;display:block;}
.pres-header{text-align:center;padding-top:8px;}
.pres-logo{font-family:'Oswald',sans-serif;font-size:52px;font-weight:700;color:var(--white);letter-spacing:10px;text-transform:uppercase;display:block;width:100%;}
.pres-logo .vip{color:var(--gold);}.pres-line{height:1px;background:linear-gradient(to right,transparent,var(--gold),transparent);margin:12px auto;max-width:100%;}
.live-badge{display:inline-flex;align-items:center;gap:6px;border:1px solid #2a2a2a;border-radius:20px;padding:4px 14px;font-size:11px;color:#777;letter-spacing:1px;margin-top:6px;}
.live-dot{width:6px;height:6px;border-radius:50%;background:#3a9a5a;animation:blink 1.5s infinite;}
.ranking-wrap{max-width:100%;margin:0;padding:0 10px;}
.rank-header{display:grid;grid-template-columns:100px 1fr 130px 180px;background:#0d0d0d;border:1px solid #2a2a2a;border-radius:8px 8px 0 0;padding:12px 30px;margin-bottom:3px;}
.rank-header span{font-family:'Oswald',sans-serif;font-size:13px;font-weight:500;letter-spacing:2px;text-transform:uppercase;color:var(--gold-dim);}
.rank-header .col-r{text-align:right;}
.rank-rows{display:flex;flex-direction:column;gap:4px;}
.rank-row{display:grid;grid-template-columns:100px 1fr 130px 180px;align-items:center;background:var(--surface);border:1px solid #1e1e1e;border-radius:6px;padding:20px 30px;transition:border-color .3s,background .4s;}
.rank-row.rank-1{background:var(--surface-gold);border-color:var(--gold-dim);}
.rank-row.nueva{animation:entradaFila .55s cubic-bezier(.22,1,.36,1) both;}
@keyframes entradaFila{from{opacity:0;transform:translateX(-32px)}to{opacity:1;transform:none}}
@keyframes crownGlow{0%{box-shadow:0 0 0px rgba(201,162,39,0);}40%{box-shadow:0 0 32px rgba(201,162,39,.55);}100%{box-shadow:none;}}
.rank-row.rank-1.ascendio{animation:crownGlow .9s ease both;}
.col-puesto{font-family:'Oswald',sans-serif;font-size:28px;font-weight:700;color:#444;}
.rank-row.rank-1 .col-puesto{color:var(--gold);font-size:34px;}.rank-row.rank-2 .col-puesto{color:#aaa;}.rank-row.rank-3 .col-puesto{color:#8a6a40;}
.col-nombre{font-family:'Oswald',sans-serif;font-size:30px;font-weight:600;color:var(--white);}.rank-row.rank-1 .col-nombre{font-size:36px;}
.col-mesa{font-family:'Oswald',sans-serif;font-size:24px;font-weight:700;color:#e8e8e8;letter-spacing:1px;}
.col-total{font-family:'Oswald',sans-serif;font-size:30px;font-weight:700;color:var(--gold);text-align:right;}.rank-row.rank-1 .col-total{font-size:38px;}
.miles-lbl{font-size:0.45em;opacity:0.55;letter-spacing:2px;margin-left:5px;vertical-align:middle;font-weight:600;}
.empty-msg{text-align:center;color:#222;font-size:15px;padding:70px 20px;letter-spacing:2px;font-family:'Oswald',sans-serif;}
.premio-wrap{text-align:center;margin-top:36px;padding-bottom:20px;}
.premio-box{display:inline-block;background:var(--gold);color:#000;font-family:'Rajdhani',sans-serif;font-size:22px;font-weight:700;letter-spacing:.5px;padding:14px 44px;border-radius:6px;}.premio-box:empty{display:none;}
#winner-overlay{display:none;position:fixed;inset:0;z-index:11000;background:rgba(0,0,0,0.97);flex-direction:column;align-items:center;justify-content:center;text-align:center;}
#winner-overlay.show{display:flex;}
.winner-rays{position:absolute;inset:0;overflow:hidden;pointer-events:none;}
.winner-content{position:relative;z-index:2;padding:0 20px;}
.winner-corona{display:block;font-size:72px;margin-bottom:12px;animation:bounce 1s infinite;}
@keyframes bounce{0%,100%{transform:scale(1);}50%{transform:scale(1.1);}}
.winner-titulo{font-family:'Oswald',sans-serif;font-size:clamp(20px,3vw,32px);color:var(--gold);letter-spacing:4px;font-weight:700;margin-bottom:16px;}
.winner-nombre{font-family:'Oswald',sans-serif;font-size:clamp(52px,9vw,110px);font-weight:700;color:#fff;letter-spacing:4px;line-height:1;}
.winner-line{height:2px;background:linear-gradient(to right,transparent,var(--gold),transparent);margin:18px auto;max-width:500px;}
.winner-info-row{display:flex;gap:48px;justify-content:center;margin-bottom:14px;}
.winner-info-block{text-align:center;}.winner-info-label{font-size:12px;color:#555;letter-spacing:2px;text-transform:uppercase;}
.winner-info-val{font-family:'Oswald',sans-serif;font-size:36px;color:var(--gold);font-weight:700;}
.winner-close{position:fixed;top:20px;right:24px;background:transparent;border:none;color:#333;font-size:28px;cursor:pointer;z-index:11001;}
.confetti-piece{position:absolute;top:-20px;animation:confettiFall linear infinite;}
@keyframes confettiFall{to{transform:translateY(110vh) rotate(720deg);opacity:0;}}
.conf-header{margin-bottom:20px;}.conf-title{font-family:'Oswald',sans-serif;font-size:22px;font-weight:700;color:var(--gold);letter-spacing:3px;text-transform:uppercase;margin-bottom:4px;}
.conf-sub{font-size:12px;color:var(--text-dim);letter-spacing:1px;}
.custom-section{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:18px 20px;margin-bottom:16px;}
.custom-section-title{font-family:'Oswald',sans-serif;font-size:13px;color:var(--gold-dim);letter-spacing:2px;text-transform:uppercase;margin-bottom:14px;}
.custom-text-row{display:flex;align-items:center;gap:10px;margin-bottom:10px;}
.custom-text-label{font-size:12px;color:var(--text-dim);width:180px;flex-shrink:0;letter-spacing:1px;}
.custom-text-input{flex:1;background:#0d0d0d;border:1px solid var(--border);border-radius:6px;color:var(--text);padding:8px 12px;font-family:'Rajdhani',sans-serif;font-size:14px;}
.custom-text-input:focus{outline:none;border-color:var(--gold);}
.btn-custom-save{background:var(--gold);color:#000;border:none;border-radius:7px;padding:12px 28px;font-family:'Rajdhani',sans-serif;font-size:15px;font-weight:700;letter-spacing:1px;cursor:pointer;transition:background .15s;}
.btn-custom-save:hover{background:var(--gold-light);}
.btn-custom-reset{background:transparent;color:#555;border:1px solid #222;border-radius:7px;padding:12px 20px;font-family:'Rajdhani',sans-serif;font-size:14px;cursor:pointer;transition:all .15s;}
.btn-custom-reset:hover{border-color:var(--danger);color:#cc4444;}
.color-grid{display:flex;flex-direction:column;gap:10px;margin-top:4px;}
.color-item{display:flex;align-items:center;gap:12px;background:#0d0d0d;border:1px solid var(--border);border-radius:8px;padding:10px 14px;}
.color-swatch{width:36px;height:36px;border-radius:6px;flex-shrink:0;overflow:hidden;position:relative;cursor:pointer;}
.color-swatch input[type=color]{position:absolute;inset:-4px;width:calc(100%+8px);height:calc(100%+8px);opacity:0;cursor:pointer;}
.color-label{font-size:12px;color:var(--text);letter-spacing:.5px;margin-bottom:2px;}.color-hex{font-size:11px;color:var(--text-dim);font-family:monospace;}
.preview-bar{background:#0a0a0a;border:1px solid #1a1a1a;border-radius:8px;padding:18px 24px;margin-top:12px;text-align:center;}
.preview-label{font-size:10px;color:var(--text-dim);letter-spacing:2px;text-transform:uppercase;margin-bottom:10px;}
.preview-logo{font-family:'Oswald',sans-serif;font-size:36px;font-weight:700;color:var(--white);letter-spacing:8px;}.vip-prev{color:var(--gold);}
#tema-overlay{position:fixed;inset:0;pointer-events:none;z-index:2;overflow:hidden;opacity:0;transition:opacity 1s;}
#efectos-overlay{position:fixed;inset:0;pointer-events:none;z-index:3;overflow:hidden;}
.toast{position:fixed;bottom:24px;right:24px;background:#1a1a1a;color:var(--text);border:1px solid var(--border);border-radius:8px;padding:10px 20px;font-family:'Rajdhani',sans-serif;font-size:14px;letter-spacing:1px;opacity:0;transform:translateY(10px);transition:all .25s;z-index:99999;pointer-events:none;}
.toast.show{opacity:1;transform:translateY(0);}.toast.error{border-color:#6a2020;color:#e74c3c;}
/* TEMA CSS */
body.tema-jagger12 .rank-row{border-color:#1e1e1e;}
body.tema-jagger12 .rank-row.rank-1{background:#1a1200;border-color:#c9a227;}
body.tema-jagger12 .rank-row.rank-2{background:#0e0e0e;border-color:#777;}
body.tema-jagger12 .rank-row.rank-3{background:#0e0800;border-color:#7a4a20;}
body.tema-jagger12 .rank-row.rank-1 .col-puesto{color:#c9a227;font-size:34px;}
body.tema-jagger12 .rank-row.rank-2 .col-puesto{color:#aaaaaa;}
body.tema-jagger12 .rank-row.rank-3 .col-puesto{color:#cd7f32;}
body.tema-jagger12 .rank-row.rank-1 .col-total{color:#e8c84a;}
body.tema-jagger12 .rank-row.rank-2 .col-total{color:#cccccc;}
body.tema-jagger12 .rank-row.rank-3 .col-total{color:#cd7f32;}
body.tema-jagger12 .col-total{color:#ccc;}
body.tema-jagger12 .pres-line{background:linear-gradient(to right,transparent,#c9a227,transparent);}
body.tema-jagger12 .live-dot{background:#c9a227;}
body.tema-jagger12 .col-nombre{color:#fff !important;}
body.tema-jagger12 .col-mesa{color:#ddd !important;}

body.tema-touchofpink .rank-header{background:#2d0022;border-color:#6a2050;}
body.tema-touchofpink .rank-header span{color:#f472b6;}
body.tema-touchofpink .rank-row{border-color:#8a3070;background:#3d002c;}
body.tema-touchofpink .rank-row.rank-1{background:#5a0042;border-color:#f472b6;box-shadow:0 0 20px rgba(244,114,182,.2);}
body.tema-touchofpink .rank-row.rank-1 .col-puesto{color:#f472b6;font-size:34px;}
body.tema-touchofpink .rank-row.rank-2 .col-puesto{color:#ffffff;}
body.tema-touchofpink .rank-row.rank-3 .col-puesto{color:#fbb6ce;}
body.tema-touchofpink .rank-row.rank-1 .col-total{color:#fce7f3;}
body.tema-touchofpink .col-total{color:#fbb6ce;}
body.tema-touchofpink .col-nombre{color:#ffffff !important;}
body.tema-touchofpink .col-mesa{color:#fce7f3 !important;}
body.tema-touchofpink .col-puesto{color:#eeaad8;}
body.tema-touchofpink .pres-line{background:linear-gradient(to right,transparent,#f472b6,transparent);}
body.tema-touchofpink .live-dot{background:#f472b6;}
body.tema-touchofpink .live-badge{border-color:#8a3070;color:#f472b6;}
body.tema-touchofpink.pink-claro{--black:#3a0028;--surface:#580040;--border:#a04080;--gold:#f472b6;--text:#ffe8f5;}
@keyframes bottleBounce{0%,100%{transform:translateY(0) rotate(-5deg) scale(1);}20%{transform:translateY(-18px) rotate(5deg) scale(1.05);}40%{transform:translateY(-6px) rotate(-3deg) scale(0.98);}60%{transform:translateY(-14px) rotate(4deg) scale(1.03);}80%{transform:translateY(-4px) rotate(-2deg) scale(0.99);}}
@keyframes winnerEntrada{from{opacity:0;transform:scale(0.7) translateY(40px)}to{opacity:1;transform:scale(1) translateY(0)}}
@keyframes rayPulse2{0%{opacity:0.2;transform:rotate(var(--r,0deg)) scaleY(0.5)}100%{opacity:0.7;transform:rotate(var(--r,0deg)) scaleY(1)}}
@keyframes goldTwinkle{0%,100%{opacity:0.1;transform:scale(1)}50%{opacity:0.28;transform:scale(1.4)}}
@keyframes bubbleRise{0%{opacity:0.6;transform:translateY(0) scale(1)}50%{opacity:0.3}100%{opacity:0;transform:translateY(-100vh) scale(0.5)}}
@keyframes copoFall{0%{transform:translateY(-40px) rotate(0deg);opacity:1}100%{transform:translateY(105vh) rotate(360deg);opacity:0}}
@keyframes sparkBoxP{0%{opacity:1;transform:scale(1)}100%{opacity:0;transform:translate(var(--bx),var(--by)) scale(0)}}
@keyframes guanteFallP{0%{opacity:0.8;transform:translateY(-40px) rotate(0deg)}100%{opacity:0;transform:translateY(105vh) rotate(360deg)}}
body.modo-presentacion .tabs-bar{display:none;}
body.modo-presentacion .config-panel{display:none;}
.btn-exit-pres{display:none;position:fixed;top:14px;right:18px;z-index:9999;background:transparent;color:#2a2a2a;border:none;font-size:22px;cursor:pointer;padding:4px 8px;transition:color .3s;line-height:1;opacity:0.4;}
.btn-exit-pres:hover{color:#888;opacity:1;}
body.modo-presentacion .btn-exit-pres{display:block;}
</style>
</head>
<body>
<div id="tema-overlay"></div>
<div id="efectos-overlay"></div>
<button class="btn-exit-pres" onclick="salirPresentacion()">✕</button>
<div class="tabs-bar">
  <button class="tab-btn active" onclick="showTab('pantalla')" id="tbtn-pantalla">📺 Pantalla <span class="dot"></span></button>
  <button class="tab-btn" onclick="showTab('diseno')" id="tbtn-diseno">🎨 Diseño</button>
  <a href="/" class="nav-link">← Hub</a>
</div>

<!-- PANTALLA TAB -->
<div id="tab-pantalla" class="screen active">
  <div class="config-panel">
    <div class="config-title">Pantalla</div>
    <div class="config-row">
      <button class="btn-pres" onclick="activarPresentacion()">⛶ Modo Presentacion</button>
      <button class="btn-show-winner" onclick="mostrarGanadorManual()">🏆 Mostrar Ganador Ahora</button>
      <button id="btn-quitar-cartel" onclick="quitarCartel()" style="display:none;background:transparent;color:#e74c3c;border:1px solid #5a1a1a;border-radius:6px;padding:9px 16px;font-family:'Rajdhani',sans-serif;font-size:13px;font-weight:700;cursor:pointer;letter-spacing:1px;white-space:nowrap;transition:all .15s;" onmouseover="this.style.background='#1a0000'" onmouseout="this.style.background='transparent'">✕ Quitar cartel</button>
    </div>
  </div>
  <div class="pres-header-wrap">
    <div class="pres-clock">
      <div class="pres-clock-hora">Hora</div>
      <div class="pres-clock-time" id="clock-display">00:00</div>
      <div class="pres-clock-fin">Finaliza</div>
      <span class="pres-clock-fin-val" id="clock-fin">05:30</span>
    </div>
    <div class="pres-header">
      <div class="pres-logo" id="main-logo">RANKING <span class="vip" id="logo-vip">VIP</span></div>
      <div id="tema-tagline" style="font-family:'Rajdhani',sans-serif;font-size:28px;font-weight:600;color:#555;letter-spacing:5px;text-transform:uppercase;text-align:center;margin-top:2px;min-height:0;">JAGGER CLUB</div>
      <div class="pres-line"></div>
      <div class="live-badge"><span class="live-dot"></span> EN VIVO</div>
    </div>
  </div>
  <div class="ranking-wrap">
    <div class="rank-header" id="rank-header" style="display:none">
      <span>PUESTO</span><span>NOMBRE</span><span>MESA</span><span class="col-r">TOTAL</span>
    </div>
    <div class="rank-rows" id="rank-rows"></div>
    <div id="empty-msg" class="empty-msg">Aun no hay consumos registrados</div>
  </div>
  <div class="premio-wrap"><div class="premio-box" id="premio-box"></div></div>
</div>

<!-- DISEÑO TAB -->
<div id="tab-diseno" class="screen">
  <div class="conf-header">
    <div class="conf-title">🎨 Personalización</div>
    <div class="conf-sub">Cambia colores, textos y temas. Los cambios se sincronizan con la pantalla en tiempo real.</div>
  </div>
  <div class="custom-section">
    <div class="custom-section-title">🎉 Temas de noche especial</div>
    <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(180px,1fr));gap:10px;margin-bottom:14px;">
      <button onclick="aplicarTema('default')" style="background:#111;border:1px solid #2a2a2a;border-radius:10px;padding:16px 10px;cursor:pointer;text-align:center;color:#888;font-family:'Rajdhani',sans-serif;font-weight:600;letter-spacing:1px;font-size:13px;transition:all .2s;" onmouseover="this.style.borderColor='#555'" onmouseout="this.style.borderColor='#2a2a2a'">
        <div style="font-size:28px;margin-bottom:6px;">⬛</div>DEFAULT
      </button>
      <button onclick="aplicarTema('jagger12')" style="background:#000;border:2px solid #333;border-radius:10px;padding:16px 10px;cursor:pointer;text-align:center;color:#fff;font-family:'Rajdhani',sans-serif;font-weight:700;letter-spacing:1px;font-size:13px;transition:all .2s;" onmouseover="this.style.borderColor='#888'" onmouseout="this.style.borderColor='#333'">
        <div style="font-size:28px;margin-bottom:6px;">🥂</div>JAGGER 12 AÑOS
      </button>

      <button onclick="aplicarTema('touchofpink')" style="background:linear-gradient(135deg,#140010,#2a0020);border:2px solid #9d174d;border-radius:10px;padding:16px 10px;cursor:pointer;text-align:center;color:#f472b6;font-family:'Rajdhani',sans-serif;font-weight:700;letter-spacing:1px;font-size:13px;transition:all .2s;" onmouseover="this.style.borderColor='#f472b6'" onmouseout="this.style.borderColor='#9d174d'">
        <div style="font-size:28px;margin-bottom:6px;">🌸</div>TURNS PINK
      </button>
    </div>
    <div id="tema-deco-toggle" style="display:none;background:#0a0a0a;border:1px solid #222;border-radius:8px;padding:12px 16px;margin-top:8px;align-items:center;gap:12px;flex-wrap:wrap;">
      <label id="deco-main-label" style="display:flex;align-items:center;gap:8px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;color:#888;letter-spacing:1px;">
        <input type="checkbox" id="toggle-deco" onchange="toggleDecoActual(this.checked)" style="width:16px;height:16px;accent-color:#c9a227;" checked />
        <span id="toggle-deco-label">Activar decoraciones animadas</span>
      </label>
      <div id="show-12-toggle" style="display:none;">
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;color:#888;letter-spacing:1px;">
          <input type="checkbox" id="toggle-12" onchange="mostrar12Fondo=this.checked;reiniciarDeco12();fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({mostrar12:this.checked})}).catch(()=>{});" style="width:16px;height:16px;accent-color:#c9a227;" checked />
          <span>Mostrar "12" de fondo</span>
        </label>
      </div>
      <div id="falling-gloves-toggle" style="display:none;">
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;color:#888;letter-spacing:1px;">
          <input type="checkbox" id="toggle-falling-gloves" onchange="fallingGlovesActivos=this.checked;fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({falling_gloves:this.checked})}).catch(()=>{});" style="width:16px;height:16px;accent-color:#ff2222;" checked />
          <span>🥊 Guantes cayendo</span>
        </label>
      </div>
      <div id="pink-petalos-toggle" style="display:none;">
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;color:#cc88bb;letter-spacing:1px;">
          <input type="checkbox" id="toggle-pink-petalos" onchange="pinkPetalosActivos=this.checked;if(this.checked){iniciarPetalos();}else{const w=document.getElementById('petalos-wrap');if(w)w.innerHTML='';}fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({petals_activos:this.checked})}).catch(()=>{});" style="width:16px;height:16px;accent-color:#f472b6;" checked />
          <span>🌸 Pétalos cayendo</span>
        </label>
      </div>
      <div id="pink-modo-toggle" style="display:none;flex-direction:column;gap:6px;">
        <div style="font-family:'Rajdhani',sans-serif;font-size:11px;color:#cc88bb;letter-spacing:1px;text-transform:uppercase;margin-bottom:2px;">Fondo</div>
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;color:#cc88bb;letter-spacing:1px;">
          <input type="radio" name="pink-modo" value="oscuro" checked style="accent-color:#f472b6;" onchange="document.body.classList.remove('pink-claro');pinkModoClaro=false;" />
          <span>🌙 Rosa oscuro</span>
        </label>
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;color:#cc88bb;letter-spacing:1px;">
          <input type="radio" name="pink-modo" value="claro" style="accent-color:#f472b6;" onchange="document.body.classList.add('pink-claro');pinkModoClaro=true;" />
          <span>🌸 Rosa medio</span>
        </label>
      </div>
      <div id="ko-anim-toggle" style="display:none;">
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;color:#888;letter-spacing:1px;">
          <input type="checkbox" id="toggle-ko-anim" onchange="koAnimActiva=this.checked;" style="width:16px;height:16px;accent-color:#ff2222;" checked />
          <span>🥊 Animación KO al cambiar de posición</span>
        </label>
      </div>
    </div>
  </div>
  <div class="custom-section">
    <div class="custom-section-title">✨ Efectos de fondo</div>
    <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(140px,1fr));gap:8px;margin-bottom:10px;" id="efectos-grid">
      <button onclick="aplicarEfecto('ninguno')" id="efecto-btn-ninguno" style="background:#0a0a0a;border:1px solid #2a2a2a;border-radius:8px;padding:12px 8px;cursor:pointer;text-align:center;color:#555;font-family:'Rajdhani',sans-serif;font-weight:600;font-size:12px;letter-spacing:1px;transition:all .2s;"><div style="font-size:22px;margin-bottom:4px;">⬛</div>NINGUNO</button>
      <button onclick="aplicarEfecto('burbujas')" id="efecto-btn-burbujas" style="background:#0a0a0a;border:1px solid #2a2a2a;border-radius:8px;padding:12px 8px;cursor:pointer;text-align:center;color:#555;font-family:'Rajdhani',sans-serif;font-weight:600;font-size:12px;letter-spacing:1px;transition:all .2s;"><div style="font-size:22px;margin-bottom:4px;">🫧</div>BURBUJAS</button>
      <button onclick="aplicarEfecto('estrellas')" id="efecto-btn-estrellas" style="background:#0a0a0a;border:1px solid #2a2a2a;border-radius:8px;padding:12px 8px;cursor:pointer;text-align:center;color:#555;font-family:'Rajdhani',sans-serif;font-weight:600;font-size:12px;letter-spacing:1px;transition:all .2s;"><div style="font-size:22px;margin-bottom:4px;">⭐</div>ESTRELLAS</button>
    </div>
    <div class="preview-bar">
      <div class="preview-label">Vista previa del logo</div>
      <div class="preview-logo" id="prev-logo">RANKING <span class="vip-prev" id="prev-vip">VIP</span></div>
    </div>
  </div>
  <div class="custom-section">
    <div class="custom-section-title">🕐 Tamaño del reloj (pantalla)</div>
    <div style="display:flex;gap:8px;margin-top:6px;">
      <button id="cs-chico" onclick="setClockSize('chico')" style="flex:1;background:#0a0a0a;border:1px solid #2a2a2a;border-radius:8px;padding:10px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-weight:600;font-size:13px;color:#555;letter-spacing:1px;transition:all .2s;">Chico</button>
      <button id="cs-mediano" onclick="setClockSize('mediano')" style="flex:1;background:#0a0a0a;border:1px solid var(--gold);border-radius:8px;padding:10px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-weight:600;font-size:13px;color:var(--gold);letter-spacing:1px;transition:all .2s;">Mediano</button>
      <button id="cs-grande" onclick="setClockSize('grande')" style="flex:1;background:#0a0a0a;border:1px solid #2a2a2a;border-radius:8px;padding:10px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-weight:600;font-size:13px;color:#555;letter-spacing:1px;transition:all .2s;">Grande</button>
    </div>
  </div>
  <div class="custom-section">
    <div class="custom-section-title">Textos</div>
    <div class="custom-text-row">
      <span class="custom-text-label">Mensaje del premio (aparece en pantalla)</span>
      <input class="custom-text-input" id="ct-premio" type="text" placeholder="Ej: El ganador se lleva una botella gratis" oninput="updateMsg()" />
    </div>
    <div class="custom-text-row">
      <span class="custom-text-label">Texto junto al nombre (VIP, GOLD...)</span>
      <input class="custom-text-input" id="ct-vip" type="text" placeholder="VIP" value="VIP" oninput="previewTextos()" />
    </div>
    <div class="custom-text-row">
      <span class="custom-text-label">Mensaje ganador</span>
      <input class="custom-text-input" id="ct-winner-msg" type="text" placeholder="¡EL GANADOR DE LA NOCHE!" value="¡EL GANADOR DE LA NOCHE!" />
    </div>
    <div class="custom-text-row">
      <span class="custom-text-label">Subtítulo ganador</span>
      <input class="custom-text-input" id="ct-winner-sub" type="text" placeholder="Ej: ¡Se lleva la botella!" value="" />
    </div>
    <div style="border-top:1px solid #1a1a1a;margin-top:14px;padding-top:14px;">
      <div style="font-size:10px;color:#555;letter-spacing:2px;text-transform:uppercase;margin-bottom:10px;">Texto debajo de "RANKING VIP"</div>
      <div class="custom-text-row">
        <span class="custom-text-label">Texto</span>
        <input class="custom-text-input" id="ct-tagline" type="text" placeholder="JAGGER CLUB" value="JAGGER CLUB" oninput="previewTagline()" />
      </div>
      <div class="custom-text-row" style="align-items:center;">
        <span class="custom-text-label">Color</span>
        <input type="color" id="ct-tagline-color" value="#555555" oninput="previewTagline()" style="width:38px;height:28px;border:none;background:none;cursor:pointer;padding:0;flex-shrink:0;" />
      </div>
      <div class="custom-text-row" style="align-items:center;gap:8px;">
        <span class="custom-text-label">Brillo glow</span>
        <input type="range" id="ct-tagline-brightness" min="0" max="1" step="0.05" value="0" oninput="previewTagline();document.getElementById('ct-tagline-brightness-val').textContent=parseFloat(this.value).toFixed(2)" style="flex:1;" />
        <span id="ct-tagline-brightness-val" style="font-size:11px;color:#555;width:32px;text-align:right;flex-shrink:0;">0.00</span>
      </div>
      <div class="custom-text-row" style="align-items:center;">
        <span class="custom-text-label">Fuente</span>
        <select id="ct-tagline-font" onchange="previewTagline()" style="flex:1;background:#0d0d0d;border:1px solid #2a2a2a;border-radius:6px;color:#f0ece0;padding:7px 8px;font-size:13px;outline:none;">
          <option value="'Rajdhani',sans-serif">Rajdhani</option>
          <option value="'Oswald',sans-serif">Oswald</option>
          <option value="Impact,sans-serif">Impact</option>
          <option value="Arial,sans-serif">Arial</option>
        </select>
      </div>
    </div>
  </div>
  <div class="custom-section">
    <div class="custom-section-title">🏆 Opciones del ganador</div>
    <div style="background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:14px 16px;">
      <div style="font-family:'Rajdhani',sans-serif;font-size:12px;color:#666;letter-spacing:2px;text-transform:uppercase;margin-bottom:10px;">Efecto de partículas</div>
      <div style="display:flex;flex-direction:column;gap:8px;">
        <label style="display:flex;align-items:center;gap:10px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:14px;color:#aaa;letter-spacing:1px;">
          <input type="radio" name="tipo-particula" value="confetti" checked style="accent-color:#c9a227;" onchange="tipoParticula='confetti';confettiGanadorActivo=true;" /><span>🎊 Confetti de colores</span>
        </label>
        <label style="display:flex;align-items:center;gap:10px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:14px;color:#aaa;letter-spacing:1px;">
          <input type="radio" name="tipo-particula" value="billetes" style="accent-color:#c9a227;" onchange="tipoParticula='billetes';confettiGanadorActivo=true;" /><span>💵 Lluvia de billetes</span>
        </label>
        <label style="display:flex;align-items:center;gap:10px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:14px;color:#aaa;letter-spacing:1px;">
          <input type="radio" name="tipo-particula" value="ninguno" style="accent-color:#c9a227;" onchange="tipoParticula='ninguno';confettiGanadorActivo=false;" /><span>Sin partículas</span>
        </label>
      </div>
    </div>
  </div>
  <div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:8px;">
    <button class="btn-custom-save" onclick="aplicarPersonalizacion()">✓ Aplicar cambios</button>
  </div>
</div>

<!-- Winner overlay -->
<div id="winner-overlay">
  <div class="winner-rays"></div>
  <div class="confetti-wrap" id="confetti-wrap"></div>
  <button class="winner-close" onclick="cerrarGanador()">✕</button>
  <div class="winner-content">
    <span class="winner-corona" id="winner-corona">👑</span>
    <div class="winner-titulo" id="winner-titulo">GANADOR DE LA NOCHE</div>
    <div class="winner-nombre" id="winner-nombre">—</div>
    <div class="winner-line"></div>
    <div class="winner-info-row">
      <div class="winner-info-block"><div class="winner-info-label">Mesa</div><div class="winner-info-val" id="winner-mesa">—</div></div>
      <div class="winner-info-block"><div class="winner-info-label">Total</div><div class="winner-info-val" id="winner-total">—</div></div>
    </div>
    <div id="winner-mensaje" style="font-family:'Rajdhani',sans-serif;font-size:20px;color:var(--gold);margin-top:8px;display:none;"></div>
  </div>
</div>

<div id="toast" class="toast"></div>

<script>
let txData = [], horaFin = '05:30', ganadorMostrado = false;
let cartelActivo = false;
let temaActual = 'default', decoActiva = true, koAnimActiva = true;
let fallingGlovesActivos = true, mostrar12Fondo = true;
let confettiGanadorActivo = true, tipoParticula = 'confetti';
let pinkPetalosActivos = true, pinkModoClaro = false;
let efectoActual = 'ninguno', efectoInterval = null;
let lastWinnerTs = 0, prevRankOrder = [], knownNames = new Set();

const COLOR_DEFS = [
  {key:'--gold',       label:'Dorado principal', default:'#c9a227'},
  {key:'--gold-light', label:'Dorado claro (hover)', default:'#e8c84a'},
  {key:'--gold-dim',   label:'Dorado oscuro', default:'#7a6010'},
  {key:'--black',      label:'Fondo principal', default:'#080808'},
  {key:'--surface',    label:'Superficies', default:'#111111'},
  {key:'--border',     label:'Bordes', default:'#2a2a2a'},
  {key:'--text',       label:'Texto principal', default:'#f0ece0'},
  {key:'--text-dim',   label:'Texto secundario', default:'#555555'},
  {key:'--white',      label:'Blanco', default:'#ffffff'},
  {key:'--green',      label:'Verde', default:'#2ecc71'},
  {key:'--danger',     label:'Rojo / peligro', default:'#a83030'},
];
let customColors = {};
COLOR_DEFS.forEach(c => customColors[c.key] = c.default);

const TEMAS_EXTRA = {
  jagger12: {colors:{'--black':'#000000','--surface':'#0a0a0a','--border':'#333333','--gold':'#ffffff','--gold-light':'#dddddd','--gold-dim':'#888888','--text':'#e8e8e8','--text-dim':'#666666','--white':'#ffffff'},bodyClass:'tema-jagger12'},

  touchofpink: {colors:{'--black':'#2d0020','--surface':'#480035','--border':'#8a3070','--gold':'#f472b6','--gold-light':'#fbb6ce','--gold-dim':'#e896cc','--text':'#ffe8f5','--text-dim':'#ddaacc','--white':'#ffffff'},bodyClass:'tema-touchofpink'}
};

function showTab(id) {
  document.querySelectorAll('.screen').forEach(s => s.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
  document.getElementById('tab-'+id).classList.add('active');
  document.getElementById('tbtn-'+id).classList.add('active');
}

function showToast(msg, error=false) {
  const t = document.getElementById('toast');
  t.textContent = msg; t.className = 'toast show' + (error ? ' error' : '');
  setTimeout(() => t.className = 'toast', 2800);
}

function fmt(n) { return '$' + Number(n).toLocaleString('es-AR',{minimumFractionDigits:0,maximumFractionDigits:0}); }

function tickClock() {
  const now = new Date();
  const el = document.getElementById('clock-display');
  if (el) el.textContent = String(now.getHours()).padStart(2,'0') + ':' + String(now.getMinutes()).padStart(2,'0');
  if (horaFin && !ganadorMostrado && txData.length > 0) {
    const [fh, fm] = horaFin.split(':').map(Number);
    if (now.getHours() === fh && now.getMinutes() === fm && now.getSeconds() < 30) {
      ganadorMostrado = true; mostrarGanador(); fetch('/api/winner/show',{method:'POST'});
    }
  }
}
setInterval(tickClock, 1000); tickClock();

function updateMsg() {
  const el = document.getElementById('ct-premio') || document.getElementById('msg-input');
  const val = el ? el.value.trim() : '';
  const pb = document.getElementById('premio-box'); if (pb) pb.textContent = val;
  fetch('/api/state',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({premio:val})});
  fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({premio:val})}).catch(()=>{});
}
function updatePremioSize() { const pb=document.getElementById('premio-box'); const ps=document.getElementById('premio-size'); if(pb&&ps) pb.style.fontSize = ps.value; }

function updateHoraFin() {
  horaFin = document.getElementById('hora-fin-input').value || '05:30';
  ganadorMostrado = false;
  const el = document.getElementById('clock-fin'); if (el) el.textContent = horaFin;
  fetch('/api/state',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({hora_fin:horaFin})});
  fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({hora_fin:horaFin})}).catch(()=>{});
}

async function resetNoche() {
  if (!confirm('Resetear todos los consumos? Los saldos de tarjetas vuelven al valor inicial.')) return;
  await fetch('/api/reset',{method:'POST'});
  ganadorMostrado = false; knownNames.clear(); prevRankOrder = [];
  await loadData(); showToast('Noche reseteada');
}

async function cerrarNoche() {
  if (!confirm('¿Cerrar la noche y guardar en el historial?')) return;
  try {
    const r = await fetch('/api/cerrar_noche',{method:'POST'}); const d = await r.json();
    if (d.ok) { showToast('✓ Noche guardada'); if (confirm('¿Exportar a Excel?')) window.open('/api/export/excel','_blank'); }
    else showToast(d.error||'Error', true);
  } catch(e) { showToast('Error de conexión', true); }
}

function activarPresentacion() {
  document.body.classList.add('modo-presentacion');
  showTab('pantalla');
}
function salirPresentacion() {
  document.body.classList.remove('modo-presentacion');
}

function generarParticulas() {
  const wrap = document.getElementById('confetti-wrap'); wrap.innerHTML = '';
  if (tipoParticula === 'ninguno') return;
  if (tipoParticula === 'billetes') {
    for (let i=0;i<55;i++) { const el=document.createElement('span'); el.style.cssText=`position:absolute;top:-60px;left:${Math.random()*100}vw;font-size:${20+Math.random()*18}px;animation:confettiFall ${4+Math.random()*4}s linear ${Math.random()*5}s infinite;`; el.textContent='💵'; wrap.appendChild(el); }
    return;
  }
  const colors = temaActual==='touchofpink' ? ['#f472b6','#fbb6ce','#ffffff','#ec4899','#fce7f3'] : ['#c9a227','#e8c84a','#fff','#2ecc71','#e74c3c','#3498db','#9b59b6'];
  for (let i=0;i<120;i++) { const el=document.createElement('div'); el.className='confetti-piece'; el.style.left=Math.random()*100+'vw'; el.style.background=colors[Math.floor(Math.random()*colors.length)]; el.style.width=(6+Math.random()*10)+'px'; el.style.height=(10+Math.random()*16)+'px'; el.style.animationDuration=(3+Math.random()*5)+'s'; el.style.animationDelay=(Math.random()*4)+'s'; el.style.borderRadius=Math.random()>.5?'50%':'2px'; wrap.appendChild(el); }
}

function mostrarGanador() {
  const totals={}, mesas={};
  txData.forEach(t => { totals[t.name]=(totals[t.name]||0)+t.amount; if(t.mesa&&!mesas[t.name]) mesas[t.name]=t.mesa; });
  const nombres = Object.keys(totals).sort((a,b)=>totals[b]-totals[a]);
  if (!nombres.length) { showToast('No hay consumos aun', true); return; }
  const ganador = nombres[0];
  
  const wmsg = (document.getElementById('ct-winner-msg').value||'¡EL GANADOR DE LA NOCHE!').toUpperCase();
  const wsub = document.getElementById('ct-winner-sub').value||'';
  const premio = document.getElementById('msg-input').value.trim();
  const coronaEl = document.getElementById('winner-corona'); if (coronaEl) coronaEl.textContent = '👑';
  const _wt=document.getElementById('winner-titulo'); if(_wt) _wt.textContent = wmsg;
  document.getElementById('winner-mesa').textContent = mesas[ganador]||'—';
  document.getElementById('winner-nombre').textContent = ganador.toUpperCase();
  document.getElementById('winner-total').textContent = fmt(totals[ganador]);
  const msgEl = document.getElementById('winner-mensaje'); const textoFinal = wsub||premio||'';
  msgEl.textContent = textoFinal; msgEl.style.display = textoFinal ? 'block' : 'none';
  if (confettiGanadorActivo) generarParticulas();
  // Rayos girando
  // Sin rayos wrap
  if (false) {
    // Sin rayos — diseño limpio como la foto
  }
  document.getElementById('winner-overlay').classList.add('show');
}

function mostrarGanadorManual() {
  ganadorMostrado = true;
  mostrarGanador(); // Mostrar overlay localmente en esta pantalla también
  fetch('/api/winner/show',{method:'POST'}).then(()=>showToast('🏆 Ganador enviado a pantalla')).catch(()=>showToast('Error de conexión',true));
}
function cerrarGanador() { document.getElementById('winner-overlay').classList.remove('show'); fetch('/api/winner/hide',{method:'POST'}); }
function quitarCartel() { fetch('/api/cartel/hide',{method:'POST'}).then(()=>{ cartelActivo=false; const btn=document.getElementById('btn-quitar-cartel'); if(btn) btn.style.display='none'; showToast('Cartel quitado'); }).catch(()=>showToast('Error de conexión',true)); }

function mesaMinMgr(a, b) {
  const na = parseInt(a), nb = parseInt(b);
  if (!isNaN(na) && !isNaN(nb)) return na < nb ? a : b;
  return a < b ? a : b;
}

function renderPantalla() {
  const header=document.getElementById('rank-header'), rows=document.getElementById('rank-rows'), empty=document.getElementById('empty-msg');
  const totals={}, mesas={};
  txData.forEach(t => {
    totals[t.name]=(totals[t.name]||0)+t.amount;
    if(t.mesa) mesas[t.name] = mesas[t.name] ? mesaMinMgr(mesas[t.name], t.mesa) : t.mesa;
  });
  const names = Object.keys(totals).sort((a,b)=>totals[b]-totals[a]).slice(0,5);
  if (!names.length) { header.style.display='none'; rows.innerHTML=''; empty.style.display='block'; knownNames.clear(); prevRankOrder=[]; return; }
  header.style.display='grid'; empty.style.display='none';
  const esc=s=>String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
  rows.innerHTML = names.map((name,i) => { const pos=i+1, rc=pos<=3?'rank-'+pos:'', esNuevo=!knownNames.has(name);
    const mesaRaw=mesas[name]||'—'; const mesaNum=mesaRaw.replace?mesaRaw.replace(/^Mesa\s+/i,''):'—';
    return `<div class="rank-row ${rc}${esNuevo?' nueva':''}" data-name="${esc(name)}"><div class="col-puesto">#${pos}</div><div class="col-nombre">${esc(name)}</div><div class="col-mesa">${esc(mesaNum)}</div><div class="col-total">${fmt(totals[name])}</div></div>`;
  }).join('');
  prevRankOrder=[...names]; knownNames=new Set(names);
}

async function loadData() {
  try {
    const [r1,r2] = await Promise.all([fetch('/api/tx'),fetch('/api/state')]);
    txData = await r1.json(); const st = await r2.json();
    renderPantalla();
    if (st.hora_fin && st.hora_fin !== horaFin) { horaFin=st.hora_fin; const hfi=document.getElementById('hora-fin-input'); if(hfi) hfi.value=horaFin; const hfe=document.getElementById('clock-fin'); if(hfe) hfe.textContent=horaFin; }
    if (st.premio !== undefined) { const mi=document.getElementById('msg-input'); if(mi&&mi.value!==st.premio) mi.value=st.premio; const pb=document.getElementById('premio-box'); if(pb) pb.textContent=st.premio; }
    // Trackear estado del cartel para mostrar/ocultar botón Quitar cartel
    const nuevoCartelActivo = !!(st.cartel_show);
    if (nuevoCartelActivo !== cartelActivo) {
      cartelActivo = nuevoCartelActivo;
      const btn = document.getElementById('btn-quitar-cartel');
      if (btn) btn.style.display = cartelActivo ? 'inline-block' : 'none';
    }
    // El ganador solo se muestra en /pantalla, no en pantalla-mgr
  } catch(e) {}
}
setInterval(loadData, 2000); loadData();

function aplicarTema(nombre) {
  const temaExtra = TEMAS_EXTRA[nombre];
  document.body.classList.remove('tema-jagger12','tema-touchofpink','pink-claro');
  pinkModoClaro = false;
  const overlay = document.getElementById('tema-overlay');
  overlay.innerHTML = ''; overlay.style.opacity = '0';
  const tl = document.getElementById('tema-tagline');
  if (temaExtra) {
    COLOR_DEFS.forEach(c => { const v=temaExtra.colors[c.key]||c.default; document.documentElement.style.setProperty(c.key,v); customColors[c.key]=v; });
    document.documentElement.style.setProperty('--surface-gold', blendSurfaceGold(temaExtra.colors['--surface']||'#111'));
    document.body.classList.add(temaExtra.bodyClass);
    temaActual = nombre;
    const fallingToggle = document.getElementById('falling-gloves-toggle'); if(fallingToggle) fallingToggle.style.display = 'none';
    const koToggle = document.getElementById('ko-anim-toggle'); if(koToggle) koToggle.style.display = 'none';
    const show12Toggle = document.getElementById('show-12-toggle'); if(show12Toggle) show12Toggle.style.display = nombre==='jagger12'?'block':'none';
    const pinkPetTog = document.getElementById('pink-petalos-toggle'); if(pinkPetTog) pinkPetTog.style.display = nombre==='touchofpink'?'block':'none';
    const pinkModoTog = document.getElementById('pink-modo-toggle'); if(pinkModoTog) pinkModoTog.style.display = nombre==='touchofpink'?'flex':'none';
    const savedVip = (()=>{try{return localStorage.getItem('rankingVIP_vip')||'VIP';}catch(e){return 'VIP';}})();
    const mainLogo = document.getElementById('main-logo'); if(mainLogo) mainLogo.innerHTML = `RANKING <span class="vip" id="logo-vip">${savedVip}</span>`;
    if (tl) {
      if (nombre==='touchofpink') tl.innerHTML=`<div style="text-align:center;"><div style="font-family:'Oswald',sans-serif;font-size:clamp(18px,2.8vw,30px);font-weight:700;letter-spacing:6px;color:#fff;text-shadow:0 0 12px rgba(255,255,255,0.9);">JAGGER CLUB</div><div style="font-family:'Oswald',sans-serif;font-size:clamp(13px,2vw,22px);font-weight:600;letter-spacing:5px;color:#f472b6;margin-top:3px;text-shadow:0 0 10px rgba(244,114,182,1);">TURNS PINK</div></div>`;
      else tl.innerHTML=`<div style="text-align:center;"><div style="font-family:'Oswald',sans-serif;font-size:clamp(14px,2.2vw,24px);font-weight:700;letter-spacing:4px;color:#e8c84a;text-shadow:0 0 14px rgba(232,200,74,0.9);">JAGGER CLUB · 12 AÑOS</div><div style="font-family:'Oswald',sans-serif;font-size:clamp(11px,1.6vw,17px);font-weight:600;letter-spacing:3px;color:#e8c84a;opacity:0.75;margin-top:2px;">12 AÑOS DE HISTORIA NO SON PARA CUALQUIERA</div></div>`;
    }
    const toggleWrap = document.getElementById('tema-deco-toggle');
    if (toggleWrap) {
      toggleWrap.style.display = 'flex';
      const decoLabel = nombre==='jagger12'?'Activar burbujas de champagne':'Activar decoraciones animadas';
      document.getElementById('toggle-deco-label').textContent = decoLabel;
      document.getElementById('toggle-deco').checked = decoActiva;
      const dl = document.getElementById('deco-main-label'); if(dl) dl.style.display = nombre==='touchofpink'?'none':'';
    }
    if (decoActiva) iniciarDecoTema(nombre);
    buildColorGrid();
    try { localStorage.setItem('rankingVIP_tema', nombre); } catch(e) {}
    showToast('Tema ' + nombre.toUpperCase() + ' aplicado');
    fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({tema:nombre,colores:customColors})}).catch(()=>{});
    return;
  }
  if (nombre === 'default') {
    COLOR_DEFS.forEach(c => { document.documentElement.style.setProperty(c.key,c.default); customColors[c.key]=c.default; });
    document.documentElement.style.setProperty('--surface-gold','#0d0b00');
    const ml = document.getElementById('main-logo'); if(ml) ml.innerHTML=`RANKING <span class="vip" id="logo-vip">VIP</span>`;
    if (tl) { tl.textContent='JAGGER CLUB'; tl.style.fontSize='28px'; tl.style.fontWeight='600'; tl.style.letterSpacing='5px'; tl.style.color='#555'; }
  }
  temaActual = nombre;
  const toggleWrap = document.getElementById('tema-deco-toggle'); if(toggleWrap) toggleWrap.style.display='none';
  buildColorGrid();
  try { localStorage.setItem('rankingVIP_tema', nombre); } catch(e) {}
  showToast('Tema ' + nombre.toUpperCase() + ' aplicado');
  fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({tema:nombre,colores:customColors})}).catch(()=>{});
}

function toggleDecoActual(checked) {
  decoActiva = checked;
  fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({deco_activa:checked})}).catch(()=>{});
  if (false) {
  } else if (temaActual === 'jagger12') {
    const wrap = document.getElementById('jagger12-particles');
    if (!checked) { if(wrap) wrap.innerHTML=''; } else if(wrap) {
      wrap.innerHTML = '';
      function lanzarB() { if(!decoActiva) return; const w=document.getElementById('jagger12-particles'); if(!w) return; const el=document.createElement('div'); const sz=4+Math.random()*12, isG=Math.random()>.7; el.style.cssText=`position:absolute;bottom:-20px;left:${Math.random()*100}%;width:${sz}px;height:${sz}px;border-radius:50%;border:1px solid rgba(${isG?'201,162,39':'255,255,255'},${isG?.35:.2});background:rgba(${isG?'201,162,39':'255,255,255'},${isG?.06:.03});animation:bubbleRise ${4+Math.random()*6}s ease-in forwards;pointer-events:none;`; w.appendChild(el); setTimeout(()=>el.remove(),11000); setTimeout(lanzarB,200+Math.random()*600); }
      for(let i=0;i<8;i++) setTimeout(lanzarB,i*120);
    }
  } else { const o=document.getElementById('tema-overlay'); if(!checked){o.innerHTML='';o.style.opacity='0';}else iniciarDecoTema(temaActual); }
}

function reiniciarDeco12() {
  if (temaActual==='jagger12') iniciarJagger12Deco();

}

function iniciarDecoTema(nombre) {
  const overlay = document.getElementById('tema-overlay'); overlay.style.opacity='1';
  if (nombre==='jagger12') { overlay.innerHTML=''; iniciarJagger12Deco(); }

  if (nombre==='touchofpink') {
    overlay.innerHTML=`<div style="position:absolute;top:12%;left:8%;font-size:14px;color:rgba(255,255,255,0.45);pointer-events:none;animation:goldTwinkle 3s ease-in-out infinite;">✦</div><div style="position:absolute;top:18%;right:10%;font-size:11px;color:rgba(255,255,255,0.38);pointer-events:none;animation:goldTwinkle 4s ease-in-out 1s infinite;">✦</div><div style="position:absolute;top:50%;left:5%;font-size:12px;color:rgba(255,255,255,0.35);pointer-events:none;animation:goldTwinkle 5s ease-in-out 2s infinite;">✦</div><div style="position:absolute;top:65%;right:6%;font-size:15px;color:rgba(255,255,255,0.4);pointer-events:none;animation:goldTwinkle 3.5s ease-in-out .5s infinite;">✦</div><div style="position:absolute;top:33%;left:14%;font-size:9px;color:rgba(251,182,206,0.55);pointer-events:none;animation:goldTwinkle 4.5s ease-in-out 1.5s infinite;">✦</div><div style="position:absolute;top:42%;right:16%;font-size:10px;color:rgba(251,182,206,0.5);pointer-events:none;animation:goldTwinkle 3.8s ease-in-out .8s infinite;">✦</div><div id="petalos-wrap" style="position:absolute;inset:0;overflow:hidden;pointer-events:none;"></div>`;
    if (pinkPetalosActivos) iniciarPetalos();
  }
}

function iniciarJagger12Deco() {
  const overlay = document.getElementById('tema-overlay');
  const svg12 = mostrar12Fondo ? `<svg width="100%" height="100%" viewBox="0 0 1000 600" style="position:absolute;inset:0;pointer-events:none;opacity:0.13;" xmlns="http://www.w3.org/2000/svg" preserveAspectRatio="xMidYMid meet"><defs><filter id="blur12j"><feGaussianBlur stdDeviation="8"/></filter><filter id="glow12j"><feGaussianBlur stdDeviation="18" result="blur"/><feMerge><feMergeNode in="blur"/><feMergeNode in="SourceGraphic"/></feMerge></filter></defs><text x="500" y="380" text-anchor="middle" dominant-baseline="middle" font-family="Oswald,Arial" font-weight="700" font-size="560" fill="rgba(201,162,39,0.15)" letter-spacing="-10" filter="url(#blur12j)">12</text><text x="500" y="380" text-anchor="middle" dominant-baseline="middle" font-family="Oswald,Arial" font-weight="700" font-size="560" fill="white" letter-spacing="-10" filter="url(#glow12j)">12</text></svg>` : '';
  overlay.innerHTML = `<div id="jagger12-particles" style="position:absolute;inset:0;overflow:hidden;pointer-events:none;"></div>${svg12}<div style="position:absolute;top:15%;left:5%;font-size:10px;color:#c9a227;opacity:0.2;pointer-events:none;animation:goldTwinkle 3s ease-in-out infinite;">✦</div><div style="position:absolute;top:25%;right:7%;font-size:8px;color:#c9a227;opacity:0.18;pointer-events:none;animation:goldTwinkle 4s ease-in-out 1s infinite;">✦</div><div style="position:absolute;top:60%;left:3%;font-size:12px;color:#c9a227;opacity:0.15;pointer-events:none;animation:goldTwinkle 5s ease-in-out 2s infinite;">✦</div><div style="position:absolute;top:70%;right:4%;font-size:9px;color:#c9a227;opacity:0.18;pointer-events:none;animation:goldTwinkle 3.5s ease-in-out 0.5s infinite;">✦</div>`;
  if (decoActiva) {
    function lanzarBurbuja() { if(!decoActiva) return; const wrap=document.getElementById('jagger12-particles'); if(!wrap) return; const el=document.createElement('div'); const sz=4+Math.random()*12,isGold=Math.random()>.7; el.style.cssText=`position:absolute;bottom:-20px;left:${Math.random()*100}%;width:${sz}px;height:${sz}px;border-radius:50%;border:1px solid rgba(${isGold?'201,162,39':'255,255,255'},${isGold?.35:.2});background:rgba(${isGold?'201,162,39':'255,255,255'},${isGold?.06:.03});animation:bubbleRise ${4+Math.random()*6}s ease-in forwards;pointer-events:none;`; wrap.appendChild(el); setTimeout(()=>el.remove(),11000); setTimeout(lanzarBurbuja,200+Math.random()*600); }
    for(let i=0;i<8;i++) setTimeout(lanzarBurbuja,i*120);
  }
}

  lanzarGuante(); setTimeout(lanzarGuante,800); setTimeout(lanzarGuante,1600);
}

function iniciarPetalos() {
  const wrap=document.getElementById('petalos-wrap'); if(!wrap) return; wrap.innerHTML='';
  const shapes=['🌸','🌸','🌺','🌷','💮'];
  for(let i=0;i<35;i++) { const el=document.createElement('div'); el.textContent=shapes[Math.floor(Math.random()*shapes.length)]; el.style.cssText=`position:absolute;top:-40px;left:${Math.random()*100}vw;font-size:${12+Math.random()*16}px;opacity:${.25+Math.random()*.45};animation:copoFall ${6+Math.random()*9}s linear ${Math.random()*8}s infinite;pointer-events:none;`; wrap.appendChild(el); }
}

function aplicarEfecto(nombre) {
  efectoActual=nombre; const overlay=document.getElementById('efectos-overlay'); if(overlay) overlay.innerHTML='';
  if(efectoInterval){clearInterval(efectoInterval);efectoInterval=null;}
  document.querySelectorAll('[id^="efecto-btn-"]').forEach(b=>{b.style.borderColor='#2a2a2a';b.style.color='#555';});
  const activeBtn=document.getElementById('efecto-btn-'+nombre); if(activeBtn){activeBtn.style.borderColor='#c9a227';activeBtn.style.color='#c9a227';}
  try{localStorage.setItem('rankingVIP_efecto',nombre);}catch(e){}
  fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({efecto:nombre})}).catch(()=>{});
  if(nombre==='ninguno') return;
  if(nombre==='burbujas') iniciarEfectoBurbujas();
  if(nombre==='estrellas') iniciarEfectoEstrellas();
  showToast('Efecto '+nombre.toUpperCase()+' activado');
}

function iniciarEfectoBurbujas() {
  const overlay=document.getElementById('efectos-overlay'); if(!overlay) return;
  function crear() { if(!document.getElementById('efectos-overlay')) return; const el=document.createElement('div'); const sz=5+Math.random()*20,isGold=Math.random()>.6; el.style.cssText=`position:absolute;bottom:-30px;left:${Math.random()*100}%;width:${sz}px;height:${sz}px;border-radius:50%;border:1px solid rgba(${isGold?'201,162,39':'255,255,255'},${isGold?.4:.2});background:rgba(${isGold?'201,162,39':'255,255,255'},${isGold?.06:.03});animation:bubbleRise ${5+Math.random()*8}s ease-in forwards;`; overlay.appendChild(el); setTimeout(()=>el.remove(),14000); }
  for(let i=0;i<10;i++) setTimeout(crear,i*200); efectoInterval=setInterval(crear,350);
}

function iniciarEfectoEstrellas() {
  const overlay=document.getElementById('efectos-overlay'); if(!overlay) return;
  if(!document.getElementById('kf-efecto-star')){const s=document.createElement('style');s.id='kf-efecto-star';s.textContent=`@keyframes starFloat{0%{opacity:0;transform:translateY(0) rotate(0deg) scale(0)}20%{opacity:0.8}80%{opacity:0.5}100%{opacity:0;transform:translateY(-80vh) rotate(360deg) scale(0.5)}}`;document.head.appendChild(s);}
  const syms=['★','✦','✧','✶','✸'];
  function crear() { if(!document.getElementById('efectos-overlay')) return; const el=document.createElement('div'); el.textContent=syms[Math.floor(Math.random()*syms.length)]; const sz=10+Math.random()*18; const colors=['#c9a227','#e8c84a','#fff','#f0ece0']; el.style.cssText=`position:absolute;bottom:-30px;left:${Math.random()*100}%;font-size:${sz}px;color:${colors[Math.floor(Math.random()*colors.length)]};opacity:0;animation:starFloat ${6+Math.random()*6}s ease-in-out forwards;pointer-events:none;`; overlay.appendChild(el); setTimeout(()=>el.remove(),13000); }
  for(let i=0;i<8;i++) setTimeout(crear,i*300); efectoInterval=setInterval(crear,500);
}

function setClockSize(size) {
  ['chico','mediano','grande'].forEach(s=>{const btn=document.getElementById('cs-'+s);if(!btn)return;btn.style.borderColor=s===size?'var(--gold)':'#2a2a2a';btn.style.color=s===size?'var(--gold)':'#555';});
  fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({clock_size:size})}).catch(()=>{});
  showToast('Reloj: '+size);
}

function blendSurfaceGold(h) { try{const r=parseInt(h.slice(1,3),16),g=parseInt(h.slice(3,5),16),b=parseInt(h.slice(5,7),16);const nr=Math.min(255,Math.round(r*.85+0xc9*.15)),ng=Math.min(255,Math.round(g*.88+0xa2*.12)),nb=Math.min(255,Math.round(b*.95+0x27*.05));return'#'+nr.toString(16).padStart(2,'0')+ng.toString(16).padStart(2,'0')+nb.toString(16).padStart(2,'0');}catch(e){return h;} }

function buildColorGrid() {
  const grid=document.getElementById('color-grid'); if(!grid) return;
  grid.innerHTML=COLOR_DEFS.map(c=>{const val=customColors[c.key];const hexId='hex-'+c.key.replace(/--/g,'').replace(/-/g,'_');const inputId='inp-'+c.key.replace(/--/g,'').replace(/-/g,'_');const swatchId='sw-'+c.key.replace(/--/g,'').replace(/-/g,'_');
    return `<div class="color-item"><div class="color-swatch" id="${swatchId}" style="background:${val}"><input type="color" id="${inputId}" value="${val}" data-key="${c.key}" oninput="onColorChange(this)" /></div><div style="flex:1"><div class="color-label">${c.label}</div><div class="color-hex" id="${hexId}">${val}</div></div><button onclick="resetColorSingle('${c.key}')" style="background:none;border:1px solid #2a2a2a;color:#444;border-radius:5px;padding:4px 8px;font-size:11px;cursor:pointer;font-family:'Rajdhani',sans-serif;letter-spacing:1px;white-space:nowrap;" onmouseover="this.style.borderColor='#c9a227';this.style.color='#c9a227'" onmouseout="this.style.borderColor='#2a2a2a';this.style.color='#444'">↺ Default</button></div>`;
  }).join('');
}

function resetColorSingle(key) {
  const def=COLOR_DEFS.find(c=>c.key===key); if(!def) return;
  customColors[key]=def.default; document.documentElement.style.setProperty(key,def.default);
  const sw=document.getElementById('sw-'+key.replace(/--/g,'').replace(/-/g,'_')); if(sw) sw.style.background=def.default;
  const inp=document.getElementById('inp-'+key.replace(/--/g,'').replace(/-/g,'_')); if(inp) inp.value=def.default;
  const hexEl=document.getElementById('hex-'+key.replace(/--/g,'').replace(/-/g,'_')); if(hexEl) hexEl.textContent=def.default;
  if(key==='--gold'){const pv=document.getElementById('prev-vip');if(pv)pv.style.color=def.default;}
  if(key==='--surface') document.documentElement.style.setProperty('--surface-gold',blendSurfaceGold(def.default));
  showToast('Color restablecido');
}

function onColorChange(el) {
  const key=el.dataset.key, val=el.value; customColors[key]=val; el.parentElement.style.background=val;
  const hexEl=document.getElementById('hex-'+key.replace(/--/g,'').replace(/-/g,'_')); if(hexEl) hexEl.textContent=val;
  document.documentElement.style.setProperty(key,val);
  if(key==='--surface') document.documentElement.style.setProperty('--surface-gold',blendSurfaceGold(val));
  if(key==='--gold'){const pv=document.getElementById('prev-vip');if(pv)pv.style.color=val;}
}

function applyTaglineStyle(el,text,color,glow,font) { if(!el)return; el.textContent=text; el.style.color=color; el.style.fontFamily=font; const g=parseFloat(glow); el.style.textShadow=g>0?`0 0 ${Math.round(g*30)}px ${color},0 0 ${Math.round(g*60)}px ${color}`:''; }

function previewTextos() { const vip=document.getElementById('ct-vip').value||'VIP'; const pv=document.getElementById('prev-vip');if(pv)pv.textContent=vip; const lv=document.getElementById('logo-vip');if(lv)lv.textContent=vip; }
function previewTagline() { applyTaglineStyle(document.getElementById('tema-tagline'),document.getElementById('ct-tagline').value||'JAGGER CLUB',document.getElementById('ct-tagline-color').value,document.getElementById('ct-tagline-brightness').value,document.getElementById('ct-tagline-font').value); }

function aplicarPersonalizacion() {
  COLOR_DEFS.forEach(c=>document.documentElement.style.setProperty(c.key,customColors[c.key]));
  if(customColors['--surface']) document.documentElement.style.setProperty('--surface-gold',blendSurfaceGold(customColors['--surface']));
  const vip=document.getElementById('ct-vip').value||'VIP'; const lv=document.getElementById('logo-vip');if(lv)lv.textContent=vip;
  const tT=document.getElementById('ct-tagline').value||'JAGGER CLUB', tC=document.getElementById('ct-tagline-color').value, tG=document.getElementById('ct-tagline-brightness').value, tF=document.getElementById('ct-tagline-font').value;
  applyTaglineStyle(document.getElementById('tema-tagline'),tT,tC,tG,tF);
  const wm=document.getElementById('ct-winner-msg').value, ws=document.getElementById('ct-winner-sub').value;
  try{localStorage.setItem('rankingVIP_colors',JSON.stringify(customColors));localStorage.setItem('rankingVIP_vip',vip);localStorage.setItem('rankingVIP_wmsg',wm);localStorage.setItem('rankingVIP_wsub',ws);localStorage.setItem('rankingVIP_tagline_text',tT);localStorage.setItem('rankingVIP_tagline_color',tC);localStorage.setItem('rankingVIP_tagline_glow',tG);localStorage.setItem('rankingVIP_tagline_font',tF);}catch(e){}
  const premioCurrent = (document.getElementById('ct-premio')||{}).value||'';
  fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({tema:temaActual,colores:customColors,vip:vip,tagline:tT,tagline_color:tC,tagline_glow:tG,winner_msg:wm,winner_sub:ws,premio:premioCurrent,hora_fin:horaFin})}).catch(()=>{});
  showToast('Personalización aplicada');
}

function resetPersonalizacion() {
  if(!confirm('Restaurar todos los colores y textos al default?')) return;
  COLOR_DEFS.forEach(c=>{customColors[c.key]=c.default;document.documentElement.style.setProperty(c.key,c.default);});
  document.documentElement.style.setProperty('--surface-gold','#0d0b00');
  document.getElementById('ct-vip').value='VIP'; const lv=document.getElementById('logo-vip');if(lv)lv.textContent='VIP';
  document.getElementById('ct-tagline').value='JAGGER CLUB'; document.getElementById('ct-tagline-color').value='#555555';
  document.getElementById('ct-tagline-brightness').value='0'; document.getElementById('ct-tagline-brightness-val').textContent='0.00';
  applyTaglineStyle(document.getElementById('tema-tagline'),'JAGGER CLUB','#555','0',"'Rajdhani',sans-serif");
  buildColorGrid(); previewTextos();
  fetch('/api/design',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({tema:'default',colores:{},vip:'VIP',tagline:'JAGGER CLUB',tagline_color:'#555555',tagline_glow:'0',winner_msg:'¡EL GANADOR DE LA NOCHE!',winner_sub:''})}).catch(()=>{});
  showToast('Colores restaurados');
}

function cargarPersonalizacionGuardada() {
  try {
    const sc=localStorage.getItem('rankingVIP_colors'); if(sc){customColors=JSON.parse(sc);COLOR_DEFS.forEach(c=>{if(customColors[c.key])document.documentElement.style.setProperty(c.key,customColors[c.key]);});}
    const vip=localStorage.getItem('rankingVIP_vip'); if(vip){const lv=document.getElementById('logo-vip');if(lv)lv.textContent=vip;const ctvip=document.getElementById('ct-vip');if(ctvip)ctvip.value=vip;const pv=document.getElementById('prev-vip');if(pv)pv.textContent=vip;}
    const wm=localStorage.getItem('rankingVIP_wmsg'),ws=localStorage.getItem('rankingVIP_wsub'); if(wm)document.getElementById('ct-winner-msg').value=wm; if(ws)document.getElementById('ct-winner-sub').value=ws;
    const tT=localStorage.getItem('rankingVIP_tagline_text'),tC=localStorage.getItem('rankingVIP_tagline_color'),tG=localStorage.getItem('rankingVIP_tagline_glow'),tF=localStorage.getItem('rankingVIP_tagline_font');
    if(tT||tC||tG||tF){if(tT)document.getElementById('ct-tagline').value=tT;if(tC)document.getElementById('ct-tagline-color').value=tC;if(tG){document.getElementById('ct-tagline-brightness').value=tG;document.getElementById('ct-tagline-brightness-val').textContent=parseFloat(tG).toFixed(2);}if(tF)document.getElementById('ct-tagline-font').value=tF;applyTaglineStyle(document.getElementById('tema-tagline'),tT||'JAGGER CLUB',tC||'#555555',tG||'0',tF||"'Rajdhani',sans-serif");}
  } catch(e) {}
  buildColorGrid();
}

function cargarTemaGuardado() { try{const t=localStorage.getItem('rankingVIP_tema');if(t)aplicarTema(t);}catch(e){} }

cargarPersonalizacionGuardada();
cargarTemaGuardado();
// Sincronizar tema y efecto desde servidor al arrancar
fetch('/api/design').then(r=>r.json()).then(d=>{
  if(d.tema && d.tema !== temaActual) aplicarTema(d.tema);
  if(d.efecto && d.efecto !== 'ninguno'){
    document.querySelectorAll('[id^="efecto-btn-"]').forEach(b=>{b.style.borderColor='#2a2a2a';b.style.color='#555';});
    const ab=document.getElementById('efecto-btn-'+d.efecto); if(ab){ab.style.borderColor='#c9a227';ab.style.color='#c9a227';}
  }
}).catch(()=>{});
// Cargar premio actual desde servidor
fetch('/api/state').then(r=>r.json()).then(st=>{
  if(st.premio){const el=document.getElementById('ct-premio');if(el)el.value=st.premio;const pb=document.getElementById('premio-box');if(pb)pb.textContent=st.premio;}
}).catch(()=>{});
</script>
</body>
</html>"""

LOGIN_HTML = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Acceso — {{ title }}</title>
<link href="https://fonts.googleapis.com/css2?family=Oswald:wght@400;700&family=Rajdhani:wght@400;600;700&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0;}
body{background:#080808;color:#e0e0e0;font-family:'Rajdhani',sans-serif;min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;}
.login-card{background:#0f0f0f;border:1px solid #2a2a2a;border-radius:16px;padding:36px 32px 32px;width:100%;max-width:340px;display:flex;flex-direction:column;align-items:center;gap:20px;}
.login-title{font-family:'Oswald',sans-serif;font-size:22px;font-weight:700;letter-spacing:4px;color:#c9a227;text-transform:uppercase;text-align:center;}
.login-sub{font-size:12px;color:#555;letter-spacing:2px;text-transform:uppercase;text-align:center;margin-top:-10px;}
.pin-dots{display:flex;gap:12px;justify-content:center;height:18px;}
.pin-dot{width:14px;height:14px;border-radius:50%;background:#1a1a1a;border:2px solid #333;transition:all .15s;}
.pin-dot.filled{background:#c9a227;border-color:#c9a227;}
.pin-keypad{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;width:100%;}
.pin-key{background:#111;border:1px solid #222;border-radius:10px;padding:16px 0;font-family:'Oswald',sans-serif;font-size:22px;font-weight:600;color:#e0e0e0;cursor:pointer;text-align:center;transition:all .1s;user-select:none;}
.pin-key:hover{background:#191919;border-color:#3a3a3a;}
.pin-key:active{background:#c9a227;color:#000;transform:scale(.96);}
.pin-key.back{font-size:18px;color:#777;}
.pin-key.ok{background:#c9a227;color:#000;font-size:14px;font-weight:700;letter-spacing:2px;}
.pin-key.ok:hover{background:#e8c84a;}
.login-err{color:#e74c3c;font-size:13px;letter-spacing:1px;min-height:18px;text-align:center;}
.back-link{color:#444;font-size:11px;letter-spacing:2px;text-decoration:none;text-transform:uppercase;margin-top:4px;}
.back-link:hover{color:#c9a227;}
</style>
</head>
<body>
<div class="login-card">
  <div class="login-title">{{ title }}</div>
  <div class="login-sub">Ingresá el PIN de acceso</div>
  <div class="pin-dots">
    <div class="pin-dot" id="d0"></div>
    <div class="pin-dot" id="d1"></div>
    <div class="pin-dot" id="d2"></div>
    <div class="pin-dot" id="d3"></div>
  </div>
  <div class="pin-keypad">
    {% for n in ['1','2','3','4','5','6','7','8','9','←','0','OK'] %}
    <div class="pin-key {% if n == '←' %}back{% elif n == 'OK' %}ok{% endif %}"
         onclick="pinPress('{{ n }}')">{{ n }}</div>
    {% endfor %}
  </div>
  <div class="login-err" id="login-err"></div>
  <a href="/" class="back-link">← volver al hub</a>
</div>
<script>
let pinVal = '';
const role = '{{ role }}';
const redirectTo = '{{ redirect_to }}';
function updateDots() {
  for (let i=0;i<4;i++) document.getElementById('d'+i).className='pin-dot'+(pinVal.length>i?' filled':'');
}
function pinPress(k) {
  if (k==='←') { pinVal=pinVal.slice(0,-1); updateDots(); return; }
  if (k==='OK') { doLogin(); return; }
  if (pinVal.length<4) { pinVal+=k; updateDots(); if(pinVal.length===4) setTimeout(doLogin,120); }
}
async function doLogin() {
  const err = document.getElementById('login-err');
  err.textContent='';
  try {
    const r = await fetch('/api/login',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({role,pin:pinVal})});
    const d = await r.json();
    if (d.ok) { window.location.href = redirectTo; }
    else { err.textContent = d.error||'PIN incorrecto'; pinVal=''; updateDots(); }
  } catch(e) { err.textContent='Error de conexión'; pinVal=''; updateDots(); }
}
document.addEventListener('keydown', e=>{
  if(e.key>='0'&&e.key<='9') pinPress(e.key);
  else if(e.key==='Backspace') pinPress('←');
  else if(e.key==='Enter') pinPress('OK');
});
</script>
</body>
</html>"""

PANTALLA_HTML = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Pantalla — Jagger VIP</title>
<link href="https://fonts.googleapis.com/css2?family=Oswald:wght@400;600;700&family=Rajdhani:wght@400;600;700&display=swap" rel="stylesheet">
<style>
:root{--gold:#c9a227;--gold2:#e8c84a;--gold-dim:#7a6010;--black:#080808;--surface:#111111;--surface-gold:#0d0b00;--border:#2a2a2a;--text:#f0ece0;--text-dim:#555555;--white:#ffffff;--green:#2ecc71;}
*{box-sizing:border-box;margin:0;padding:0;}
html,body{height:100%;overflow:hidden;}
body{background:var(--black);color:var(--text);font-family:'Rajdhani',sans-serif;padding:24px 28px 16px;}
.pres-header-wrap{position:relative;margin-bottom:24px;}
.pres-clock{position:absolute;top:0;left:0;}
.pres-clock-hora{font-family:'Oswald',sans-serif;font-size:13px;color:var(--text-dim);letter-spacing:1px;text-transform:uppercase;}
.pres-clock-time{font-family:'Oswald',sans-serif;font-size:32px;color:#fff;font-weight:700;line-height:1.1;}
.pres-clock-fin{font-size:11px;color:var(--gold-dim);letter-spacing:1px;margin-top:6px;text-transform:uppercase;}
.pres-clock-fin-val{font-family:'Oswald',sans-serif;font-size:26px;color:var(--gold);font-weight:700;line-height:1.1;display:block;}
.pres-header{text-align:center;padding-top:6px;display:flex;flex-direction:column;align-items:center;}
.pres-logo{font-family:'Oswald',sans-serif;font-size:52px;font-weight:700;color:#fff;letter-spacing:10px;text-transform:uppercase;display:block;width:100%;text-align:center;}
.pres-logo .vip{color:var(--gold);}
.pres-line{height:1px;background:linear-gradient(to right,transparent,var(--gold),transparent);margin:10px auto;max-width:100%;}
.live-badge{display:inline-flex;align-items:center;gap:6px;border:1px solid #2a2a2a;border-radius:20px;padding:4px 14px;font-size:11px;color:#777;letter-spacing:1px;margin-top:4px;}
.live-dot{width:6px;height:6px;border-radius:50%;background:var(--green);animation:blink 1.5s infinite;}
@keyframes blink{0%,100%{opacity:1;}50%{opacity:0.3;}}
.ranking-wrap{max-width:100%;padding:0;}
.rank-header{display:grid;grid-template-columns:100px 1fr 130px 180px;background:#0d0d0d;border:1px solid #2a2a2a;border-radius:8px 8px 0 0;padding:12px 30px;margin-bottom:3px;}
.rank-header span{font-family:'Oswald',sans-serif;font-size:13px;font-weight:500;letter-spacing:2px;text-transform:uppercase;color:var(--gold-dim);}
.rank-header .col-r{text-align:right;}
.rank-rows{display:flex;flex-direction:column;gap:4px;position:relative;}
.rank-row{position:relative;}
.rank-row{display:grid;grid-template-columns:100px 1fr 130px 180px;align-items:center;background:var(--surface);border:1px solid #1e1e1e;border-radius:6px;padding:20px 30px;transition:all .4s;}
.rank-row.rank-1{background:var(--surface-gold);border-color:var(--gold-dim);}
.rank-row.nueva{animation:entradaFila 1.2s cubic-bezier(.16,1,.3,1) both;}
.rank-row.rank-1.nueva{animation:entradaFila 1.2s cubic-bezier(.16,1,.3,1) both,brilloOro 2s ease .4s both;}
.rank-row.rank-2.nueva{animation:entradaFila 1.2s cubic-bezier(.16,1,.3,1) both,brilloPlata 2s ease .4s both;}
.rank-row.rank-3.nueva{animation:entradaFila 1.2s cubic-bezier(.16,1,.3,1) both,brilloBronce 2s ease .4s both;}
@keyframes entradaFila{from{opacity:0;transform:translateX(-50px) scale(0.96)}to{opacity:1;transform:none}}
@keyframes brilloOro{0%{box-shadow:0 0 0 rgba(201,162,39,0)}40%{box-shadow:0 0 50px rgba(201,162,39,.85),0 0 22px rgba(201,162,39,.6)}100%{box-shadow:0 0 14px rgba(201,162,39,.22)}}
@keyframes brilloPlata{0%{box-shadow:0 0 0 rgba(190,190,190,0)}40%{box-shadow:0 0 45px rgba(210,210,210,.8),0 0 20px rgba(200,200,200,.55)}100%{box-shadow:0 0 10px rgba(160,160,160,.18)}}
@keyframes brilloBronce{0%{box-shadow:0 0 0 rgba(180,100,30,0)}40%{box-shadow:0 0 45px rgba(205,127,50,.85),0 0 20px rgba(180,100,30,.55)}100%{box-shadow:0 0 10px rgba(180,100,30,.18)}}
.col-puesto{font-family:'Oswald',sans-serif;font-size:28px;font-weight:700;color:#444;}
.rank-row.rank-1 .col-puesto{color:var(--gold);font-size:34px;}
.rank-row.rank-2 .col-puesto{color:#aaa;}
.rank-row.rank-3 .col-puesto{color:#8a6a40;}
.col-nombre{font-family:'Oswald',sans-serif;font-size:30px;font-weight:600;color:#fff;}
.rank-row.rank-1 .col-nombre{font-size:36px;}
.col-mesa{font-family:'Oswald',sans-serif;font-size:24px;font-weight:700;color:#e8e8e8;letter-spacing:1px;}
.col-total{font-family:'Oswald',sans-serif;font-size:30px;font-weight:700;color:var(--gold);text-align:right;}
.rank-row.rank-1 .col-total{font-size:38px;}
.miles-lbl{font-size:0.45em;opacity:0.55;letter-spacing:2px;margin-left:5px;vertical-align:middle;font-weight:600;}
.empty-msg{text-align:center;color:#222;font-size:15px;padding:60px 20px;letter-spacing:2px;font-family:'Oswald',sans-serif;}
.premio-wrap{text-align:center;margin-top:24px;}
.premio-box{display:inline-block;background:var(--gold);color:#000;font-family:'Rajdhani',sans-serif;font-size:22px;font-weight:700;letter-spacing:.5px;padding:12px 40px;border-radius:6px;}
.premio-box:empty{display:none;}
/* Winner */
#winner-overlay{display:none;position:fixed;inset:0;z-index:11000;background:rgba(0,0,0,0.97);flex-direction:column;align-items:center;justify-content:center;text-align:center;}
#winner-overlay.show{display:flex;animation:winnerFadeIn .5s ease both;}
@keyframes winnerFadeIn{from{opacity:0}to{opacity:1}}
.winner-rays{position:absolute;inset:0;overflow:hidden;pointer-events:none;}
.winner-content{position:relative;z-index:2;padding:0 20px;animation:winnerEntrada .7s cubic-bezier(.22,1,.36,1) both .15s;}
.winner-corona{display:block;font-size:80px;margin-bottom:10px;animation:bounce 1.2s ease-in-out infinite;}
@keyframes bounce{0%,100%{transform:translateY(0);}50%{transform:translateY(-10px);}}
.winner-titulo{font-family:'Oswald',sans-serif;font-size:clamp(18px,2.8vw,32px);color:var(--gold);letter-spacing:6px;font-weight:700;margin-bottom:10px;text-shadow:0 0 20px rgba(201,162,39,0.5);}
.winner-nombre{font-family:'Oswald',sans-serif;font-size:clamp(70px,12vw,150px);font-weight:700;color:#fff;letter-spacing:3px;line-height:1;text-shadow:0 0 40px rgba(255,255,255,0.15),0 0 80px rgba(201,162,39,0.25);}
.winner-line{height:2px;background:linear-gradient(to right,transparent,var(--gold),transparent);margin:16px auto;max-width:650px;width:90%;}
.winner-info-row{display:flex;gap:60px;justify-content:center;margin-bottom:20px;}
.winner-info-block{text-align:center;}
.winner-info-label{font-size:11px;color:#666;letter-spacing:3px;text-transform:uppercase;margin-bottom:4px;}
.winner-info-val{font-family:'Oswald',sans-serif;font-size:clamp(32px,4.5vw,58px);color:var(--gold);font-weight:700;line-height:1.1;}
#winner-mensaje{display:inline-block;background:#0a0a08;border:1px solid var(--gold);border-radius:6px;padding:14px 44px;font-family:'Oswald',sans-serif !important;font-size:clamp(20px,2.8vw,38px) !important;font-weight:700 !important;color:#fff !important;letter-spacing:2px;text-shadow:none;margin-top:8px;}
.winner-close{position:fixed;top:20px;right:24px;background:transparent;border:none;color:#333;font-size:28px;cursor:pointer;z-index:11001;transition:color .2s;}
.winner-close:hover{color:#888;}
/* Cartel */
#cartel-overlay{display:none;position:fixed;inset:0;z-index:11000;background:rgba(0,0,0,0.98);flex-direction:column;align-items:center;justify-content:center;overflow:hidden;}
/* Publicidad */
#pub-overlay{display:none;position:fixed;inset:0;z-index:10500;background:#000;align-items:center;justify-content:center;flex-direction:column;}
#pub-overlay.show{display:flex;}
#pub-video{width:100%;height:100%;object-fit:contain;display:none;}
#pub-iframe{width:100%;height:100%;border:none;display:none;}
/* Tema overlay */
#tema-overlay{position:fixed;inset:0;pointer-events:none;z-index:2;overflow:hidden;opacity:0;transition:opacity 1s;}
#efectos-overlay{position:fixed;inset:0;pointer-events:none;z-index:3;overflow:hidden;}
/* Confetti */
.confetti-piece{position:absolute;top:-20px;animation:confettiFall linear infinite;}
@keyframes confettiFall{to{transform:translateY(110vh) rotate(720deg);opacity:0;}}
/* Bottom link */
.hub-link{position:fixed;bottom:12px;left:16px;color:#1a1a1a;font-size:11px;letter-spacing:2px;text-decoration:none;text-transform:uppercase;transition:color .2s;}
.hub-link:hover{color:#555;}
@keyframes bottleBounce{0%,100%{transform:translateY(0) rotate(-5deg) scale(1);}20%{transform:translateY(-18px) rotate(5deg) scale(1.05);}40%{transform:translateY(-6px) rotate(-3deg) scale(0.98);}60%{transform:translateY(-14px) rotate(4deg) scale(1.03);}80%{transform:translateY(-4px) rotate(-2deg) scale(0.99);}}
@keyframes emojiGlow{from{filter:drop-shadow(0 0 20px rgba(201,162,39,0.4))}to{filter:drop-shadow(0 0 50px rgba(201,162,39,0.9))}}
@keyframes winnerEntrada{from{opacity:0;transform:scale(0.7) translateY(40px)}to{opacity:1;transform:scale(1) translateY(0)}}
@keyframes rayPulse2{0%{opacity:0.2;transform:rotate(var(--r,0deg)) scaleY(0.5)}100%{opacity:0.7;transform:rotate(var(--r,0deg)) scaleY(1)}}
@keyframes goldTwinkle{0%,100%{opacity:0.1;transform:scale(1)}50%{opacity:0.28;transform:scale(1.4)}}
@keyframes bubbleRise{0%{opacity:0.6;transform:translateY(0) scale(1)}50%{opacity:0.3}100%{opacity:0;transform:translateY(-100vh) scale(0.5)}}
@keyframes copoFall{0%{transform:translateY(-40px) rotate(0deg);opacity:1}100%{transform:translateY(105vh) rotate(360deg);opacity:0}}
@keyframes sparkBoxP{0%{opacity:1;transform:scale(1)}100%{opacity:0;transform:translate(var(--bx),var(--by)) scale(0)}}
@keyframes roundPulseP{0%,100%{opacity:0.18}50%{opacity:0.32}}
@keyframes guanteFloatP{0%,100%{transform:rotate(-10deg) translateY(0)}50%{transform:rotate(-10deg) translateY(-12px)}}
@keyframes guanteFallP{0%{opacity:0.8;transform:translateY(-40px) rotate(0deg)}100%{opacity:0;transform:translateY(105vh) rotate(360deg)}}
/* ══ TEMA JAGGER 12 AÑOS ══ */
body.tema-jagger12 .rank-row{border-color:#1e1e1e;}
body.tema-jagger12 .rank-row.rank-1{background:#1a1200;border-color:#c9a227;}
body.tema-jagger12 .rank-row.rank-2{background:#0e0e0e;border-color:#777;}
body.tema-jagger12 .rank-row.rank-3{background:#0e0800;border-color:#7a4a20;}
body.tema-jagger12 .rank-row.rank-1 .col-puesto{color:#c9a227;font-size:34px;}
body.tema-jagger12 .rank-row.rank-2 .col-puesto{color:#aaaaaa;}
body.tema-jagger12 .rank-row.rank-3 .col-puesto{color:#cd7f32;}
body.tema-jagger12 .rank-row.rank-1 .col-total{color:#e8c84a;}
body.tema-jagger12 .rank-row.rank-2 .col-total{color:#cccccc;}
body.tema-jagger12 .rank-row.rank-3 .col-total{color:#cd7f32;}
body.tema-jagger12 .col-total{color:#ccc;}
body.tema-jagger12 .pres-line{background:linear-gradient(to right,transparent,#c9a227,transparent);}
body.tema-jagger12 .live-dot{background:#c9a227;}
body.tema-jagger12 .col-nombre{color:#fff !important;}
body.tema-jagger12 .col-mesa{color:#ddd !important;}

/* ══ TEMA A TOUCH OF PINK ══ */
body.tema-touchofpink .rank-header{background:#2d0022;border-color:#6a2050;}
body.tema-touchofpink .rank-header span{color:#f472b6;text-shadow:0 0 8px rgba(244,114,182,0.4);}
body.tema-touchofpink .rank-row{border-color:#8a3070;background:#3d002c;}
body.tema-touchofpink .rank-row.rank-1{background:#5a0042;border-color:#f472b6;box-shadow:0 0 20px rgba(244,114,182,.2);}
body.tema-touchofpink .rank-row.rank-1 .col-puesto{color:#f472b6;font-size:34px;}
body.tema-touchofpink .rank-row.rank-2 .col-puesto{color:#ffffff;}
body.tema-touchofpink .rank-row.rank-3 .col-puesto{color:#fbb6ce;}
body.tema-touchofpink .rank-row.rank-1 .col-total{color:#fce7f3;}
body.tema-touchofpink .col-total{color:#fbb6ce;}
body.tema-touchofpink .col-nombre{color:#ffffff !important;}
body.tema-touchofpink .col-mesa{color:#fce7f3 !important;}
body.tema-touchofpink .col-puesto{color:#eeaad8;}
body.tema-touchofpink .pres-line{background:linear-gradient(to right,transparent,#f472b6,transparent);}
body.tema-touchofpink .live-dot{background:#f472b6;}
body.tema-touchofpink .live-badge{border-color:#8a3070;color:#f472b6;}
body.tema-touchofpink.pink-claro{--black:#3a0028;--surface:#580040;--border:#a04080;--gold:#f472b6;--gold-light:#fbb6ce;--gold-dim:#e896cc;--text:#ffe8f5;--text-dim:#ddaacc;--white:#ffffff;}
body.tema-touchofpink.pink-claro .rank-header{background:#4a0035;border-color:#a04080;}
body.tema-touchofpink.pink-claro .rank-row{background:#580040;border-color:#a04080;}
body.tema-touchofpink.pink-claro .rank-row.rank-1{background:#7a0058;border-color:#f472b6;box-shadow:0 0 24px rgba(244,114,182,.3);}
body.tema-touchofpink.pink-claro .rank-row.rank-1 .col-puesto{color:#f472b6;}
body.tema-touchofpink.pink-claro .rank-row.rank-2 .col-puesto{color:#ffffff;}
body.tema-touchofpink.pink-claro .rank-row.rank-3 .col-puesto{color:#fbb6ce;}
body.tema-touchofpink.pink-claro .rank-row.rank-1 .col-total{color:#fce7f3;}
body.tema-touchofpink.pink-claro .col-total{color:#fbb6ce;}
body.tema-touchofpink.pink-claro .col-nombre{color:#ffffff !important;}
body.tema-touchofpink.pink-claro .col-mesa{color:#fce7f3 !important;}
body.tema-touchofpink.pink-claro .col-puesto{color:#eeaad8;}
</style>
</head>
<body>
<div id="tema-overlay"></div>
<div id="efectos-overlay"></div>

<div class="pres-header-wrap">
  <div class="pres-clock">
    <div class="pres-clock-hora">Hora</div>
    <div class="pres-clock-time" id="clock-display">00:00</div>
    <div class="pres-clock-fin">Finaliza</div>
    <span class="pres-clock-fin-val" id="clock-fin">05:30</span>
  </div>
  <div class="pres-header">
    <div class="pres-logo" id="main-logo">RANKING <span class="vip" id="logo-vip">VIP</span></div>
    <div id="tema-tagline" style="font-family:'Rajdhani',sans-serif;font-size:26px;font-weight:600;color:#555;letter-spacing:5px;text-transform:uppercase;text-align:center;min-height:0;width:100%;display:block;">JAGGER CLUB</div>
    <div class="pres-line"></div>
    <div class="live-badge"><span class="live-dot"></span> EN VIVO</div>
  </div>
</div>

<div class="ranking-wrap">
  <div class="rank-header" id="rank-header" style="display:none">
    <span>PUESTO</span><span>NOMBRE</span><span>MESA</span><span class="col-r">TOTAL</span>
  </div>
  <div class="rank-rows" id="rank-rows"></div>
  <div id="empty-msg" class="empty-msg">Aún no hay consumos registrados</div>
</div>
<div class="premio-wrap"><div class="premio-box" id="premio-box"></div></div>

<!-- Winner overlay -->
<div id="winner-overlay">
  <div class="confetti-wrap" id="confetti-wrap"></div>
  <button class="winner-close" onclick="cerrarGanador()">✕</button>
  <div class="winner-content">
    <span class="winner-corona" id="winner-corona">👑</span>
    <div class="winner-titulo" id="winner-titulo">¡EL GANADOR DE LA NOCHE!</div>
    <div class="winner-nombre" id="winner-nombre">—</div>
    <div class="winner-line"></div>
    <div class="winner-info-row">
      <div class="winner-info-block">
        <div class="winner-info-label">Mesa</div>
        <div class="winner-info-val" id="winner-mesa">—</div>
      </div>
      <div class="winner-info-block">
        <div class="winner-info-label">Total consumido</div>
        <div class="winner-info-val" id="winner-total">—</div>
      </div>
    </div>
    <div id="winner-mensaje" style="display:none;"></div>
  </div>
</div>

<!-- Cartel overlay -->
<div id="cartel-overlay">
  <div id="cartel-tema-bg" style="position:absolute;inset:0;pointer-events:none;z-index:0;"></div>
  <div id="cartel-rays" style="position:absolute;inset:0;pointer-events:none;z-index:1;overflow:hidden;"></div>
  <div id="cartel-content" style="position:relative;z-index:2;text-align:center;max-width:90vw;padding:0 24px;">
    <div id="cartel-emoji-big" style="font-size:110px;margin-bottom:6px;display:block;animation:bottleBounce 1.2s cubic-bezier(.36,.07,.19,.97) infinite,emojiGlow 2s ease-in-out infinite alternate;filter:drop-shadow(0 0 30px rgba(201,162,39,0.6));"></div>
    <div id="cartel-nombre-display" style="font-family:'Oswald',sans-serif;font-size:clamp(48px,8vw,100px);font-weight:700;color:#fff;letter-spacing:4px;text-transform:uppercase;line-height:1;text-shadow:0 0 40px rgba(201,162,39,.5);"></div>
    <div id="cartel-mesa-display" style="font-family:'Oswald',sans-serif;font-size:clamp(20px,3vw,36px);color:#888;letter-spacing:3px;margin-top:8px;"></div>
    <div style="height:2px;background:linear-gradient(to right,transparent,var(--gold),transparent);margin:18px auto;max-width:600px;width:80%;"></div>
    <div id="cartel-frase-display" style="font-family:'Rajdhani',sans-serif;font-size:clamp(26px,4.5vw,56px);font-weight:700;color:#fff;letter-spacing:2px;text-transform:uppercase;line-height:1.3;padding:18px 36px;border:2px solid var(--gold);border-radius:8px;background:rgba(201,162,39,0.1);max-width:85vw;display:inline-block;text-shadow:0 0 20px rgba(201,162,39,0.3);box-shadow:0 0 40px rgba(201,162,39,0.15),inset 0 0 40px rgba(201,162,39,0.05);"></div>
  </div>
</div>

<!-- Publicidad overlay -->
<div id="pub-overlay">
  <button onclick="cerrarPublicidad()" style="position:fixed;top:18px;right:22px;z-index:10502;background:rgba(0,0,0,0.7);color:#aaa;border:1px solid #444;font-size:22px;cursor:pointer;border-radius:8px;width:40px;height:40px;line-height:1;transition:color .2s;" onmouseover="this.style.color='#fff'" onmouseout="this.style.color='#aaa'">✕</button>
  <video id="pub-video" autoplay playsinline style="display:none;width:100%;height:100%;object-fit:contain;background:#000;"></video>
  <iframe id="pub-iframe" allow="autoplay;fullscreen" allowfullscreen style="display:none;width:100%;height:100%;border:none;"></iframe>
</div>

<a href="/" class="hub-link">← Hub</a>

<script>
let txData = [];
let lastWinnerTs = 0;
let lastCartelTs = 0;
let prevRankMap = {};
let horaFin = '05:30';
let horaFinMostrada = false;
let temaActual = 'default';
let pubActiva = false;
let _designWinnerMsg = '¡EL GANADOR DE LA NOCHE!';
let _designWinnerSub = '';
let pubFrecuenciaMs = 15 * 60 * 1000;
let pubLastShown = parseInt(localStorage.getItem('pubLastShown') || '0');

// El navegador bloquea el autoplay con sonido hasta que haya una interacción en la página.
// Con el primer clic o tecla en esta pantalla habilitamos el sonido para los videos.
let pubAudioUnlocked = false;
function _unlockPubAudio(){
  pubAudioUnlocked = true;
  const v = document.getElementById('pub-video');
  if (v && !v.paused) { v.muted = false; }
}
document.addEventListener('click', _unlockPubAudio);
document.addEventListener('keydown', _unlockPubAudio);
document.addEventListener('touchstart', _unlockPubAudio, {passive:true});

// Reproduce el video intentando con sonido; si el navegador lo bloquea, reintenta en mudo
// para que al menos se vea (evita la pantalla en negro).
function _playPubVideo(video, offsetSegundos){
  if (offsetSegundos > 0 && offsetSegundos < (video.duration || Infinity)) {
    try { video.currentTime = offsetSegundos; } catch(e){}
  }
  video.muted = false;
  const p = video.play();
  if (p && p.catch) {
    p.catch(() => {
      video.muted = true;            // autoplay con sonido bloqueado → reintentar en mudo
      video.play().catch(()=>{});
    });
  }
}

function fmt(n){return '$'+Number(n).toLocaleString('es-AR',{minimumFractionDigits:0,maximumFractionDigits:0});}

function esc(s){const d=document.createElement('div');d.textContent=s||'';return d.innerHTML;}

// ── Clock ──
function updateClock() {
  const now = new Date();
  const h = String(now.getHours()).padStart(2,'0');
  const m = String(now.getMinutes()).padStart(2,'0');
  document.getElementById('clock-display').textContent = h+':'+m;
  if (horaFin && !horaFinMostrada && txData && txData.length > 0) {
    const [fh, fm] = horaFin.split(':').map(Number);
    if (now.getHours() === fh && now.getMinutes() === fm) {
      horaFinMostrada = true;
      // Notificar al servidor — loadData() lo detectará y llamará mostrarGanador()
      fetch('/api/winner/show', {method:'POST'}).catch(()=>{});
    }
  }
}
setInterval(updateClock, 1000);
updateClock();

// ── Render ranking ──
function fmtMiles(n){
  if(n>=1000000){const m=n/1000000;return fmt(n)+'<span class="miles-lbl">'+(m%1===0?m:m.toFixed(1))+(m<2?' MILLÓN':' MILLONES')+'</span>';}
  if(n>=1000){const k=n/1000;return fmt(n)+'<span class="miles-lbl">'+(k%1===0?k:k.toFixed(1))+(k<2?' MIL':' MILES')+'</span>';}
  return fmt(n);
}

function mesaMin(a, b) {
  const na = parseInt(a), nb = parseInt(b);
  if (!isNaN(na) && !isNaN(nb)) return na < nb ? a : b;
  return a < b ? a : b;
}

function render() {
  const totals={}, mesas={};
  txData.forEach(t=>{
    totals[t.name]=(totals[t.name]||0)+t.amount;
    if(t.mesa) mesas[t.name] = mesas[t.name] ? mesaMin(mesas[t.name], t.mesa) : t.mesa;
  });
  const sorted = Object.entries(totals).sort((a,b)=>b[1]-a[1]);
  const rankHeader = document.getElementById('rank-header');
  const rankRows = document.getElementById('rank-rows');
  const emptyMsg = document.getElementById('empty-msg');
  if (!sorted.length) {
    rankHeader.style.display='none'; rankRows.innerHTML=''; emptyMsg.style.display='block';
    prevRankMap={};
    return;
  }
  rankHeader.style.display='grid'; emptyMsg.style.display='none';
  const top8 = sorted.slice(0,5);
  const newMap = {};
  top8.forEach(([n],i) => newMap[n]=i+1);

  // FLIP: capturar posiciones FIRST de los elementos actuales
  const firstPositions = {};
  rankRows.querySelectorAll('.rank-row[data-name]').forEach(el => {
    firstPositions[el.dataset.name] = el.getBoundingClientRect().top;
  });

  // Actualizar DOM (LAST)
  rankRows.innerHTML = top8.map(([name,total],i)=>{
    const pos = i+1;
    const posClass = pos<=3?'rank-'+pos:'';
    const prev = prevRankMap[name];
    const animClass = prev===undefined ? 'nueva' : '';
    const mesaRaw = mesas[name]||'';
    const mesaNum = mesaRaw.replace(/^Mesa\s+/i,'') || '—';
    const stagger = animClass==='nueva' ? `animation-delay:${i*0.18}s;` : '';
    return `<div class="rank-row ${posClass} ${animClass}" data-name="${esc(name)}" style="${stagger}">
      <div class="col-puesto">${pos}</div>
      <div class="col-nombre">${esc(name)}</div>
      <div class="col-mesa">${mesaNum}</div>
      <div class="col-total">${fmt(total)}</div>
    </div>`;
  }).join('');

  // PLAY: animar — quien sube pasa encima. Brillo según posición destino.
  rankRows.querySelectorAll('.rank-row[data-name]').forEach(el => {
    const name = el.dataset.name;
    if (firstPositions[name] !== undefined) {
      const lastTop = el.getBoundingClientRect().top;
      const delta = firstPositions[name] - lastTop;
      if (Math.abs(delta) > 1) {
        const sube = delta > 0;
        const posNum = parseInt(el.className.match(/rank-(\d+)/)?.[1] || '0');
        el.style.zIndex = sube ? '10' : '1';
        el.style.transition = 'none';
        el.style.transform = `translateY(${delta}px)`;
        if (sube) { el.style.boxShadow = 'none'; }
        requestAnimationFrame(() => {
          requestAnimationFrame(() => {
            const dur = sube ? '2.0s' : '2.5s';
            const ease = sube ? 'cubic-bezier(.16,.8,.3,1)' : 'cubic-bezier(.16,.35,.3,1)';
            const durMs = parseFloat(dur) * 1000;
            if (sube) {
              // Brillo DESDE EL INICIO según rango destino
              const glows = {
                1: '0 0 55px rgba(201,162,39,1), 0 0 110px rgba(201,162,39,0.55)',
                2: '0 0 45px rgba(220,220,220,0.95), 0 0 90px rgba(200,200,200,0.45)',
                3: '0 0 45px rgba(215,135,50,0.95), 0 0 90px rgba(190,110,30,0.45)',
              };
              el.style.boxShadow = glows[posNum] || '0 0 30px rgba(255,255,255,0.25)';
              el.style.transition = `transform ${dur} ${ease}, box-shadow ${dur} ease`;
              el.style.transform = 'translateY(0)';
              setTimeout(() => {
                el.style.transition = 'box-shadow 1.8s ease';
                el.style.boxShadow = '';
                setTimeout(() => { el.style.zIndex=''; el.style.transition=''; }, 1900);
              }, durMs);
            } else {
              el.style.transition = `transform ${dur} ${ease}`;
              el.style.transform = 'translateY(0)';
              setTimeout(() => { el.style.zIndex=''; el.style.transition=''; }, durMs+200);
            }
          });
        });
      }
    }
  });

  prevRankMap = newMap;
}

// ── Load data + state ──
async function loadData() {
  try {
    const [r1,r2] = await Promise.all([fetch('/api/tx'),fetch('/api/state')]);
    txData = await r1.json();
    const st = await r2.json();
    render();
    if (st.hora_fin) { horaFin=st.hora_fin; document.getElementById('clock-fin').textContent=horaFin; }
    // Premio lo maneja loadDesign con tamaño — loadData no sobreescribe
    // Winner
    if (st.winner_show && st.winner_ts && st.winner_ts !== lastWinnerTs) {
      lastWinnerTs = st.winner_ts;
      mostrarGanador();
    } else if (!st.winner_show && !horaFinMostrada) {
      // Solo cerrar si el servidor dice false Y no fue disparado localmente por horaFin
      document.getElementById('winner-overlay').classList.remove('show');
    }
    // Cartel
    if (st.cartel_show && st.cartel_ts && st.cartel_ts !== lastCartelTs) {
      lastCartelTs = st.cartel_ts;
      const cd = st.cartel_data || {};
      const nd=document.getElementById('cartel-nombre-display');
      const md=document.getElementById('cartel-mesa-display');
      const fd=document.getElementById('cartel-frase-display');
      const ed=document.getElementById('cartel-emoji-big');
      if(nd){nd.textContent=cd.nombre||'';nd.style.display=cd.nombre?'block':'none';}
      if(md) md.textContent=cd.mesa?'MESA '+cd.mesa:'';
      if(fd) fd.textContent=cd.frase||'';
      if(ed) ed.textContent=cd.emoji||'🍾';
      // Rayos según tema
      const rays = document.getElementById('cartel-rays');
      if (rays) {
        rays.innerHTML = '';
        const rayColor = temaActual==='touchofpink'?'rgba(244,114,182,0.12)':'rgba(201,162,39,0.12)';
        for (let i=0;i<12;i++) {
          const r=document.createElement('div');
          r.style.cssText=`position:absolute;left:50%;top:50%;width:1px;height:55vh;background:linear-gradient(to bottom,${rayColor},transparent);transform-origin:0% 0%;transform:rotate(${i*30}deg);opacity:0.5;animation:rayPulse2 ${2+i*0.15}s ease-in-out ${i*0.1}s infinite alternate;`;
          r.style.setProperty('--r',(i*30)+'deg');
          rays.appendChild(r);
        }
      }
      // Fondo según tema
      const bg=document.getElementById('cartel-tema-bg');
      if(bg){
        bg.innerHTML=
          temaActual==='jagger12'?
          `<div style="position:absolute;inset:0;background:radial-gradient(ellipse at center,#111 0%,#000 100%);"></div><div style="position:absolute;inset:0;background:radial-gradient(ellipse at 50% 50%,rgba(201,162,39,0.05) 0%,transparent 70%);"></div>`:
          temaActual==='touchofpink'?
          `<div style="position:absolute;inset:0;background:radial-gradient(ellipse at center,#1e0018 0%,#080005 100%);"></div><div style="position:absolute;inset:0;background:radial-gradient(ellipse at 50% 50%,rgba(244,114,182,0.07) 0%,transparent 70%);"></div>`:
          `<div style="position:absolute;inset:0;background:radial-gradient(ellipse at center,#0a0a0a 0%,#000 100%);"></div>`;
      }
      const overlay=document.getElementById('cartel-overlay');
      if(overlay){
        overlay.style.display='flex';
        const c=document.getElementById('cartel-content');
        if(c){c.style.animation='none';c.style.opacity='0';requestAnimationFrame(()=>{requestAnimationFrame(()=>{c.style.animation='winnerEntrada 0.8s cubic-bezier(.22,1,.36,1) forwards';c.style.opacity='';});});}
      }
    } else if (!st.cartel_show) {
      document.getElementById('cartel-overlay').style.display='none';
    }
  } catch(e){}
}
setInterval(loadData, 2000);
loadData();

// ── Design sync ──
let temaAnterior = '';
let _loadDesignTs = 0;
async function loadDesign() {
  try {
    const r = await fetch('/api/design');
    const d = await r.json();
    const ts = d._ts || 0;
    const tsChanged = (ts > 0 && ts !== _loadDesignTs);
    if (ts > 0) _loadDesignTs = ts;

    // ── Colores ── aplicar siempre si hay datos
    if (d.colores) {
      const cols = d.colores;
      const keys = Object.keys(cols);
      if (keys.length > 0) {
        keys.forEach(k => document.documentElement.style.setProperty(k, cols[k]));
      }
    }

    // ── VIP label ──
    if (d.vip !== undefined) {
      const el = document.getElementById('logo-vip');
      if (el) el.textContent = d.vip || 'VIP';
    }

    // ── Tagline (texto debajo del ranking) ── siempre aplicar
    if (d.tagline !== undefined) {
      const tl = document.getElementById('tema-tagline');
      if (tl) {
        tl.textContent = d.tagline || '';
        tl.style.color = d.tagline_color || '#555555';
        tl.style.fontFamily = d.tagline_font || "'Rajdhani',sans-serif";
        const gv = parseFloat(d.tagline_glow || 0);
        tl.style.textShadow = gv > 0 ? `0 0 ${Math.round(gv*40)}px currentColor` : '';
      }
    }

    // ── Premio ──
    if (d.premio !== undefined) {
      const pb = document.getElementById('premio-box');
      if (pb) {
        pb.textContent = d.premio || '';
        pb.style.fontSize = (d.premio_size || '22') + 'px';
      }
    }

    // ── Hora fin ──
    if (d.hora_fin) { horaFin = d.hora_fin; const cf = document.getElementById('clock-fin'); if(cf) cf.textContent = d.hora_fin; }

    // ── Winner msg/sub ──
    if (d.winner_msg !== undefined) { _designWinnerMsg = d.winner_msg || '¡EL GANADOR DE LA NOCHE!'; const wm = document.getElementById('winner-titulo'); if(wm) wm.textContent = _designWinnerMsg; }
    if (d.winner_sub !== undefined) { _designWinnerSub = d.winner_sub || ''; }

    // ── Tipo particula ──
    if (d.tipo_particula !== undefined) _tipoParticulaPantalla = d.tipo_particula;

    // ── Clock size ──
    if (d.clock_size) applyClockSize(d.clock_size);

    // ── Tema (body class + decoraciones) ──
    if (d.tema !== undefined) {
      document.body.classList.remove('tema-fullblack','tema-navidad','tema-anonuevo','tema-halloween','tema-touchofpink','pink-claro','tema-jagger12');
      if (d.tema && d.tema !== 'default') document.body.classList.add('tema-'+d.tema);
      temaActual = d.tema || 'default';
      const decoChanged = (
        (d.deco_activa !== undefined && !!d.deco_activa !== decoActivaPantalla) ||
        (d.petals_activos !== undefined && !!d.petals_activos !== petalosActivosPantalla) ||
        (d.falling_gloves !== undefined && !!d.falling_gloves !== glovesActivosPantalla) ||
        (d.mostrar12 !== undefined && !!d.mostrar12 !== mostrar12FondoPantalla)
      );
      if (d.deco_activa !== undefined) decoActivaPantalla = !!d.deco_activa;
      if (d.petals_activos !== undefined) petalosActivosPantalla = !!d.petals_activos;
      if (d.falling_gloves !== undefined) glovesActivosPantalla = !!d.falling_gloves;
      if (d.mostrar12 !== undefined) mostrar12FondoPantalla = !!d.mostrar12;
      if (d.tema !== temaAnterior || tsChanged || decoChanged) {
        temaAnterior = d.tema;
        const overlay = document.getElementById('tema-overlay');
        if (overlay) { overlay.innerHTML = ''; overlay.style.opacity = '0'; }
        if (d.tema && d.tema !== 'default') iniciarDecoTemaPantalla(d.tema);
      }
    }

    // ── Efecto de fondo ── re-aplicar si cambió el efecto O si hubo cambio de design
    if (d.efecto !== undefined) {
      const tsNow = d._ts || 0;
      if (d.efecto !== _efectoPantallaAnterior || tsNow !== _designTsAnterior) {
        _efectoPantallaAnterior = d.efecto;
        _designTsAnterior = tsNow;
        const efOv = document.getElementById('efectos-overlay');
        if (efOv) efOv.innerHTML = '';
        if (_efectoPantallaInterval) { clearInterval(_efectoPantallaInterval); _efectoPantallaInterval = null; }
        if (d.efecto === 'burbujas') _pantallaEfectoBurbujas();
        else if (d.efecto === 'estrellas') _pantallaEfectoEstrellas();
      }
    }
  } catch(e) { console.error('loadDesign error:', e); }
}
let _efectoPantallaAnterior = '';
let _efectoPantallaInterval = null;
let _designTsAnterior = 0;
let _tipoParticulaPantalla = 'confetti';
function _pantallaEfectoBurbujas(){const o=document.getElementById('efectos-overlay');if(!o)return;function c(){if(!document.getElementById('efectos-overlay'))return;const e=document.createElement('div');const s=5+Math.random()*20,g=Math.random()>.6;e.style.cssText=`position:absolute;bottom:-30px;left:${Math.random()*100}%;width:${s}px;height:${s}px;border-radius:50%;border:1px solid rgba(${g?'201,162,39':'255,255,255'},${g?.4:.2});background:rgba(${g?'201,162,39':'255,255,255'},${g?.06:.03});animation:bubbleRise ${5+Math.random()*8}s ease-in forwards;`;o.appendChild(e);setTimeout(()=>e.remove(),14000);}for(let i=0;i<10;i++)setTimeout(c,i*200);_efectoPantallaInterval=setInterval(c,350);}
function _pantallaEfectoEstrellas(){const o=document.getElementById('efectos-overlay');if(!o)return;if(!document.getElementById('kf-ep-st')){const s=document.createElement('style');s.id='kf-ep-st';s.textContent='@keyframes epStFloat{0%{opacity:0;transform:translateY(0) rotate(0) scale(0)}20%{opacity:.8}80%{opacity:.5}100%{opacity:0;transform:translateY(-80vh) rotate(360deg) scale(.5)}}';document.head.appendChild(s);}const sym=['★','✦','✧','✶'],col=['#c9a227','#e8c84a','#fff'];function c(){if(!document.getElementById('efectos-overlay'))return;const e=document.createElement('div');e.textContent=sym[Math.floor(Math.random()*sym.length)];e.style.cssText=`position:absolute;bottom:-30px;left:${Math.random()*100}%;font-size:${10+Math.random()*18}px;color:${col[Math.floor(Math.random()*col.length)]};opacity:0;animation:epStFloat ${6+Math.random()*6}s ease-in-out forwards;pointer-events:none;`;o.appendChild(e);setTimeout(()=>e.remove(),13000);}for(let i=0;i<8;i++)setTimeout(c,i*300);_efectoPantallaInterval=setInterval(c,500);}

// ── Decoraciones de tema para /pantalla ──
let decoActivaPantalla = true;
let petalosActivosPantalla = true;
let glovesActivosPantalla = true;

function iniciarDecoTemaPantalla(nombre) {
  const overlay = document.getElementById('tema-overlay');
  if (!overlay) return;
  overlay.style.opacity = '1';
  if (!decoActivaPantalla) { overlay.innerHTML = ''; overlay.style.opacity = '0'; return; }
  if (nombre === 'jagger12') iniciarJagger12DecoPantalla();

  if (nombre === 'touchofpink') iniciarTouchOfPinkDecoPantalla();
}

let mostrar12FondoPantalla = true;
function iniciarJagger12DecoPantalla() {
  const overlay = document.getElementById('tema-overlay');
  if (!overlay) return;
  const svg12 = mostrar12FondoPantalla ? `
    <svg width="100%" height="100%" viewBox="0 0 1000 600" style="position:absolute;inset:0;pointer-events:none;opacity:0.1;" xmlns="http://www.w3.org/2000/svg" preserveAspectRatio="xMidYMid meet">
      <defs><filter id="blur12p"><feGaussianBlur stdDeviation="8"/></filter><filter id="glow12p"><feGaussianBlur stdDeviation="18" result="blur"/><feMerge><feMergeNode in="blur"/><feMergeNode in="SourceGraphic"/></feMerge></filter></defs>
      <text x="500" y="380" text-anchor="middle" dominant-baseline="middle" font-family="Oswald,Arial" font-weight="700" font-size="560" fill="rgba(201,162,39,0.15)" letter-spacing="-10" filter="url(#blur12p)">12</text>
      <text x="500" y="380" text-anchor="middle" dominant-baseline="middle" font-family="Oswald,Arial" font-weight="700" font-size="560" fill="white" letter-spacing="-10" filter="url(#glow12p)">12</text>
    </svg>` : '';
  overlay.innerHTML = `
    <div id="p12-particles" style="position:absolute;inset:0;overflow:hidden;pointer-events:none;"></div>
    ${svg12}
    <div style="position:absolute;top:15%;left:5%;font-size:10px;color:#c9a227;opacity:0.2;pointer-events:none;animation:goldTwinkle 3s ease-in-out infinite;">✦</div>
    <div style="position:absolute;top:25%;right:7%;font-size:8px;color:#c9a227;opacity:0.18;pointer-events:none;animation:goldTwinkle 4s ease-in-out 1s infinite;">✦</div>
    <div style="position:absolute;top:60%;left:3%;font-size:12px;color:#c9a227;opacity:0.15;pointer-events:none;animation:goldTwinkle 5s ease-in-out 2s infinite;">✦</div>
    <div style="position:absolute;top:70%;right:4%;font-size:9px;color:#c9a227;opacity:0.18;pointer-events:none;animation:goldTwinkle 3.5s ease-in-out 0.5s infinite;">✦</div>`;
  function lanzarBurbuja() {
    const wrap=document.getElementById('p12-particles'); if(!wrap) return;
    const el=document.createElement('div');
    const sz=4+Math.random()*12, isGold=Math.random()>0.7;
    el.style.cssText=`position:absolute;bottom:-20px;left:${Math.random()*100}%;width:${sz}px;height:${sz}px;border-radius:50%;border:1px solid rgba(${isGold?'201,162,39':'255,255,255'},${isGold?0.35:0.2});background:rgba(${isGold?'201,162,39':'255,255,255'},${isGold?0.06:0.03});animation:bubbleRise ${4+Math.random()*6}s ease-in forwards;pointer-events:none;`;
    wrap.appendChild(el);
    setTimeout(()=>el.remove(),11000);
    setTimeout(lanzarBurbuja,200+Math.random()*600);
  }
  for(let i=0;i<8;i++) setTimeout(lanzarBurbuja,i*120);
}

function iniciarTouchOfPinkDecoPantalla() {
  const overlay = document.getElementById('tema-overlay');
  if (!overlay) return;
  overlay.innerHTML = `
    <div style="position:absolute;top:12%;left:8%;font-size:14px;color:rgba(255,255,255,0.45);pointer-events:none;animation:goldTwinkle 3s ease-in-out infinite;">✦</div>
    <div style="position:absolute;top:18%;right:10%;font-size:11px;color:rgba(255,255,255,0.38);pointer-events:none;animation:goldTwinkle 4s ease-in-out 1s infinite;">✦</div>
    <div style="position:absolute;top:50%;left:5%;font-size:12px;color:rgba(255,255,255,0.35);pointer-events:none;animation:goldTwinkle 5s ease-in-out 2s infinite;">✦</div>
    <div style="position:absolute;top:65%;right:6%;font-size:15px;color:rgba(255,255,255,0.4);pointer-events:none;animation:goldTwinkle 3.5s ease-in-out .5s infinite;">✦</div>
    <div style="position:absolute;top:33%;left:14%;font-size:9px;color:rgba(251,182,206,0.55);pointer-events:none;animation:goldTwinkle 4.5s ease-in-out 1.5s infinite;">✦</div>
    <div style="position:absolute;top:42%;right:16%;font-size:10px;color:rgba(251,182,206,0.5);pointer-events:none;animation:goldTwinkle 3.8s ease-in-out .8s infinite;">✦</div>
    <div id="ppetalos-wrap" style="position:absolute;inset:0;overflow:hidden;pointer-events:none;"></div>`;
  const wrap=document.getElementById('ppetalos-wrap'); if(!wrap) return;
  if (!petalosActivosPantalla) return;
  const shapes=['🌸','🌸','🌺','🌷','💮'];
  for(let i=0;i<35;i++){
    const el=document.createElement('div');
    el.textContent=shapes[Math.floor(Math.random()*shapes.length)];
    el.style.cssText=`position:absolute;top:-40px;left:${Math.random()*100}vw;font-size:${12+Math.random()*16}px;opacity:${0.25+Math.random()*0.45};animation:copoFall ${6+Math.random()*9}s linear ${Math.random()*8}s infinite;pointer-events:none;`;
    wrap.appendChild(el);
  }
}

function applyClockSize(size) {
  const sizes = {chico:{time:'20px',fin:'18px'},mediano:{time:'32px',fin:'26px'},grande:{time:'52px',fin:'40px'}};
  const s = sizes[size] || sizes.mediano;
  const te = document.querySelector('.pres-clock-time');
  const fe = document.querySelector('.pres-clock-fin-val');
  if (te) te.style.fontSize = s.time;
  if (fe) fe.style.fontSize = s.fin;
}
setInterval(loadDesign, 2000);
loadDesign();

// ── Publicidad (frecuencia) ──
let _pubMostrarTsAnterior = 0;
async function checkPublicidad() {
  try {
    const r = await fetch('/api/publicidad/estado');
    const d = await r.json();
    pubFrecuenciaMs = (parseInt(d.frecuencia)||15) * 60 * 1000;
    if (!d.url) { pubActiva = false; cerrarPublicidad(); return; }

    // Bug fix 1: chequear activa ANTES de procesar mostrar_ts
    if (!d.activa) {
      if (pubActiva) cerrarPublicidad(); // Bug fix 2: cerrar si estaba activo localmente
      pubActiva = false;
      return;
    }

    pubActiva = true;

    // Mostrar sincronizado: "Mostrar ahora" desde el manager
    const ts = d.mostrar_ts || 0;
    if (ts > 0 && ts !== _pubMostrarTsAnterior) {
      _pubMostrarTsAnterior = ts;
      const serverNow = d.server_time || (Date.now() / 1000);
      const offset = Math.max(0, serverNow - ts);
      mostrarPublicidadSync(d.url, offset);
      return;
    }

    // Mostrar por frecuencia programada (solo si ya se mostro antes y paso el intervalo)
    const now = Date.now();
    if (pubLastShown > 0 && (now - pubLastShown) >= pubFrecuenciaMs) {
      mostrarPublicidad(d.url);
    }
  } catch(e){}
}
setInterval(checkPublicidad, 3000); // poll más frecuente para sincronizar mejor
checkPublicidad();

function mostrarPublicidad(url) {
  mostrarPublicidadSync(url, 0);
}

function mostrarPublicidadSync(url, offsetSegundos) {
  const overlay = document.getElementById('pub-overlay');
  const video = document.getElementById('pub-video');
  const iframe = document.getElementById('pub-iframe');
  if (!overlay) return;
  pubLastShown = Date.now();
  localStorage.setItem('pubLastShown', pubLastShown);
  if (iframe) { iframe.style.display='none'; iframe.src=''; }
  video.style.display='block';
  overlay.classList.add('show');
  // Si ya tenemos la misma URL cargada y el offset está dentro de la duración, hacer seek
  if (video.src && video.src.endsWith(url.replace(/^.*\//,'')) && video.readyState >= 2) {
    _playPubVideo(video, offsetSegundos);
  } else {
    video.src = url;
    video.oncanplay = function() {
      _playPubVideo(video, offsetSegundos);
      video.oncanplay = null;
    };
    video.load();
  }
  video.onended = cerrarPublicidad;
  video.onerror = function() { cerrarPublicidad(); };
  // Safety timeout: auto-close after 3 minutes max
  clearTimeout(window._pubSafetyTimer);
  window._pubSafetyTimer = setTimeout(cerrarPublicidad, 3 * 60 * 1000);
}

function cerrarPublicidad() {
  clearTimeout(window._pubSafetyTimer);
  const overlay = document.getElementById('pub-overlay');
  const video = document.getElementById('pub-video');
  const iframe = document.getElementById('pub-iframe');
  if (overlay) overlay.classList.remove('show');
  if (video) { video.pause(); video.onended=null; video.src=''; video.load(); video.style.display='none'; }
  if (iframe) { iframe.src=''; iframe.style.display='none'; }
  pubActiva = false;
}

// ── Winner ──
function mostrarGanador() {
  const totals={}, mesas={};
  txData.forEach(t=>{ totals[t.name]=(totals[t.name]||0)+t.amount; if(t.mesa&&!mesas[t.name])mesas[t.name]=t.mesa; });
  const sorted = Object.entries(totals).sort((a,b)=>b[1]-a[1]);
  if (!sorted.length) return;
  const [ganador,total] = sorted[0];
  document.getElementById('winner-titulo').textContent=_designWinnerMsg||'¡EL GANADOR DE LA NOCHE!';
  document.getElementById('winner-nombre').textContent=ganador.toUpperCase();
  document.getElementById('winner-mesa').textContent=mesas[ganador]||'—';
  document.getElementById('winner-total').textContent=fmt(total);
  // Subtítulo
  const msgEl = document.getElementById('winner-mensaje');
  if (msgEl) {
    const sub = _designWinnerSub || '';
    msgEl.textContent = sub;
    msgEl.style.display = sub ? 'block' : 'none';
  }
  // Partículas según tipo
  if(_tipoParticulaPantalla==='billetes') generarBilletes();
  else if(_tipoParticulaPantalla==='champagne') _generarEmojiRain('🍾',50);
  else if(_tipoParticulaPantalla==='confetti_dorado') generarConfettiDorado();
  else if(_tipoParticulaPantalla==='ninguno') { const w=document.getElementById('confetti-wrap'); if(w) w.innerHTML=''; }
  else generarConfetti();
  document.getElementById('winner-overlay').classList.add('show');
}
function generarBilletes(){
  _generarEmojiRain('💵',60);
}
function _generarEmojiRain(emoji, count){
  const w=document.getElementById('confetti-wrap');if(!w)return;w.innerHTML='';
  for(let i=0;i<count;i++){const e=document.createElement('div');e.textContent=emoji;e.style.cssText=`position:absolute;top:-30px;left:${Math.random()*100}vw;font-size:${20+Math.random()*20}px;animation:confettiFall ${3+Math.random()*5}s linear ${Math.random()*4}s infinite;pointer-events:none;`;w.appendChild(e);}
}
function generarConfettiDorado(){
  const w=document.getElementById('confetti-wrap');if(!w)return;w.innerHTML='';
  const colors=['#c9a227','#e8c84a','#f0d060','#fff8dc','#ffd700'];
  for(let i=0;i<90;i++){const e=document.createElement('div');e.className='confetti-piece';e.style.left=Math.random()*100+'vw';e.style.background=colors[Math.floor(Math.random()*colors.length)];e.style.width=(5+Math.random()*8)+'px';e.style.height=(8+Math.random()*14)+'px';e.style.animationDuration=(3+Math.random()*5)+'s';e.style.animationDelay=(Math.random()*4)+'s';e.style.borderRadius=Math.random()>.5?'50%':'2px';e.style.boxShadow=`0 0 4px ${colors[0]}`;w.appendChild(e);}
}
function cerrarGanador() {
  document.getElementById('winner-overlay').classList.remove('show');
  fetch('/api/winner/hide',{method:'POST'}).catch(()=>{});
}
function generarConfetti() {
  const wrap=document.getElementById('confetti-wrap');
  if (!wrap) return;
  wrap.innerHTML='';
  const colors=['#c9a227','#e8c84a','#fff','#f0ece0','#2ecc71'];
  for(let i=0;i<80;i++){
    const el=document.createElement('div');
    el.className='confetti-piece';
    el.style.left=Math.random()*100+'vw';
    el.style.background=colors[Math.floor(Math.random()*colors.length)];
    el.style.width=(6+Math.random()*10)+'px';
    el.style.height=(10+Math.random()*16)+'px';
    el.style.animationDuration=(3+Math.random()*5)+'s';
    el.style.animationDelay=(Math.random()*4)+'s';
    el.style.borderRadius=Math.random()>.5?'50%':'2px';
    wrap.appendChild(el);
  }
}
</script>
</body>
</html>"""

TARJETAS_HTML = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Tarjetas — Jagger VIP</title>
<link href="https://fonts.googleapis.com/css2?family=Oswald:wght@400;600;700&family=Rajdhani:wght@400;600;700&display=swap" rel="stylesheet">
<style>
:root{--gold:#c9a227;--gold2:#e8c84a;--black:#080808;--surface:#111;--border:#2a2a2a;--text:#e0e0e0;}
*{box-sizing:border-box;margin:0;padding:0;}
body{background:var(--black);color:var(--text);font-family:'Rajdhani',sans-serif;padding:0;}
.top-bar{display:flex;align-items:center;gap:12px;padding:12px 16px;background:var(--surface);border-bottom:1px solid var(--border);position:sticky;top:0;z-index:100;}
.top-title{font-family:'Oswald',sans-serif;font-size:18px;font-weight:700;color:var(--gold);letter-spacing:2px;flex:1;}
.back-btn{color:#555;font-size:12px;letter-spacing:2px;text-decoration:none;text-transform:uppercase;padding:6px 12px;border:1px solid #222;border-radius:6px;transition:all .15s;}
.back-btn:hover{color:var(--gold);border-color:var(--gold);}
.actions-bar{display:flex;gap:12px;padding:16px 20px;border-bottom:1px solid #111;flex-wrap:wrap;align-items:center;}
.btn-guardar{background:var(--gold);color:#000;border:none;border-radius:8px;padding:14px 28px;font-family:'Rajdhani',sans-serif;font-size:16px;font-weight:700;letter-spacing:2px;cursor:pointer;transition:background .15s;text-transform:uppercase;}
.btn-guardar:hover{background:var(--gold2);}
.btn-limpiar{background:transparent;color:#555;border:1px solid #222;border-radius:8px;padding:14px 20px;font-family:'Rajdhani',sans-serif;font-size:15px;cursor:pointer;transition:all .15s;}
.btn-limpiar:hover{border-color:#a83030;color:#a83030;}
.scan-badge{display:inline-flex;align-items:center;gap:6px;background:#0a1a0a;border:1px solid #1a3a1a;border-radius:20px;padding:6px 14px;font-size:12px;color:#3a9a5a;letter-spacing:1px;}
.scan-badge.active{background:#1a3a1a;border-color:#3a9a5a;animation:pulse 1s infinite;}
@keyframes pulse{0%,100%{opacity:1;}50%{opacity:0.6;}}
.tarjetas-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(290px,1fr));gap:18px;padding:20px;}
.tarjeta-conf{background:#0d0d0d;border:2px solid #2a2a2a;border-radius:12px;padding:20px 22px;transition:border-color .2s;}
.tarjeta-conf.configurada{border-color:#3a5a0a;background:#0a0f00;}
.tc-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:14px;}
.tc-num{font-family:'Oswald',sans-serif;font-size:16px;color:#888;letter-spacing:2px;}
.tarjeta-conf.configurada .tc-num{color:var(--gold);}
.tc-btns{display:flex;gap:8px;align-items:center;}
.tc-status{font-size:12px;letter-spacing:1px;color:#333;font-weight:600;}
.tarjeta-conf.configurada .tc-status{color:#4aaa2a;}
.tc-scan-btn{background:#111;border:1px solid #333;border-radius:6px;color:#aaa;font-size:13px;padding:8px 14px;font-family:'Rajdhani',sans-serif;font-weight:700;letter-spacing:1px;cursor:pointer;transition:all .15s;text-transform:uppercase;}
.tc-scan-btn:hover,.tc-scan-btn.activo{border-color:var(--gold);color:var(--gold);}
.tc-confirm-btn{background:var(--gold);border:none;border-radius:6px;color:#000;font-size:13px;padding:8px 14px;font-family:'Rajdhani',sans-serif;font-weight:700;letter-spacing:1px;cursor:pointer;display:none;text-transform:uppercase;}
.tc-clear-btn{background:transparent;border:1px solid #333;border-radius:6px;color:#555;font-size:14px;cursor:pointer;transition:color .15s;padding:6px 10px;}
.tc-clear-btn:hover{color:#a83030;border-color:#a83030;}
.tc-field{margin-bottom:12px;}
.tc-label{font-size:12px;color:#aaa;letter-spacing:1px;text-transform:uppercase;display:block;margin-bottom:6px;font-weight:600;}
.tc-input{width:100%;background:#0a0a0a;border:1px solid #333;border-radius:7px;color:#f0ece0;padding:11px 14px;font-family:'Rajdhani',sans-serif;font-size:17px;outline:none;transition:border .15s;}
.tc-input:focus{border-color:var(--gold);}
.tc-code{font-size:12px;color:#3a3a3a;letter-spacing:1px;margin-top:8px;}
.tc-saldo-bar{margin-top:10px;}
.tc-saldo-info{display:flex;justify-content:space-between;font-size:12px;margin-bottom:6px;font-family:'Rajdhani',sans-serif;}
.tc-saldo-used{color:#c84a4a;font-weight:700;}
.tc-saldo-left{color:#3a9a5a;font-weight:700;}
.tc-bar-wrap{height:6px;background:#1a1a1a;border-radius:3px;overflow:hidden;}
.tc-bar-fill{height:100%;background:linear-gradient(to right,#c9a227,#e8c84a);border-radius:3px;transition:width .5s;}
#toast-box{position:fixed;bottom:20px;right:20px;z-index:9999;display:flex;flex-direction:column;gap:8px;pointer-events:none;}
.toast{background:#111;border:1px solid #333;border-radius:8px;padding:10px 16px;font-size:13px;letter-spacing:1px;color:var(--text);animation:fadeIn .2s;}
.toast.err{border-color:#5a2020;color:#e74c3c;}
@keyframes fadeIn{from{opacity:0;transform:translateY(10px)}to{opacity:1;transform:translateY(0)}}
</style>
</head>
<body>
<div style="background:#fbbf24;color:#000;font-family:'Rajdhani',sans-serif;font-size:18px;font-weight:700;padding:18px 24px;border-radius:0;width:100%;line-height:1.5;letter-spacing:0.5px;">
  ⚠️ IMPORTANTE: GUARDAR SIEMPRE SINO NO SE RECONOCERA LA TARJETA
</div>
<div class="top-bar">
  <div class="top-title">💳 Configuración de Tarjetas</div>
  <a href="/" class="back-btn">← Hub</a>
</div>
<div id="unsaved-warning" style="display:none;background:#1a0e00;border:1px solid #c9a227;border-radius:8px;padding:10px 16px;margin-bottom:12px;font-family:'Rajdhani',sans-serif;font-size:13px;color:#e8c84a;letter-spacing:1px;">
  ⚠ Tenés cambios sin guardar — hacé click en <strong>Guardar configuración</strong> para no perderlos
</div>
<div class="actions-bar">
  <button class="btn-guardar" onclick="guardarTarjetas()">💾 Guardar configuración</button>
  <button class="btn-limpiar" onclick="limpiarTarjetas()">Limpiar todo</button>
  <div class="scan-badge" id="scan-badge">▤ Lector listo</div>
</div>
<!-- ── TUTORIAL TARJETAS ── -->
<div style="margin:12px 20px 0;">
  <button onclick="(function(b,p){b._open=!b._open;p.style.display=b._open?'block':'none';b.querySelector('.tut-arrow').textContent=b._open?'▲':'▼';})(this,document.getElementById('tut-tarjetas'))" style="width:100%;display:flex;justify-content:space-between;align-items:center;background:#0a1500;border:1px solid #2a4a00;border-radius:8px;padding:12px 18px;font-family:'Rajdhani',sans-serif;font-size:14px;font-weight:700;color:#8acc50;letter-spacing:1px;cursor:pointer;text-align:left;">
    <span>📖 Cómo usar — Tarjetas</span><span class="tut-arrow" style="font-size:12px;color:#556;">▼</span>
  </button>
  <div id="tut-tarjetas" style="display:none;background:#080f00;border:1px solid #1a3a00;border-radius:0 0 8px 8px;padding:16px 20px;font-family:'Rajdhani',sans-serif;font-size:14px;color:#aaa;line-height:1.8;letter-spacing:0.3px;">
    <div style="display:flex;flex-direction:column;gap:10px;">
      <div style="display:flex;gap:12px;align-items:flex-start;"><span style="color:#8acc50;font-size:18px;flex-shrink:0;">1</span><span><strong style="color:#c9a227;">Asignale un slot a cada tarjeta:</strong> Cada tarjeta física corresponde a una mesa. Usá el botón <strong>LEER</strong> en el slot correcto y pasá la tarjeta por el lector.</span></div>
      <div style="display:flex;gap:12px;align-items:flex-start;"><span style="color:#8acc50;font-size:18px;flex-shrink:0;">2</span><span><strong style="color:#c9a227;">Confirmá la vinculación:</strong> Después de leer la tarjeta, aparece el botón <strong>CONFIRMAR</strong>. Apretalo para registrar la tarjeta en ese slot.</span></div>
      <div style="display:flex;gap:12px;align-items:flex-start;"><span style="color:#8acc50;font-size:18px;flex-shrink:0;">3</span><span><strong style="color:#c9a227;">Completá nombre y saldo inicial:</strong> Ingresá el nombre del cliente y el saldo inicial (crédito disponible para esa mesa).</span></div>
      <div style="display:flex;gap:12px;align-items:flex-start;"><span style="color:#8acc50;font-size:18px;flex-shrink:0;">4</span><span><strong style="color:#c9a227;">⚠ SIEMPRE GUARDAR:</strong> Tocá <strong>💾 Guardar configuración</strong> después de cada cambio. Sin guardar, el sistema no reconocerá la tarjeta.</span></div>
      <div style="display:flex;gap:12px;align-items:flex-start;"><span style="color:#8acc50;font-size:18px;flex-shrink:0;">5</span><span><strong style="color:#c9a227;">Verificación:</strong> Podés pasar cualquier tarjeta sin tener ningún slot activo para verificar si está vinculada.</span></div>
    </div>
  </div>
</div>
<div style="background:#0d1500;border:1px solid #2a4a00;border-radius:8px;padding:12px 20px;margin:12px 20px 0;font-family:'Rajdhani',sans-serif;font-size:14px;color:#8acc50;letter-spacing:0.5px;line-height:1.6;">
  💡 <strong>Nota multi-tarjeta:</strong> Si una persona tiene más de una tarjeta registrada con el mismo nombre, sus consumos se suman en el ranking y se muestra la mesa de número más bajo.
</div>
<div id="scan-invalid-msg" style="display:none;background:#1a0000;border:1px solid #6a0000;border-radius:8px;padding:12px 20px;margin:10px 20px 0;font-family:'Rajdhani',sans-serif;font-size:16px;color:#ff4444;font-weight:700;letter-spacing:1px;">
  ❌ Tarjeta no reconocida — no está vinculada a ninguna mesa
</div>
<div class="tarjetas-grid" id="tarjetas-grid"></div>
<div id="toast-box"></div>
<script>
let confTarjetas = Array.from({length:30},(_,i)=>({slot:i+1,codigo:'',saldo_inicial:'',nombre_cliente:''}));
let tarjetasData = {};
let pendingCodigos = {};
let scanSlotActivo = null;
let globalBuffer = '';
let globalTimeout = null;
const READER_SPEED_MS = 80;
let lastKeyTime = 0;
let isDirty = false;
function markDirty() { isDirty=true; const w=document.getElementById('unsaved-warning'); if(w) w.style.display='block'; }

function fmt(n){return '$'+Number(n).toLocaleString('es-AR',{minimumFractionDigits:0,maximumFractionDigits:0});}
function showToast(msg,err=false){
  const box=document.getElementById('toast-box');
  const t=document.createElement('div');t.className='toast'+(err?' err':'');t.textContent=msg;
  box.appendChild(t);setTimeout(()=>t.remove(),3000);
}

// ── Barcode scanner ──
document.addEventListener('keydown', e => {
  if(['INPUT','TEXTAREA'].includes(document.activeElement.tagName)) return;
  const now = Date.now();
  lastKeyTime = now;
  if (e.key==='Enter') {
    if (globalBuffer.length>1) handleScan(globalBuffer.trim());
    globalBuffer=''; return;
  }
  if (e.key.length===1) globalBuffer+=e.key;
  clearTimeout(globalTimeout);
  globalTimeout=setTimeout(()=>{globalBuffer='';lastKeyTime=0;},500);
});

function handleScan(codigo) {
  const invalidMsg = document.getElementById('scan-invalid-msg');
  if (scanSlotActivo===null) {
    // Modo verificación: chequear si la tarjeta está vinculada
    const encontrada = confTarjetas.find(t => t.codigo === codigo);
    if (encontrada) {
      if (invalidMsg) invalidMsg.style.display='none';
      showToast('Tarjeta reconocida → Slot ' + encontrada.slot + ' (Mesa ' + encontrada.slot + ')');
    } else {
      if (invalidMsg) { invalidMsg.style.display='block'; setTimeout(()=>{ if(invalidMsg) invalidMsg.style.display='none'; },4000); }
      showToast('❌ Tarjeta no reconocida',true);
    }
    return;
  }
  if (invalidMsg) invalidMsg.style.display='none';
  const idx=scanSlotActivo;
  pendingCodigos[idx]=codigo;
  renderConfigTarjetas();
  showToast('Tarjeta leída: '+codigo+' → presioná CONFIRMAR');
}

function iniciarScan(idx) {
  if(scanSlotActivo!==null){
    const pb=document.getElementById('scan-btn-'+scanSlotActivo);
    if(pb){pb.textContent='LEER';pb.classList.remove('activo');}
  }
  scanSlotActivo=idx;
  const btn=document.getElementById('scan-btn-'+idx);
  btn.textContent='ESPERANDO...';btn.classList.add('activo');
  const badge=document.getElementById('scan-badge');
  if(badge){badge.classList.add('active');badge.textContent='⟳ Esperando tarjeta para slot '+(idx+1);}
  showToast('Pasa la tarjeta por el lector');
  setTimeout(()=>{
    if(scanSlotActivo===idx){
      scanSlotActivo=null;
      btn.textContent='LEER';btn.classList.remove('activo');
      if(badge){badge.classList.remove('active');badge.textContent='▤ Lector listo';}
    }
  },10000);
}

function confirmarVinculo(idx) {
  const codigo=pendingCodigos[idx];
  if(!codigo) return;
  confTarjetas[idx].codigo=codigo;
  delete pendingCodigos[idx];
  renderConfigTarjetas();
  markDirty();
  showToast('Tarjeta vinculada al slot '+(idx+1));
}

function clearSlot(idx){confTarjetas[idx].codigo='';delete pendingCodigos[idx];renderConfigTarjetas();markDirty();}

function renderConfigTarjetas() {
  const grid=document.getElementById('tarjetas-grid');
  grid.innerHTML=confTarjetas.map((t,i)=>{
    const configurada=t.codigo;
    const saldoInfo=tarjetasData[t.codigo];
    const saldoIni=parseFloat(t.saldo_inicial||0);
    const saldoAct=saldoInfo!==undefined?saldoInfo.saldo_actual:saldoIni;
    const pct=saldoIni>0?Math.max(0,Math.round((saldoAct/saldoIni)*100)):0;
    const hasPending=pendingCodigos[i]!==undefined;
    return `<div class="tarjeta-conf ${configurada?'configurada':''}" id="tc-${i}">
      <div class="tc-header">
        <span class="tc-num">Tarjeta ${t.slot} — Mesa ${t.slot}</span>
        <div class="tc-btns">
          <span class="tc-status" style="${configurada?'color:#3a9a5a':''}">${configurada?'Vinculada':'Sin vincular'}</span>
          <button class="tc-scan-btn" id="scan-btn-${i}" onclick="iniciarScan(${i})">LEER</button>
          <button class="tc-confirm-btn" id="confirm-btn-${i}" onclick="confirmarVinculo(${i})" style="display:${hasPending?'inline-block':'none'}">CONFIRMAR</button>
          <button class="tc-clear-btn" onclick="clearSlot(${i})">✕</button>
        </div>
      </div>
      <div class="tc-field">
        <span class="tc-label">Nombre del cliente</span>
        <input class="tc-input" id="tc-nombre-${i}" type="text" placeholder="Nombre..." value="${t.nombre_cliente||''}"
          oninput="confTarjetas[${i}].nombre_cliente=this.value;markDirty()" />
      </div>
      <div class="tc-field">
        <span class="tc-label">Saldo inicial ($)</span>
        <input class="tc-input" id="tc-saldo-${i}" type="number" placeholder="0" value="${t.saldo_inicial}"
          oninput="confTarjetas[${i}].saldo_inicial=this.value;markDirty()" />
      </div>
      <div class="tc-code">${hasPending?'Leída: '+pendingCodigos[i]:(t.codigo?'Cód: '+t.codigo:'Presioná LEER y pasá la tarjeta')}</div>
      ${saldoIni>0?`<div style="margin-top:10px;">
        <div style="display:flex;justify-content:space-between;font-size:13px;font-family:'Rajdhani',sans-serif;letter-spacing:.5px;margin-bottom:6px;">
          <span style="color:#c84a4a;font-weight:700;">Gastado: ${fmt(Math.max(0,saldoIni-saldoAct))}</span>
          <span style="color:#3a9a5a;font-weight:700;">Disponible: ${fmt(Math.max(0,saldoAct))}</span>
        </div>
        <div style="height:8px;background:#111;border-radius:4px;overflow:hidden;display:flex;">
          <div style="height:100%;width:${Math.max(0,Math.min(100,100-pct))}%;background:linear-gradient(to right,#6b1010,#c84a4a);"></div>
          <div style="height:100%;flex:1;background:linear-gradient(to right,#1a5a2a,#3a9a5a);"></div>
        </div>
      </div>`:''}
      ${(()=>{const recargas=saldoInfo?saldoInfo.recargas||[]:[];if(!recargas.length)return '';return `<div style="margin-top:8px;border-top:1px solid #1a1a1a;padding-top:6px;"><div style="font-size:10px;color:#555;letter-spacing:2px;text-transform:uppercase;margin-bottom:5px;">Recargas</div>${recargas.map(r=>`<div style="display:flex;align-items:center;gap:6px;padding:4px 0;border-bottom:1px solid #0f0f0f;"><span style="font-size:11px;color:#666;min-width:50px;">${r.fecha||''} ${r.hora||''}</span><span style="flex:1;font-family:'Oswald',sans-serif;font-size:14px;color:#3a9a5a;font-weight:700;">${fmt(r.monto)}</span><button onclick="editarRecarga('${t.codigo}','${r.id}',${r.monto})" style="background:transparent;border:1px solid #2a2a2a;border-radius:4px;padding:2px 7px;color:#777;font-size:11px;cursor:pointer;">✎</button><button onclick="eliminarRecarga('${t.codigo}','${r.id}')" style="background:transparent;border:1px solid #2a1a1a;border-radius:4px;padding:2px 7px;color:#883333;font-size:11px;cursor:pointer;">✕</button></div>`).join('')}</div>`;})()}
    </div>`;
  }).join('');
}

async function eliminarRecarga(codigo, recargaId) {
  if (!confirm('¿Eliminar esta recarga?')) return;
  try {
    const r = await fetch(`/api/tarjetas/recarga/${codigo}/${recargaId}`, {method:'DELETE'});
    const d = await r.json();
    if (d.ok) { showToast('Recarga eliminada — nuevo saldo: '+fmt(d.nuevo_saldo)); await cargarDatos(); renderConfigTarjetas(); }
    else showToast(d.error||'Error', true);
  } catch(e) { showToast('Error al eliminar', true); }
}

async function editarRecarga(codigo, recargaId, montoActual) {
  const nuevoStr = prompt('Editar monto de recarga (actual: '+fmt(montoActual)+'):', montoActual);
  if (nuevoStr === null) return;
  const nuevo = parseFloat(nuevoStr);
  if (!nuevo || nuevo <= 0) { showToast('Monto inválido', true); return; }
  try {
    const r = await fetch(`/api/tarjetas/recarga/${codigo}/${recargaId}`, {
      method:'PUT', headers:{'Content-Type':'application/json'}, body:JSON.stringify({monto:nuevo})
    });
    const d = await r.json();
    if (d.ok) { showToast('Recarga actualizada — nuevo saldo: '+fmt(d.nuevo_saldo)); await cargarDatos(); renderConfigTarjetas(); }
    else showToast(d.error||'Error', true);
  } catch(e) { showToast('Error al editar', true); }
}

async function guardarTarjetas() {
  try {
    const res=await fetch('/api/tarjetas/config',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(confTarjetas)});
    if(res.ok){isDirty=false;const w=document.getElementById('unsaved-warning');if(w)w.style.display='none';showToast('Configuración guardada');await cargarDatos();renderConfigTarjetas();}
    else{const err=await res.json();showToast(err.error||'Error al guardar',true);}
  }catch(e){showToast('Error al guardar',true);}
}

async function limpiarTarjetas() {
  if(!confirm('¿Limpiar toda la configuración?')) return;
  confTarjetas=Array.from({length:30},(_,i)=>({slot:i+1,codigo:'',saldo_inicial:'',nombre_cliente:''}));
  await fetch('/api/tarjetas/limpiar',{method:'POST'});
  renderConfigTarjetas();showToast('Tarjetas limpiadas');
}

async function cargarDatos() {
  try{
    const [r1,r2]=await Promise.all([fetch('/api/tarjetas/config'),fetch('/api/tarjetas')]);
    if(r1.ok){const d=await r1.json();if(d&&d.length)confTarjetas=d;}
    if(r2.ok) tarjetasData=await r2.json();
  }catch(e){}
}

cargarDatos().then(renderConfigTarjetas);
setInterval(()=>fetch('/api/tarjetas').then(r=>r.json()).then(d=>{tarjetasData=d;renderConfigTarjetas();}).catch(()=>{}), 15000);
</script>
</body>
</html>"""

CAJA_HTML = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0,maximum-scale=1">
<title>Caja {{ caja_nombre }} — Jagger VIP</title>
<link href="https://fonts.googleapis.com/css2?family=Oswald:wght@400;600;700&family=Rajdhani:wght@400;600;700&display=swap" rel="stylesheet">
<style>
:root{--gold:#c9a227;--gold2:#e8c84a;--black:#080808;--surface:#0f0f0f;--border:#2a2a2a;--text:#e0e0e0;--green:#2ecc71;--danger:#e74c3c;}
*{box-sizing:border-box;margin:0;padding:0;}
body{background:var(--black);color:var(--text);font-family:'Rajdhani',sans-serif;min-height:100vh;padding-bottom:120px;}
.top-bar{display:flex;align-items:center;gap:12px;padding:12px 16px;background:var(--surface);border-bottom:1px solid var(--border);position:sticky;top:0;z-index:100;}
.caja-badge{font-family:'Oswald',sans-serif;font-size:13px;font-weight:700;letter-spacing:3px;color:var(--gold);text-transform:uppercase;background:#1a1500;border:1px solid #3a3000;border-radius:6px;padding:4px 12px;}
.caja-title{font-family:'Oswald',sans-serif;font-size:17px;font-weight:700;letter-spacing:2px;color:var(--text);flex:1;}
.back-btn{color:#555;font-size:12px;letter-spacing:2px;text-decoration:none;text-transform:uppercase;padding:6px 12px;border:1px solid #222;border-radius:6px;transition:all .15s;}
.back-btn:hover{color:var(--gold);border-color:var(--gold);}
/* Tabs */
.modo-tabs{display:flex;gap:0;border-bottom:1px solid var(--border);background:var(--surface);}
.modo-tab{padding:12px 18px;font-family:'Rajdhani',sans-serif;font-size:13px;font-weight:600;letter-spacing:2px;color:#555;cursor:pointer;border-bottom:2px solid transparent;transition:all .15s;text-transform:uppercase;user-select:none;}
.modo-tab:hover{color:var(--text);}
.modo-tab.active{color:var(--gold);border-bottom-color:var(--gold);}
.modo-content{display:none;}
.modo-content.active{display:block;}
/* Scan zone */
.scan-zone{padding:12px 16px;background:#0a0a0a;border-bottom:1px solid #1a1a1a;}
.scan-hint{display:flex;align-items:center;gap:10px;padding:12px 14px;background:var(--surface);border:1px solid var(--border);border-radius:8px;font-size:14px;letter-spacing:1px;font-weight:600;}
.scan-hint.esperando{border-color:#554400;color:#c9a227;background:#0d0c00;animation:pulse 2s infinite;}
.scan-hint.activa{border-color:#1a5a2a;color:#2ecc71;background:#071207;}
@keyframes pulse{0%,100%{border-color:#222;}50%{border-color:#3a3a3a;}}
.scan-icon{font-size:20px;color:#444;}
.tarjeta-card{background:#0a0a08;border:1px solid #3a3000;border-radius:8px;padding:14px 16px;display:none;}
.tarjeta-card.visible{display:block;}
.tarjeta-card.sin-saldo{border-color:#5a2020;}
.tarjeta-top{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:8px;}
.tarjeta-mesa-label,.tarjeta-saldo-label{font-size:12px;color:#aaa;letter-spacing:2px;text-transform:uppercase;}
.tarjeta-mesa-num{font-family:'Oswald',sans-serif;font-size:32px;color:var(--gold);font-weight:700;}
.tarjeta-saldo-wrap{text-align:right;}
.tarjeta-saldo{font-family:'Oswald',sans-serif;font-size:22px;color:var(--gold);font-weight:700;}
.tarjeta-bar-wrap{height:7px;background:#1a2a1a;border-radius:4px;margin-bottom:6px;overflow:hidden;}
.tarjeta-bar{height:100%;background:linear-gradient(to right,#1a6a2a,#2ecc71);border-radius:4px;transition:width .6s cubic-bezier(.16,1,.3,1);}
.tarjeta-nombre{font-size:12px;color:#888;letter-spacing:1px;}
.saldo-bajo-warn{font-size:11px;color:#c9a227;border:1px solid #3a3000;border-radius:6px;padding:5px 10px;margin-top:6px;letter-spacing:1px;}
/* Category bar */
.cat-bar{display:flex;gap:8px;padding:10px 16px;overflow-x:auto;border-bottom:1px solid #111;background:#0a0a0a;}
.cat-bar::-webkit-scrollbar{height:2px;}
.cat-bar::-webkit-scrollbar-thumb{background:#333;}
.cat-chip{flex-shrink:0;background:var(--surface);border:1px solid var(--border);border-radius:20px;padding:5px 14px;font-size:12px;color:#777;cursor:pointer;transition:all .15s;letter-spacing:1px;text-transform:uppercase;white-space:nowrap;}
.cat-chip.active,.cat-chip:hover{background:#1a1500;border-color:var(--gold);color:var(--gold);}
.cat-chip.varios{border-color:#3a4a2a;color:#5a8a4a;}
.cat-chip.varios.active{background:#0a1a0a;border-color:#5a8a4a;color:var(--green);}
/* Product grid */
.prod-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(150px,1fr));gap:10px;padding:12px 16px;}
.prod-btn{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:12px 10px;cursor:pointer;transition:all .15s;text-align:left;position:relative;user-select:none;}
.prod-btn:active{transform:scale(0.97);}
.prod-btn:hover{border-color:var(--gold);background:#111;}
.prod-btn.selected{border-color:var(--gold);background:#1a1500;}
.prod-cat{font-size:10px;color:#666;letter-spacing:2px;text-transform:uppercase;margin-bottom:3px;}
.prod-name{font-size:15px;font-weight:700;color:#f0ece0;margin-bottom:5px;line-height:1.3;}
.prod-price{font-family:'Oswald',sans-serif;font-size:19px;color:var(--gold);font-weight:700;}
.prod-qty-badge{position:absolute;top:6px;right:8px;background:var(--gold);color:#000;font-family:'Oswald',sans-serif;font-size:13px;font-weight:700;border-radius:50%;width:22px;height:22px;display:flex;align-items:center;justify-content:center;}
/* Varios panel */
.varios-panel{padding:12px 16px;background:#0a0f0a;border-bottom:1px solid #1a2a1a;display:none;}
.varios-panel.show{display:block;}
.varios-row{display:flex;gap:10px;align-items:flex-end;}
.field-label{display:block;font-size:10px;color:#666;letter-spacing:2px;text-transform:uppercase;margin-bottom:5px;}
.field-input{width:100%;background:#0d0d0d;border:1px solid var(--border);border-radius:7px;color:var(--text);padding:10px 13px;font-family:'Rajdhani',sans-serif;font-size:15px;outline:none;transition:border .15s;}
.field-input:focus{border-color:var(--gold);}
.btn-add-varios{background:#1a3a1a;border:1px solid #3a6a3a;color:var(--green);border-radius:7px;padding:10px 16px;font-family:'Rajdhani',sans-serif;font-size:14px;font-weight:700;cursor:pointer;white-space:nowrap;transition:all .15s;}
.btn-add-varios:hover{background:#2a4a2a;}
.hint-miles{font-size:11px;color:#555;letter-spacing:1px;margin-top:3px;}
.hint-miles.ok{color:var(--green);}
/* Order panel */
.order-panel{background:#0a0a0a;border-top:2px solid #2a2a2a;padding:14px 16px;display:none;}
.order-panel.show{display:block;}
.order-header{font-size:11px;color:#c9a227;letter-spacing:3px;text-transform:uppercase;margin-bottom:10px;font-family:'Oswald',sans-serif;font-weight:600;}
.order-item{display:flex;align-items:center;gap:8px;padding:7px 0;border-bottom:1px solid #111;}
.oi-name{flex:1;font-size:15px;font-weight:700;color:#f0ece0;}
.oi-price{font-family:'Oswald',sans-serif;font-size:16px;color:var(--gold);font-weight:700;}
.qty-btn{background:#1a1a1a;border:1px solid #333;border-radius:5px;width:26px;height:26px;color:var(--text);font-size:16px;cursor:pointer;display:flex;align-items:center;justify-content:center;transition:all .1s;font-family:'Oswald',sans-serif;line-height:1;}
.qty-btn:hover{border-color:var(--gold);color:var(--gold);}
.qty-val{font-family:'Oswald',sans-serif;font-size:15px;min-width:22px;text-align:center;}
.order-summary{display:flex;justify-content:space-between;align-items:center;padding:10px 0 8px;margin-top:4px;}
.order-total-val{font-family:'Oswald',sans-serif;font-size:22px;color:var(--gold);font-weight:700;}
/* Confirm bar */
.confirm-bar{padding:10px 14px;background:#0a0a08;border-top:2px solid #1e1e1e;display:none;position:sticky;bottom:50px;z-index:100;gap:8px;align-items:center;}
.confirm-bar.show{display:flex;}
.name-input{flex:1;min-width:100px;background:#0d0d0d;border:1px solid var(--border);border-radius:7px;color:var(--text);padding:9px 12px;font-family:'Rajdhani',sans-serif;font-size:15px;outline:none;transition:border .15s;}
.cb-tag{font-family:'Oswald',sans-serif;font-size:13px;color:#c9a227;background:#1a1500;border:1px solid #3a3000;border-radius:6px;padding:4px 10px;white-space:nowrap;flex-shrink:0;letter-spacing:1px;}
.cb-total{font-family:'Oswald',sans-serif;font-size:20px;color:#fff;font-weight:700;white-space:nowrap;flex-shrink:0;}
.btn-cancel{background:transparent;color:#cc3333;border:1px solid #4a1a1a;border-radius:7px;padding:10px 14px;font-size:18px;cursor:pointer;transition:all .15s;flex-shrink:0;line-height:1;}
.btn-cancel:hover{background:#1a0808;border-color:#cc3333;}
.btn-cobrar-ok{background:#1a4a1a;color:#2ecc71;border:1px solid #2ecc71;border-radius:7px;padding:10px 18px;font-family:'Rajdhani',sans-serif;font-size:15px;font-weight:700;letter-spacing:1px;cursor:pointer;transition:all .15s;white-space:nowrap;flex-shrink:0;}
.btn-cobrar-ok:hover:not(:disabled){background:#2a6a2a;}
.btn-cobrar-ok:disabled{opacity:.35;cursor:not-allowed;}
.name-input:focus{border-color:var(--gold);}
.btn-cobrar{background:var(--gold);color:#000;border:none;border-radius:7px;padding:11px 22px;font-family:'Rajdhani',sans-serif;font-size:15px;font-weight:700;letter-spacing:1px;cursor:pointer;transition:background .15s;white-space:nowrap;}
.btn-cobrar:hover:not(:disabled){background:var(--gold2);}
.btn-cobrar:disabled{opacity:.4;cursor:not-allowed;}
.btn-clear{background:transparent;color:#555;border:1px solid #222;border-radius:7px;padding:11px 14px;font-family:'Rajdhani',sans-serif;font-size:14px;cursor:pointer;transition:all .15s;white-space:nowrap;}
.btn-clear:hover{border-color:#555;color:#aaa;}
/* Manual form */
.form-card{padding:16px;}
.form-section{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:16px;margin-bottom:12px;}
.field-input.wide{margin-bottom:10px;}
.btn-submit{background:var(--gold);color:#000;border:none;border-radius:7px;padding:12px 24px;font-family:'Rajdhani',sans-serif;font-size:14px;font-weight:700;letter-spacing:1px;cursor:pointer;transition:background .15s;}
.btn-submit:hover{background:var(--gold2);}
/* TX list */
.section-label{font-size:10px;color:#444;letter-spacing:3px;text-transform:uppercase;padding:10px 16px 6px;border-top:1px solid #111;}
.tx-list{padding:0 16px 8px;}
.tx-item{display:flex;justify-content:space-between;align-items:flex-start;padding:8px 0;border-bottom:1px solid #111;font-size:14px;}
.tx-name{color:#f0ece0;font-weight:700;font-size:16px;letter-spacing:.5px;}
.tx-meta{font-size:13px;color:#888;letter-spacing:1px;}
.tx-amount{font-family:'Oswald',sans-serif;font-size:20px;color:var(--gold);font-weight:700;}
.tx-del{color:#333;cursor:pointer;padding:0 4px;font-size:16px;transition:color .15s;}
.tx-del:hover{color:var(--danger);}
.caja-total-bar{position:fixed;bottom:0;left:0;right:0;display:flex;justify-content:space-between;align-items:center;padding:8px 16px;background:#0a0a08;border-top:1px solid var(--border);z-index:50;min-height:50px;}
.caja-total-label{font-size:11px;color:#555;letter-spacing:3px;text-transform:uppercase;}
.caja-total-val{font-family:'Oswald',sans-serif;font-size:20px;color:var(--gold);font-weight:700;}
#toast-box{position:fixed;bottom:60px;right:16px;z-index:9999;display:flex;flex-direction:column;gap:8px;pointer-events:none;}
.toast{background:#111;border:1px solid #333;border-radius:8px;padding:10px 16px;font-size:13px;letter-spacing:1px;color:var(--text);animation:fadeIn .2s;}
.toast.err{border-color:#5a2020;color:#e74c3c;}
@keyframes fadeIn{from{opacity:0;transform:translateY(8px)}to{opacity:1;transform:translateY(0)}}
</style>
</head>
<body>
<div class="top-bar">
  <div class="caja-badge">Caja {{ caja_num }}</div>
  <div class="caja-title">{{ caja_nombre }}</div>
  <button onclick="abrirCartelModal()" style="background:#1a1500;border:1px solid var(--gold);color:var(--gold);border-radius:6px;padding:6px 12px;font-family:'Rajdhani',sans-serif;font-size:12px;font-weight:700;letter-spacing:2px;cursor:pointer;text-transform:uppercase;flex-shrink:0;">🍾 Cartel</button>
  <button onclick="cerrarGanadorDesdeCaja()" style="background:#0d0800;border:1px solid #3a2a00;color:#7a6010;border-radius:6px;padding:6px 12px;font-family:'Rajdhani',sans-serif;font-size:12px;font-weight:700;letter-spacing:2px;cursor:pointer;text-transform:uppercase;flex-shrink:0;transition:all .15s;" onmouseover="this.style.borderColor='#c9a227';this.style.color='#c9a227'" onmouseout="this.style.borderColor='#3a2a00';this.style.color='#7a6010'">🏆 Cerrar Ganador</button>
  <button onclick="cerrarCartelPantalla()" style="background:#150a0a;border:1px solid #5a2020;color:#e74c3c;border-radius:6px;padding:6px 12px;font-family:'Rajdhani',sans-serif;font-size:12px;font-weight:700;letter-spacing:2px;cursor:pointer;text-transform:uppercase;flex-shrink:0;">✕ Cerrar</button>
  <a href="/" class="back-btn">← Hub</a>
</div>

<div class="modo-tabs">
  <div class="modo-tab active" id="mtab-tarjeta" onclick="setModo('tarjeta')">\U0001f4b3 Con Tarjeta</div>
  <div class="modo-tab" id="mtab-manual" onclick="setModo('manual')">✎ Sin Tarjeta</div>
  <div class="modo-tab" id="mtab-recargar" onclick="setModo('recargar')" style="color:#3a9a5a;">⊕ Recargar</div>
</div>

<!-- ── TUTORIAL CAJA ── -->
<div style="margin:8px 12px 0;">
  <button onclick="(function(b,p){b._open=!b._open;p.style.display=b._open?'block':'none';b.querySelector('.tut-arrow').textContent=b._open?'▲':'▼';})(this,document.getElementById('tut-caja'))" style="width:100%;display:flex;justify-content:space-between;align-items:center;background:#0a0f00;border:1px solid #2a2a00;border-radius:8px;padding:10px 16px;font-family:'Rajdhani',sans-serif;font-size:13px;font-weight:700;color:#c9a227;letter-spacing:1px;cursor:pointer;text-align:left;">
    <span>📖 Cómo usar esta caja</span><span class="tut-arrow" style="font-size:12px;color:#556;">▼</span>
  </button>
  <div id="tut-caja" style="display:none;background:#080800;border:1px solid #1a1a00;border-radius:0 0 8px 8px;padding:14px 18px;font-family:'Rajdhani',sans-serif;font-size:13px;color:#aaa;line-height:1.8;letter-spacing:0.3px;">
    <div style="display:flex;flex-direction:column;gap:8px;">
      <div style="font-size:12px;color:#c9a227;letter-spacing:2px;text-transform:uppercase;font-weight:700;margin-bottom:4px;">💳 Con Tarjeta</div>
      <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">①</span><span>Pasá la tarjeta por el lector — se carga automáticamente la mesa y el saldo.</span></div>
      <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">②</span><span>Seleccioná los productos del menú. Al tocar un producto lo agregás al pedido. Tocá de nuevo para sumar más unidades.</span></div>
      <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">③</span><span>En el cuadrado del producto: si hay <strong>1 unidad</strong>, aparece una ✕ para eliminar directo. Con <strong>2 o más</strong>, podés sumar o restar con los botones + / −.</span></div>
      <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">④</span><span>Tocá <strong>✓ Cobrar</strong> o presioná <strong>Enter</strong> para confirmar el pedido.</span></div>
      <div style="display:flex;gap:10px;"><span style="color:#c9a227;flex-shrink:0;">🗂</span><span><strong style="color:#fff;">Seleccionar mesa manualmente:</strong> tocá el botón <strong>🗂</strong> (arriba a la derecha, junto al lector) para elegir la mesa directamente de la lista. <strong style="color:#e8c84a;">No es necesario pasar la tarjeta por el lector</strong> — al elegir la mesa se carga igual que si la hubieras pasado.</span></div>
      <div style="font-size:12px;color:#3a9a5a;letter-spacing:2px;text-transform:uppercase;font-weight:700;margin-top:6px;margin-bottom:4px;">✎ Sin Tarjeta</div>
      <div style="display:flex;gap:10px;"><span style="color:#3a9a5a;flex-shrink:0;">①</span><span>Ingresá nombre y mesa del cliente, luego seleccioná productos igual que con tarjeta.</span></div>
      <div style="display:flex;gap:10px;"><span style="color:#3a9a5a;flex-shrink:0;">②</span><span>El pedido se registra sin descontar saldo de tarjeta, pero suma al ranking.</span></div>
      <div style="font-size:12px;color:#3a6a9a;letter-spacing:2px;text-transform:uppercase;font-weight:700;margin-top:6px;margin-bottom:4px;">⊕ Recargar</div>
      <div style="display:flex;gap:10px;"><span style="color:#3a6a9a;flex-shrink:0;">①</span><span>Pasá la tarjeta y luego ingresá el monto a agregar al saldo de esa mesa.</span></div>
    </div>
  </div>
</div>

<!-- ── CON TARJETA ── -->
<div class="modo-content active" id="mc-tarjeta">
  <!-- Scan zone -->
  <div class="scan-zone">
    <div style="display:flex;align-items:center;gap:8px;position:relative;">
      <div class="scan-hint esperando" id="scan-hint-tarjeta" style="flex:1;">
        <span class="scan-icon">▤</span>
        <span id="scan-txt">Pasá la tarjeta por el lector para comenzar</span>
      </div>
      <button onclick="toggleMesasPopupTarjeta()" id="btn-mesas-tarjeta" title="Seleccionar mesa manualmente" style="background:#0f0e05;border:1px solid #c9a22766;border-radius:8px;padding:10px 13px;font-size:18px;cursor:pointer;color:#c9a227;flex-shrink:0;line-height:1;transition:border-color .15s;" onmouseover="this.style.borderColor='#c9a227'" onmouseout="this.style.borderColor='#c9a22766'">🗂</button>
      <div id="mesas-popup-tarjeta" style="display:none;position:absolute;top:calc(100% + 6px);right:0;z-index:300;background:#111;border:1px solid #c9a22744;border-radius:12px;min-width:260px;max-width:320px;box-shadow:0 8px 40px rgba(0,0,0,0.9);overflow:hidden;">
        <div style="display:flex;justify-content:space-between;align-items:center;padding:10px 14px 6px;border-bottom:1px solid #1e1e1e;">
          <span style="font-family:'Oswald',sans-serif;font-size:11px;color:#c9a227;letter-spacing:2px;text-transform:uppercase;">Seleccionar mesa</span>
          <button onclick="document.getElementById('mesas-popup-tarjeta').style.display='none'" style="background:none;border:none;color:#444;font-size:16px;cursor:pointer;line-height:1;padding:0 2px;">✕</button>
        </div>
        <div id="mesas-popup-list-tarjeta" style="max-height:240px;overflow-y:auto;"></div>
      </div>
    </div>
    <div class="tarjeta-card" id="tarjeta-card"></div>
  </div>

  <!-- Category bar (disabled until card scanned) -->
  <div id="menu-area-tarjeta" style="opacity:.3;pointer-events:none;">
    <div style="padding:8px 12px 4px;">
      <input id="search-tarjeta" type="text" placeholder="🔍 Buscar producto..." autocomplete="off" oninput="searchTermT=this.value;renderMenuProds('tarjeta')" style="width:100%;background:#111;border:1px solid #2a2a2a;border-radius:8px;color:#f0ece0;padding:9px 14px;font-family:'Rajdhani',sans-serif;font-size:15px;outline:none;" onfocus="this.style.borderColor='#c9a227'" onblur="this.style.borderColor='#2a2a2a'">
    </div>
    <div class="cat-bar" id="cats-tarjeta"></div>
    <div class="prod-grid" id="prods-tarjeta"></div>
    <!-- Varios panel -->
    <div class="varios-panel" id="varios-panel-tarjeta">
      <div class="varios-row">
        <div style="flex:1;">
          <label class="field-label">Descripción (opcional)</label>
          <input class="field-input" id="varios-desc-tarjeta" type="text" placeholder="Ej: Extra cervezas..." autocomplete="off">
        </div>
        <div>
          <label class="field-label">Monto $</label>
          <input class="field-input" id="varios-monto-tarjeta" type="number" min="0" step="100" placeholder="0" style="width:130px;">
        </div>
        <button class="btn-add-varios" onclick="addVariosTarjeta()">+ Agregar</button>
      </div>
      <div class="hint-miles" id="varios-hint-tarjeta"></div>
    </div>
    <!-- Order panel -->
    <div class="order-panel" id="order-panel-tarjeta">
      <div class="order-header">Pedido actual</div>
      <div id="order-items-tarjeta"></div>
      <div class="order-summary">
        <span style="font-size:11px;color:#555;letter-spacing:2px;">TOTAL</span>
        <span class="order-total-val" id="order-total-tarjeta">$0</span>
      </div>
    </div>
    <!-- Confirm bar -->
    <div class="confirm-bar" id="confirm-bar-tarjeta">
      <span class="cb-tag" id="cb-mesa-tarjeta">▤ —</span>
      <span class="cb-total" id="cb-total-tarjeta" style="flex:1;">$0</span>
      <button class="btn-cancel" onclick="limpiarPedido('tarjeta')" title="Cancelar">✕</button>
      <button class="btn-cobrar-ok" id="btn-cobrar-tarjeta" onclick="confirmarTarjeta()" disabled>✓ Cobrar</button>
    </div>
  </div>
</div>

<!-- ── SIN TARJETA ── -->
<div class="modo-content" id="mc-manual">
  <div class="form-card">
    <div class="form-section">
      <label class="field-label">Nombre del cliente</label>
      <input class="field-input wide" id="mname" type="text" placeholder="Nombre..." autocomplete="off">
      <label class="field-label">Mesa (opcional)</label>
      <div style="display:flex;gap:8px;align-items:center;position:relative;margin-bottom:10px;">
        <input class="field-input" id="mmesa" type="text" placeholder="Ej: 5" style="flex:1;margin-bottom:0;">
        <button onclick="toggleMesasPopupCaja()" id="btn-mesas-popup" title="Ver mesas de esta noche" style="background:#0f0e05;border:1px solid #c9a22766;border-radius:8px;padding:10px 13px;font-size:18px;cursor:pointer;color:#c9a227;flex-shrink:0;line-height:1;transition:border-color .15s;" onmouseover="this.style.borderColor='#c9a227'" onmouseout="this.style.borderColor='#c9a22766'">🗂</button>
        <div id="mesas-popup-caja" style="display:none;position:absolute;top:calc(100% + 6px);right:0;z-index:300;background:#111;border:1px solid #c9a22744;border-radius:12px;min-width:260px;max-width:320px;box-shadow:0 8px 40px rgba(0,0,0,0.9);overflow:hidden;">
          <div style="display:flex;justify-content:space-between;align-items:center;padding:10px 14px 6px;border-bottom:1px solid #1e1e1e;">
            <span style="font-family:'Oswald',sans-serif;font-size:11px;color:#c9a227;letter-spacing:2px;text-transform:uppercase;">Mesas esta noche</span>
            <button onclick="document.getElementById('mesas-popup-caja').style.display='none'" style="background:none;border:none;color:#444;font-size:16px;cursor:pointer;line-height:1;padding:0 2px;">✕</button>
          </div>
          <div id="mesas-popup-list-caja" style="max-height:240px;overflow-y:auto;"></div>
        </div>
      </div>
    </div>
  </div>
  <div style="padding:8px 12px 4px;">
    <input id="search-manual" type="text" placeholder="🔍 Buscar producto..." autocomplete="off" oninput="searchTermM=this.value;renderMenuProds('manual')" style="width:100%;background:#111;border:1px solid #2a2a2a;border-radius:8px;color:#f0ece0;padding:9px 14px;font-family:'Rajdhani',sans-serif;font-size:15px;outline:none;" onfocus="this.style.borderColor='#c9a227'" onblur="this.style.borderColor='#2a2a2a'">
  </div>
  <div class="cat-bar" id="cats-manual"></div>
  <div class="prod-grid" id="prods-manual"></div>
  <div class="varios-panel" id="varios-panel-manual">
    <div class="varios-row">
      <div style="flex:1;">
        <label class="field-label">Descripción (opcional)</label>
        <input class="field-input" id="varios-desc-manual" type="text" placeholder="Ej: Extra cervezas..." autocomplete="off">
      </div>
      <div>
        <label class="field-label">Monto $</label>
        <input class="field-input" id="varios-monto-manual" type="number" min="0" step="100" placeholder="0" style="width:130px;">
      </div>
      <button class="btn-add-varios" onclick="addVariosManual()">+ Agregar</button>
    </div>
    <div class="hint-miles" id="varios-hint-manual"></div>
  </div>
  <div class="order-panel" id="order-panel-manual">
    <div class="order-header">Pedido actual</div>
    <div id="order-items-manual"></div>
    <div class="order-summary">
      <span style="font-size:11px;color:#555;letter-spacing:2px;">TOTAL</span>
      <span class="order-total-val" id="order-total-manual">$0</span>
    </div>
  </div>
  <div class="confirm-bar" id="confirm-bar-manual">
    <span class="cb-tag">✎ Sin tarjeta</span>
    <span class="cb-total" id="cb-total-manual">$0</span>
    <button class="btn-cancel" onclick="limpiarPedido('manual')" title="Cancelar">✕</button>
    <button class="btn-cobrar-ok" id="btn-cobrar-manual" onclick="confirmarManual()">✓ Registrar</button>
  </div>
</div>

<!-- ── RECARGAR ── -->
<div class="modo-content" id="mc-recargar">
  <div class="scan-zone">
    <div class="scan-hint esperando" id="scan-hint-recargar">
      <span class="scan-icon">▤</span>
      <span id="rscan-txt">Pasá la tarjeta para recargar</span>
    </div>
    <div class="tarjeta-card" id="tarjeta-card-r"></div>
  </div>
  <div class="form-card" id="form-recargar" style="opacity:.35;pointer-events:none;">
    <div class="form-section">
      <label class="field-label" style="color:#3a9a5a;">Monto a recargar ($)</label>
      <input class="field-input wide" id="recarga" type="number" min="0" step="100" placeholder="0" style="color:#3a9a5a;">
      <div class="hint-miles" id="rhint">Ingresá el monto a agregar</div>
      <button class="btn-submit" id="btn-recargar" style="background:#2a6a3a;border:1px solid #3a9a5a;" onclick="recargarTarjeta()" disabled>⊕ Recargar saldo</button>
    </div>
    <!-- Historial de recargas -->
    <div id="recargas-list" style="display:none;margin-top:16px;border-top:1px solid #1a2a1a;padding-top:12px;">
      <div style="font-size:10px;color:#3a9a5a;letter-spacing:3px;text-transform:uppercase;margin-bottom:8px;">Historial de recargas</div>
      <div id="recargas-items"></div>
    </div>
  </div>
</div>

<!-- TX list -->
<div class="section-label">Últimas operaciones — Caja {{ caja_nombre }}</div>
<div class="tx-list" id="txlist"></div>

<div class="caja-total-bar">
  <span style="font-size:10px;color:#555;letter-spacing:3px;text-transform:uppercase;">TOTAL CAJA</span>
  <span class="caja-total-val" id="ctotal">$0</span>
</div>
<div id="cartel-modal" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,0.92);z-index:500;align-items:center;justify-content:center;">
  <div style="background:#111;border:1px solid #3a3000;border-radius:14px;padding:20px;width:340px;max-width:93vw;max-height:90vh;overflow-y:auto;">
    <div style="text-align:center;font-size:56px;line-height:1;margin-bottom:4px;">🍾</div>
    <div style="font-family:'Oswald',sans-serif;font-size:16px;color:#c9a227;letter-spacing:2px;margin-bottom:14px;text-align:center;text-transform:uppercase;">Sacó una botella</div>
    <div style="margin-bottom:9px;">
      <label style="font-size:10px;color:#555;letter-spacing:2px;text-transform:uppercase;display:block;margin-bottom:4px;">Nombre</label>
      <input id="cartel-nombre-caja" type="text" placeholder="Nombre del cliente" style="width:100%;background:#0d0d0d;border:1px solid #2a2a2a;border-radius:7px;color:#f0ece0;padding:9px 12px;font-family:'Rajdhani',sans-serif;font-size:15px;font-weight:600;outline:none;" onfocus="this.style.borderColor='#c9a227'" onblur="this.style.borderColor='#2a2a2a'" />
    </div>
    <div style="margin-bottom:9px;">
      <label style="font-size:10px;color:#555;letter-spacing:2px;text-transform:uppercase;display:block;margin-bottom:4px;">Mesa</label>
      <input id="cartel-mesa-caja" type="text" placeholder="Número de mesa" style="width:100%;background:#0d0d0d;border:1px solid #2a2a2a;border-radius:7px;color:#f0ece0;padding:9px 12px;font-family:'Rajdhani',sans-serif;font-size:15px;outline:none;" onfocus="this.style.borderColor='#c9a227'" onblur="this.style.borderColor='#2a2a2a'" />
    </div>
    <div style="margin-bottom:9px;">
      <label style="font-size:10px;color:#555;letter-spacing:2px;text-transform:uppercase;display:block;margin-bottom:4px;">Mensaje personalizable</label>
      <input id="cartel-frase-caja" type="text" placeholder="Ej: ¡SACÓ UN NUVO CON BENGALAS!" style="width:100%;background:#0d0d0d;border:1px solid #2a2a2a;border-radius:7px;color:#f0ece0;padding:9px 12px;font-family:'Rajdhani',sans-serif;font-size:14px;outline:none;" onfocus="this.style.borderColor='#c9a227'" onblur="this.style.borderColor='#2a2a2a'" />
    </div>
    <!-- Tipo de cartel -->
    <div style="margin-bottom:9px;">
      <label style="font-size:10px;color:#555;letter-spacing:2px;text-transform:uppercase;display:block;margin-bottom:6px;">Tipo</label>
      <div style="display:flex;gap:6px;">
        <button id="ctipo-virtual" onclick="selTipoCartel('virtual')" style="flex:1;background:#0a150a;border:2px solid #3a9a5a;color:#2ecc71;border-radius:7px;padding:8px 4px;font-family:'Rajdhani',sans-serif;font-size:12px;font-weight:700;cursor:pointer;transition:all .15s;">
          📺 Virtual<br><span id="cprecio-virtual" style="font-size:10px;opacity:0.7;"></span>
        </button>
        <button id="ctipo-fisico" onclick="selTipoCartel('fisico')" style="flex:1;background:#0a0a0a;border:2px solid #333;color:#888;border-radius:7px;padding:8px 4px;font-family:'Rajdhani',sans-serif;font-size:12px;font-weight:700;cursor:pointer;transition:all .15s;">
          🖨 Físico<br><span id="cprecio-fisico" style="font-size:10px;opacity:0.7;"></span>
        </button>
        <button id="ctipo-combo" onclick="selTipoCartel('combo')" style="flex:1;background:#0a0a0a;border:2px solid #333;color:#888;border-radius:7px;padding:8px 4px;font-family:'Rajdhani',sans-serif;font-size:12px;font-weight:700;cursor:pointer;transition:all .15s;">
          ⭐ Combo<br><span id="cprecio-combo" style="font-size:10px;opacity:0.7;"></span>
        </button>
      </div>
    </div>
    <!-- Descuento tarjeta -->
    <div id="cartel-descuento-row" style="display:none;margin-bottom:9px;background:#0a0a08;border:1px solid #3a3000;border-radius:7px;padding:8px 12px;">
      <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;color:#c9a227;">
        <input type="checkbox" id="cartel-descontar" checked style="width:15px;height:15px;accent-color:#c9a227;" />
        Descontar <span id="cartel-monto-label" style="font-weight:700;"></span> de la tarjeta
      </label>
    </div>
    <div style="display:flex;gap:8px;">
      <button onclick="emitirCartel()" style="flex:1;background:#c9a227;color:#000;border:none;border-radius:7px;padding:11px;font-family:'Rajdhani',sans-serif;font-size:13px;font-weight:700;letter-spacing:1px;cursor:pointer;">🍾 MOSTRAR</button>
      <button onclick="cerrarCartelModal()" style="background:transparent;border:1px solid #333;color:#555;border-radius:7px;padding:11px 13px;font-family:'Rajdhani',sans-serif;font-size:13px;cursor:pointer;">✕</button>
    </div>
  </div>
</div>
<div id="toast-box"></div>

<script>
const CAJA = {{ caja_num }};
let cajaState = null;
let txData = [];
let tarjetasData = {};
let menuData = [];
let orderT = {};
let variosT = [];
let orderM = {};
let variosM = [];
let activeCatT = 'Todos';
let activeCatM = 'Todos';
let searchTermT = '';
let searchTermM = '';
let variosOpenT = false;
let totalGastadoSesion = 0;
let variosOpenM = false;
let globalBuffer = '';
let lastKeyTime = 0;
let globalTimeout = null;
const READER_SPEED_MS = 80;

function fmt(n){return '$'+Number(n).toLocaleString('es-AR',{minimumFractionDigits:0,maximumFractionDigits:0});}
function fmtLabel(n){
  if(n>=1000000){const m=n/1000000;return(m%1===0?m:m.toFixed(1))+(m<2?' MILLÓN':' MILLONES');}
  if(n>=1000){const k=n/1000;return(k%1===0?k:k.toFixed(1))+(k<2?' MIL':' MILES');}
  return '';
}
function esc(s){const d=document.createElement('div');d.textContent=s||'';return d.innerHTML;}
function now(){return new Date().toLocaleTimeString('es-AR',{hour:'2-digit',minute:'2-digit',hour12:false});}
function showToast(msg,err=false){
  const box=document.getElementById('toast-box');
  const t=document.createElement('div');
  t.className='toast'+(err?' err':'');
  t.textContent=msg;
  box.appendChild(t);
  setTimeout(()=>t.remove(),3000);
}

// Barcode scanner
document.addEventListener('keydown', e => {
  if(['INPUT','TEXTAREA','SELECT'].includes(document.activeElement.tagName)) return;
  const now2 = Date.now();
  lastKeyTime = now2;
  if (e.key === 'Enter') {
    if (globalBuffer.length > 1) { handleScan(globalBuffer.trim()); globalBuffer = ''; return; }
    globalBuffer = '';
    // Enter sin buffer: confirmar pedido activo si hay items
    const modoT = document.getElementById('mtab-tarjeta')?.classList.contains('active');
    const modoM = document.getElementById('mtab-manual')?.classList.contains('active');
    if (modoT) {
      const btn = document.getElementById('btn-cobrar-tarjeta');
      if (btn && !btn.disabled) { e.preventDefault(); confirmarTarjeta(); }
    } else if (modoM) {
      const bar = document.getElementById('confirm-bar-manual');
      if (bar && bar.classList.contains('show')) { e.preventDefault(); confirmarManual(); }
    }
    return;
  }
  if (e.key.length === 1) globalBuffer += e.key;
  clearTimeout(globalTimeout);
  globalTimeout = setTimeout(() => { globalBuffer = ''; lastKeyTime = 0; }, 500);
});

async function handleScan(codigo) {
  try {
    const [r1,r2] = await Promise.all([fetch('/api/tarjetas'),fetch('/api/tarjetas/config')]);
    tarjetasData = await r1.json();
    const conf_list = await r2.json();
    const conf = conf_list.find(c => c.codigo === codigo);
    if (!conf) { showToast('Tarjeta no reconocida: '+codigo, true); return; }
    const mesa = conf.slot != null ? 'Mesa '+conf.slot : '?';
    const saldoInicial = conf.saldo_inicial ? parseFloat(conf.saldo_inicial) : 0;
    const tdEntry = tarjetasData[codigo];
    const saldoActual = tdEntry ? tdEntry.saldo_actual : saldoInicial;
    const nombre = conf.nombre_cliente || (tdEntry ? tdEntry.nombre || '' : '');
    cajaState = {codigo, mesa, slot: conf.slot, saldo_inicial: saldoInicial, saldo_actual: saldoActual, nombre};
    totalGastadoSesion = 0;
    setModo('tarjeta');
    renderTarjeta();
    const menuArea = document.getElementById('menu-area-tarjeta');
    if (menuArea) { menuArea.style.opacity='1'; menuArea.style.pointerEvents='auto'; }
    const ni = document.getElementById('name-tarjeta');
    if (ni) ni.value = nombre || '';
    if (saldoActual <= 0) {
      showToast('Mesa '+conf.slot+' — '+nombre+' — SIN SALDO', true);
    } else {
      showToast('Tarjeta — Mesa '+conf.slot+' — Saldo: '+fmt(saldoActual));
    }
  } catch(e) { showToast('Error leyendo tarjeta', true); }
}

function calcTotalActual() {
  return calcTotal('tarjeta');
}

function renderTarjeta() {
  ['tarjeta-card','tarjeta-card-r'].forEach(id => {
    const el = document.getElementById(id);
    if (!el) return;
    const ta = cajaState;
    if (!ta) { el.className='tarjeta-card'; el.innerHTML=''; return; }
    const pedidoActual = calcTotalActual();
    // saldo_actual ya refleja lo gastado anteriormente (se descuenta al confirmar)
    // solo hay que ver si el PEDIDO ACTUAL excede lo que queda
    const techoReal = Math.max(ta.saldo_inicial, ta.saldo_actual);
    const excede = pedidoActual > 0 && pedidoActual > ta.saldo_actual;
    const sinSaldo = ta.saldo_actual <= 0;
    const saldoTrasBar = Math.max(0, ta.saldo_actual - pedidoActual);
    const pct = techoReal > 0 ? Math.max(0, Math.min(100, Math.round((saldoTrasBar / techoReal) * 100))) : 0;
    const saldoTrasComp = ta.saldo_actual - pedidoActual;
    const pctRest = techoReal > 0 ? ta.saldo_actual / techoReal : 1;
    const warnBajo = !excede && !sinSaldo && pctRest <= 0.3;
    const saldoColor = excede ? '#ff4444' : warnBajo ? '#ffaa00' : 'var(--gold)';
    el.className = 'tarjeta-card visible'+(sinSaldo?' sin-saldo':'');
    el.innerHTML = `<div class="tarjeta-top">
      <div><div class="tarjeta-mesa-label">Mesa</div><div class="tarjeta-mesa-num">${esc(ta.mesa).replace(/^Mesa\s+/i,'')}</div></div>
      <div class="tarjeta-saldo-wrap">
        <div class="tarjeta-saldo-label">Saldo disponible</div>
        <div class="tarjeta-saldo" style="color:${saldoColor};">${fmt(ta.saldo_actual)}</div>
      </div>
    </div>
    ${ta.nombre?`<div style="font-family:'Oswald',sans-serif;font-size:20px;font-weight:700;color:#f0ece0;letter-spacing:2px;margin-bottom:6px;">${esc(ta.nombre.toUpperCase())}</div>`:''}
    <div class="tarjeta-bar-wrap"><div class="tarjeta-bar" style="width:${pct}%;background:${warnBajo?'#886600':excede?'#8b1a1a':'var(--gold)'}"></div></div>
    ${pedidoActual > 0 ? `
    <div style="margin-top:4px;padding:10px 14px;border-radius:8px;background:${excede?'#1a0000':warnBajo?'#1a1200':'#0a1000'};border:1px solid ${excede?'#aa2020':warnBajo?'#886600':'#1a4a1a'};display:flex;justify-content:space-between;align-items:center;">
      <span style="font-size:11px;color:#888;letter-spacing:2px;text-transform:uppercase;">Quedaría</span>
      <span style="font-family:'Oswald',sans-serif;font-size:28px;color:${excede?'#ff3333':warnBajo?'#ffaa00':'#4aaa4a'};font-weight:700;">${fmt(saldoTrasComp)}</span>
    </div>` :
    `<div style="margin-top:8px;padding:10px 14px;border-radius:8px;background:#0a1000;border:1px solid #1a4a1a;display:flex;justify-content:space-between;align-items:center;min-height:42px;">
      <span style="font-size:11px;color:#888;letter-spacing:2px;text-transform:uppercase;">Saldo disponible</span>
      <span style="font-family:'Oswald',sans-serif;font-size:26px;color:${sinSaldo?'#a83030':pctRest<=0.3?'#ffaa00':'#4aaa4a'};font-weight:700;">${fmt(ta.saldo_actual)}</span>
    </div>`}
    ${excede?`<div class="saldo-bajo-warn" style="border-color:#a83030;color:#ff4444;margin-top:4px;">✕ Saldo insuficiente — faltan ${fmt(Math.abs(saldoTrasComp))}</div>`:''}
    ${warnBajo&&!pedidoActual?`<div class="saldo-bajo-warn" style="border-color:#886600;color:#ffaa00;margin-top:4px;">⚠ Saldo bajo — queda ${Math.round(pctRest*100)}%</div>`:''}
    ${sinSaldo?`<div class="saldo-bajo-warn" style="border-color:#a83030;color:#ff4444;">✕ Sin saldo disponible</div>`:''}`;
  });
  updateGastoSesion('tarjeta');
  const scanT = document.getElementById('scan-hint-tarjeta');
  const scanR = document.getElementById('scan-hint-recargar');
  const formR = document.getElementById('form-recargar');
  const btnR = document.getElementById('btn-recargar');
  if (cajaState) {
    if (scanT) { scanT.className='scan-hint activa'; document.getElementById('scan-txt').textContent='● Tarjeta activa — pasá otra para cambiar'; }
    if (scanR) { scanR.className='scan-hint activa'; document.getElementById('rscan-txt').textContent='● Tarjeta activa — pasá otra para cambiar'; }
    if (formR) { formR.style.opacity='1'; formR.style.pointerEvents='auto'; }
    if (btnR) btnR.disabled = false;
  } else {
    if (scanT) { scanT.className='scan-hint esperando'; document.getElementById('scan-txt').textContent='▸ Pasá la tarjeta por el lector para comenzar'; }
    if (scanR) { scanR.className='scan-hint esperando'; document.getElementById('rscan-txt').textContent='▸ Pasá la tarjeta para recargar'; }
    if (formR) { formR.style.opacity='.35'; formR.style.pointerEvents='none'; }
    if (btnR) btnR.disabled = true;
  }
}

// ── Popup mesas (modo sin tarjeta) ──────────────────
function toggleMesasPopupCaja() {
  const popup = document.getElementById('mesas-popup-caja');
  if (!popup) return;
  if (popup.style.display === 'none') {
    popup.style.display = 'block';
    buildMesasPopupCaja();
    setTimeout(() => {
      document.addEventListener('click', function closePop(e) {
        if (!popup.contains(e.target) && e.target.id !== 'btn-mesas-popup') {
          popup.style.display = 'none';
          document.removeEventListener('click', closePop);
        }
      });
    }, 10);
  } else {
    popup.style.display = 'none';
  }
}

async function buildMesasPopupCaja() {
  const list = document.getElementById('mesas-popup-list-caja');
  if (!list) return;
  list.innerHTML = '<div style="padding:16px;color:#555;text-align:center;">Cargando...</div>';
  try {
    const r = await fetch('/api/tarjetas/config');
    const conf_list = await r.json();
    const tarjetas = conf_list.filter(t => t.slot != null);
    tarjetas.sort((a, b) => Number(a.slot) - Number(b.slot));
    if (!tarjetas.length) {
      list.innerHTML = '<div style="padding:20px;color:#444;font-size:15px;text-align:center;">Sin tarjetas configuradas</div>';
      return;
    }
    list.innerHTML = tarjetas.map(t => {
      const nombre = t.nombre_cliente || '—';
      return `<div onclick="seleccionarMesaCaja('${t.slot}')" style="padding:16px 20px;cursor:pointer;border-bottom:1px solid #1a1a1a;transition:background .12s;" onmouseover="this.style.background='#1a1600'" onmouseout="this.style.background='transparent'">
        <div style="font-family:'Oswald',sans-serif;font-size:32px;color:#c9a227;font-weight:700;letter-spacing:2px;line-height:1;">MESA ${esc(String(t.slot))}</div>
        <div style="font-family:'Rajdhani',sans-serif;font-size:16px;color:#e0e0e0;font-weight:600;margin-top:4px;letter-spacing:.5px;">${esc(nombre)}</div>
      </div>`;
    }).join('');
  } catch(e) {
    list.innerHTML = '<div style="padding:16px;color:#555;text-align:center;">Error cargando mesas</div>';
  }
}

function seleccionarMesaCaja(mesa) {
  const input = document.getElementById('mmesa');
  if (input) input.value = mesa;
  const popup = document.getElementById('mesas-popup-caja');
  if (popup) popup.style.display = 'none';
}

function toggleMesasPopupTarjeta() {
  const popup = document.getElementById('mesas-popup-tarjeta');
  if (!popup) return;
  if (popup.style.display === 'none') {
    popup.style.display = 'block';
    buildMesasPopupTarjeta();
    setTimeout(() => {
      document.addEventListener('click', function closePop(e) {
        if (!popup.contains(e.target) && e.target.id !== 'btn-mesas-tarjeta') {
          popup.style.display = 'none';
          document.removeEventListener('click', closePop);
        }
      });
    }, 10);
  } else {
    popup.style.display = 'none';
  }
}

async function buildMesasPopupTarjeta() {
  const list = document.getElementById('mesas-popup-list-tarjeta');
  if (!list) return;
  list.innerHTML = '<div style="padding:16px;color:#555;text-align:center;">Cargando...</div>';
  try {
    const [r1, r2] = await Promise.all([fetch('/api/tarjetas'), fetch('/api/tarjetas/config')]);
    tarjetasData = await r1.json();
    const conf_list = await r2.json();
    const tarjetas = conf_list.filter(t => t.codigo && t.slot != null);
    tarjetas.sort((a, b) => Number(a.slot) - Number(b.slot));
    if (!tarjetas.length) {
      list.innerHTML = '<div style="padding:20px;color:#444;font-size:15px;text-align:center;">Sin tarjetas configuradas</div>';
      return;
    }
    list.innerHTML = tarjetas.map(t => {
      const nombre = t.nombre_cliente || '—';
      const td = tarjetasData[t.codigo] || {};
      const saldo = td.saldo_actual !== undefined ? td.saldo_actual : parseFloat(t.saldo_inicial || 0);
      const sinSaldo = saldo <= 0;
      const clickHandler = `seleccionarMesaTarjeta('${esc(t.codigo)}')`;
      return `<div onclick="${clickHandler}" style="padding:14px 20px;cursor:pointer;border-bottom:1px solid #1a1a1a;transition:background .12s;${sinSaldo?'opacity:.45':''}" onmouseover="this.style.background='#1a1600'" onmouseout="this.style.background='transparent'">
        <div style="display:flex;justify-content:space-between;align-items:center;">
          <div>
            <div style="font-family:'Oswald',sans-serif;font-size:28px;color:${sinSaldo?'#555':'#c9a227'};font-weight:700;letter-spacing:2px;line-height:1;">MESA ${esc(String(t.slot))}</div>
            <div style="font-family:'Rajdhani',sans-serif;font-size:15px;color:${sinSaldo?'#444':'#e0e0e0'};font-weight:600;margin-top:3px;">${esc(nombre)}</div>
          </div>
          <div style="text-align:right;">
            <div style="font-family:'Oswald',sans-serif;font-size:16px;color:${sinSaldo?'#a83030':'#4aaa4a'};font-weight:700;">${fmt(saldo)}</div>
            <div style="font-size:11px;color:#444;letter-spacing:1px;">${sinSaldo?'SIN SALDO':'SALDO'}</div>
          </div>
        </div>
      </div>`;
    }).join('');
  } catch(e) {
    list.innerHTML = '<div style="padding:16px;color:#555;text-align:center;">Error cargando mesas</div>';
  }
}

async function seleccionarMesaTarjeta(codigo) {
  document.getElementById('mesas-popup-tarjeta').style.display = 'none';
  // Simula exactamente lo mismo que handleScan pero sin lector físico
  await handleScan(codigo);
}

function setModo(m) {
  ['tarjeta','manual','recargar'].forEach(x => {
    document.getElementById('mtab-'+x)?.classList.toggle('active', x===m);
    document.getElementById('mc-'+x)?.classList.toggle('active', x===m);
  });
  if (m==='manual') renderMenuCats('manual');
  if (m==='recargar') renderRecargasList();
  updateGastoSesion(m);
}

function updateGastoSesion(modo) {
  const el = document.getElementById('caja-gasto-sesion');
  if (!el) return;
  if (modo === 'tarjeta' || (!modo && document.getElementById('mtab-tarjeta')?.classList.contains('active'))) {
    if (cajaState) {
      const gastado = totalGastadoSesion + calcTotal('tarjeta');
      el.textContent = gastado > 0 ? 'Tarjeta · ' + fmt(gastado) + ' gastado' : 'Tarjeta activa';
      el.style.color = '#c9a227';
    } else {
      el.textContent = 'Esperando tarjeta';
      el.style.color = '#444';
    }
  } else {
    el.textContent = 'Sin tarjeta';
    el.style.color = '#555';
  }
}

async function loadMenu() {
  try {
    const r = await fetch('/api/menu');
    menuData = await r.json();
    renderMenuCats('tarjeta');
    renderMenuCats('manual');
  } catch(e){ menuData=[]; }
}

function getCats() {
  return ['Todos', ...new Set(menuData.map(p=>p.categoria))];
}

function renderMenuCats(side) {
  const el = document.getElementById('cats-'+side);
  if (!el) return;
  const activeCat = side==='tarjeta' ? activeCatT : activeCatM;
  const cats = getCats();
  el.innerHTML = cats.map(c =>
    `<div class="cat-chip${c===activeCat?' active':''}" onclick="selectCat('${side}','${c}')">${c}</div>`
  ).join('') + `<div class="cat-chip varios${(side==='tarjeta'?variosOpenT:variosOpenM)?' active':''}" onclick="toggleVarios('${side}')">+ Varios / Otro</div>`;
  renderMenuProds(side);
}

function selectCat(side, cat) {
  if (side==='tarjeta') activeCatT=cat; else activeCatM=cat;
  renderMenuCats(side);
}

function toggleVarios(side) {
  if (side==='tarjeta') variosOpenT = !variosOpenT; else variosOpenM = !variosOpenM;
  const panel = document.getElementById('varios-panel-'+side);
  if (panel) panel.classList.toggle('show', side==='tarjeta'?variosOpenT:variosOpenM);
  renderMenuCats(side);
}

function renderMenuProds(side) {
  const el = document.getElementById('prods-'+side);
  if (!el) return;
  const activeCat = side==='tarjeta' ? activeCatT : activeCatM;
  const order = side==='tarjeta' ? orderT : orderM;
  const searchTerm = (side==='tarjeta' ? searchTermT : searchTermM).toLowerCase().trim();
  let filtered = searchTerm
    ? menuData.filter(p => p.nombre.toLowerCase().includes(searchTerm) || p.categoria.toLowerCase().includes(searchTerm))
    : menuData.filter(p => activeCat==='Todos' || p.categoria===activeCat);
  if (!filtered.length) { el.innerHTML='<div style="padding:16px;color:#444;font-size:13px;letter-spacing:1px;grid-column:1/-1;">Sin resultados</div>'; return; }
  el.innerHTML = filtered.map(p => {
    const qty = order[p.id]||0;
    let overlay = '';
    if (qty === 1) {
      overlay = `<div onclick="event.stopPropagation();changeQty('${side}',${p.id},-1)" style="position:absolute;top:5px;right:6px;width:22px;height:22px;background:#3a0000;border:1px solid #a83030;border-radius:50%;display:flex;align-items:center;justify-content:center;cursor:pointer;font-size:13px;color:#e74c3c;line-height:1;z-index:10;" title="Eliminar">✕</div>`;
    } else if (qty >= 2) {
      overlay = `<div style="position:absolute;top:5px;right:6px;display:flex;align-items:center;gap:3px;z-index:10;" onclick="event.stopPropagation()">
        <button onclick="changeQty('${side}',${p.id},-1)" style="width:20px;height:20px;background:#1a1a1a;border:1px solid #555;border-radius:4px;color:#fff;font-size:14px;cursor:pointer;display:flex;align-items:center;justify-content:center;line-height:1;font-family:'Oswald',sans-serif;padding:0;">−</button>
        <button onclick="changeQty('${side}',${p.id},1)" style="width:20px;height:20px;background:#1a3a00;border:1px solid #3a6a00;border-radius:4px;color:#8acc50;font-size:14px;cursor:pointer;display:flex;align-items:center;justify-content:center;line-height:1;font-family:'Oswald',sans-serif;padding:0;">+</button>
      </div>`;
    }
    return `<div class="prod-btn${qty>0?' selected':''}" onclick="addToOrder('${side}',${p.id})">
      ${qty>0?`<div class="prod-qty-badge">${qty}</div>`:''}
      ${overlay}
      <div class="prod-cat">${esc(p.categoria)}</div>
      <div class="prod-name">${esc(p.nombre)}</div>
      <div class="prod-price">${fmt(p.precio)}</div>
    </div>`;
  }).join('');
}

function addToOrder(side, id) {
  if (side === 'tarjeta' && cajaState) {
    const p = menuData.find(x => x.id === id);
    if (p) {
      const totalActual = calcTotal('tarjeta');
      const nuevoPrecio = p.precio * ((orderT[id]||0) + 1) - p.precio * (orderT[id]||0);
      if (totalActual + nuevoPrecio > cajaState.saldo_actual) {
        showToast('Saldo insuficiente para agregar '+p.nombre+' — Disponible: '+fmt(cajaState.saldo_actual), true);
        return;
      }
    }
  }
  if (side==='tarjeta') orderT[id]=(orderT[id]||0)+1;
  else orderM[id]=(orderM[id]||0)+1;
  renderMenuProds(side);
  renderOrder(side);
  if (side === 'tarjeta') renderTarjeta();
}

function changeQty(side, id, delta) {
  const order = side==='tarjeta' ? orderT : orderM;
  if (delta > 0 && side === 'tarjeta' && cajaState) {
    const p = menuData.find(x => x.id === id);
    if (p) {
      const totalActual = calcTotal('tarjeta');
      if (totalActual + p.precio > cajaState.saldo_actual) {
        showToast('Saldo insuficiente para agregar '+p.nombre, true);
        return;
      }
    }
  }
  order[id] = Math.max(0,(order[id]||0)+delta);
  if (!order[id]) delete order[id];
  renderMenuProds(side);
  renderOrder(side);
  if (side === 'tarjeta') renderTarjeta();
}

function changeVarQty(side, idx, delta) {
  const arr = side==='tarjeta' ? variosT : variosM;
  arr[idx].qty = Math.max(0,(arr[idx].qty||1)+delta);
  if (arr[idx].qty === 0) arr.splice(idx,1);
  renderOrder(side);
}

function addVariosTarjeta() {
  const monto = parseFloat(document.getElementById('varios-monto-tarjeta').value);
  if (!monto||monto<=0) { showToast('Ingresá un monto', true); return; }
  const desc = document.getElementById('varios-desc-tarjeta').value.trim() || 'Varios';
  variosT.push({desc, monto, qty:1});
  document.getElementById('varios-monto-tarjeta').value='';
  document.getElementById('varios-desc-tarjeta').value='';
  document.getElementById('varios-hint-tarjeta').textContent='';
  renderOrder('tarjeta');
}

function addVariosManual() {
  const monto = parseFloat(document.getElementById('varios-monto-manual').value);
  if (!monto||monto<=0) { showToast('Ingresá un monto', true); return; }
  const desc = document.getElementById('varios-desc-manual').value.trim() || 'Varios';
  variosM.push({desc, monto, qty:1});
  document.getElementById('varios-monto-manual').value='';
  document.getElementById('varios-desc-manual').value='';
  document.getElementById('varios-hint-manual').textContent='';
  renderOrder('manual');
}

function renderOrder(side) {
  const order = side==='tarjeta' ? orderT : orderM;
  const varios = side==='tarjeta' ? variosT : variosM;
  const ids = Object.keys(order).map(Number);
  const panel = document.getElementById('order-panel-'+side);
  const itemsEl = document.getElementById('order-items-'+side);
  const totalEl = document.getElementById('order-total-'+side);
  const confirmBar = document.getElementById('confirm-bar-'+side);
  const hasItems = ids.length > 0 || varios.length > 0;
  if (panel) panel.classList.toggle('show', hasItems);
  if (confirmBar) confirmBar.classList.toggle('show', hasItems);
  if (!hasItems) return;
  let total = 0;
  const menuRows = ids.map(id => {
    const p = menuData.find(x=>x.id===id);
    if (!p) return '';
    const sub = p.precio * order[id];
    total += sub;
    return `<div class="order-item">
      <div class="oi-name">${esc(p.nombre)}</div>
      <div style="display:flex;align-items:center;gap:5px;">
        <button class="qty-btn" onclick="changeQty('${side}',${id},-1)">−</button>
        <span class="qty-val">${order[id]}</span>
        <button class="qty-btn" onclick="changeQty('${side}',${id},1)">+</button>
      </div>
      <span class="oi-price">${fmt(sub)}</span>
    </div>`;
  }).join('');
  const variosRows = varios.map((v,i) => {
    const sub = v.monto * (v.qty||1);
    total += sub;
    return `<div class="order-item">
      <div class="oi-name" style="color:#8aaa7a;">${esc(v.desc)}</div>
      <div style="display:flex;align-items:center;gap:5px;">
        <button class="qty-btn" onclick="changeVarQty('${side}',${i},-1)">−</button>
        <span class="qty-val">${v.qty||1}</span>
        <button class="qty-btn" onclick="changeVarQty('${side}',${i},1)">+</button>
      </div>
      <span class="oi-price">${fmt(sub)}</span>
    </div>`;
  }).join('');
  if (itemsEl) itemsEl.innerHTML = menuRows + variosRows;
  if (totalEl) totalEl.textContent = fmt(total);
  if (side==='tarjeta') {
    const btn = document.getElementById('btn-cobrar-tarjeta');
    if (btn) btn.disabled = !hasItems || !cajaState;
    // Actualizar total en confirm-bar
    const cbTotal = document.getElementById('cb-total-tarjeta');
    if (cbTotal) cbTotal.textContent = fmt(total);
    // Actualizar tag mesa
    const cbMesa = document.getElementById('cb-mesa-tarjeta');
    if (cbMesa && cajaState) cbMesa.textContent = '▤ ' + cajaState.mesa;
    renderTarjeta();
  } else {
    const cbTotal = document.getElementById('cb-total-manual');
    if (cbTotal) cbTotal.textContent = fmt(total);
  }
}

function limpiarPedido(side) {
  if (side==='tarjeta') {
    orderT={}; variosT=[]; searchTermT='';
    const s=document.getElementById('search-tarjeta'); if(s) s.value='';
    // Ocultar confirm-bar y limpiar barra inferior
    const cb = document.getElementById('confirm-bar-tarjeta');
    if (cb) cb.classList.remove('show');
    const op = document.getElementById('order-panel-tarjeta');
    if (op) op.classList.remove('show');
    const cbTotal = document.getElementById('cb-total-tarjeta');
    if (cbTotal) cbTotal.textContent = '$0';
    renderTarjeta();
  } else {
    orderM={}; variosM=[]; searchTermM='';
    const s=document.getElementById('search-manual'); if(s) s.value='';
    const cb = document.getElementById('confirm-bar-manual');
    if (cb) cb.classList.remove('show');
    const op = document.getElementById('order-panel-manual');
    if (op) op.classList.remove('show');
    const cbTotal = document.getElementById('cb-total-manual');
    if (cbTotal) cbTotal.textContent = '$0';
  }
  renderMenuProds(side);
}

function buildItems(side) {
  const order = side==='tarjeta' ? orderT : orderM;
  const varios = side==='tarjeta' ? variosT : variosM;
  const items = [];
  Object.keys(order).forEach(id => {
    const p = menuData.find(x=>x.id===parseInt(id));
    if (p) items.push({id:p.id, nombre:p.nombre, precio:p.precio, qty:order[id]});
  });
  varios.forEach(v => items.push({id:null, nombre:v.desc, precio:v.monto, qty:v.qty||1}));
  return items;
}

function calcTotal(side) {
  const order = side==='tarjeta' ? orderT : orderM;
  const varios = side==='tarjeta' ? variosT : variosM;
  let t = 0;
  Object.keys(order).forEach(id => { const p=menuData.find(x=>x.id===parseInt(id)); if(p) t+=p.precio*order[id]; });
  varios.forEach(v => t += v.monto*(v.qty||1));
  return t;
}

function showToastCobro(msg) {
  const box = document.getElementById('toast-box');
  const t = document.createElement('div');
  t.style.cssText = 'background:#071a07;border:2px solid #2ecc71;border-radius:10px;padding:14px 22px;font-family:Oswald,sans-serif;font-size:18px;font-weight:700;color:#2ecc71;letter-spacing:1px;animation:fadeIn .2s;';
  t.textContent = msg;
  box.appendChild(t);
  setTimeout(() => t.remove(), 3500);
}

async function confirmarTarjeta() {
  const ta = cajaState;
  if (!ta) { showToast('Pasá una tarjeta primero', true); return; }
  const items = buildItems('tarjeta');
  if (!items.length) { showToast('El pedido está vacío', true); return; }
  const total = calcTotal('tarjeta');
  if (total > ta.saldo_actual) { showToast('Saldo insuficiente. Disponible: '+fmt(ta.saldo_actual), true); return; }
  const name = ta.nombre || 'Cliente';
  // Deshabilitar botón mientras se procesa
  const btnCobrar = document.getElementById('btn-cobrar-tarjeta');
  if (btnCobrar) btnCobrar.disabled = true;
  try {
    const res = await fetch('/api/tx',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({
      name, amount:total, caja:CAJA, mesa:ta.mesa, tarjeta_codigo:ta.codigo, client_time:now(), items
    })});
    let d = {};
    try { d = await res.json(); } catch(_) {}
    if (res.ok && d.ok !== false) {
      // Éxito: limpiar pedido y actualizar saldo
      ta.saldo_actual -= total;
      ta.nombre = name;
      totalGastadoSesion += total;
      orderT = {}; variosT = [];
      // Ocultar confirm-bar explícitamente
      const cb = document.getElementById('confirm-bar-tarjeta');
      if (cb) cb.classList.remove('show');
      const panel = document.getElementById('order-panel-tarjeta');
      if (panel) panel.classList.remove('show');
      renderTarjeta();
      renderMenuProds('tarjeta');
      renderOrder('tarjeta');
      await loadData();
      showToastCobro(fmt(total)+' cobrado ✓' + (ta.saldo_actual > 0 ? ' — Saldo: '+fmt(ta.saldo_actual) : ' — Sin saldo restante'));
    } else {
      showToast(d.error || 'Error al registrar — intentá de nuevo', true);
      if (btnCobrar) btnCobrar.disabled = false;
    }
  } catch(e) {
    showToast('Error al enviar — verificá la conexión', true);
    if (btnCobrar) btnCobrar.disabled = false;
  }
}

async function confirmarManual() {
  const name = document.getElementById('mname').value.trim();
  const mesa = document.getElementById('mmesa').value.trim();
  if (!name) { document.getElementById('mname').focus(); showToast('Falta el nombre', true); return; }
  const items = buildItems('manual');
  if (!items.length) { showToast('El pedido está vacío', true); return; }
  const total = calcTotal('manual');
  try {
    const res = await fetch('/api/tx',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({
      name, amount:total, caja:CAJA, mesa, tarjeta_codigo:'', client_time:now(), items
    })});
    if (res.ok) {
      showToastCobro('Registrado ✓ — '+fmt(total)+' para '+name);
      orderM={}; variosM=[];
      renderMenuProds('manual');
      renderOrder('manual');
      await loadData();
    } else { const err=await res.json().catch(()=>({})); showToast(err.error||'Error', true); }
  } catch(e){ showToast('Error de conexión', true); }
}

async function recargarTarjeta() {
  const ta = cajaState;
  if (!ta) { showToast('Pasá una tarjeta primero', true); return; }
  const monto = parseFloat(document.getElementById('recarga').value);
  if (!monto||monto<=0) { document.getElementById('recarga').focus(); showToast('Falta el monto', true); return; }
  try {
    const r = await fetch('/api/tarjetas/recargar',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({codigo:ta.codigo,monto})});
    const d = await r.json();
    if (d.ok) {
      cajaState.saldo_actual = d.nuevo_saldo;
      totalGastadoSesion = 0; // reset al recargar (saldo fresco)
      document.getElementById('recarga').value='';
      showToast('Recargado — Nuevo saldo: '+fmt(d.nuevo_saldo));
      renderTarjeta();
      await loadData();
      renderRecargasList();
    } else { showToast(d.error||'Error', true); }
  } catch(e){ showToast('Error de conexión', true); }
}

async function renderRecargasList() {
  const ta = cajaState;
  const listWrap = document.getElementById('recargas-list');
  const listEl = document.getElementById('recargas-items');
  if (!listWrap || !listEl || !ta) return;
  try {
    const r = await fetch('/api/tarjetas');
    const data = await r.json();
    const td = data[ta.codigo];
    const recargas = td ? (td.recargas || []) : [];
    if (!recargas.length) { listWrap.style.display='none'; return; }
    listWrap.style.display='block';
    listEl.innerHTML = recargas.slice().reverse().map(rec => `
      <div style="display:flex;align-items:center;gap:8px;padding:8px 0;border-bottom:1px solid #0f1a0f;">
        <span style="font-size:11px;color:#555;min-width:60px;">${rec.fecha||''} ${rec.hora||''}</span>
        <span style="flex:1;font-family:'Oswald',sans-serif;font-size:18px;color:#3a9a5a;font-weight:700;">${fmt(rec.monto)}</span>
        <button onclick="editarRecargaCaja('${ta.codigo}','${rec.id}',${rec.monto})" style="background:transparent;border:1px solid #2a3a2a;border-radius:5px;padding:4px 10px;color:#5a8a5a;font-size:12px;cursor:pointer;letter-spacing:1px;">✎ Editar</button>
        <button onclick="eliminarRecargaCaja('${ta.codigo}','${rec.id}')" style="background:transparent;border:1px solid #3a1a1a;border-radius:5px;padding:4px 10px;color:#883333;font-size:12px;cursor:pointer;letter-spacing:1px;">✕ Quitar</button>
      </div>`).join('');
  } catch(e) {}
}

async function eliminarRecargaCaja(codigo, recargaId) {
  if (!confirm('¿Eliminar esta recarga?')) return;
  try {
    const r = await fetch(`/api/tarjetas/recarga/${codigo}/${recargaId}`, {method:'DELETE'});
    const d = await r.json();
    if (d.ok) {
      if (cajaState && cajaState.codigo === codigo) {
        cajaState.saldo_actual = d.nuevo_saldo;
        totalGastadoSesion = 0;
        renderTarjeta();
      }
      showToast('Recarga eliminada — nuevo saldo: '+fmt(d.nuevo_saldo));
      await loadData();
      renderRecargasList();
    } else showToast(d.error||'Error', true);
  } catch(e) { showToast('Error al eliminar', true); }
}

async function editarRecargaCaja(codigo, recargaId, montoActual) {
  const nuevoStr = prompt('Editar monto de recarga (actual: '+fmt(montoActual)+'):', montoActual);
  if (nuevoStr === null) return;
  const nuevo = parseFloat(nuevoStr);
  if (!nuevo || nuevo <= 0) { showToast('Monto inválido', true); return; }
  try {
    const r = await fetch(`/api/tarjetas/recarga/${codigo}/${recargaId}`, {
      method:'PUT', headers:{'Content-Type':'application/json'}, body:JSON.stringify({monto:nuevo})
    });
    const d = await r.json();
    if (d.ok) {
      if (cajaState && cajaState.codigo === codigo) {
        cajaState.saldo_actual = d.nuevo_saldo;
        totalGastadoSesion = 0;
        renderTarjeta();
      }
      showToast('Recarga actualizada — nuevo saldo: '+fmt(d.nuevo_saldo));
      await loadData();
      renderRecargasList();
    } else showToast(d.error||'Error', true);
  } catch(e) { showToast('Error al editar', true); }
}

async function loadData() {
  try {
    const [r1,r2] = await Promise.all([fetch('/api/tx'),fetch('/api/tarjetas')]);
    txData = await r1.json();
    tarjetasData = await r2.json();
    renderList();
  } catch(e){}
}

function renderList() {
  const myTxs = txData.filter(t => t.caja == CAJA).reverse();
  const list = document.getElementById('txlist');
  if (!list) return;
  if (!myTxs.length) { list.innerHTML='<div style="padding:16px;color:#444;font-size:13px;letter-spacing:1px;">Sin operaciones esta noche</div>'; return; }
  list.innerHTML = myTxs.map(t => {
    const itemsStr = t.items ? t.items.map(i=>`${i.qty}x ${esc(i.nombre)}`).join(', ') : '';
    return `<div class="tx-item">
      <div>
        <div class="tx-name">${esc(t.name)}</div>
        <div class="tx-meta">${esc(t.mesa||'')}${t.mesa&&t.time?' · ':''}${esc(t.time||'')}${itemsStr?' · '+itemsStr:''}</div>
      </div>
      <div style="display:flex;align-items:center;gap:10px;">
        <span class="tx-amount">${fmt(t.amount)}</span>
        <span class="tx-del" onclick="deleteTx(${t.id})">✕</span>
      </div>
    </div>`;
  }).join('');
  const total = txData.filter(t => t.caja == CAJA).reduce((s,t)=>s+t.amount, 0);
  const el = document.getElementById('ctotal');
  if (el) el.textContent = fmt(total);
}

async function cerrarGanadorDesdeCaja() {
  try {
    const st = await fetch('/api/state').then(r=>r.json()).catch(()=>({}));
    if (!st.winner_show) { showToast('El ganador todavía no apareció en pantalla', false); return; }
    await fetch('/api/winner/hide', {method:'POST'});
    showToast('Mensaje de ganador cerrado ✓');
  } catch(e) { showToast('Error de conexión', true); }
}

async function deleteTx(id) {
  await fetch('/api/tx/'+id,{method:'DELETE'});
  await loadData();
  // Si hay una tarjeta activa, actualizar su saldo desde los datos frescos
  if (cajaState && cajaState.codigo) {
    const td = tarjetasData[cajaState.codigo];
    if (td && td.saldo_actual !== undefined) {
      cajaState.saldo_actual = td.saldo_actual;
      renderTarjeta();
    }
  }
}

['varios-monto-tarjeta','varios-monto-manual'].forEach(id => {
  const el = document.getElementById(id);
  const hintId = id.includes('tarjeta') ? 'varios-hint-tarjeta' : 'varios-hint-manual';
  if (!el) return;
  el.addEventListener('input', () => {
    const v = parseFloat(el.value);
    const hint = document.getElementById(hintId);
    if (!hint) return;
    if (v>0) { const lbl=fmtLabel(v); hint.textContent='= '+fmt(v)+(lbl?' · '+lbl:''); hint.className='hint-miles ok'; }
    else { hint.textContent=''; hint.className='hint-miles'; }
  });
  el.addEventListener('keydown', e => { if(e.key==='Enter'){e.preventDefault(); id.includes('tarjeta')?addVariosTarjeta():addVariosManual();} });
});

document.getElementById('recarga')?.addEventListener('input', function(){
  const v = parseFloat(this.value);
  const hint = document.getElementById('rhint');
  if (!hint) return;
  if (v>0) { const lbl=fmtLabel(v); hint.textContent='= '+fmt(v)+(lbl?' · '+lbl:''); hint.className='hint-miles ok'; }
  else { hint.textContent='Ingresá el monto a agregar'; hint.className='hint-miles'; }
});
document.getElementById('recarga')?.addEventListener('keydown', e=>{ if(e.key==='Enter'){e.preventDefault();recargarTarjeta();} });

let cartelPrecios = {virtual:0, fisico:0, combo:0};
let cartelTipoSel = 'virtual';

async function cargarCartelPreciosCaja() {
  try {
    const r = await fetch('/api/cartel/precios');
    cartelPrecios = await r.json();
    const fmt = n => n>0 ? '$'+Number(n).toLocaleString('es-AR') : 'Sin cargo';
    const pv=document.getElementById('cprecio-virtual'); if(pv) pv.textContent=fmt(cartelPrecios.virtual||0);
    const pf=document.getElementById('cprecio-fisico');  if(pf) pf.textContent=fmt(cartelPrecios.fisico||0);
    const pc=document.getElementById('cprecio-combo');   if(pc) pc.textContent=fmt(cartelPrecios.combo||0);
  } catch(e) {}
}

function selTipoCartel(tipo) {
  cartelTipoSel = tipo;
  ['virtual','fisico','combo'].forEach(t => {
    const btn = document.getElementById('ctipo-'+t);
    if (!btn) return;
    if (t === tipo) {
      btn.style.background = t==='virtual'?'#0a150a':t==='fisico'?'#151515':'#15100a';
      btn.style.borderColor = t==='virtual'?'#2ecc71':t==='fisico'?'#c9a227':'#e8c84a';
      btn.style.color = t==='virtual'?'#2ecc71':t==='fisico'?'#c9a227':'#e8c84a';
    } else {
      btn.style.background = '#0a0a0a'; btn.style.borderColor = '#333'; btn.style.color = '#555';
    }
  });
  const precio = cartelPrecios[tipo] || 0;
  const descRow = document.getElementById('cartel-descuento-row');
  const montoLbl = document.getElementById('cartel-monto-label');
  if (descRow && cajaState) {
    if (precio > 0) {
      descRow.style.display = 'block';
      if (montoLbl) montoLbl.textContent = '$'+Number(precio).toLocaleString('es-AR');
    } else {
      descRow.style.display = 'none';
    }
  }
}

function abrirCartelModal() {
  const nombre = cajaState ? (cajaState.nombre || '') : '';
  const mesa = cajaState ? String(cajaState.slot || '') : '';
  const ni = document.getElementById('cartel-nombre-caja');
  const mi = document.getElementById('cartel-mesa-caja');
  const fi = document.getElementById('cartel-frase-caja');
  if (ni) ni.value = nombre;
  if (mi) mi.value = mesa;
  if (fi) fi.value = '';
  cargarCartelPreciosCaja().then(() => selTipoCartel('virtual'));
  document.getElementById('cartel-modal').style.display = 'flex';
}
function cerrarCartelModal() {
  document.getElementById('cartel-modal').style.display = 'none';
}
async function emitirCartel() {
  const nombre = (document.getElementById('cartel-nombre-caja')?.value || '').trim().toUpperCase();
  const mesa = (document.getElementById('cartel-mesa-caja')?.value || '').trim();
  const frase = (document.getElementById('cartel-frase-caja')?.value || '').trim().toUpperCase();
  if (!frase) { showToast('Escribí un mensaje para el cartel', true); return; }
  const tipo = cartelTipoSel;
  const precio = cartelPrecios[tipo] || 0;
  const descontar = document.getElementById('cartel-descontar')?.checked;
  cerrarCartelModal();
  // Descontar de tarjeta si corresponde
  if (precio > 0 && descontar && cajaState) {
    if (cajaState.saldo_actual < precio) {
      showToast('Saldo insuficiente para descontar el cartel ($'+Number(precio).toLocaleString('es-AR')+')', true); return;
    }
    try {
      const r = await fetch('/api/tx', {method:'POST', headers:{'Content-Type':'application/json'},
        body: JSON.stringify({
          name: nombre || cajaState.nombre || 'Cartel',
          amount: precio, caja: CAJA,
          mesa: mesa || String(cajaState.slot||''),
          tarjeta_codigo: cajaState.codigo,
          items: [{nombre:'Cartel '+tipo, cantidad:1, precio}]
        })});
      const d = await r.json();
      if (!d.ok) { showToast(d.error||'Error descontando', true); return; }
      cajaState.saldo_actual -= precio;
      renderTarjeta(); await loadData();
    } catch(e) { showToast('Error descontando de tarjeta', true); return; }
  }
  // Mostrar en pantalla (virtual o combo)
  if (tipo === 'virtual' || tipo === 'combo') {
    try {
      await fetch('/api/cartel/show', {method:'POST', headers:{'Content-Type':'application/json'},
        body: JSON.stringify({nombre, mesa, frase, emoji: '🍾'})});
      showToast('Cartel enviado a pantalla');
    } catch(e) { showToast('Error enviando cartel', true); }
  } else {
    showToast('🖨 Cartel físico — Mesa '+(mesa||cajaState?.slot||'—'));
  }
}

async function cerrarCartelPantalla() {
  try {
    await fetch('/api/cartel/hide', {method:'POST'});
    showToast('Cartel cerrado en pantalla');
  } catch(e) { showToast('Error cerrando cartel', true); }
}

loadMenu();
loadData();
setInterval(loadData, 3000);
</script>
</body>
</html>"""


HISTORIAL_HTML = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Historial — Jagger VIP</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
@import url('https://fonts.googleapis.com/css2?family=Oswald:wght@400;600;700&family=Rajdhani:wght@400;600&display=swap');
*{box-sizing:border-box;margin:0;padding:0;}
body{background:#080808;color:#f0ece0;font-family:'Rajdhani',sans-serif;min-height:100vh;}
.top-bar{background:#0a0a0a;border-bottom:1px solid #1a1a1a;padding:14px 20px;display:flex;align-items:center;justify-content:space-between;}
.logo{font-family:'Oswald',sans-serif;font-size:22px;color:#c9a227;letter-spacing:3px;}
.back-btn{color:#555;text-decoration:none;font-size:13px;letter-spacing:2px;border:1px solid #222;border-radius:6px;padding:7px 14px;transition:all .15s;}
.back-btn:hover{color:#c9a227;border-color:#c9a227;}
.nav{display:flex;gap:0;background:#0a0a0a;border-bottom:1px solid #1a1a1a;overflow-x:auto;}
.nav-btn{padding:12px 22px;cursor:pointer;font-family:'Rajdhani',sans-serif;font-size:13px;font-weight:600;letter-spacing:2px;text-transform:uppercase;color:#555;background:none;border:none;border-bottom:2px solid transparent;transition:all .2s;white-space:nowrap;}
.nav-btn.active{color:#c9a227;border-bottom-color:#c9a227;}
.page{display:none;padding:20px;max-width:1100px;margin:0 auto;}
.page.active{display:block;}
.kpi-row{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:24px;}
.kpi{flex:1;min-width:140px;background:#111;border:1px solid #1a1a1a;border-radius:10px;padding:16px 18px;}
.kpi-lbl{font-size:9px;letter-spacing:2px;color:#999;text-transform:uppercase;margin-bottom:6px;}
.kpi-val{font-family:'Oswald',sans-serif;font-size:26px;font-weight:700;color:#c9a227;}
.section-title{font-family:'Oswald',sans-serif;font-size:15px;color:#c9a227;letter-spacing:2px;margin:24px 0 12px;text-transform:uppercase;}
table{width:100%;border-collapse:collapse;font-size:14px;}
th{background:#0d0d0d;color:#aaa;font-size:10px;letter-spacing:2px;text-transform:uppercase;padding:10px 12px;text-align:left;border-bottom:1px solid #222;}
td{padding:10px 12px;border-bottom:1px solid #181818;color:#e0ddd0;}
tr:hover td{background:#111;}
.gold{color:#c9a227;font-weight:700;}
.rank-badge{display:inline-block;width:24px;height:24px;border-radius:50%;text-align:center;line-height:24px;font-size:11px;font-weight:700;}
.r1{background:#c9a227;color:#000;}
.r2{background:#888;color:#000;}
.r3{background:#6a3a00;color:#e8c84a;}
.noche-row{cursor:pointer;}
.noche-row:hover td{background:#111;}
.detail-panel{display:none;background:#0d0d0d;border:1px solid #222;border-radius:8px;padding:16px;margin:4px 0 16px;}
.detail-panel.open{display:block;}
.trim-tabs{display:flex;gap:8px;margin-bottom:16px;flex-wrap:wrap;}
.trim-tab{padding:7px 16px;border:1px solid #2a2a2a;border-radius:20px;cursor:pointer;font-size:13px;letter-spacing:1px;color:#aaa;transition:all .15s;}
.trim-tab.active{border-color:#c9a227;color:#c9a227;background:#0d0b00;}
.empty{color:#444;text-align:center;padding:60px 20px;font-size:14px;letter-spacing:2px;}
.dl-btn{display:inline-block;background:transparent;color:#c9a227;border:1px solid #c9a227;border-radius:7px;padding:8px 18px;font-family:'Rajdhani',sans-serif;font-size:13px;font-weight:700;letter-spacing:1px;cursor:pointer;text-decoration:none;transition:all .15s;}
.dl-btn:hover{background:#c9a227;color:#000;}
.charts-grid{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin:16px 0 24px;}
@media(max-width:640px){.charts-grid{grid-template-columns:1fr;}}
.chart-card{background:#0d0d0d;border:1px solid #1a1a1a;border-radius:10px;padding:16px;}
.chart-card-title{font-size:10px;letter-spacing:2px;color:#999;text-transform:uppercase;margin-bottom:12px;}
.chart-card.wide{grid-column:1/-1;}
.hora-row{display:flex;align-items:center;gap:10px;padding:5px 0;border-bottom:1px solid #181818;}
.hora-lbl{width:36px;font-size:12px;color:#aaa;flex-shrink:0;}
.hora-bar-wrap{flex:1;background:#111;border-radius:3px;height:8px;overflow:hidden;}
.hora-bar{height:8px;background:#c9a227;border-radius:3px;transition:width .3s;}
.hora-val{width:80px;text-align:right;font-size:12px;color:#e0ddd0;flex-shrink:0;}
</style>
</head>
<body>
<div class="top-bar">
  <span class="logo">RANKING VIP — HISTORIAL</span>
  <div style="display:flex;gap:10px;align-items:center;">
    <a class="dl-btn" href="/api/export/excel" target="_blank">⬇ Descargar Excel</a>
    <a class="back-btn" href="/">← Volver</a>
  </div>
</div>
<div class="nav">
  <button class="nav-btn active" onclick="showPage('general',this)">General</button>
  <button class="nav-btn" onclick="showPage('noches',this)">Noches</button>
  <button class="nav-btn" onclick="showPage('ranking',this)">Ranking</button>
</div>

<!-- GENERAL -->
<div id="page-general" class="page active">
  <div id="kpis" class="kpi-row"></div>
  <div class="charts-grid">
    <div class="chart-card wide">
      <div class="chart-card-title">Evolución noche a noche</div>
      <canvas id="chart-trend"></canvas>
    </div>
    <div class="chart-card">
      <div class="chart-card-title">Facturación por mes</div>
      <canvas id="chart-meses"></canvas>
    </div>
    <div class="chart-card">
      <div class="chart-card-title">Distribución por caja (acumulado)</div>
      <canvas id="chart-cajas-total"></canvas>
    </div>
  </div>
  <div class="section-title">Totales por mes</div>
  <table id="tbl-meses"><thead><tr><th>Mes</th><th>Noches</th><th>Total</th><th>Ops</th></tr></thead><tbody></tbody></table>
</div>

<!-- NOCHES -->
<div id="page-noches" class="page">
  <div class="section-title">Todas las noches</div>
  <table><thead><tr><th>Fecha</th><th>Cierre</th><th>Total</th><th>Ops</th><th>1°</th><th>2°</th><th>3°</th></tr></thead>
  <tbody id="tbl-noches"></tbody></table>
</div>

<!-- RANKING -->
<div id="page-ranking" class="page">
  <div class="section-title">Período</div>
  <div class="trim-tabs" id="trim-tabs"></div>
  <table><thead><tr><th>#</th><th>Nombre</th><th>Noches</th><th>Total</th><th>Promedio / noche</th></tr></thead>
  <tbody id="tbl-ranking"></tbody></table>
</div>

<script>
let historial = [];
let periodoActivo = 'all';
let chartTrend=null, chartMeses=null, chartCajas=null;

const GOLD = '#c9a227', GOLD2 = '#e8c84a', DARK = '#0d0d0d';
const CAJA_COLORS = ['#c9a227','#3a9a5a','#3a6ac9'];
Chart.defaults.color = '#999';
Chart.defaults.borderColor = '#222';

async function init() {
  const r = await fetch('/api/historial');
  historial = await r.json();
  renderGeneral();
  renderNoches();
  renderTrimTabs();
  renderRanking('all');
}

function fmt(n) { return '$'+Number(n).toLocaleString('es-AR',{minimumFractionDigits:0,maximumFractionDigits:0}); }
function esc(s){ return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;'); }

function showPage(id, btn) {
  document.querySelectorAll('.page').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.nav-btn').forEach(b=>b.classList.remove('active'));
  document.getElementById('page-'+id).classList.add('active');
  if(btn) btn.classList.add('active');
}

function renderGeneral() {
  const total = historial.reduce((s,n)=>s+n.total,0);
  const ops   = historial.reduce((s,n)=>s+n.operaciones,0);
  const rk = {}; historial.forEach(n=>(n.ranking||[]).forEach(r=>{ rk[r.name]=(rk[r.name]||0)+r.total; }));
  const best = Object.entries(rk).sort((a,b)=>b[1]-a[1])[0];
  const mejorNoche = historial.length ? historial.reduce((mx,n)=>n.total>mx.total?n:mx) : null;
  document.getElementById('kpis').innerHTML = [
    ['Total acumulado', fmt(total)],
    ['Noches', historial.length],
    ['Operaciones', ops],
    ['Cliente top', best ? best[0] : '—'],
    ['Mejor noche', mejorNoche ? fmt(mejorNoche.total) : '—'],
  ].map(([l,v])=>`<div class="kpi"><div class="kpi-lbl">${l}</div><div class="kpi-val" style="font-size:${String(v).length>10?'16px':'26px'}">${v}</div></div>`).join('');

  // Tabla mensual
  const por_mes = {};
  historial.forEach(n=>{ const m=n.fecha.slice(0,7); if(!por_mes[m]) por_mes[m]={total:0,ops:0,noches:0}; por_mes[m].total+=n.total; por_mes[m].ops+=n.operaciones; por_mes[m].noches++; });
  document.querySelector('#tbl-meses tbody').innerHTML = Object.entries(por_mes).sort((a,b)=>b[0].localeCompare(a[0]))
    .map(([m,v])=>`<tr><td class="gold">${m}</td><td>${v.noches}</td><td class="gold">${fmt(v.total)}</td><td>${v.ops}</td></tr>`).join('') || '<tr><td colspan="4" class="empty">Sin datos</td></tr>';

  // Gráfico tendencia noche a noche (line)
  const nochesSorted = [...historial].sort((a,b)=>a.fecha.localeCompare(b.fecha));
  const trendLabels = nochesSorted.map(n=>n.fecha.slice(5));
  const trendData   = nochesSorted.map(n=>n.total);
  if(chartTrend) chartTrend.destroy();
  const fmtTick = v => '$'+Number(v).toLocaleString('es-AR');
  const tooltipFmt = { callbacks:{ label: ctx => ' '+fmt(ctx.raw) } };
  const xStyle = { color:'#bbb', font:{size:11} };
  const yStyle = { color:'#bbb', font:{size:11}, callback: fmtTick };
  const axisTitle = txt => ({ display:true, text:txt, color:'#aaa', font:{size:11} });

  chartTrend = new Chart(document.getElementById('chart-trend'),{
    type:'line',
    data:{labels:trendLabels,datasets:[{label:'Total por noche',data:trendData,borderColor:GOLD,backgroundColor:'rgba(201,162,39,0.10)',tension:0.35,pointBackgroundColor:GOLD,pointRadius:5,pointHoverRadius:7,fill:true}]},
    options:{
      plugins:{
        legend:{display:true,labels:{color:'#e0ddd0',font:{size:12},boxWidth:14}},
        title:{display:true,text:'Evolución del total por noche',color:'#e0ddd0',font:{size:13},padding:{bottom:10}},
        tooltip:{callbacks:{label: ctx=>' '+fmt(ctx.raw)}}
      },
      scales:{
        x:{title:axisTitle('Fecha'),ticks:xStyle},
        y:{title:axisTitle('Total recaudado ($)'),ticks:yStyle}
      },
      maintainAspectRatio:true,aspectRatio:3.2
    }
  });

  // Gráfico facturación por mes (bar)
  const mesesSorted = Object.entries(por_mes).sort((a,b)=>a[0].localeCompare(b[0]));
  if(chartMeses) chartMeses.destroy();
  chartMeses = new Chart(document.getElementById('chart-meses'),{
    type:'bar',
    data:{labels:mesesSorted.map(([m])=>m),datasets:[{label:'Total mensual',data:mesesSorted.map(([,v])=>v.total),backgroundColor:GOLD+'aa',borderColor:GOLD,borderWidth:1,borderRadius:4}]},
    options:{
      plugins:{
        legend:{display:true,labels:{color:'#e0ddd0',font:{size:12},boxWidth:14}},
        title:{display:true,text:'Facturación total por mes',color:'#e0ddd0',font:{size:13},padding:{bottom:10}},
        tooltip:tooltipFmt
      },
      scales:{
        x:{title:axisTitle('Mes'),ticks:xStyle},
        y:{title:axisTitle('Total ($)'),ticks:yStyle}
      },
      maintainAspectRatio:true,aspectRatio:2
    }
  });

  // Gráfico cajas total (doughnut)
  const cajasTotals = [1,2,3].map(c=>historial.reduce((s,n)=>s+(n.por_caja?.[c]||n.por_caja?.[String(c)]||0),0));
  if(chartCajas) chartCajas.destroy();
  chartCajas = new Chart(document.getElementById('chart-cajas-total'),{
    type:'doughnut',
    data:{labels:['Abajo','Extendido','VIP'],datasets:[{data:cajasTotals,backgroundColor:CAJA_COLORS,borderColor:'#080808',borderWidth:3}]},
    options:{
      plugins:{
        legend:{position:'bottom',labels:{color:'#e0ddd0',font:{size:13},padding:16,boxWidth:14}},
        title:{display:true,text:'Distribución por caja (acumulado)',color:'#e0ddd0',font:{size:13},padding:{bottom:10}},
        tooltip:{callbacks:{label: ctx=>' '+ctx.label+': '+fmt(ctx.raw)+' ('+Math.round(ctx.parsed/cajasTotals.reduce((a,b)=>a+b,0)*100)+'%)'}}
      },
      maintainAspectRatio:true,aspectRatio:1.5
    }
  });
}

function horasHTML(txs) {
  if (!txs || !txs.length) return '';
  const byHour = {};
  txs.forEach(t => {
    const h = (t.time||'00:00').split(':')[0].padStart(2,'0');
    byHour[h] = (byHour[h]||0) + t.amount;
  });
  const horas = Object.entries(byHour).sort((a,b)=>a[0].localeCompare(b[0]));
  if (!horas.length) return '';
  const maxVal = Math.max(...horas.map(([,v])=>v));
  return `<div style="margin-top:12px;">
    <div style="font-size:10px;letter-spacing:2px;color:#999;text-transform:uppercase;margin-bottom:8px;">Consumo por hora</div>
    ${horas.map(([h,v])=>`<div class="hora-row">
      <span class="hora-lbl">${h}h</span>
      <div class="hora-bar-wrap"><div class="hora-bar" style="width:${Math.round(v/maxVal*100)}%"></div></div>
      <span class="hora-val">${fmt(v)}</span>
    </div>`).join('')}
  </div>`;
}

function renderNoches() {
  document.getElementById('tbl-noches').innerHTML = [...historial].reverse().map((n,idx)=>{
    const rnk=n.ranking||[];
    const id='dp-'+idx;
    return `<tr class="noche-row" onclick="toggleDetail('${id}')">
      <td class="gold">${n.fecha}</td><td>${n.hora_cierre||'—'}</td>
      <td class="gold">${fmt(n.total)}</td><td>${n.operaciones}</td>
      <td>${esc(rnk[0]?.name||'—')}</td><td>${esc(rnk[1]?.name||'—')}</td><td>${esc(rnk[2]?.name||'—')}</td>
    </tr>
    <tr><td colspan="7" style="padding:0;border:none">
      <div class="detail-panel" id="${id}">
        <div style="font-size:11px;color:#aaa;letter-spacing:2px;margin-bottom:10px;">
          ABAJO: <span class="gold">${fmt(n.por_caja?.[1]||n.por_caja?.['1']||0)}</span> &nbsp;·&nbsp;
          EXTENDIDO: <span class="gold">${fmt(n.por_caja?.[2]||n.por_caja?.['2']||0)}</span> &nbsp;·&nbsp;
          VIP: <span class="gold">${fmt(n.por_caja?.[3]||n.por_caja?.['3']||0)}</span>
        </div>
        <div style="font-size:10px;letter-spacing:2px;color:#999;text-transform:uppercase;margin-bottom:6px;">Ranking de la noche</div>
        ${rnk.slice(0,5).map((r,i)=>`<div style="display:flex;justify-content:space-between;padding:6px 0;border-bottom:1px solid #181818;font-size:14px;">
          <span><span class="rank-badge ${i<3?'r'+(i+1):''}" style="${i>=3?'background:#1a1a1a;color:#aaa;width:24px;height:24px;border-radius:50%;display:inline-block;text-align:center;line-height:24px;font-size:11px;':''}">${i+1}</span> &nbsp;${esc(r.name)}</span>
          <span class="gold">${fmt(r.total)}</span>
        </div>`).join('')}
        ${horasHTML(n.transactions)}
      </div>
    </td></tr>`;
  }).join('') || `<tr><td colspan="7" class="empty">Sin noches registradas</td></tr>`;
}

function toggleDetail(id) { document.getElementById(id).classList.toggle('open'); }

function renderTrimTabs() {
  const ahora = new Date();
  const tabs = [
    {id:'all', label:'Todo'},
    {id:'year', label:String(ahora.getFullYear())},
    {id:'q'+Math.ceil((ahora.getMonth()+1)/3), label:'Trimestre '+Math.ceil((ahora.getMonth()+1)/3)},
    {id:'month', label:ahora.toLocaleString('es',{month:'long'})},
  ];
  document.getElementById('trim-tabs').innerHTML = tabs.map(t=>
    `<div class="trim-tab${t.id===periodoActivo?' active':''}" onclick="selectPeriodo('${t.id}',this)">${t.label}</div>`
  ).join('');
}

function selectPeriodo(id, el) {
  periodoActivo = id;
  document.querySelectorAll('.trim-tab').forEach(x=>x.classList.remove('active'));
  el.classList.add('active');
  renderRanking(id);
}

function filtrarNoches(periodo) {
  const ahora = new Date();
  const año = ahora.getFullYear(), mes = ahora.getMonth()+1;
  return historial.filter(n=>{
    if (periodo==='all') return true;
    const [y,m] = n.fecha.split('-').map(Number);
    if (periodo==='year') return y===año;
    if (periodo==='month') return y===año && m===mes;
    if (periodo.startsWith('q')) { const q=parseInt(periodo[1]); return y===año && Math.ceil(m/3)===q; }
    return true;
  });
}

function renderRanking(periodo) {
  const noches = filtrarNoches(periodo);
  const rk = {};
  noches.forEach(n=>(n.ranking||[]).forEach(r=>{ if(!rk[r.name]) rk[r.name]={total:0,noches:0}; rk[r.name].total+=r.total; rk[r.name].noches++; }));
  const sorted = Object.entries(rk).sort((a,b)=>b[1].total-a[1].total);
  document.getElementById('tbl-ranking').innerHTML = sorted.map(([nm,v],i)=>{
    const badge = i<3 ? `<span class="rank-badge r${i+1}">${i+1}</span>` : `<span style="color:#555;padding:0 6px;">${i+1}</span>`;
    return `<tr><td>${badge}</td><td>${esc(nm)}</td><td>${v.noches}</td><td class="gold">${fmt(v.total)}</td><td>${fmt(Math.round(v.total/v.noches))}</td></tr>`;
  }).join('') || `<tr><td colspan="5" class="empty">Sin datos para este período</td></tr>`;
}

init();

</script>
</body>
</html>"""

@app.route('/historial')
def historial_page():
    return render_template_string(HISTORIAL_HTML)

@app.route('/')
def hub():
    return render_template_string(HUB_HTML)

@app.route('/pantalla')
def pantalla_page():
    return render_template_string(PANTALLA_HTML)

@app.route('/celular')
def celular_page():
    login = require_auth('manager', 'Celular', '/celular')
    if login: return login
    return render_template_string(MONITOR_HTML)

MONITOR_HTML = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1, maximum-scale=1">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<title>Monitor — Jagger VIP</title>
<link href="https://fonts.googleapis.com/css2?family=Oswald:wght@400;700&family=Rajdhani:wght@400;600;700&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0;-webkit-tap-highlight-color:transparent;}
html,body{background:#080808;color:#f0ece0;font-family:'Rajdhani',sans-serif;min-height:100vh;overscroll-behavior:none;}
body{padding:0 0 32px;}

/* Header */
.mon-header{background:#0d0c00;border-bottom:1px solid #2a2200;padding:14px 16px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:10;}
.mon-logo{font-family:'Oswald',sans-serif;font-size:18px;font-weight:700;color:#c9a227;letter-spacing:3px;}
.mon-sub{font-size:11px;color:#555;letter-spacing:2px;text-transform:uppercase;margin-top:2px;}
.mon-live{display:flex;align-items:center;gap:6px;font-size:11px;color:#3a9a5a;letter-spacing:1px;font-weight:700;}
.mon-live-dot{width:7px;height:7px;border-radius:50%;background:#3a9a5a;animation:pulse 1.8s ease-in-out infinite;}
@keyframes pulse{0%,100%{opacity:1;transform:scale(1);}50%{opacity:.4;transform:scale(.7);}}

/* KPIs */
.kpi-row{display:grid;grid-template-columns:1fr 1fr;gap:8px;padding:12px 12px 0;}
.kpi-box{background:#0d0c00;border:1px solid #2a2200;border-radius:10px;padding:12px 14px;}
.kpi-label{font-size:10px;color:#555;letter-spacing:2px;text-transform:uppercase;margin-bottom:4px;}
.kpi-val{font-family:'Oswald',sans-serif;font-size:22px;font-weight:700;color:#c9a227;}

/* Ranking */
.ranking-title{font-size:10px;color:#444;letter-spacing:3px;text-transform:uppercase;padding:16px 14px 6px;font-weight:700;}

.rank-item{display:flex;align-items:center;gap:12px;padding:12px 14px;border-bottom:1px solid #111;transition:background .15s;}
.rank-item:active{background:#111;}
.rank-pos{font-family:'Oswald',sans-serif;font-size:22px;font-weight:700;color:#333;width:28px;text-align:center;flex-shrink:0;}
.rank-pos.gold{color:#c9a227;}
.rank-pos.silver{color:#aaaaaa;}
.rank-pos.bronze{color:#cd7f32;}
.rank-info{flex:1;min-width:0;}
.rank-name{font-family:'Oswald',sans-serif;font-size:19px;font-weight:700;color:#f0ece0;letter-spacing:1px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.rank-mesa{font-size:12px;color:#555;letter-spacing:1px;margin-top:1px;}
.rank-amount{font-family:'Oswald',sans-serif;font-size:20px;font-weight:700;color:#c9a227;flex-shrink:0;text-align:right;}

/* Barra de progreso relativa al 1ro */
.rank-bar-wrap{height:3px;background:#1a1a1a;border-radius:2px;margin-top:6px;}
.rank-bar{height:3px;border-radius:2px;background:linear-gradient(90deg,#c9a227,#e8c84a);transition:width .4s ease;}

/* Saldo tarjetas */
.saldo-title{font-size:10px;color:#444;letter-spacing:3px;text-transform:uppercase;padding:16px 14px 6px;font-weight:700;}
.saldo-item{display:flex;align-items:center;gap:10px;padding:10px 14px;border-bottom:1px solid #111;}
.saldo-mesa{font-family:'Oswald',sans-serif;font-size:15px;font-weight:700;color:#888;width:56px;flex-shrink:0;}
.saldo-nombre{flex:1;font-size:14px;color:#ccc;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.saldo-val{font-family:'Oswald',sans-serif;font-size:15px;font-weight:700;flex-shrink:0;}
.saldo-val.ok{color:#4aaa4a;}
.saldo-val.bajo{color:#ffaa00;}
.saldo-val.cero{color:#a83030;}
.saldo-bar-wrap{height:2px;background:#1a1a1a;border-radius:1px;margin-top:4px;}
.saldo-bar{height:2px;border-radius:1px;transition:width .4s ease;}

/* Empty */
.empty{padding:32px 20px;text-align:center;color:#333;font-size:14px;letter-spacing:1px;}

/* Update indicator */
.update-row{padding:8px 14px;text-align:right;font-size:10px;color:#2a2a2a;letter-spacing:1px;}
</style>
</head>
<body>

<div class="mon-header">
  <div>
    <div class="mon-logo">RANKING VIP</div>
    <div class="mon-sub">Monitor en vivo</div>
  </div>
  <div class="mon-live"><div class="mon-live-dot"></div>EN VIVO</div>
</div>

<div class="kpi-row">
  <div class="kpi-box">
    <div class="kpi-label">Total noche</div>
    <div class="kpi-val" id="kpi-total">$0</div>
  </div>
  <div class="kpi-box">
    <div class="kpi-label">Operaciones</div>
    <div class="kpi-val" id="kpi-ops">0</div>
  </div>
</div>

<div class="ranking-title">🏆 RANKING</div>
<div id="ranking-list"></div>

<div class="saldo-title">💳 SALDOS DE TARJETAS</div>
<div id="saldo-list"></div>

<div class="update-row" id="update-label">—</div>

<script>
function fmt(n){return '$'+Number(n).toLocaleString('es-AR',{minimumFractionDigits:0,maximumFractionDigits:0});}
function esc(s){return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}

async function refresh() {
  try {
    const [r1, r2, r3] = await Promise.all([
      fetch('/api/tx'),
      fetch('/api/tarjetas'),
      fetch('/api/tarjetas/config')
    ]);
    const txs = await r1.json();
    const saldos = await r2.json();
    const conf = await r3.json();

    // ── KPIs ──
    const total = txs.reduce((s,t) => s + t.amount, 0);
    document.getElementById('kpi-total').textContent = fmt(total);
    document.getElementById('kpi-ops').textContent = txs.length;

    // ── Ranking ──
    const totals = {}, mesas = {};
    txs.forEach(t => {
      totals[t.name] = (totals[t.name]||0) + t.amount;
      if (t.mesa) mesas[t.name] = t.mesa;
    });
    const ranking = Object.entries(totals).sort((a,b) => b[1]-a[1]);
    const max = ranking[0]?.[1] || 1;
    const medals = ['🥇','🥈','🥉'];
    const posClass = ['gold','silver','bronze'];
    const rl = document.getElementById('ranking-list');
    if (!ranking.length) {
      rl.innerHTML = '<div class="empty">Sin consumos registrados</div>';
    } else {
      rl.innerHTML = ranking.map(([name, amount], i) => {
        const pct = Math.round(amount / max * 100);
        const pc = posClass[i] || '';
        const medal = medals[i] || (i+1);
        return `<div class="rank-item">
          <div class="rank-pos ${pc}">${medal}</div>
          <div class="rank-info">
            <div class="rank-name">${esc(name)}</div>
            <div class="rank-mesa">Mesa ${esc(mesas[name]||'—')}</div>
            <div class="rank-bar-wrap"><div class="rank-bar" style="width:${pct}%"></div></div>
          </div>
          <div class="rank-amount">${fmt(amount)}</div>
        </div>`;
      }).join('');
    }

    // ── Saldos ──
    const activas = conf.filter(t => t.codigo).sort((a,b) => Number(a.slot)-Number(b.slot));
    const sl = document.getElementById('saldo-list');
    if (!activas.length) {
      sl.innerHTML = '<div class="empty">Sin tarjetas configuradas</div>';
    } else {
      sl.innerHTML = activas.map(t => {
        const td = saldos[t.codigo] || {};
        const ini = parseFloat(t.saldo_inicial||0);
        const act = td.saldo_actual !== undefined ? td.saldo_actual : ini;
        const techo = Math.max(ini, act);
        const pct = techo > 0 ? Math.min(100, Math.round(act/techo*100)) : 0;
        const cls = act <= 0 ? 'cero' : pct <= 25 ? 'bajo' : 'ok';
        const barColor = act <= 0 ? '#a83030' : pct <= 25 ? '#ffaa00' : '#4aaa4a';
        return `<div class="saldo-item">
          <div>
            <div class="saldo-mesa">Mesa ${esc(String(t.slot))}</div>
          </div>
          <div class="rank-info">
            <div class="saldo-nombre">${esc(t.nombre_cliente||'—')}</div>
            <div class="saldo-bar-wrap"><div class="saldo-bar" style="width:${pct}%;background:${barColor};"></div></div>
          </div>
          <div class="saldo-val ${cls}">${fmt(act)}</div>
        </div>`;
      }).join('');
    }

    // timestamp
    const now = new Date();
    document.getElementById('update-label').textContent =
      'Actualizado ' + now.getHours().toString().padStart(2,'0') + ':' +
      now.getMinutes().toString().padStart(2,'0') + ':' +
      now.getSeconds().toString().padStart(2,'0');

  } catch(e) {
    document.getElementById('update-label').textContent = 'Sin conexión...';
  }
}

refresh();
setInterval(refresh, 4000);
</script>
</body>
</html>
"""

@app.route('/tarjetas')
def tarjetas_page():
    login = require_auth('tarjetas', 'Tarjetas', '/tarjetas')
    if login: return login
    return render_template_string(TARJETAS_HTML)

def check_auth(role):
    if role != 'manager' and session.get('manager_auth') == True:
        cfg_v = load_config()
        if session.get('session_version', -1) == cfg_v.get('session_version', 0):
            return True
    if session.get(role + '_auth') != True:
        return False
    cfg_v = load_config()
    if session.get('session_version', -1) != cfg_v.get('session_version', 0):
        session.clear()
        return False
    return True

def require_auth(role, title, redirect_to):
    if check_auth(role):
        return None
    return render_template_string(LOGIN_HTML, title=title, role=role, redirect_to=redirect_to)

@app.route('/manager')
def manager_page():
    login = require_auth('manager', 'Manager', '/manager')
    if login: return login
    ensure_menu()
    with lock:
        _menu_data = load_data().get('menu', [])
    import json as _json
    menu_json = _json.dumps(_menu_data, ensure_ascii=False)
    return render_template_string(HTML, menu_json=menu_json)

@app.route('/pantalla-mgr')
def pantalla_mgr_page():
    login = require_auth('manager', 'Manager', '/pantalla-mgr')
    if login: return login
    return render_template_string(PANTALLA_MGR_HTML)

@app.route('/cajaabajo')
def cajaabajo_page():
    login = require_auth('cajaabajo', 'Caja Abajo', '/cajaabajo')
    if login: return login
    return render_template_string(CAJA_HTML, caja_num=1, caja_nombre='Abajo')

@app.route('/cajaextendido')
def cajaextendido_page():
    login = require_auth('cajaextendido', 'Caja Extendido', '/cajaextendido')
    if login: return login
    return render_template_string(CAJA_HTML, caja_num=2, caja_nombre='Extendido')

@app.route('/cajavip')
def cajavip_page():
    login = require_auth('cajavip', 'Caja VIP', '/cajavip')
    if login: return login
    return render_template_string(CAJA_HTML, caja_num=3, caja_nombre='VIP')

@app.route('/api/tx', methods=['GET'])
def get_tx():
    with lock:
        data = load_data()
    return jsonify(data['transactions'])

@app.route('/api/names', methods=['GET'])
def get_names():
    data = load_data()
    names = sorted(list({t['name'] for t in data.get('transactions', []) if t.get('name')}), key=str.lower)
    return jsonify(names)

@app.route('/api/tx', methods=['POST'])
def add_tx():
    try:
        with lock:
            data = load_data()
            body = request.get_json()
            if not body or 'name' not in body or 'amount' not in body or 'caja' not in body:
                return jsonify({'ok': False, 'error': 'Faltan campos requeridos'}), 400
            amount = float(body['amount'])
            codigo = str(body.get('tarjeta_codigo',''))

            # Normalizar nombre: si ya existe alguien con ese nombre (distinto case), usar el canónico
            raw_name = str(body['name']).strip()
            existing_names = {t['name'] for t in data.get('transactions', [])}
            canonical_name = next((n for n in existing_names if n.lower() == raw_name.lower()), raw_name)

            if codigo:
                if 'tarjetas' not in data: data['tarjetas'] = {}
                if codigo not in data['tarjetas']:
                    conf_list = data.get('tarjetas_conf', [])
                    conf = next((t for t in conf_list if t.get('codigo')==codigo), None)
                    try:
                        saldo_ini = float(conf['saldo_inicial']) if conf and conf.get('saldo_inicial') not in ('', None) else 0
                    except (ValueError, TypeError):
                        saldo_ini = 0
                    data['tarjetas'][codigo] = {'saldo_actual': saldo_ini, 'nombre': canonical_name}
                saldo_disponible = data['tarjetas'][codigo]['saldo_actual']
                if saldo_disponible < amount:
                    return jsonify({'ok': False, 'error': 'Saldo insuficiente. Disponible: $' + str(int(saldo_disponible))}), 400
            data['tx_id_counter'] += 1
            tx = {
                'id': data['tx_id_counter'],
                'name': canonical_name,
                'amount': amount,
                'caja': int(body['caja']),
                'mesa': str(body.get('mesa','')),
                'tarjeta_codigo': codigo,
                'time': str(body.get('client_time','')) or datetime.now().strftime('%H:%M'),
                'items': body.get('items', []),
            }
            data['transactions'].append(tx)
            if codigo:
                data['tarjetas'][codigo]['saldo_actual'] -= amount
                data['tarjetas'][codigo]['nombre'] = canonical_name
            save_data(data)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/tx/<int:tid>', methods=['PUT'])
def edit_tx(tid):
    try:
        with lock:
            data = load_data()
            tx = next((t for t in data['transactions'] if t['id']==tid), None)
            if not tx:
                return jsonify({'ok': False, 'error': 'Transacción no encontrada'}), 404
            body = request.get_json() or {}
            new_amount = float(body.get('amount', tx['amount']))
            diff = new_amount - tx['amount']
            codigo = tx.get('tarjeta_codigo', '')
            if codigo and diff != 0 and 'tarjetas' in data and codigo in data['tarjetas']:
                nuevo_saldo = data['tarjetas'][codigo]['saldo_actual'] - diff
                if nuevo_saldo < 0:
                    return jsonify({'ok': False, 'error': 'Saldo insuficiente para este ajuste'}), 400
                data['tarjetas'][codigo]['saldo_actual'] = nuevo_saldo
            tx['amount'] = new_amount
            if 'name' in body and str(body['name']).strip():
                tx['name'] = str(body['name']).strip()
            save_data(data)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/tx/<int:tid>', methods=['DELETE'])
def delete_tx(tid):
    with lock:
        data = load_data()
        tx = next((t for t in data['transactions'] if t['id']==tid), None)
        if tx:
            codigo = tx.get('tarjeta_codigo','')
            if codigo and 'tarjetas' in data and codigo in data['tarjetas']:
                data['tarjetas'][codigo]['saldo_actual'] += tx['amount']
            data['transactions'] = [t for t in data['transactions'] if t['id']!=tid]
            save_data(data)
    return jsonify({'ok': True})

@app.route('/api/tarjetas', methods=['GET'])
def get_tarjetas():
    with lock:
        data = load_data()
    return jsonify(data.get('tarjetas', {}))

@app.route('/api/tarjetas/config', methods=['GET'])
def get_tarjetas_conf():
    with lock:
        data = load_data()
    return jsonify(data.get('tarjetas_conf', []))

@app.route('/api/tarjetas/config', methods=['POST'])
def set_tarjetas_conf():
    try:
        with lock:
            data = load_data()
            conf = request.get_json()
            if not conf:
                return jsonify({'ok': False, 'error': 'Configuracion invalida'}), 400
            codigos_vistos = {}
            for t in conf:
                codigo = t.get('codigo','')
                if codigo:
                    if codigo in codigos_vistos:
                        slot_anterior = codigos_vistos[codigo]
                        return jsonify({'ok': False, 'error': f'La tarjeta ya esta asignada a Mesa {slot_anterior}'}), 400
                    codigos_vistos[codigo] = t.get('slot', '?')
            conf_vieja = {t.get('codigo',''): t for t in data.get('tarjetas_conf', []) if t.get('codigo','')}
            data['tarjetas_conf'] = conf
            if 'tarjetas' not in data: data['tarjetas'] = {}
            for t in conf:
                codigo = t.get('codigo','')
                if codigo and t.get('saldo_inicial'):
                    nuevo_saldo_ini = float(t['saldo_inicial'])
                    if codigo not in data['tarjetas']:
                        data['tarjetas'][codigo] = {'saldo_actual': nuevo_saldo_ini, 'nombre':''}
                    else:
                        old_conf = conf_vieja.get(codigo)
                        old_ini = float(old_conf['saldo_inicial']) if old_conf and old_conf.get('saldo_inicial') else None
                        if old_ini is not None and old_ini != nuevo_saldo_ini:
                            gastado = old_ini - data['tarjetas'][codigo]['saldo_actual']
                            nuevo_saldo_act = max(0, nuevo_saldo_ini - gastado)
                            data['tarjetas'][codigo]['saldo_actual'] = nuevo_saldo_act
                        elif old_ini is None:
                            data['tarjetas'][codigo]['saldo_actual'] = nuevo_saldo_ini
            save_data(data)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/tarjetas/limpiar', methods=['POST'])
def limpiar_tarjetas():
    with lock:
        data = load_data()
        data['tarjetas_conf'] = []
        data['tarjetas'] = {}
        save_data(data)
    return jsonify({'ok': True})

@app.route('/api/reset', methods=['POST'])
def reset():
    with lock:
        data = load_data()
        data['transactions'] = []
        data['tx_id_counter'] = 0
        conf_list = data.get('tarjetas_conf', [])
        data['tarjetas'] = {}
        for t in conf_list:
            codigo = t.get('codigo','')
            if codigo and t.get('saldo_inicial'):
                data['tarjetas'][codigo] = {'saldo_actual': float(t['saldo_inicial']), 'nombre':''}
        save_data(data)
    return jsonify({'ok': True})

@app.route('/api/state', methods=['GET'])
def get_state():
    return jsonify(_state)

@app.route('/api/state', methods=['POST'])
def set_state():
    import time
    body = request.get_json() or {}
    with lock:
        if 'hora_fin' in body: _state['hora_fin'] = str(body['hora_fin'])
        if 'premio' in body: _state['premio'] = str(body['premio'])
        save_state()
    return jsonify({'ok': True})

@app.route('/api/winner/show', methods=['POST'])
def winner_show():
    import time
    with lock:
        _state['winner_show'] = True
        _state['winner_ts'] = int(time.time() * 1000)
    return jsonify({'ok': True})

@app.route('/api/winner/hide', methods=['POST'])
def winner_hide():
    with lock:
        _state['winner_show'] = False
    return jsonify({'ok': True})

@app.route('/api/cartel/show', methods=['POST'])
def cartel_show():
    import time
    body = request.get_json() or {}
    with lock:
        _state['cartel_show'] = True
        _state['cartel_ts'] = int(time.time() * 1000)
        _state['cartel_data'] = {
            'nombre': str(body.get('nombre','')),
            'mesa': str(body.get('mesa','')),
            'frase': str(body.get('frase','')),
            'emoji': str(body.get('emoji','🍾')),
        }
    return jsonify({'ok': True})

@app.route('/api/cartel/hide', methods=['POST'])
def cartel_hide():
    with lock:
        _state['cartel_show'] = False
    return jsonify({'ok': True})

@app.route('/api/cartel/precios', methods=['GET'])
def get_cartel_precios():
    return jsonify(_state.get('cartel_precios', {'virtual': 0, 'fisico': 0, 'combo': 0}))

@app.route('/api/cartel/precios', methods=['POST'])
def set_cartel_precios():
    body = request.get_json() or {}
    with lock:
        if 'cartel_precios' not in _state:
            _state['cartel_precios'] = {'virtual': 0, 'fisico': 0, 'combo': 0}
        for k in ('virtual', 'fisico', 'combo'):
            if k in body:
                try:
                    _state['cartel_precios'][k] = float(body[k])
                except (ValueError, TypeError):
                    pass
        save_state()
    return jsonify({'ok': True})

@app.route('/api/cerrar_noche', methods=['POST'])
def cerrar_noche():
    with lock:
        data = load_data()
        txs = data['transactions']
        if not txs:
            return jsonify({'ok': False, 'error': 'No hay operaciones esta noche'}), 400
        totals, mesas, por_caja = {}, {}, {1:0, 2:0, 3:0}
        for t in txs:
            totals[t['name']] = totals.get(t['name'], 0) + t['amount']
            if t.get('mesa') and t['name'] not in mesas:
                mesas[t['name']] = t['mesa']
            c = int(t.get('caja', 1))
            if 1 <= c <= 3:
                por_caja[c] += t['amount']
        ranking = sorted(
            [{'name': n, 'total': v, 'mesa': mesas.get(n, '')} for n, v in totals.items()],
            key=lambda x: -x['total']
        )
        noche = {
            'id': datetime.now().strftime('%Y%m%d_%H%M%S'),
            'fecha': datetime.now().strftime('%Y-%m-%d'),
            'hora_cierre': datetime.now().strftime('%H:%M'),
            'hora_fin': _state.get('hora_fin', '05:30'),
            'total': sum(t['amount'] for t in txs),
            'operaciones': len(txs),
            'por_caja': por_caja,
            'ranking': ranking,
            'transactions': list(txs),
        }
        historial = load_historial()
        historial.append(noche)
        save_historial(historial)
    return jsonify({'ok': True, 'noche_id': noche['id']})

@app.route('/api/historial', methods=['GET'])
def get_historial():
    return jsonify(load_historial())

@app.route('/api/export/excel')
def export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series
        from openpyxl.chart.series import DataPoint
        from openpyxl.utils import get_column_letter
        from io import BytesIO
    except ImportError:
        return jsonify({'error': 'Instalá openpyxl: pip install openpyxl'}), 500

    historial = load_historial()
    if not historial:
        return jsonify({'error': 'No hay noches registradas aún'}), 400

    wb = Workbook()

    # ── Paleta única coherente ────────────────────────────────
    GOLD    = 'C9A227'
    GOLD_LT = 'F5E6A0'
    BLACK   = '111111'
    DARK    = '1E1E1E'
    WHITE   = 'FFFFFF'
    LGRAY   = 'F2F2F2'
    MGRAY   = 'D9D9D9'
    DGRAY   = '666666'
    GREEN   = '1E6B3C'
    GREEN_LT= 'D6EFDF'
    RED     = 'A83030'
    RED_LT  = 'F8D7DA'
    BLUE    = '1A4A8A'
    BLUE_LT = 'D0E4F7'

    def fill(c): return PatternFill('solid', fgColor=c)
    def side(c='CCCCCC', s='thin'): return Side(style=s, color=c)
    def border(c='CCCCCC'):
        s = side(c); return Border(left=s, right=s, top=s, bottom=s)
    def font(bold=False, color=BLACK, size=10, name='Calibri'):
        return Font(bold=bold, color=color, size=size, name=name)
    def align(h='left', v='center', wrap=False, indent=0):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

    BDR      = border('CCCCCC')
    BDR_GOLD = border(GOLD)
    BDR_DK   = border('999999')

    def cell(ws, r, c, val, f=None, bg=None, bdr=None, al=None, fmt=None):
        cc = ws.cell(row=r, column=c, value=val)
        if f:   cc.font      = f
        if bg:  cc.fill      = fill(bg)
        if bdr: cc.border    = bdr
        if al:  cc.alignment = al
        if fmt: cc.number_format = fmt
        return cc

    def row_h(ws, r, h): ws.row_dimensions[r].height = h
    def col_w(ws, c, w): ws.column_dimensions[get_column_letter(c)].width = w

    def title_row(ws, r, text, sub=''):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        cc = ws.cell(row=r, column=1, value=text)
        cc.font = Font(bold=True, color=GOLD, size=18, name='Calibri')
        cc.fill = fill(BLACK); cc.alignment = align('left', indent=2)
        row_h(ws, r, 40)
        if sub:
            ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=12)
            sc = ws.cell(row=r+1, column=1, value=sub)
            sc.font = Font(color=DGRAY, size=9, name='Calibri')
            sc.fill = fill(BLACK); sc.alignment = align('left', indent=2)
            row_h(ws, r+1, 15)

    def section(ws, r, text, col_start=1, col_end=8):
        ws.merge_cells(start_row=r, start_column=col_start, end_row=r, end_column=col_end)
        cc = ws.cell(row=r, column=col_start, value=text)
        cc.font = Font(bold=True, color=WHITE, size=10, name='Calibri')
        cc.fill = fill(DARK); cc.alignment = align('left', indent=1)
        row_h(ws, r, 20)

    def header_row(ws, r, labels, col_start=1, bg=GOLD):
        for ci, lbl in enumerate(labels, col_start):
            cc = ws.cell(row=r, column=ci, value=lbl)
            cc.font = Font(bold=True, color=BLACK, size=9, name='Calibri')
            cc.fill = fill(bg); cc.border = BDR_GOLD
            cc.alignment = align('center')
        row_h(ws, r, 20)

    def data_cell(ws, r, c, val, bold=False, alt=False, fmt=None, color=None, h='center'):
        bg = LGRAY if alt else WHITE
        f = font(bold=bold, color=color or BLACK)
        al = align(h)
        cc = ws.cell(row=r, column=c, value=val)
        cc.font = f; cc.fill = fill(bg); cc.border = BDR; cc.alignment = al
        if fmt: cc.number_format = fmt
        row_h(ws, r, 17)
        return cc

    def money(ws, r, c, val, alt=False, bold=False, color=None):
        data_cell(ws, r, c, val, bold=bold, alt=alt, fmt='#.##0', h='right', color=color or GREEN)

    def pct(ws, r, c, val, alt=False):
        data_cell(ws, r, c, val, alt=alt, fmt='0.0%', h='center')

    def kpi_block(ws, r, c, label, value, fmt='#.##0', val_color=GREEN):
        lc = ws.cell(row=r, column=c, value=label)
        lc.font = Font(bold=False, color=DGRAY, size=8, name='Calibri')
        lc.fill = fill(LGRAY); lc.border = border('E0E0E0')
        lc.alignment = align('left', indent=1); row_h(ws, r, 22)
        vc = ws.cell(row=r, column=c+1, value=value)
        vc.font = Font(bold=True, color=val_color, size=12, name='Calibri')
        vc.fill = fill(WHITE); vc.border = border('E0E0E0')
        vc.alignment = align('right', indent=1)
        if fmt: vc.number_format = fmt

    # ── Pre-cálculos ─────────────────────────────────────────
    total_all  = sum(n['total'] for n in historial)
    total_ops  = sum(n['operaciones'] for n in historial)
    avg_noche  = total_all / len(historial) if historial else 0
    mejor      = max(n['total'] for n in historial) if historial else 0

    rk_all = {}
    noches_cliente = {}
    for n in historial:
        vistos = set()
        for rv in n.get('ranking', []):
            nm = rv['name']
            rk_all[nm] = rk_all.get(nm, 0) + rv['total']
            if nm not in vistos:
                noches_cliente[nm] = noches_cliente.get(nm, 0) + 1
                vistos.add(nm)
    rk_sorted = sorted(rk_all.items(), key=lambda x: -x[1])

    por_mes = {}
    noches_mes = {}
    for n in historial:
        mes = n['fecha'][:7]
        por_mes[mes] = por_mes.get(mes, 0) + n['total']
        noches_mes[mes] = noches_mes.get(mes, 0) + 1
    meses = sorted(por_mes.keys())

    ul = historial[-1]
    pc3 = ul.get('por_caja', {})
    cj  = [pc3.get(i, pc3.get(str(i), 0)) for i in [1,2,3]]
    if not any(cj):
        ops_cj = [{},{},[0,0,0]][0]
        for t in ul.get('transactions', []):
            k = int(t.get('caja', 1)) - 1
            if 0 <= k <= 2: cj[k] = cj[k] + t['amount']

    ops_cj = [0, 0, 0]
    for t in ul.get('transactions', []):
        k = int(t.get('caja', 1)) - 1
        if 0 <= k <= 2: ops_cj[k] += 1

    medal_color = {1: GOLD, 2: '9E9E9E', 3: 'CD7F32'}

    # ════════════════════════════════════════════════════════
    # HOJA 1 — RESUMEN GENERAL
    # ════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Resumen'
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = GOLD
    ws1.freeze_panes = 'A4'

    title_row(ws1, 1,
        'JAGGER CLUB — RESUMEN GENERAL',
        f'Exportado el {datetime.now().strftime("%d/%m/%Y %H:%M")}  ·  {len(historial)} noche(s) registrada(s)')

    # KPIs fila 3 (3 pares lado a lado)
    ws1.row_dimensions[3].height = 6
    kpi_block(ws1, 4, 1, 'TOTAL FACTURADO',    total_all,   '#.##0')
    kpi_block(ws1, 4, 3, 'NOCHES',             len(historial), 'General', BLUE)
    kpi_block(ws1, 4, 5, 'OPERACIONES',        total_ops,   'General', BLUE)
    kpi_block(ws1, 5, 1, 'PROMEDIO / NOCHE',   avg_noche,   '#.##0')
    kpi_block(ws1, 5, 3, 'PROM / OPERACIÓN',   total_all/total_ops if total_ops else 0, '#.##0')
    kpi_block(ws1, 5, 5, 'MEJOR NOCHE',        mejor,       '#.##0')
    ws1.row_dimensions[6].height = 8

    # Ranking general
    section(ws1, 7, '  RANKING GENERAL — TOP 20', 1, 5)
    header_row(ws1, 8, ['#', 'CLIENTE', 'TOTAL ($)', '% DEL TOTAL', 'NOCHES'], 1, GOLD)
    for i, (nm, tot) in enumerate(rk_sorted[:20], 1):
        alt = (i % 2 == 0)
        mc = medal_color.get(i)
        data_cell(ws1, 8+i, 1, f'#{i}', bold=i<=3, alt=alt, h='center', color=mc)
        data_cell(ws1, 8+i, 2, nm,      bold=i<=3, alt=alt, h='left',   color=mc)
        money(ws1, 8+i, 3, tot, alt=alt, bold=i<=3, color=mc)
        pct(ws1, 8+i, 4, tot/total_all if total_all else 0, alt=alt)
        data_cell(ws1, 8+i, 5, noches_cliente.get(nm, 0), alt=alt, h='center')

    base = 10 + len(rk_sorted[:20])
    ws1.row_dimensions[base].height = 10

    # Por mes
    section(ws1, base+1, '  FACTURACIÓN POR MES', 1, 5)
    header_row(ws1, base+2, ['MES', 'TOTAL ($)', '% DEL TOTAL', 'NOCHES', 'PROM/NOCHE'], 1, GOLD)
    ms = base + 3
    for i, mes in enumerate(meses):
        alt = (i%2==0)
        data_cell(ws1, ms+i, 1, mes, alt=alt)
        money(ws1, ms+i, 2, por_mes[mes], alt=alt)
        pct(ws1, ms+i, 3, por_mes[mes]/total_all if total_all else 0, alt=alt)
        data_cell(ws1, ms+i, 4, noches_mes.get(mes,0), alt=alt, h='center')
        money(ws1, ms+i, 5, por_mes[mes]/noches_mes[mes] if noches_mes.get(mes) else 0, alt=alt)
    ms_end = ms + len(meses) - 1

    # Evolución noche a noche — tabla completa
    ev = ms_end + 3
    ws1.row_dimensions[ev].height = 10
    section(ws1, ev+1, '  EVOLUCIÓN NOCHE A NOCHE', 1, 5)
    header_row(ws1, ev+2, ['FECHA', 'TOTAL ($)', 'OPS', 'PROM/OP ($)', 'CAJA TOP'], 1, GOLD)
    ev_d = ev + 3
    for i, n in enumerate(historial):
        alt = (i%2==0)
        data_cell(ws1, ev_d+i, 1, n['fecha'], alt=alt)
        money(ws1, ev_d+i, 2, n['total'], alt=alt)
        data_cell(ws1, ev_d+i, 3, n['operaciones'], alt=alt, h='center')
        money(ws1, ev_d+i, 4, n['total']/n['operaciones'] if n['operaciones'] else 0, alt=alt)
        pc = n.get('por_caja', {}); top = max(pc, key=lambda k: pc[k], default='—')
        data_cell(ws1, ev_d+i, 5, CAJA_NOMBRES.get(int(top) if str(top).isdigit() else 0, str(top)), alt=alt)
    ev_end = ev_d + len(historial) - 1

    # ── Gráfico evolución — UNA línea, un punto por noche, colores distintos por marcador ──
    # El gráfico va DEBAJO de todas las tablas
    chart_anchor_row = ev_end + 3
    ws1.row_dimensions[chart_anchor_row].height = 8

    EVOL_COLORS = ['C9A227','1A4A8A','1E6B3C','A83030','7B4EA0',
                   '4A6B8A','C87320','1E5A4A','6B1A3A','2A3A6B',
                   '3A6A9A','4A9A3A','C94A27','9A27C9','6B3A1A']

    if len(historial) >= 1:
        # Una sola serie = la línea que sube/baja
        ch_l = LineChart()
        ch_l.title = 'Evolución noche a noche'
        ch_l.style = 10
        ch_l.width = 28; ch_l.height = 14
        ch_l.y_axis.numFmt = '#,##0'
        ch_l.y_axis.title = 'Total ($)'
        ch_l.x_axis.title = 'Noche'

        # Datos: todos los totales en una sola serie
        ch_l.add_data(Reference(ws1, min_col=2, min_row=ev+2, max_row=ev_end),
                      titles_from_data=True)
        ch_l.set_categories(Reference(ws1, min_col=1, min_row=ev_d, max_row=ev_end))

        # Línea dorada
        s = ch_l.series[0]
        s.graphicalProperties.line.solidFill = GOLD
        s.graphicalProperties.line.width = 22000

        # Marcadores — color diferente por punto usando DataPoint
        s.marker.symbol = 'circle'
        s.marker.size = 10
        s.marker.graphicalProperties.solidFill = GOLD
        s.marker.graphicalProperties.ln.solidFill = BLACK

        for i in range(len(historial)):
            pt = DataPoint(idx=i)
            c = EVOL_COLORS[i % len(EVOL_COLORS)]
            pt.marker = s.marker.__class__()
            pt.marker.symbol = 'circle'
            pt.marker.size = 10
            pt.marker.graphicalProperties.solidFill = c
            pt.marker.graphicalProperties.ln.solidFill = BLACK
            s.dPt.append(pt)

        ws1.add_chart(ch_l, f'A{chart_anchor_row}')

    # Anchos generosos — elimina todos los ####
    for c, w in enumerate([18, 16, 18, 16, 18, 16, 1, 18, 18, 18, 18, 18], 1):
        col_w(ws1, c, w)

    # ════════════════════════════════════════════════════════
    # HOJA 2 — HISTORIAL DE NOCHES
    # ════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Historial')
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = DARK
    ws2.freeze_panes = 'A4'

    title_row(ws2, 1, 'HISTORIAL DE NOCHES',
        f'{len(historial)} noche(s)  ·  Acumulado: ${total_all:,.0f}'.replace(',','.'))

    cols2 = ['FECHA', 'HORA CIERRE', 'TOTAL ($)', 'OPS', 'PROM/OP ($)', '1° LUGAR', '2° LUGAR', '3° LUGAR', 'CAJA ABAJO ($)', 'CAJA EXT ($)', 'CAJA VIP ($)']
    header_row(ws2, 3, cols2, 1, DARK)
    # fix: re-apply white font since DARK bg
    for ci in range(1, len(cols2)+1):
        ws2.cell(row=3, column=ci).font = Font(bold=True, color=WHITE, size=9, name='Calibri')

    for idx, n in enumerate(reversed(historial)):
        r = 4 + idx; alt = (idx%2==0)
        pc = n.get('por_caja', {})
        rk = sorted(n.get('ranking', []), key=lambda x: -x['total'])
        def top(pos): return rk[pos]['name'] if len(rk) > pos else '—'
        vals = [
            n['fecha'], n.get('hora_cierre','—'), n['total'], n['operaciones'],
            n['total']/n['operaciones'] if n['operaciones'] else 0,
            top(0), top(1), top(2),
            pc.get(1, pc.get('1',0)), pc.get(2, pc.get('2',0)), pc.get(3, pc.get('3',0))
        ]
        fmts = [None, None, '#.##0', 'General', '#.##0', None, None, None, '#.##0', '#.##0', '#.##0']
        cols_color = [None, None, GREEN, BLUE, GREEN, medal_color.get(1), medal_color.get(2), medal_color.get(3), GREEN, GREEN, GREEN]
        for ci, (v, fmt, cc) in enumerate(zip(vals, fmts, cols_color), 1):
            bg = LGRAY if alt else WHITE
            c2 = ws2.cell(row=r, column=ci, value=v)
            c2.font = Font(bold=ci in [1,3], color=cc or BLACK, size=9, name='Calibri')
            c2.fill = fill(bg); c2.border = BDR
            c2.alignment = align('right' if fmt=='#.##0' else 'center')
            if fmt: c2.number_format = fmt
            row_h(ws2, r, 17)

    for c, w in enumerate([13, 12, 16, 7, 16, 20, 20, 20, 16, 16, 16], 1):
        col_w(ws2, c, w)

    # ════════════════════════════════════════════════════════
    # HOJAS POR NOCHE — una hoja por cada noche del mes actual
    # ════════════════════════════════════════════════════════
    mes_actual = datetime.now().strftime('%Y-%m')
    noches_mes = [n for n in historial if n['fecha'].startswith(mes_actual)]
    if not noches_mes:
        noches_mes = historial  # fallback: todas las noches

    def make_noche_sheet(n, sheet_name):
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = GOLD
        ws.freeze_panes = 'A5'

        # Datos de la noche
        pc = n.get('por_caja', {})
        cj_n = [pc.get(i, pc.get(str(i), 0)) for i in [1,2,3]]
        t_n  = n['total']; ops_n = n['operaciones']
        if not any(cj_n):
            for t in n.get('transactions', []):
                k = int(t.get('caja', 1)) - 1
                if 0 <= k <= 2: cj_n[k] += t['amount']
        ops_cj_n = [0, 0, 0]
        for t in n.get('transactions', []):
            k = int(t.get('caja', 1)) - 1
            if 0 <= k <= 2: ops_cj_n[k] += 1

        title_row(ws, 1, f'DETALLE — {n["fecha"]}',
            f'Cierre: {n.get("hora_cierre","—")}  ·  Total: ${t_n:,.0f}  ·  {ops_n} ops  ·  Prom/op: ${t_n/ops_n:,.0f}'.replace(',','.'))
        ws.row_dimensions[3].height = 6

        # KPIs — etiquetas completas sin truncar
        kpi_block(ws, 4, 1, 'TOTAL NOCHE',     t_n,   '#.##0')
        kpi_block(ws, 4, 3, 'OPERACIONES',      ops_n, 'General', BLUE)
        kpi_block(ws, 4, 5, 'PROMEDIO / OP',    t_n/ops_n if ops_n else 0, '#.##0')
        kpi_block(ws, 5, 1, 'CAJA ABAJO',       cj_n[0], '#.##0')
        kpi_block(ws, 5, 3, 'CAJA EXTENDIDO',   cj_n[1], '#.##0')
        kpi_block(ws, 5, 5, 'CAJA VIP',         cj_n[2], '#.##0')
        ws.row_dimensions[6].height = 8

        # ── Ranking de la noche ──────────────────────────────
        section(ws, 7, '  RANKING DE LA NOCHE', 1, 5)
        header_row(ws, 8, ['#', 'CLIENTE', 'MESA', 'TOTAL ($)', '% DEL TOTAL'], 1, GOLD)
        rk_n = n.get('ranking', [])
        for i, rv in enumerate(rk_n, 1):
            alt = (i%2==0); mc = medal_color.get(i)
            data_cell(ws, 8+i, 1, f'#{i}',           bold=i<=3, alt=alt, h='center', color=mc)
            data_cell(ws, 8+i, 2, rv['name'],         bold=i<=3, alt=alt, h='left',   color=mc)
            data_cell(ws, 8+i, 3, rv.get('mesa','—'), alt=alt, h='center')
            money(ws, 8+i, 4, rv['total'], alt=alt, bold=i<=3, color=mc)
            pct(ws, 8+i, 5, rv['total']/t_n if t_n else 0, alt=alt)

        # ── Distribución por caja — con pie chart bien posicionado ──
        dist_start = 10 + len(rk_n)
        ws.row_dimensions[dist_start].height = 8
        section(ws, dist_start+1, '  DISTRIBUCIÓN POR CAJA', 1, 5)
        header_row(ws, dist_start+2, ['CAJA', 'TOTAL ($)', '% DEL TOTAL', 'OPS', 'PROM/OP ($)'], 1, GOLD)
        cs = dist_start + 3
        caja_names_l  = ['Caja Abajo', 'Caja Extendido', 'Caja VIP']
        caja_colors_l = [BLUE, GREEN, GOLD]
        for i in range(3):
            alt = (i%2==0)
            data_cell(ws, cs+i, 1, caja_names_l[i], bold=True, alt=alt, h='left', color=caja_colors_l[i])
            money(ws, cs+i, 2, cj_n[i], alt=alt, color=caja_colors_l[i])
            pct(ws, cs+i, 3, cj_n[i]/t_n if t_n else 0, alt=alt)
            data_cell(ws, cs+i, 4, ops_cj_n[i], alt=alt, h='center')
            money(ws, cs+i, 5, cj_n[i]/ops_cj_n[i] if ops_cj_n[i] else 0, alt=alt)

        # ── Todas las operaciones ────────────────────────────
        op_start = cs + 5
        ws.row_dimensions[op_start].height = 8
        section(ws, op_start+1, '  TODAS LAS OPERACIONES', 1, 6)
        header_row(ws, op_start+2, ['HORA', 'CLIENTE', 'MESA', 'MONTO ($)', 'CAJA', 'PRODUCTOS'], 1, DARK)
        for ci in range(1, 7):
            ws.cell(row=op_start+2, column=ci).font = Font(bold=True, color=WHITE, size=9, name='Calibri')
        txs_n = n.get('transactions', [])
        for i, t in enumerate(txs_n):
            alt = (i%2==0)
            data_cell(ws, op_start+3+i, 1, t.get('time','—'), alt=alt)
            data_cell(ws, op_start+3+i, 2, t['name'], bold=True, alt=alt, h='left', color=BLUE)
            data_cell(ws, op_start+3+i, 3, t.get('mesa','—'), alt=alt)
            money(ws, op_start+3+i, 4, t['amount'], alt=alt)
            data_cell(ws, op_start+3+i, 5, CAJA_NOMBRES.get(int(t.get('caja',1)), '—'), alt=alt)
            istr = ' · '.join(f"{it.get('cantidad',it.get('qty',1))}x {it['nombre']}" for it in t.get('items',[])) or '—'
            cc2 = ws.cell(row=op_start+3+i, column=6, value=istr)
            cc2.font = Font(color=DGRAY, size=9, name='Calibri')
            cc2.fill = fill(LGRAY if alt else WHITE); cc2.border = BDR
            cc2.alignment = align('left', indent=1); row_h(ws, op_start+3+i, 17)

        # Pie chart — debajo de TODAS las tablas y operaciones
        pie_anchor = op_start + 3 + len(txs_n) + 3
        if any(cj_n):
            ch_pie = PieChart(); ch_pie.title = 'Distribución por caja'; ch_pie.style = 10
            ch_pie.width = 18; ch_pie.height = 12
            ch_pie.add_data(Reference(ws, min_col=2, min_row=dist_start+2, max_row=cs+2), titles_from_data=True)
            ch_pie.set_categories(Reference(ws, min_col=1, min_row=cs, max_row=cs+2))
            for idx, color in enumerate([BLUE, GREEN, GOLD]):
                pt = DataPoint(idx=idx)
                pt.graphicalProperties.solidFill = color
                ch_pie.series[0].dPt.append(pt)
            ws.add_chart(ch_pie, f'A{pie_anchor}')

        # Anchos — generosos para evitar palabras cortadas y ####
        for c, w in enumerate([7, 24, 8, 16, 14, 44, 1, 18, 18, 18, 18, 18], 1):
            col_w(ws, c, w)
        # Columnas KPI (1-6): label + value pairs
        for c, w in enumerate([18, 16, 18, 16, 18, 16], 1):
            col_w(ws, c, w)

    # Generar una hoja por cada noche del mes (más reciente primero)
    for n in reversed(noches_mes):
        try:
            from datetime import datetime as _dt
            d = _dt.strptime(n['fecha'], '%Y-%m-%d')
            sname = d.strftime('%d-%m')
        except Exception:
            sname = n['fecha'][-5:].replace('-','/')
        make_noche_sheet(n, sname)

    # ════════════════════════════════════════════════════════
    # HOJA — CLIENTES (siempre al final)
    # ════════════════════════════════════════════════════════
    ws4 = wb.create_sheet('Clientes')
    ws4.sheet_view.showGridLines = False
    ws4.sheet_properties.tabColor = BLUE
    ws4.freeze_panes = 'A4'

    title_row(ws4, 1, 'RANKING DE CLIENTES — HISTÓRICO',
        f'{len(rk_sorted)} clientes únicos  ·  Total facturado: ${total_all:,.0f}'.replace(',','.'))

    ws4.row_dimensions[3].height = 6
    header_row(ws4, 3, ['#', 'CLIENTE', 'TOTAL HISTÓRICO ($)', '% DEL TOTAL', 'NOCHES', 'PROM/NOCHE ($)'], 1, DARK)
    for ci in range(1, 7):
        ws4.cell(row=3, column=ci).font = Font(bold=True, color=WHITE, size=9, name='Calibri')
    for i, (nm, tot) in enumerate(rk_sorted, 1):
        alt = (i%2==0); mc = medal_color.get(i); nn = noches_cliente.get(nm, 1)
        data_cell(ws4, 3+i, 1, f'#{i}', bold=i<=3, alt=alt, h='center', color=mc)
        data_cell(ws4, 3+i, 2, nm,      bold=i<=3, alt=alt, h='left',   color=mc)
        money(ws4, 3+i, 3, tot, alt=alt, bold=i<=3, color=mc)
        pct(ws4, 3+i, 4, tot/total_all if total_all else 0, alt=alt)
        data_cell(ws4, 3+i, 5, nn, alt=alt, h='center')
        money(ws4, 3+i, 6, tot/nn, alt=alt)
    for c, w in enumerate([6, 24, 18, 12, 8, 14], 1): col_w(ws4, c, w)

    # ── Generar y enviar ──────────────────────────────────────
    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    from flask import send_file
    fecha_str = datetime.now().strftime('%Y%m%d_%H%M')
    return send_file(buf, as_attachment=True,
                     download_name=f'jagger_vip_{fecha_str}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
@app.route('/api/tarjetas/recargar', methods=['POST'])
def recargar_tarjeta():
    body = request.get_json() or {}
    codigo = str(body.get('codigo', ''))
    try:
        monto = float(body.get('monto', 0))
    except (ValueError, TypeError):
        return jsonify({'ok': False, 'error': 'Monto inválido'}), 400
    if not codigo or monto <= 0:
        return jsonify({'ok': False, 'error': 'Datos inválidos'}), 400
    with lock:
        data = load_data()
        if codigo not in data.get('tarjetas', {}):
            return jsonify({'ok': False, 'error': 'Tarjeta no encontrada'}), 404
        import time as _time
        recarga_id = str(int(_time.time()*1000))
        data['tarjetas'][codigo]['saldo_actual'] += monto
        nuevo_saldo = data['tarjetas'][codigo]['saldo_actual']
        if 'recargas' not in data['tarjetas'][codigo]:
            data['tarjetas'][codigo]['recargas'] = []
        from datetime import datetime as _dt
        data['tarjetas'][codigo]['recargas'].append({
            'id': recarga_id, 'monto': monto,
            'hora': _dt.now().strftime('%H:%M'), 'fecha': _dt.now().strftime('%d/%m')
        })
        save_data(data)
    return jsonify({'ok': True, 'nuevo_saldo': nuevo_saldo, 'recarga_id': recarga_id})

@app.route('/api/tarjetas/recarga/<codigo>/<recarga_id>', methods=['DELETE'])
def eliminar_recarga(codigo, recarga_id):
    with lock:
        data = load_data()
        t = data.get('tarjetas', {}).get(codigo)
        if not t:
            return jsonify({'ok': False, 'error': 'Tarjeta no encontrada'}), 404
        recargas = t.get('recargas', [])
        rec = next((r for r in recargas if r['id'] == recarga_id), None)
        if not rec:
            return jsonify({'ok': False, 'error': 'Recarga no encontrada'}), 404
        t['saldo_actual'] = max(0, t['saldo_actual'] - rec['monto'])
        t['recargas'] = [r for r in recargas if r['id'] != recarga_id]
        nuevo_saldo = t['saldo_actual']
        save_data(data)
    return jsonify({'ok': True, 'nuevo_saldo': nuevo_saldo})

@app.route('/api/tarjetas/recarga/<codigo>/<recarga_id>', methods=['PUT'])
def editar_recarga(codigo, recarga_id):
    body = request.get_json() or {}
    try:
        nuevo_monto = float(body.get('monto', 0))
    except (ValueError, TypeError):
        return jsonify({'ok': False, 'error': 'Monto inválido'}), 400
    if nuevo_monto <= 0:
        return jsonify({'ok': False, 'error': 'Monto inválido'}), 400
    with lock:
        data = load_data()
        t = data.get('tarjetas', {}).get(codigo)
        if not t:
            return jsonify({'ok': False, 'error': 'Tarjeta no encontrada'}), 404
        recargas = t.get('recargas', [])
        rec = next((r for r in recargas if r['id'] == recarga_id), None)
        if not rec:
            return jsonify({'ok': False, 'error': 'Recarga no encontrada'}), 404
        diff = nuevo_monto - rec['monto']
        t['saldo_actual'] = max(0, t['saldo_actual'] + diff)
        rec['monto'] = nuevo_monto
        nuevo_saldo = t['saldo_actual']
        save_data(data)
    return jsonify({'ok': True, 'nuevo_saldo': nuevo_saldo})

@app.route('/api/auth', methods=['POST'])
def auth():
    body = request.get_json() or {}
    cfg = load_config()
    if str(body.get('password', '')) == str(cfg.get('password', '1212')):
        return jsonify({'ok': True})
    return jsonify({'ok': False, 'error': 'Contraseña incorrecta'}), 401

@app.route('/api/auth/change', methods=['POST'])
def auth_change():
    body = request.get_json() or {}
    cfg = load_config()
    if str(body.get('current', '')) != str(cfg.get('password', '1212')):
        return jsonify({'ok': False, 'error': 'PIN actual incorrecto'}), 401
    nueva = str(body.get('new', '')).strip()
    if not nueva.isdigit() or len(nueva) != 4:
        return jsonify({'ok': False, 'error': 'El PIN debe tener exactamente 4 dígitos numéricos'}), 400
    cfg['password'] = nueva
    save_config(cfg)
    return jsonify({'ok': True})

# ── Login / Logout ──────────────────────────────────────────────────────────

@app.route('/api/login', methods=['POST'])
def api_login():
    body = request.get_json() or {}
    role = str(body.get('role', ''))
    pin = str(body.get('pin', ''))
    cfg = load_config()
    valid_roles = ['manager', 'cajaabajo', 'cajaextendido', 'cajavip', 'tarjetas']
    if role not in valid_roles:
        return jsonify({'ok': False, 'error': 'Rol inválido'}), 400
    cfg_key = 'pin_' + role
    if pin == str(cfg.get(cfg_key, '1234')):
        cfg2 = load_config()
        session.permanent = True
        session[role + '_auth'] = True
        session['session_version'] = cfg2.get('session_version', 0)
        if role == 'manager':
            session['cajaabajo_auth'] = True
            session['cajaextendido_auth'] = True
            session['cajavip_auth'] = True
        return jsonify({'ok': True})
    return jsonify({'ok': False, 'error': 'PIN incorrecto'}), 401

@app.route('/api/session', methods=['GET'])
def get_session_info():
    return jsonify({
        'manager': bool(session.get('manager_auth')),
        'cajaabajo': bool(session.get('cajaabajo_auth')),
        'cajaextendido': bool(session.get('cajaextendido_auth')),
        'cajavip': bool(session.get('cajavip_auth')),
    })

@app.route('/api/logout', methods=['POST'])
def api_logout():
    body = request.get_json() or {}
    role = body.get('role')
    if role:
        session.pop(role + '_auth', None)
    else:
        session.clear()
    return jsonify({'ok': True})

# ── Menu CRUD ───────────────────────────────────────────────────────────────

DEFAULT_MENU = [
    # CHAMPAGNE
    {"categoria":"Champagne","nombre":"Norton Cosecha Tardia","precio":50000},
    {"categoria":"Champagne","nombre":"Mumm","precio":60000},
    {"categoria":"Champagne","nombre":"Chandon","precio":80000},
    {"categoria":"Champagne","nombre":"Baron B","precio":100000},
    # BOTTLE SERVICE
    {"categoria":"Bottle Service","nombre":"Sernova","precio":85000},
    {"categoria":"Bottle Service","nombre":"Absolut","precio":105000},
    {"categoria":"Bottle Service","nombre":"Red Label","precio":120000},
    {"categoria":"Bottle Service","nombre":"Red Label Litro","precio":140000},
    {"categoria":"Bottle Service","nombre":"Black Label","precio":170000},
    {"categoria":"Bottle Service","nombre":"Jagermeister","precio":120000},
    {"categoria":"Bottle Service","nombre":"Beefeater","precio":120000},
    {"categoria":"Bottle Service","nombre":"Gin Blu","precio":100000},
    {"categoria":"Bottle Service","nombre":"Malibu","precio":100000},
    {"categoria":"Bottle Service","nombre":"Fernet Branca","precio":90000},
    {"categoria":"Bottle Service","nombre":"Ramazotti","precio":90000},
    # IMPORTADOS
    {"categoria":"Importados","nombre":"Belvedere Luminous","precio":180000},
    {"categoria":"Importados","nombre":"Belvedere Luminous X2","precio":340000},
    {"categoria":"Importados","nombre":"Cliquot","precio":220000},
    {"categoria":"Importados","nombre":"Moet Nectar","precio":240000},
    {"categoria":"Importados","nombre":"Moet Ice","precio":250000},
    {"categoria":"Importados","nombre":"Nuvo","precio":180000},
    {"categoria":"Importados","nombre":"Nuvo X2","precio":340000},
    # BEBIDAS
    {"categoria":"Bebidas","nombre":"Speed","precio":7000},
    {"categoria":"Bebidas","nombre":"Speed X6","precio":35000},
    {"categoria":"Bebidas","nombre":"Jugo","precio":7000},
    {"categoria":"Bebidas","nombre":"Lata Coca","precio":6500},
    {"categoria":"Bebidas","nombre":"Lata Ton","precio":7000},
    {"categoria":"Bebidas","nombre":"Agua","precio":5500},
    {"categoria":"Bebidas","nombre":"Budweiser","precio":7000},
    {"categoria":"Bebidas","nombre":"Budweiser X6","precio":35000},
    # TRAGOS
    {"categoria":"Tragos","nombre":"Sernova","precio":8000},
    {"categoria":"Tragos","nombre":"Absolut","precio":9500},
    {"categoria":"Tragos","nombre":"Hodlmoser","precio":9000},
    {"categoria":"Tragos","nombre":"Fernet","precio":8000},
    {"categoria":"Tragos","nombre":"Gin","precio":9500},
    {"categoria":"Tragos","nombre":"Campari","precio":8000},
    {"categoria":"Tragos","nombre":"Malibu","precio":8500},
    # SHOTS
    {"categoria":"Shots","nombre":"Hodlmoser","precio":7000},
    {"categoria":"Shots","nombre":"Absolut","precio":7000},
]

def ensure_menu():
    """Populate default menu into _db if it's empty."""
    with lock:
        data = load_data()
        if not data.get('menu'):
            for i, item in enumerate(DEFAULT_MENU, 1):
                data['menu'].append({'id': i, 'categoria': item['categoria'], 'nombre': item['nombre'], 'precio': float(item['precio'])})
            data['menu_id_counter'] = len(DEFAULT_MENU)
            save_data(data)

@app.route('/api/menu/reset', methods=['POST'])
def reset_menu():
    with lock:
        data = load_data()
        data['menu'] = []
        for i, item in enumerate(DEFAULT_MENU, 1):
            data['menu'].append({'id':i,'categoria':item['categoria'],'nombre':item['nombre'],'precio':float(item['precio'])})
        data['menu_id_counter'] = len(DEFAULT_MENU)
        save_data(data)
    return jsonify({'ok': True, 'count': len(DEFAULT_MENU)})

@app.route('/api/menu', methods=['GET'])
def get_menu():
    ensure_menu()
    with lock:
        data = load_data()
    return jsonify(data.get('menu', []))

@app.route('/api/menu', methods=['POST'])
def add_menu_item():
    body = request.get_json() or {}
    if not body.get('nombre') or body.get('precio') is None:
        return jsonify({'ok': False, 'error': 'Faltan campos'}), 400
    with lock:
        data = load_data()
        if 'menu' not in data: data['menu'] = []
        counter = data.get('menu_id_counter', len(data['menu']))
        counter += 1
        item = {
            'id': counter,
            'categoria': str(body.get('categoria', 'Otros')),
            'nombre': str(body['nombre']),
            'precio': float(body['precio']),
        }
        data['menu'].append(item)
        data['menu_id_counter'] = counter
        save_data(data)
    return jsonify({'ok': True, 'item': item})

@app.route('/api/menu/<int:mid>', methods=['PUT'])
def update_menu_item(mid):
    body = request.get_json() or {}
    with lock:
        data = load_data()
        item = next((x for x in data.get('menu', []) if x['id'] == mid), None)
        if not item:
            return jsonify({'ok': False, 'error': 'No encontrado'}), 404
        if 'nombre' in body: item['nombre'] = str(body['nombre'])
        if 'precio' in body: item['precio'] = float(body['precio'])
        if 'categoria' in body: item['categoria'] = str(body['categoria'])
        save_data(data)
    return jsonify({'ok': True})

@app.route('/api/menu/<int:mid>', methods=['DELETE'])
def delete_menu_item(mid):
    with lock:
        data = load_data()
        data['menu'] = [x for x in data.get('menu', []) if x['id'] != mid]
        save_data(data)
    return jsonify({'ok': True})

# ── Publicidad ──────────────────────────────────────────────────────────────

import werkzeug.utils as _wu

@app.route('/api/publicidad/estado', methods=['GET'])
def pub_estado():
    import time as _t
    return jsonify({
        'mostrar_ts': _state.get('publicidad_mostrar_ts', 0),
        'activa': _state['publicidad_activa'],
        'url': _state['publicidad_url'],
        'frecuencia': _state['publicidad_frecuencia'],
        'mostrar_ts': _state.get('publicidad_mostrar_ts', 0),
        'server_time': _t.time(),
    })

@app.route('/api/publicidad/activar', methods=['POST'])
def pub_activar():
    body = request.get_json() or {}
    with lock:
        _state['publicidad_activa'] = True
        _state['publicidad_url'] = str(body.get('url', _state['publicidad_url']))
        _state['publicidad_frecuencia'] = int(body.get('frecuencia', _state['publicidad_frecuencia']))
        save_state()
    return jsonify({'ok': True})

@app.route('/api/publicidad/desactivar', methods=['POST'])
def pub_desactivar():
    with lock:
        _state['publicidad_activa'] = False
        save_state()
    return jsonify({'ok': True})

@app.route('/api/publicidad/mostrar-ahora', methods=['POST'])
def pub_mostrar_ahora():
    import time as _t
    with lock:
        _state['publicidad_mostrar_ts'] = _t.time()
    return jsonify({'ok': True})

@app.route('/api/publicidad/upload', methods=['POST'])
def pub_upload():
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'No se envió archivo'}), 400
    f = request.files['file']
    if not f.filename.lower().endswith('.mp4'):
        return jsonify({'ok': False, 'error': 'Solo se aceptan archivos mp4'}), 400
    os.makedirs(VIDEOS_DIR, exist_ok=True)
    fname = _wu.secure_filename(f.filename)
    fpath = os.path.join(VIDEOS_DIR, fname)
    f.save(fpath)
    url = '/api/publicidad/video/' + fname
    with lock:
        _state['publicidad_url'] = url
        save_state()
    return jsonify({'ok': True, 'url': url})

@app.route('/api/publicidad/video/<path:filename>')
def serve_video(filename):
    from flask import send_from_directory
    return send_from_directory(VIDEOS_DIR, filename)

# ── Config / PINs ───────────────────────────────────────────────────────────

@app.route('/api/config/pines', methods=['POST'])
def update_pines():
    body = request.get_json() or {}
    valid_keys = {'pin_manager', 'pin_cajaabajo', 'pin_cajaextendido', 'pin_cajavip', 'pin_tarjetas'}
    cfg = load_config()
    changed = False
    for k, v in body.items():
        if k in valid_keys:
            v = str(v).strip()
            if not v.isdigit() or len(v) != 4:
                return jsonify({'ok': False, 'error': f'PIN inválido para {k}'}), 400
            cfg[k] = v
            changed = True
    if changed:
        cfg['session_version'] = cfg.get('session_version', 0) + 1
    save_config(cfg)
    return jsonify({'ok': True})

@app.route('/api/config/pines', methods=['GET'])
def get_pines():
    cfg = load_config()
    return jsonify({k: cfg.get(k, '') for k in ['pin_manager','pin_cajaabajo','pin_cajaextendido','pin_cajavip']})

# ── Design state ─────────────────────────────────────────────────────────────

@app.route('/api/design', methods=['GET'])
def get_design():
    return jsonify(_state.get('design', {}))

@app.route('/api/design', methods=['POST'])
def set_design():
    body = request.get_json() or {}
    with lock:
        if 'design' not in _state:
            _state['design'] = {}
        allowed = {'tema','colores','logo','vip','tagline','tagline_color','tagline_glow',
                   'tagline_font','winner_msg','winner_sub','premio','premio_size',
                   'hora_fin','clock_size','deco_activa','petals_activos',
                   'falling_gloves','pink_modo','efecto','tipo_particula','mostrar12'}
        updated = False
        for k, v in body.items():
            if k in allowed:
                # No sobreescribir colores con diccionario vacío
                if k == 'colores' and isinstance(v, dict) and len(v) == 0:
                    continue
                _state['design'][k] = v
                updated = True
        if updated:
            import time as _t
            _state['design']['_ts'] = _t.time()
            save_state()
    return jsonify({'ok': True})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port, debug=False)
